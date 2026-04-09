import json, sys, re, webbrowser, threading, time, logging
import html as html_module
import xml.etree.ElementTree as ET
from pathlib import Path
import firebase_admin
from firebase_admin import credentials, firestore

def init_firebase():
    import glob
    import os
    cred_files = glob.glob(os.path.join(os.path.dirname(__file__), '..', '..', 'backend', 'config', 'log-solution-*-firebase-adminsdk-*.json'))
    if not firebase_admin._apps and cred_files:
        cred = credentials.Certificate(cred_files[0])
        firebase_admin.initialize_app(cred)
    elif not cred_files:
        return None
    return firestore.client()

# Flask e Flask-CORS verranno importati solo se serve avviare il server
try:
    from flask import Flask, render_template_string, request, jsonify
    from flask_cors import CORS
    HAS_FLASK = True
except ImportError:
    HAS_FLASK = False
    Flask = lambda *args, **kwargs: None
    render_template_string = request = jsonify = None

# --- CONFIGURAZIONE ---
PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"

# INSERISCI QUI LA TUA API KEY DI GOOGLE MAPS
GOOGLE_MAPS_API_KEY = "AIzaSyAHQ3HjuEEIS8bn5KMh6N3UoM6kZ2MYGL4"

if HAS_FLASK:
    app = Flask(__name__)
    CORS(app, resources={r"/*": {"origins": "*"}})
else:
    app = None

# Configurazione Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Variabili globali per tracciare i file in uso
TARGET_FILE_UNIFICATO = None
TARGET_FILE_VIAGGI = None
ZONE_LIST_CACHE = []
DATA_GIORNO = ""

def _get_color(idx):
    palette = ["#4f46e5", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6", "#ec4899", "#06b6d4", "#f97316", "#14b8a6", "#6366f1", "#a855f7", "#3b82f6", "#22c55e", "#d946ef", "#84cc16"]
    return palette[idx % len(palette)]

excel_lock = threading.Lock()

def _aggiorna_entrambi_excel(nome, lat, lon, indirizzo=None):
    """Sincronizzazione Atomica su Firebase e Excel Giornaliero."""
    import pandas as pd
    try:
        output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
        target_excel = output_base / "punti_consegna.xlsx"
        
        # 1. Aggiorna Firebase
        db = init_firebase()
        if db:
            docs = db.collection("customers").document("DNR").collection("clienti").where("nome", "==", str(nome).strip()).stream()
            for doc in docs:
                doc.reference.update({"lat": lat, "lon": lon, "lng": lon})
                logger.info(f"Aggiornato Firebase (DNR): {nome}")
        
        # 2. Aggiorna Excel Giornaliero (se esiste)
        success = True
        if target_excel.exists():
            with excel_lock:
                try:
                    df = pd.read_excel(target_excel)
                    col_nome = next((c for c in df.columns if str(c).strip().lower() in ["a chi va consegnato", "cliente", "nome", "destinatario"]), None)
                    col_ind = next((c for c in df.columns if str(c).strip().lower() in ["indirizzo", "via", "via/piazza", "sede"]), None)
                    if col_nome:
                        mask = df[col_nome].astype(str).str.strip().str.lower() == str(nome).strip().lower()
                        if indirizzo and col_ind and mask.sum() > 1:
                            mask_ind = df[col_ind].astype(str).str.strip().str.lower() == str(indirizzo).strip().lower()
                            if (mask & mask_ind).any():
                                mask = mask & mask_ind
                        if mask.any():
                            df.loc[mask, 'Latitudine'] = lat
                            df.loc[mask, 'Longitudine'] = lon
                            df.to_excel(target_excel, index=False)
                except Exception as e:
                    logger.exception(f"Errore scrittura {target_excel.name}")
                    success = False
        return success
    except Exception as e:
        logger.exception("Errore globale Firebase/Excel update")
        return False

if HAS_FLASK:
    @app.after_request
    def add_header(r):
        r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        r.headers["Pragma"] = "no-cache"
        r.headers["Expires"] = "0"
        return r

    @app.route('/')
    def index():
        is_locked_val = True
        if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
            try:
                unif = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                is_locked_val = unif.get("is_locked", True)
            except: pass
        return render_template_string(HTML_TEMPLATE, DATA_GIORNO=DATA_GIORNO, JSON_ZONE=json.dumps(ZONE_LIST_CACHE, ensure_ascii=False), GOOGLE_MAPS_API_KEY=GOOGLE_MAPS_API_KEY, IS_LOCKED_JS="true" if is_locked_val else "false")

    @app.route('/save', methods=['POST'])
    def save():
        """Salvataggio globale: Zone + Coordinate su tutti i file + aggiornamento rientri DDT."""
        global ZONE_LIST_CACHE
        try:
            data = request.json
            ZONE_LIST_CACHE = data
            output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"

            # 1. Aggiorna Excel Giornaliero e Firebase (Bulk)
            import pandas as pd
            db = init_firebase()
            ex_p = output_base / "punti_consegna.xlsx"
            if ex_p.exists():
                with excel_lock:
                    df = pd.read_excel(ex_p)
                    col_nome = next((c for c in df.columns if str(c).strip().lower() in ["a chi va consegnato", "cliente", "nome", "destinatario"]), None)
                    if col_nome:
                        for z in data:
                            for p in z.get("lista_punti", []):
                                mask = df[col_nome].astype(str).str.strip().str.lower() == str(p['nome']).strip().lower()
                                if mask.any():
                                    df.loc[mask, 'Latitudine'] = p['lat']
                                    df.loc[mask, 'Longitudine'] = p['lon']
                                    
                                if db:
                                    # Bulk Firebase save per zone assignments / repins
                                    docs = db.collection("customers").document("DNR").collection("clienti").where("nome", "==", str(p['nome']).strip()).stream()
                                    for doc in docs: doc.reference.update({"lat": p['lat'], "lon": p['lon'], "lng": p['lon']})
                    df.to_excel(ex_p, index=False)

            # 2. Aggiorna JSON Unificato
            if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
                unif = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                p_map = { (p.get("codice_univoco") or f"{p.get('codice_frutta')}_{p.get('codice_latte')}"): (z["id_zona"], p['lat'], p['lon']) for z in data for p in z["lista_punti"] }
                for p in unif.get("punti", []):
                    pid = p.get("codice_univoco") or f"{p.get('codice_frutta')}_{p.get('codice_latte')}"
                    if pid in p_map:
                        p["zona"], lat, lon = p_map[pid]
                        if lat: p["lat"] = lat
                        if lon: p["lon"] = lon
                TARGET_FILE_UNIFICATO.write_text(json.dumps(unif, indent=2, ensure_ascii=False), encoding="utf-8")

            # 3. Aggiorna JSON Viaggi
            data_f = [z for z in data if len(z.get("lista_punti", [])) > 0]
            if TARGET_FILE_VIAGGI:
                TARGET_FILE_VIAGGI.write_text(json.dumps(data_f, indent=2, ensure_ascii=False), encoding="utf-8")

            # 4. Aggiorna rientri_ddt.xlsx in base alla zona finale dei punti speciali
            _aggiorna_rientri_dopo_salvataggio(data, DATA_GIORNO)

            return jsonify({"status": "ok"})
        except Exception as e:
            logger.exception("Errore save")
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route('/save_coord', methods=['POST'])
    def save_coord():
        """Salvataggio atomico al movimento di un singolo punto."""
        try:
            p = request.json  # {nome, indirizzo, lat, lon}
            nome = p.get('nome', '')
            indirizzo = p.get('indirizzo', '')
            lat = p['lat']
            lon = p['lon']

            res = _aggiorna_entrambi_excel(nome, lat, lon, indirizzo=indirizzo)
            if res is True:
                global ZONE_LIST_CACHE
                for z in ZONE_LIST_CACHE:
                    for pt in z.get("lista_punti", []):
                        if pt.get("nome", "").lower() == nome.lower() and \
                           (not indirizzo or pt.get("indirizzo", "").lower() == indirizzo.lower()):
                            pt['lat'], pt['lon'] = lat, lon

                # Aggiorna JSON Unificato e Viaggi al volo
                for f_path, is_unif in [(TARGET_FILE_UNIFICATO, True), (TARGET_FILE_VIAGGI, False)]:
                    if f_path and f_path.exists():
                        d = json.loads(f_path.read_text(encoding='utf-8'))
                        lista = d.get('punti', []) if is_unif else [x for zz in d for x in zz.get('lista_punti', [])]
                        for pt in lista:
                            if pt.get('nome', '').lower() == nome.lower() and \
                               (not indirizzo or pt.get('indirizzo', '').lower() == indirizzo.lower()):
                                pt['lat'], pt['lon'] = lat, lon
                        f_path.write_text(json.dumps(d, indent=2, ensure_ascii=False), encoding='utf-8')
                return jsonify({"status": "ok", "msg": "Sincronizzazione completata!"})
            return jsonify({"status": "error", "msg": str(res)}), 500
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500

    @app.route('/toggle_lock', methods=['POST'])
    def toggle_lock():
        try:
            req = request.json
            action = req.get('action')
            if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
                unificato = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                is_locked = (action == 'lock')
                unificato['is_locked'] = is_locked
                TARGET_FILE_UNIFICATO.write_text(json.dumps(unificato, indent=2, ensure_ascii=False), encoding="utf-8")
                return jsonify({"status": "ok", "is_locked": is_locked})
            return jsonify({"status": "error", "message": "File non trovato"}), 404
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

def _aggiorna_rientri_dopo_salvataggio(zone_data: list, data_giorno: str):
    """Aggiorna colonna C di rientri_ddt.xlsx in base alla zona finale di OGNI punto."""
    rientri_path = BASE_DIR / "rientri_ddt.xlsx"
    if not rientri_path.exists(): return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(rientri_path)
        ws = wb.active
        
        # Mappatura Codice -> Stato finale desiderato
        status_map = {} # {codice_ddt: "allegato..." o ""}
        
        for z in zone_data:
            zid = z.get("id_zona", "")
            for p in z.get("lista_punti", []):
                # Raccogliamo tutti i possibili codici DDT del punto (da liste o campi singoli)
                codici = []
                codici.extend(p.get("codici_ddt_frutta") or [])
                codici.extend(p.get("codici_ddt_latte") or [])
                if not codici:
                    # Fallback se le liste sono vuote
                    c_f, c_l = p.get("codice_frutta"), p.get("codice_latte")
                    if c_f and c_f != "p00000": codici.append(c_f)
                    if c_l and c_l != "p00000": codici.append(c_l)
                
                # Determiniamo lo stato in base a dove si trova il punto ora
                final_status = "" if zid == "DDT_DA_INSERIRE" else f"allegato DDT {data_giorno}"
                
                for c in codici:
                    status_map[str(c).strip().lower()] = final_status

        if not status_map: return
        
        modifiche = 0
        for row in ws.iter_rows(min_row=2):
            cod_excel = str(row[0].value or '').strip().lower()
            if cod_excel in status_map:
                row[2].value = status_map[cod_excel]
                modifiche += 1
        
        if modifiche:
            wb.save(rientri_path)
            logger.info(f"📂 Excel Rientri aggiornato: {modifiche} righe modificate.")
    except Exception as e:
        logger.exception("Errore aggiornamento rientri_ddt.xlsx")

def _libera_porta_5000():
    import subprocess as _sp
    try:
        res = _sp.run(["netstat", "-ano"], capture_output=True, text=True)
        for d in res.stdout.splitlines():
            if ":5000 " in d and "LISTENING" in d:
                pid = d.strip().split()[-1]
                if pid.isdigit():
                    _sp.run(["taskkill", "/PID", pid, "/F"], capture_output=True)
                    print(f"🔪 Porta 5000 liberata (PID {pid})")
        time.sleep(0.5)
    except: pass

def _salva_kml(punti, path, data):
    root = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
    doc = ET.SubElement(root, "Document")
    ET.SubElement(doc, "name").text = f"Zone Google {data}"
    for i, p in enumerate(punti, 1):
        if not p.get('lat') or not p.get('lon'): continue
        pm = ET.SubElement(doc, "Placemark")
        ET.SubElement(pm, "name").text = f"{i}. {p.get('nome')[:80]}"
        desc = [f"<b>{p.get('nome')}</b>", f"Indirizzo: {p.get('indirizzo', '')}", f"Zona: {p.get('zona', '')}"]
        ET.SubElement(pm, "description").text = "<br>".join(desc)
        pt = ET.SubElement(pm, "Point")
        ET.SubElement(pt, "coordinates").text = f"{p['lon']},{p['lat']},0"
    tree = ET.ElementTree(root)
    if hasattr(ET, "indent"): ET.indent(tree, space="  ")
    tree.write(path, encoding="utf-8", xml_declaration=True, method="xml")

def _salva_html_fisico(output_base, template_html, data, zone_json, api_key, is_locked):
    is_locked_js = 'true' if is_locked else 'false'
    html_content = template_html.replace("{{DATA_GIORNO}}", data).replace("{{JSON_ZONE | safe}}", zone_json).replace("{{GOOGLE_MAPS_API_KEY}}", api_key).replace("{{IS_LOCKED_JS}}", is_locked_js)
    file_path = output_base / "4_mappa_zone_google.html"
    file_path.write_text(html_content, encoding="utf-8")

# --- TEMPLATE HTML RIPRISTINATO INTEGRALMENTE ---
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="utf-8">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://fonts.googleapis.com/icon?family=Material+Icons+Round" rel="stylesheet">
    <script>
        (g=>{var h,a,k,p="The Google Maps JavaScript API",c="google",l="importLibrary",q="__ib__",m=document,b=window;b=b[c]||(b[c]={});var d=b.maps||(b.maps={}),r=new Set,e=new URLSearchParams,u=()=>h||(h=new Promise(async(f,n)=>{await (a=m.createElement("script"));e.set("libraries",[...r]+"");for(k in g)e.set(k.replace(/[A-Z]/g,t=>"_"+t[0].toLowerCase()),g[k]);e.set("callback",c+".maps."+q);a.src=`https://maps.${c}apis.com/maps/api/js?`+e;d[q]=f;a.onerror=()=>h=n(Error(p+" could not load."));a.nonce=m.querySelector("script[nonce]")?.nonce||"";m.head.append(a)}));d[l]?console.warn(p+" only loads once. See https://goo.gle/js-api-loading-troubleshooting"):d[l]=(f,...n)=>r.add(f)&&u().then(()=>d[l](f,...n))})({
            key: "{{GOOGLE_MAPS_API_KEY}}",
            v: "beta"
        });
    </script>
    <style>
        :root { 
            --primary: #6366f1; --primary-dark: #4f46e5; --bg: #0f172a; 
            --sidebar-bg: #ffffff; --text-main: #1e293b; --text-muted: #64748b;
            --accent: #10b981; --card-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1);
        }
        * { box-sizing: border-box; }
        body { margin: 0; font-family: 'Inter', sans-serif; height: 100vh; display: flex; background: var(--bg); overflow: hidden; }
        #sidebar { width: 400px; height: 100%; background: var(--sidebar-bg); border-right: 1px solid #e2e8f0; display: flex; flex-direction: column; z-index: 1000; box-shadow: 10px 0 30px rgba(0,0,0,0.05); }
        #header { padding: 25px; background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%); color: white; }
        #zone-list { flex: 1; overflow-y: auto; padding: 15px; background: #f8fafc; }
        .zone-card { background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 12px; margin-bottom: 12px; cursor: pointer; transition: 0.2s; box-shadow: var(--card-shadow); }
        .zone-card.selected { border: 2px solid var(--primary); background: #f5f3ff; }
        .zone-header { display: flex; align-items: center; gap: 10px; }
        .color-pill { width: 12px; height: 12px; border-radius: 4px; }
        .zone-title { font-weight: 700; flex: 1; font-size: 0.9rem; }
        #map { flex: 1; }
        .btn { padding: 6px 10px; font-size: 0.7rem; font-weight: 700; cursor: pointer; border-radius: 6px; border: 1px solid #ddd; background: white; }
        .btn-confirm { background: #22c55e; color: white; border: none; }
        .btn-cancel { background: #ef4444; color: white; border: none; }
        .btn-lock { background: #f1f5f9; min-width: 100px; }
        .btn-lock.locked { background: #ef4444; color: white; }
        #save-status { position: fixed; top: 20px; left: 50%; transform: translateX(-50%); padding: 10px 20px; background: #10b981; color: white; border-radius: 30px; font-weight: 700; display: none; z-index: 5000; box-shadow: 0 10px 20px rgba(0,0,0,0.2); }
        .custom-marker { position: absolute; width: 32px; height: 32px; display: flex; align-items: center; justify-content: center; color: white; font-weight: 800; font-size: 11px; box-shadow: 0 4px 10px rgba(0,0,0,0.3); border: 2px solid white; transition: 0.2s; transform: translate(-50%, -100%); }
        .point-card { background: #f8fafc; padding: 7px 9px; margin-bottom: 5px; border-radius: 8px; border: 1px solid #e2e8f0; }
        .point-num { display: inline-flex; align-items: center; justify-content: center; width: 20px; height: 20px; border-radius: 50%; color: white; font-size: 0.65rem; font-weight: 800; margin-right: 6px; flex-shrink: 0; vertical-align: middle; }
        .point-name { font-size: 0.75rem; font-weight: 700; color: #1e293b; }
        .point-addr { font-size: 0.65rem; color: #64748b; margin-top: 2px; padding-left: 26px; }
        .zone-chip { display:inline-block; padding:3px 9px; border-radius:20px; border:2px solid #cbd5e1; font-size:0.6rem; font-weight:700; cursor:pointer; margin:2px 2px 0 0; color:#475569; background:white; transition:0.15s; white-space:nowrap; }
        .zone-chip.sel { color:white; border-color:transparent; }
        .chk-box { display:none; }
        .chk-lbl { display:inline-flex; width:22px; height:22px; border-radius:5px; border:2px solid #cbd5e1; align-items:center; justify-content:center; cursor:pointer; background:white; font-size:13px; color:transparent; flex-shrink:0; transition:0.15s; }
        .chk-box:checked + .chk-lbl { background:#6366f1; border-color:#6366f1; color:white; }
        .m-tonda { border-radius: 50%; }
        .m-goccia { border-radius: 50% 50% 50% 0; transform: rotate(-45deg); }
        .m-goccia span { transform: rotate(45deg); }
        .m-quadrato { border-radius: 4px; }
        .m-triangolo { background: transparent !important; border-left: 18px solid transparent; border-right: 18px solid transparent; border-bottom: 36px solid #ccc; width: 0 !important; height: 0 !important; border-radius:0; }
        .m-triangolo span { position: absolute; top: 16px; left: -8px; width: 16px; }
        /* Zona speciale DDT da inserire */
        .zone-card.speciale { border: 2px dashed #f59e0b; background: #fffbeb; }
        .zone-card.speciale .zone-title span:first-child { color: #92400e; }
        .zone-card.speciale .zone-title span:last-child { color: #b45309; }
        .badge-speciale { display:inline-block; background:#f59e0b; color:white; font-size:0.55rem; font-weight:800; padding:1px 6px; border-radius:10px; margin-left:6px; vertical-align:middle; letter-spacing:0.03em; }
    </style>
</head>
<body>
    <div id="sidebar">
        <div id="header">
            <h1 style="margin:0; font-size:1.2rem;">Gestione Zone <span id="tot-points" style="font-size:0.7rem; opacity:0.7;">0 Punti</span></h1>
            <div style="margin-top:10px; display:flex; justify-content: space-between; align-items:center;">
                <button onclick="saveAllToServer()" style="background:var(--accent); color:white; border:none; padding:8px 15px; border-radius:8px; font-weight:800; cursor:pointer;">SALVA TUTTO</button>
                <div style="text-align:right;"><small style="opacity:0.6;">{{DATA_GIORNO}}</small></div>
            </div>
            <div style="display:flex; gap:5px; margin-top:10px;">
                <button onclick="toggleLockMap()" id="btn-lock-map" class="btn btn-lock">CARICAMENTO...</button>
                <button id="btn-toggle-drag" onclick="toggleDragging()" class="btn" style="flex:1;">🔒 SPOSTA PUNTI (OFF)</button>
            </div>
        </div>
        <div id="zone-list"></div>
    </div>
    <div id="map"></div>
    <div id="save-status">💾 AGGIORNAMENTO COMPLETATO!</div>

    <script>
        let DATA_ZONE = {{JSON_ZONE | safe}};
        let map, gMarkers = [], DRAGGING_ENABLED = false, isLockedGlobal = {{IS_LOCKED_JS}};
        let activeExpandedZid = null, activeAction = null, activeSourceZid = null;
        let _hasUnsavedChanges = false;

        window.addEventListener('beforeunload', function(e) {
            if (_hasUnsavedChanges) {
                e.preventDefault();
                e.returnValue = 'Hai modifiche non salvate. Premi SALVA TUTTO prima di chiudere.';
                return e.returnValue;
            }
        });

        async function initMap() {
            const { Map } = await google.maps.importLibrary("maps");
            map = new Map(document.getElementById("map"), { center: { lat: 45.5, lng: 11.8 }, zoom: 9, mapTypeId: 'hybrid', mapId: "MY_MAP_ID" });
            _updateLockUI();
            updateTotals();
            await renderMarkers();
            renderSidebar();
            if(gMarkers.length) { 
                let b = new google.maps.LatLngBounds(); 
                gMarkers.forEach(m => b.extend(m.position)); 
                map.fitBounds(b); 
            }
        }

        function _updateLockUI() {
            const btn = document.getElementById('btn-lock-map');
            btn.innerHTML = isLockedGlobal ? 'BLOCCATA 🔒' : 'SBLOCCATA 🔓';
            btn.className = 'btn btn-lock' + (isLockedGlobal ? ' locked' : '');
        }

        async function toggleLockMap() {
            isLockedGlobal = !isLockedGlobal;
            fetch('/toggle_lock', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({action: isLockedGlobal?'lock':'unlock'}) })
            .then(() => { _updateLockUI(); renderMarkers(); renderSidebar(); });
        }

        async function toggleDragging() {
            DRAGGING_ENABLED = !DRAGGING_ENABLED;
            const btn = document.getElementById('btn-toggle-drag');
            btn.textContent = DRAGGING_ENABLED ? "🔓 SPOSTA PUNTI (ON)" : "🔒 SPOSTA PUNTI (OFF)";
            btn.style.background = DRAGGING_ENABLED ? "#10b981" : "#4b5563";
            btn.style.color = "white";
            await renderMarkers();
        }

        async function renderMarkers() {
            gMarkers.forEach(m => m.setMap(null)); gMarkers = [];
            const { AdvancedMarkerElement } = await google.maps.importLibrary("marker");
            const coordCounts = {};
            DATA_ZONE.forEach(z => {
                z.lista_punti.forEach((p, idx) => {
                    if (!p.lat) return;
                    const key = `${p.lat.toFixed(6)}_${p.lon.toFixed(6)}`;
                    coordCounts[key] = (coordCounts[key] || 0) + 1;
                    const layer = coordCounts[key];
                    let shape = 'm-goccia';
                    if (layer === 2) shape = 'm-quadrato'; else if (layer === 3) shape = 'm-triangolo'; else if (layer > 3) shape = 'm-tonda';
                    
                    const el = document.createElement("div");
                    el.className = `custom-marker ${shape}`;
                    if (shape === 'm-triangolo') el.style.borderBottomColor = z.color; else el.style.backgroundColor = z.color;
                    el.innerHTML = `<span>${idx+1}</span>`;
                    
                    const m = new AdvancedMarkerElement({ position: {lat: p.lat, lng: p.lon}, map: map, title: p.nome, content: el, gmpDraggable: DRAGGING_ENABLED && !isLockedGlobal });
                    m.addListener("dragend", () => {
                        p.lat = m.position.lat; p.lon = m.position.lng;
                        _hasUnsavedChanges = true;
                        _salvaSingolo(p.nome, p.indirizzo || '', p.lat, p.lon);
                    });
                    m.addListener("gmp-click", () => {
                        new google.maps.InfoWindow({ content: `<div style="padding:10px;"><b>${p.nome}</b><br>${p.indirizzo}<br><br><button onclick="window.open('https://www.google.com/maps/search/?api=1&query=${p.lat},${p.lon}')" style="width:100%; cursor:pointer;">VEDI SU GOOGLE</button></div>` }).open(map, m);
                    });
                    gMarkers.push(m);
                });
            });
        }

        function _salvaSingolo(nome, indirizzo, lat, lon) {
            fetch('/save_coord', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({nome, indirizzo, lat, lon}) })
            .then(r => r.json())
            .then(d => { 
                let st = document.getElementById('save-status');
                if (d.status === 'ok') {
                    _hasUnsavedChanges = false;
                    st.style.background = '#10b981';
                    st.textContent = '\u2705 Salvato: ' + nome;
                } else {
                    st.style.background = '#ef4444';
                    st.textContent = '\u274c Errore: ' + (d.msg || 'sconosciuto');
                }
                st.style.display = 'block'; 
                setTimeout(()=>{st.style.display='none';}, 3000); 
            })
            .catch(err => {
                let st = document.getElementById('save-status');
                st.style.background = '#ef4444';
                st.textContent = '\u274c Connessione persa!';
                st.style.display = 'block';
                setTimeout(()=>{st.style.display='none';}, 3000);
            });
        }

        function renderSidebar() {
            const list = document.getElementById('zone-list');
            list.innerHTML = DATA_ZONE.filter(z => z.lista_punti.length > 0).map(z => {
                const isSelected = (activeExpandedZid === z.id_zona) || (activeSourceZid === z.id_zona);
                const isSpeciale = (z.id_zona === 'DDT_DA_INSERIRE');
                return `
                <div class="zone-card ${isSelected ? 'selected' : ''} ${isSpeciale ? 'speciale' : ''}" onclick="focusZone('${z.id_zona}')">
                    <div class="zone-header">
                        <div class="color-pill" style="background: ${z.color}"></div>
                        <div class="zone-title">
                            <span style="display:block; font-size:0.95rem; font-weight:800;">${z.nome_giro || z.id_zona}${isSpeciale ? '<span class="badge-speciale">RIENTRI</span>' : ''}</span>
                            <span style="display:block; font-size:0.65rem; font-weight:600; color:#94a3b8;">${isSpeciale ? 'Assegna a un viaggio' : 'Zona ' + z.id_zona}</span>
                        </div>
                        <div style="text-align:right; flex-shrink:0;">
                            <div style="font-size:0.72rem; font-weight:800; background:${isSpeciale?'#fde68a':'#e2e8f0'}; padding:2px 7px; border-radius:20px; color:${isSpeciale?'#92400e':'#475569'}; white-space:nowrap;">${z.lista_punti.length} pt</div>
                            ${(() => { const s = calcolaValoreZona(z); return s.tot_ddt > 0 ? `<div style="font-size:0.65rem; font-weight:700; color:#10b981; margin-top:2px; white-space:nowrap;">${s.tot_ddt} DDT · € ${s.valore}</div>` : ''; })()}
                        </div>
                    </div>
                    ${isSelected ? `
                        <div style="margin-top:10px; border-top:1px solid #e2e8f0; padding-top:10px; max-height:320px; overflow-y:auto;">
                            ${z.lista_punti.map((p, idx) => {
                                const pid = (p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_');
                                const parts = (p.indirizzo || '').split(',');
                                const via = parts[0] || '';
                                const comune = parts.slice(1).join(',').trim();
                                const codiceDDT = [...(p.codici_ddt_frutta||[]), ...(p.codici_ddt_latte||[])].join(', ');
                                let ctrl = '';
                                if (activeAction === 'dividi' && !isSpeciale) ctrl = `<input type="checkbox" id="chk-${pid}" class="chk-box dividi-chk" value="${pid}" onclick="event.stopPropagation()"><label for="chk-${pid}" class="chk-lbl" onclick="event.stopPropagation()">✓</label>`;
                                else if (activeAction === 'sposta') ctrl = `<div class="zone-chips" data-pid="${pid}" style="padding-left:26px; margin-top:4px;">${DATA_ZONE.filter(o=>o.id_zona!==z.id_zona && o.id_zona!=='DDT_DA_INSERIRE').map(o=>`<span class="zone-chip" data-zone="${o.id_zona}" style="" onclick="event.stopPropagation(); _selChip(this,'${o.color}')">${o.nome_giro||o.id_zona}</span>`).join('')}</div>`;
                                return `<div class="point-card">
                                    <div style="display:flex; align-items:center; justify-content:space-between; gap:4px;">
                                        <div style="display:flex; align-items:center; flex:1; min-width:0; gap:0;">
                                            <span class="point-num" style="background:${z.color};">${idx+1}</span>
                                            <span class="point-name" style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${p.nome}</span>
                                        </div>
                                         ${activeAction === 'dividi' ? ctrl : ''}
                                    </div>
                                    ${codiceDDT ? `<div style="font-size:0.6rem; color:#f59e0b; font-weight:700; padding-left:26px; margin-top:2px;">DDT: ${codiceDDT}</div>` : ''}
                                     <div class="point-addr">
                                         <span style="display:block;">${via}</span>
                                         ${comune ? `<span style="font-weight:700; color:#334155;">${comune}</span>` : ''}
                                     </div>
                                     ${activeAction === 'sposta' ? ctrl : ''}
                                 </div>`;
                            }).join('')}
                        </div>
                        <div style="margin-top:10px; display:flex; gap:4px;">
                            ${activeAction ? `
                                <button class="btn btn-cancel" onclick="event.stopPropagation(); cancelAction()">✕ Annulla</button>
                                <button class="btn btn-confirm" style="flex:1;" onclick="event.stopPropagation(); ${activeAction==='dividi'?'executeDividi':'executeSposta'}('${z.id_zona}')">✓ Conferma</button>
                            ` : `
                                ${!isSpeciale ? `<button class="btn" style="flex:1;" ${isLockedGlobal?'disabled':''} onclick="event.stopPropagation(); startAction('dividi', '${z.id_zona}')">DIVIDI</button>` : ''}
                                <button class="btn" style="flex:1; ${isSpeciale?'background:#f59e0b;color:white;border:none;':''}" ${isLockedGlobal?'disabled':''} onclick="event.stopPropagation(); startAction('sposta', '${z.id_zona}')">${isSpeciale ? '📦 ASSEGNA A VIAGGIO' : 'SPOSTA'}</button>
                            `}
                        </div>
                    ` : ''}
                </div>`;
            }).join('');
        }

        function focusZone(zid) { if(activeAction) return; activeExpandedZid = (activeExpandedZid === zid) ? null : zid; renderSidebar(); }
        function startAction(type, zid) { activeAction = type; activeSourceZid = zid; renderSidebar(); }
        function cancelAction() { activeAction = null; activeSourceZid = null; renderSidebar(); }
        function _selChip(el, color) {
            el.closest('.zone-chips').querySelectorAll('.zone-chip').forEach(c => { c.classList.remove('sel'); c.style.background='white'; c.style.color='#475569'; });
            el.classList.add('sel'); el.style.background = color; el.style.color = 'white';
        }

        function executeDividi(sourceZid) {
            const z = DATA_ZONE.find(x => x.id_zona === sourceZid);
            const chks = document.querySelectorAll('.dividi-chk:checked');
            if(!chks.length) return alert("Seleziona punti!");
            const ids = Array.from(chks).map(c => c.value);
            const toMove = z.lista_punti.filter(p => ids.includes((p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_')));
            const newZ = { id_zona: sourceZid+"_B", nome_giro: (z.nome_giro||"Viaggio")+"/B", color: "#"+Math.floor(Math.random()*16777215).toString(16), lista_punti: toMove };
            z.lista_punti = z.lista_punti.filter(p => !toMove.includes(p));
            DATA_ZONE.push(newZ);
            cancelAction(); updateTotals(); renderMarkers(); renderSidebar();
        }

        function executeSposta(sourceZid) {
            const z = DATA_ZONE.find(x => x.id_zona === sourceZid);
            document.querySelectorAll('.zone-chips').forEach(chips => {
                const sel = chips.querySelector('.zone-chip.sel');
                if(sel) {
                    const target = DATA_ZONE.find(x => x.id_zona === sel.dataset.zone);
                    const pid = chips.dataset.pid;
                    const pIdx = z.lista_punti.findIndex(p => (p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_') === pid);
                    if(pIdx >= 0 && target) { target.lista_punti.push(z.lista_punti.splice(pIdx, 1)[0]); }
                }
            });
            cancelAction(); updateTotals(); renderMarkers(); renderSidebar();
        }

        function updateTotals() {
            const total = DATA_ZONE.reduce((acc, z) => acc + z.lista_punti.length, 0);
            document.getElementById('tot-points').textContent = `${total} Punti`;
        }

        const VALORE_DDT = 18.50;
        function calcolaValoreZona(z) {
            let totDdt = 0;
            z.lista_punti.forEach(p => {
                totDdt += (p.codici_ddt_frutta || []).length;
                totDdt += (p.codici_ddt_latte  || []).length;
                // fallback: se le liste non ci sono, conta 1 DDT se ha almeno un codice
                if (!(p.codici_ddt_frutta) && !(p.codici_ddt_latte)) {
                    if (p.codice_frutta && p.codice_frutta !== 'p00000') totDdt++;
                    if (p.codice_latte  && p.codice_latte  !== 'p00000') totDdt++;
                }
            });
            return { tot_ddt: totDdt, valore: (totDdt * VALORE_DDT).toFixed(2) };
        }

        function saveAllToServer() {
            const btn = event.target; btn.disabled = true; btn.textContent = "...";
            fetch('/save', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(DATA_ZONE) })
            .then(() => { 
                _hasUnsavedChanges = false;
                const st = document.getElementById('save-status');
                st.style.background = '#10b981';
                st.textContent = '✅ SALVATAGGIO COMPLETATO!';
                st.style.display = 'block'; 
                setTimeout(()=>{ st.style.display='none'; }, 3000); 
            }).finally(() => { btn.disabled = false; btn.textContent = "SALVA TUTTO"; });
        }

        window.onload = initMap;
    </script>
</body>
</html>"""

def _carica_e_genera(data_giorno):
    global TARGET_FILE_UNIFICATO, TARGET_FILE_VIAGGI, ZONE_LIST_CACHE, DATA_GIORNO
    DATA_GIORNO = data_giorno
    output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
    TARGET_FILE_UNIFICATO = output_base / "punti_consegna_unificati.json"
    TARGET_FILE_VIAGGI = output_base / "viaggi_giornalieri.json"
    
    if not TARGET_FILE_UNIFICATO.exists(): return False
    try:
        unif_data = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
        punti = unif_data.get("punti", [])
        zone_dict = {}
        for p in punti:
            zid = p.get("zona") or "SENZA_ZONA"
            if zid not in zone_dict:
                if zid == "DDT_DA_INSERIRE":
                    nome_z = "⚠️ DDT DA INSERIRE"
                    colore_z = "#f59e0b"  # arancio per distinguerla
                else:
                    nome_z = f"Viaggio {len([z for z in zone_dict if z != 'DDT_DA_INSERIRE'])+1}"
                    colore_z = _get_color(len(zone_dict))
                zone_dict[zid] = {"id_zona": zid, "lista_punti": [], "color": colore_z, "nome_giro": nome_z}
            zone_dict[zid]["lista_punti"].append(p)
        # Ordina: zone normali per id, DDT_DA_INSERIRE alla fine
        zone_normali = sorted([z for z in zone_dict.values() if z["id_zona"] != "DDT_DA_INSERIRE"], key=lambda x: str(x["id_zona"]))
        zone_speciali = [z for z in zone_dict.values() if z["id_zona"] == "DDT_DA_INSERIRE"]
        ZONE_LIST_CACHE = zone_normali + zone_speciali
        
        if not TARGET_FILE_VIAGGI.exists():
            TARGET_FILE_VIAGGI.write_text(json.dumps(ZONE_LIST_CACHE, indent=2, ensure_ascii=False), encoding="utf-8")
        
        _salva_html_fisico(output_base, HTML_TEMPLATE, DATA_GIORNO, json.dumps(ZONE_LIST_CACHE, ensure_ascii=False), GOOGLE_MAPS_API_KEY, unif_data.get("is_locked", True))
        _salva_kml(punti, output_base / f"zone_google_{DATA_GIORNO.replace('-', '_')}.kml", DATA_GIORNO)
        return True
    except: return False

def main():
    args = sys.argv[1:]; serve = True; data = ""
    for a in args:
        if a == "--no-serve": serve = False
        elif not a.startswith("--"): data = a
    if not data:
        f = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not f: return
        f.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = f[0].name.replace("CONSEGNE_", "")
    
    if _carica_e_genera(data) and serve:
        _libera_porta_5000()
        print(f"\n🌐 Mappa Operativa: http://127.0.0.1:5000")
        def open_b():
            time.sleep(1.2)
            webbrowser.open("http://127.0.0.1:5000")
        threading.Thread(target=open_b, daemon=True).start()
        app.run(port=5000, debug=False)

if __name__ == "__main__": main()
