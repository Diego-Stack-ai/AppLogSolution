import pandas as pd
f1 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026/punti_consegna.xlsx'
f2 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026_copia/punti_consegna.xlsx'

df_new = pd.read_excel(f1)
df_old = pd.read_excel(f2)

merged = df_old.merge(df_new, on='Nome', suffixes=('_old', '_new'))

diff_lat = (merged['Latitudine_old'].round(5) != merged['Latitudine_new'].round(5))
diff_lon = (merged['Longitudine_old'].round(5) != merged['Longitudine_new'].round(5))
diff = merged[diff_lat | diff_lon]

print('Punti analizzati:', len(merged))
print('Punti con differenze nelle coordinate:', len(diff))

for idx, row in diff.iterrows():
    print(row['Nome'], '-> VERO(copia):', row['Latitudine_old'], row['Longitudine_old'], '| SBAGLIATO(nuovo):', row['Latitudine_new'], row['Longitudine_new'])

if len(diff) > 0:
    from openpyxl import load_workbook
    map_file = 'G:/Il mio Drive/App/AppLogSolution/dati/PROGRAMMA/mappatura_destinazioni.xlsx'
    wb = load_workbook(map_file)
    ws = wb.active
    headers = [str(c.value).strip().lower() for c in ws[1]]
    col_nome = next(i for i, h in enumerate(headers) if 'a chi va' in h or h == 'nome')
    col_lat = next(i for i, h in enumerate(headers) if h == 'latitudine')
    col_lon = next(i for i, h in enumerate(headers) if h == 'longitudine')
    
    saved = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        n = str(row[col_nome].value).strip().lower()
        match = diff[diff['Nome'].str.lower().str.strip() == n]
        if not match.empty:
            ws.cell(row=r_idx, column=col_lat+1, value=match.iloc[0]['Latitudine_old'])
            ws.cell(row=r_idx, column=col_lon+1, value=match.iloc[0]['Longitudine_old'])
            saved += 1
            
    wb.save(map_file)
    print('\nAggiornati', saved, 'punti nel master Excel mappatura_destinazioni.xlsx!')
