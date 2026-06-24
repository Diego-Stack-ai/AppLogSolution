#!/usr/bin/env python3
"""
(1_2_3_4)_estrai_ddt_consegne.py
════════════════════════════════
Estrae tutti i DDT dai PDF in CONSEGNE/DDT-ORIGINALI/FRUTTA e LATTE.
Identifica nuovi clienti ed estrae automaticamente i loro dati (Indirizzo, CAP, Orari, ecc.).

Uso: py (1_2_3_4)_estrai_ddt_consegne.py [data]
"""

import re
import sys
import subprocess
import time
import json
from pathlib import Path
from datetime import datetime

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    print("pip install pypdf")
    sys.exit(1)

# --- CONFIGURAZIONE ---
PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
INPUT_FRUTTA = BASE_DIR / "FRUTTA"
INPUT_LATTE = BASE_DIR / "LATTE"
MAPPATURA_XLSX = PROG_DIR / "mappatura_destinazioni.xlsx"

# Regex per estrazione dati
DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'(?:[Ll]uogo [Dd]i [Dd]estinazione|[Cc]odice [Dd]estinazione):\s*([pP]\d{4,5})')
CAP_RE = re.compile(r"\b(\d{5})\b")
PROVINCIA_RE = re.compile(r"\(([A-Z]{2})\)")
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})(?:\s+H(\d{2}))?(?:\s+(\d{3}))?', re.I)


def _estrai_data_luogo(text: str) -> tuple[str | None, str | None]:
    """Estrae (data, luogo) da una pagina DDT. data in formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    return (data, luogo)


def _estrai_dati_consegna_da_testo(text: str, codice: str, da_frutta: bool) -> dict:
    """Estrae destinatario, indirizzo, CAP, città, provincia e orari dal testo di una pagina DDT."""
    res = {"dest": "", "ind": "", "cap": "", "cit": "", "prov": "", "om": "", "oM": "14:00"}
    if codice.lower() not in text.lower():
        return res
    
    idx_l = text.find("Luogo di destinazione")
    if idx_l < 0: return res

    # 1. Nome e Indirizzo
    if da_frutta:
        blocco = text[idx_l : idx_l + 650]
        lines = [ln.strip() for ln in blocco.split("\n") if ln.strip()]
        for i, ln in enumerate(lines):
            if LUOGO_RE.search(ln):
                if i + 1 < len(lines): res["dest"] = lines[i + 1].strip().title()
                if i + 2 < len(lines): res["ind"] = lines[i + 2].strip().title()
                break
    else:
        idx_causale = text.upper().find("CAUSALE DEL TRASPORTO")
        blocco = text[:idx_causale] if idx_causale > 0 else text[idx_l : idx_l + 900]
        for ln in blocco.split("\n"):
            ln = ln.strip()
            cf_m = re.match(r"^[Cc]\.?[Ff]\.?\s+", ln)
            if cf_m: res["dest"] = ln[cf_m.end():].strip().title()
            else:
                albo_m = re.match(r"^[Aa]lbo\s+", ln, re.I)
                if albo_m: res["ind"] = ln[albo_m.end():].strip().title()

    # 2. CAP, Provincia, Città
    idx_resp = text.upper().find("RESPONSABILE DEL TRASPORTO")
    blocco_prov = text[idx_resp:] if idx_resp >= 0 else text
    
    for prov_m in PROVINCIA_RE.finditer(blocco_prov):
        sigla = prov_m.group(1)
        if sigla == "MN" and ("Pomponesco" in blocco_prov[max(0, prov_m.start()-40):prov_m.start()] or "46030" in blocco_prov):
            continue
        res["prov"] = sigla
        caps = list(CAP_RE.finditer(blocco_prov[:prov_m.start()]))
        if caps:
            res["cap"] = caps[-1].group(1)
            pre = blocco_prov[caps[-1].end() : caps[-1].end() + 60]
            citta_m = re.search(r"\s*[-]?\s*([A-Za-zÀ-ÿ\s'.]+?)\s*\([A-Z]{2}\)", pre)
            if citta_m: res["cit"] = citta_m.group(1).strip().title()
        break
        
    # 3. Orari
    idx_c = text.upper().find("CAUSALE DEL TRASPORTO")
    if idx_c >= 0:
        sezione = text[idx_c:idx_c+150]
        m = CAUSALE_RE.search(sezione)
        if m:
            if m.group(2): res["oM"] = f"{int(m.group(2)):02d}:00"
            if m.group(3):
                s = m.group(3)
                if len(s) == 3: res["om"] = f"{int(s[0]):02d}:{int(s[1:3]):02d}"
    return res


def _ricava_date_da_pdf() -> list[str]:
    """Raccoglie TUTTE le date distinte trovate nei PDF di FRUTTA e LATTE."""
    import pdfplumber
    date_trovate: set[str] = set()
    for cart in (INPUT_FRUTTA, INPUT_LATTE):
        if not cart.exists(): continue
        for pdf_path in sorted(cart.glob("*.pdf")):
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        data, _ = _estrai_data_luogo(text)
                        if data: date_trovate.add(data)
            except: pass
    return sorted(date_trovate)


def _leggi_codici_mappatura():
    if not MAPPATURA_XLSX.exists(): return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
        ws = wb.active
        codici = set()
        for row in ws.iter_rows(min_row=2, max_col=2):
            for cell in row:
                val = str(cell.value or "").strip().lower()
                if val and val != "p00000": codici.add(val)
        wb.close()
        return codici
    except Exception as e:
        print(f"⚠️ Errore mappatura: {e}")
        return set()


def _verifica_nuovi_clienti(dati_nuovi: dict):
    if not dati_nuovi: return True
    print("\n" + "!"*60)
    print(f"🛑 RILEVATI {len(dati_nuovi)} NUOVI CODICI CLIENTE:")
    for cod, info in sorted(dati_nuovi.items()):
        print(f"   - {cod} ({info['tipo']}): {info['dest']} - {info['cit']}")
    print("!"*60 + "\n")
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nuovi Codici"
        headers = ["Codice Frutta", "Codice Latte", "A chi va consegnato", "Tipologia grado", "Indirizzo", "CAP", "Città", "Provincia",
                  "Tipologia consegna", "Tipologia consegna.1", "Email", "Sito web", "Orario min", "Orario max"]
        ws.append(headers)
        for cod, info in sorted(dati_nuovi.items()):
            row = [""] * 14
            if info["tipo"] == "FRUTTA": row[0] = cod
            else:                        row[1] = cod
            row[2], row[4], row[5], row[6], row[7], row[12], row[13] = info["dest"], info["ind"], info["cap"], info["cit"], info["prov"], info["om"], info["oM"]
            ws.append(row)
        hp = BASE_DIR / "nuovi_codici_consegna.xlsx"
        wb.save(hp)
        print(f"📄 File generato: {hp.name}\n💡 Copia i dati in mappatura_destinazioni.xlsx per proseguire.\n")
    except Exception as e: print(f"⚠️ Errore excel: {e}")
    return False


def _estrai_da_cartella(cart_in: Path, cart_out: Path, etichetta: str, date_valide: set[str], mappati: set, *, duplicata: bool = False):
    """Estrae DDT da tutti i PDF accettando qualsiasi data presente in date_valide."""
    nuovi_dati = {}
    if not cart_in.exists(): return 0, 0, nuovi_dati
    cart_out.mkdir(parents=True, exist_ok=True)
    pdf_files = list(cart_in.glob("*.pdf"))
    if not pdf_files: return 0, 0, nuovi_dati
    import pdfplumber
    creati = 0
    visti = {}
    for pdf_path in pdf_files:
        try:
            reader = PdfReader(pdf_path)
            with pdfplumber.open(pdf_path) as pdf:
                for i in range(0, len(pdf.pages), 2 if duplicata else 1):
                    text = pdf.pages[i].extract_text() or ""
                    d, l = _estrai_data_luogo(text)
                    if not d or not l or d not in date_valide: continue  # accetta tutte le date valide
                    if l not in mappati and l not in nuovi_dati:
                        nuovi_dati[l] = _estrai_dati_consegna_da_testo(text, l, duplicata)
                        nuovi_dati[l]["tipo"] = etichetta
                    chiave = (d, l)
                    cnt = visti.get(chiave, 0) + 1
                    visti[chiave] = cnt
                    fname = f"{l}_{d}_{cnt}.pdf" if cnt > 1 else f"{l}_{d}.pdf"
                    writer = PdfWriter()
                    writer.add_page(reader.pages[i])
                    with open(cart_out / fname, "wb") as f: writer.write(f)
                    creati += 1
                    if creati <= 3 or creati % 50 == 0: print(f"    {fname}")
        except Exception as e: print(f"  Errore {pdf_path.name}: {e}")
    return creati, sum(1 for c in visti.values() if c > 1), nuovi_dati


def _pulisci_sorgenti(cart_in: Path, date_valide: set[str]):
    """Elimina i PDF sorgenti che contengono pagine di qualsiasi data elaborata."""
    import pdfplumber
    rimossi = 0
    for p in cart_in.glob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                if any(_estrai_data_luogo(pg.extract_text() or "")[0] in date_valide for pg in pdf.pages):
                    pdf.close(); p.unlink(); rimossi += 1
        except: pass
    return rimossi


def _pulisci_output(base: Path, data_v: str):
    for f in [base/"punti_consegna.xlsx", base/"punti_consegna_unificati.json", base/"4_mappa_zone_google.html", base/f"zone_google_{data_v.replace('-', '_')}.kml"]:
        if f.exists():
            try: f.unlink()
            except: pass


# --- ARTICOLI NOTI ---
ARTICOLI_NOTI = {"ME-T-DI-V0-NA", "PE-T-DI-L3-NA", "10-GEL", "10-FLYER", "10-MANIFESTO", "LT-DL-02-LC", "LT-ES-04-LS",
                 "LT-ESL-IN-LB", "LT-AQ-04-LV", "YO-BI-MN-04-LB", "YO-DL-02-LC", "AP-SU-PC", "FO-DI-PV-04-LB",
                 "CA-Z-BI-L3-NA", "FO-DI-GP-01-NI", "FVNS-03-GADGET", "KI-S-BI-L3-NA", "FVNS-03-POSTER",
                 # Aggiunti 26/03/2026
                 "FI-Z-BI-L3-NA",   # Finocchio biologico da porzionare in classe
                 "ME-S-BI-L3-NA",   # Mela biologica del territorio per estratto
                 # Aggiunti 09/04/2026
                 "10-AT-01", "LNS-04-GADGET", "LNS-04-POSTER", "FVNS-03-FOLDER", "FVNS-03-MAGAZINE",
                 "FR-M-BI-L3-NI",
                 }

def _verifica_nuovi_articoli(base):
    print("Estrazione Tabella DDT in corso...")
    import pdfplumber
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    divisi = base / "DDT-ORIGINALI-DIVISI"
    if not divisi.exists(): return True

    # --- FASE 1: Raccolta righe grezze ---
    righe_grezze = []
    file_processati = 0
    for p in divisi.rglob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                for pg in pdf.pages:
                    tables = pg.extract_tables()
                    for t in tables:
                        if not t or not t[0] or not t[0][0]: continue
                        if "Cod." in str(t[0][0]) and "Articolo" in str(t[0][0]):
                            for row in t[1:]:
                                if not row or not row[0]: continue
                                raw_codice = str(row[0])
                                lines = raw_codice.split('\n')
                                clean_codice = ""
                                for ln in lines:
                                    if ln.strip().startswith("Codice:"): break
                                    m2 = re.match(r'^([A-Z0-9\-]+)', ln.strip())
                                    if m2: clean_codice += m2.group(1)
                                if re.match(r'^--\d{6}$', clean_codice):
                                    clean_codice = "10-FLYER"
                                righe_grezze.append({
                                    "codice":      clean_codice,
                                    "descrizione": str(row[1]).replace('\n', ' ') if len(row)>1 and row[1] else "",
                                    "netto_raw":   str(row[2]).replace('\n', ' ') if len(row)>2 and row[2] else "0",
                                    "qta_raw":     str(row[3]).replace('\n', ' ') if len(row)>3 and row[3] else "",
                                    "porz_raw":    str(row[4]).replace('\n', ' ') if len(row)>4 and row[4] else "0",
                                    "conf":        str(row[5]).replace('\n', ' ') if len(row)>5 and row[5] else "",
                                    "grezzo":      raw_codice.replace('\n', ' | '),
                                })
            file_processati += 1
        except Exception:
            pass

    # --- FASE 2: GroupBy (codice, descrizione, grezzo) ---
    gruppi = defaultdict(lambda: {"netto_kg": 0.0, "porzioni": 0, "qta_num": 0, "qta_unita": "", "conf": ""})
    for r in righe_grezze:
        key = (r["codice"], r["descrizione"], r["grezzo"])
        g = gruppi[key]
        try:
            g["netto_kg"] += float(r["netto_raw"].replace(',', '.'))
        except Exception:
            pass
        try:
            num = re.search(r'\d+', r["porz_raw"])
            if num: g["porzioni"] += int(num.group())
        except Exception:
            pass
        try:
            m_qta = re.match(r'^(\d+)\s*(.*)', r["qta_raw"].strip())
            if m_qta:
                g["qta_num"] += int(m_qta.group(1))
                if not g["qta_unita"]: g["qta_unita"] = m_qta.group(2).strip()
        except Exception:
            pass
        g["conf"] = r["conf"]

    # --- FASE 3: Scrittura Excel ---
    wb = Workbook()

    # Foglio 1: Dettaglio Completo
    ws_det = wb.active
    ws_det.title = "Dettaglio Completo"
    hdr1 = ["Codice Articolo Pulito", "Descrizione Natura Qualita", "Netto Kg",
            "Quantita in consegna", "Porzioni Effettive", "Confezionamento", "Codice Grezzo (PDF)"]
    ws_det.append(hdr1)
    for cell in ws_det[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4057")
    for r in righe_grezze:
        ws_det.append([r["codice"], r["descrizione"], r["netto_raw"], r["qta_raw"], r["porz_raw"], r["conf"], r["grezzo"]])

    # Foglio 2: RIEPILOGO MASTER compattato
    ws_r = wb.create_sheet("RIEPILOGO MASTER")
    hdr2 = ["Codice Articolo", "Descrizione", "TOTALE Netto Kg", "TOTALE Quantita", "TOTALE Porzioni", "Confezionamento"]
    ws_r.append(hdr2)
    for cell in ws_r[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A936F")
        cell.alignment = Alignment(horizontal="center")
    for (cod, desc, _), g in sorted(gruppi.items()):
        netto_fmt = f"{g['netto_kg']:.1f}".replace('.', ',') if g["netto_kg"] > 0 else "0"
        qta_fmt = f"{g['qta_num']} {g['qta_unita']}".strip() if g["qta_num"] > 0 else ""
        ws_r.append([cod, desc, netto_fmt, qta_fmt, g["porzioni"] or "", g["conf"]])

    # Larghezze colonne automatiche
    for ws_x in [ws_det, ws_r]:
        for col in ws_x.columns:
            max_l = max((len(str(c.value or "")) for c in col), default=10)
            ws_x.column_dimensions[col[0].column_letter].width = min(max_l + 4, 60)

    out_path = BASE_DIR / "prova_codici.xlsx"
    wb.save(out_path)
    print(f"\n[SIMULATORE] COMPLETATO! {len(righe_grezze)} righe da {file_processati} DDT.")
    print(f"[SIMULATORE] Compattate in {len(gruppi)} articoli unici nel RIEPILOGO MASTER.")
    print(f"[SIMULATORE] File Excel: {out_path}\n")
    return True


def main():
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    if not arg:
        print("❌ Passa la data del giorno. Esempio: python COPIA_TEST_estrai_codici.py 13-04-2026")
        return
        
    date_valide = set()
    for parte in arg.split("_"):
        if re.match(r"^\d{2}-\d{2}$", parte): parte = f"{parte}-2026"
        date_valide.add(parte)
        
    data_label = "_".join(sorted(date_valide))
    base = CONSEGNE_DIR / f"CONSEGNE_{data_label}"
    
    if not base.exists():
        print(f"❌ Cartella {base.name} non trovata.")
        return
        
    print(f"\n--- Avvio Estrazione Tabellare Magica ({data_label}) ---")
    _verifica_nuovi_articoli(base)

if __name__ == "__main__": main()
