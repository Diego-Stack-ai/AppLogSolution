#!/usr/bin/env python3
"""
9_genera_distinte_da_viaggi.py
══════════════════════════════
Legge viaggi_giornalieri_OTTIMIZZATO.json (generato da 8_genera_json_ottimizzato.py)
e per ogni viaggio:
  1. Trova i PDF DDT individuali in DDT-ORIGINALI-DIVISI/{FRUTTA,LATTE}/
  2. Estrae e consolida gli articoli (logica da crea_distinta_magazzino)
  3. Genera una distinta PDF per viaggio nell'ORDINE OTTIMIZZATO
  4. Assembla un Master PDF unico

NOTA ORDINE CARICO MAGAZZINO:
  - L'ordine nel JSON è l'ordine di CONSEGNA (fermata 1 = prima consegna)
  - Il magazziniere carica AL CONTRARIO: ultima fermata = primo caricato sul furgone

VERIFICA ORFANI:
  - Dopo aver processato tutti i viaggi, lo script verifica che non ci siano
    PDF DDT nella cartella che non sono stati assegnati a nessun viaggio.

Uso: py 9_genera_distinte_da_viaggi.py [data]
     es: py 9_genera_distinte_da_viaggi.py 26-03-2026
"""

import json
import re
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

# --- CONFIGURAZIONE ---
PROG_DIR     = Path(__file__).resolve().parent
BASE_DIR     = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
RIENTRI_XLSX = BASE_DIR / "rientri_ddt.xlsx"
CODICE_VUOTO = "p00000"

# Importa logica estrazione articoli da crea_distinta_magazzino
DISTINTA_MOD = BASE_DIR / "crea_distinta_magazzino.py"

# --- DATABASE ARTICOLI (SORGENTE DI VERITÀ) ---
ARTICOLI_NOTI_SET = frozenset({
    "10-FLYER", "10-GEL", "10-MANIFESTO", "10-AT-01", "10-BICC", "10-CUCCH", "10-PIATTO",
    "AP-SU-PC", "FO-DI-PV-04-LB", "FO-DI-GP-01-NI", "FVNS-03", "FVNS-03-", 
    "LT-AQ-04-LV", "LT-AQ-04-LB", "LT-AQ-04-LS", "LT-DL-02-LC", "LT-ES-04-LS", "LT-ESL-IN-LB", 
    "MA-T-LI-L3-NA", "ME-T-DI-V0-NA", "ME-S-BI-L3-NA", "PE-T-DI-L3-NA",
    "YO-BI-MN-04-LB", "YO-DL-02-LC", "FI-Z-BI-L3-NA", "FR-M-BI-L3-NI",
    "LNS-04-GADGET", "LNS-04-", "CA-Z-BI-L3-NA", "KI-S-BI-L3-NA"
})

def _is_primary_code(text):
    """Rileva se una stringa è un codice primario noto (RIGA 1)."""
    if not text: return False
    text = text.strip().upper()
    # Controllo esatto o se inizia con uno dei prefissi noti
    if text in ARTICOLI_NOTI_SET: return True
    # Caso speciale per prefissi come FVNS-03- o LNS-04-
    for prefix in ARTICOLI_NOTI_SET:
        if prefix.endswith('-') and text.startswith(prefix):
            return True
    # Pattern generico di sicurezza per codici standard
    return bool(re.match(r'^([A-Z0-9]{2,}-[A-Z0-9\-]+|--\d{6})', text))

CODICE_MAP = {
    "FVNS-03-": "FVNS-03", # Normalizzazione base
}

CONSOLIDAMENTO = {
    "LT-ES-04-LS":   ("Fardelli",  "Bottiglie", 10),
    "LT-ESL-IN-LB":  ("Fardelli",  "Bottiglie",  6),
    "LT-AQ-04-LB":   ("Fardelli",  "Bottiglie", 12),
    "LT-AQ-04-LS":   ("Fardelli",  "Bottiglie", 10),
    "LT-AQ-04-LV":   ("Fardelli",  "Bottiglie",  6),
    "YO-BI-MN-04-LB":("Cartoni",   "Cluster",   10),
    "YO-DL-02-LC":   ("Cartoni",   "Porzioni",   6),
    "AP-SU-PC":      ("Cartoni",   "Porzioni",  24),
    "FO-DI-GP-01-NI":("Colli",     "Buste",     16),
    "FO-DI-PV-04-LB":("Colli",     "Fette",     20),
}

UNITA_QTY = r"(Confezioni|Confezione|confezioni|confezione|Colli|Collo|colli|collo|Brick|brick|Fardelli|Fardello|fardelli|fardello|Bottiglie|Bottiglia|bottiglie|bottiglia|Cartoni|Cartone|cartoni|cartone|Cluster|cluster|Porzioni|Porzione|porzioni|porzione|Fascette|Fascetta|fascette|fascetta|Manifesti|Manifesto|manifesti|manifesto|Fette|Fetta|fette|fetta|Buste|Busta|buste|busta|pz)"
CONF_PATTERN = r"[\d,\.]+\s*(?:Porzioni?|Bottiglie?|Cluster|Fetta?)\s*/\s*\w+"
SCAD_RE = re.compile(r"Scad\.\s*min\.\s*(\d{2}/\d{2}/\d{4})", re.I)


# ──────────────────────────────────────────────────────────────────────────────
# UTILITÀ
# ──────────────────────────────────────────────────────────────────────────────

def _normalizza_unita(u: str) -> str:
    u = u.strip().lower()
    mapping = {
        "bottiglia": "Bottiglie", "bottiglie": "Bottiglie",
        "fardello": "Fardelli",   "fardelli": "Fardelli",
        "cartone": "Cartoni",     "cartoni": "Cartoni",
        "cluster": "Cluster",
        "porzione": "Porzioni",   "porzioni": "Porzioni",
        "collo": "Colli",         "colli": "Colli",
        "fetta": "Fette",         "fette": "Fette",
        "brick": "Brick",
        "confezione": "Confezioni", "confezioni": "Confezioni",
        "manifesto": "Manifesti", "manifesti": "Manifesti",
        "fascetta": "Fascette",
        "busta": "Buste",         "buste": "Buste",
        "pz": "pz"
    }
    return mapping.get(u, u.title() if u else u)


def _consolida_quantita(codice: str, lista_qty: list) -> tuple:
    if codice not in CONSOLIDAMENTO:
        by_unit = defaultdict(int)
        for qty, unit in lista_qty:
            by_unit[_normalizza_unita(unit)] += qty
        result = [(v, k) for k, v in sorted(by_unit.items()) if v > 0]
        return result, " ".join(f"{q} {u}" for q, u in result)

    unit_princ, unit_second, ratio = CONSOLIDAMENTO[codice]
    tot_princ = tot_second = 0
    for qty, unit in lista_qty:
        ul = unit.lower()
        if unit_princ.lower() in ul or ul in ("fardello", "fardelli", "cartoni", "cartone",
                                               "brick", "colli", "confezioni", "manifesti", "fascette"):
            tot_princ += qty
        else:
            tot_second += qty

    extra_princ   = tot_second // ratio
    resto_second  = tot_second % ratio
    tot_princ    += extra_princ

    result = []
    if tot_princ > 0:
        result.append((tot_princ, unit_princ))
    if resto_second > 0:
        result.append((resto_second, unit_second))
    display = " e ".join(f"{q} {u}" for q, u in result)
    return result, display


def _parse_quantita_da_cella(cell) -> list:
    if not cell or not str(cell).strip():
        return []
    text = str(cell).replace("\n", " ").replace("  ", " ")
    quantita = []
    for m in re.finditer(r"(?:^|e\s+)(\d+)\s+(" + UNITA_QTY + r")", text, re.I):
        quantita.append((int(m.group(1)), _normalizza_unita(m.group(2))))
    if not quantita and re.search(r"^(\d+)\s*$", text.strip()):
        quantita.append((int(text.strip()), "pz"))
    return quantita


def _normalizza_cella_codice(raw: str) -> tuple[str, str]:
    """
    Dato il contenuto grezzo di una cella "Cod. Articolo" (possibilmente multi-riga),
    restituisce (codice_base, variante_raw) senza alcuna interpretazione semantica.

    Regole:
    - Splitta per \n
    - Filtra righe vuote e righe che iniziano con "Codice:" (metadati PDF interni)
    - La PRIMA riga che corrisponde ad ARTICOLI_NOTI diventa codice_base
    - Tutte le righe successive (nella stessa cella) diventano variante_raw
    - variante_raw è normalizzata SOLO: spazi multipli -> spazio, trattini multipli -> uno
    """
    righe = [l.strip() for l in raw.split('\n')
             if l.strip() and not l.strip().startswith("Codice:")]

    if not righe:
        return "", ""

    # Cerca il codice base nella prima riga (idealmente riga 0)
    codice_base = ""
    idx_base = -1
    for i, riga in enumerate(righe):
        if _is_primary_code(riga):
            codice_base = riga.strip()
            idx_base = i
            break

    if not codice_base:
        # Nessun codice base riconosciuto: trattiamo la prima riga come base grezza
        codice_base = righe[0]
        idx_base = 0

    # Variante = righe successive alla base, nella stessa cella
    righe_variante = righe[idx_base + 1:]
    variante_raw = " ".join(righe_variante).strip()
    # Normalizzazione minima: spazi multipli e trattini doppi
    variante_raw = re.sub(r'\s+', ' ', variante_raw)
    variante_raw = re.sub(r'-{2,}', '-', variante_raw).strip('-').strip()

    return codice_base, variante_raw


def _estrai_articoli_da_tabella(page) -> list | None:
    """
    Logica column-based stabile: OGNI RIGA della tabella = UN articolo indipendente.
    Nessuna state machine tra righe diverse. Nessuna interpretazione semantica.
    """
    tables = page.extract_tables()
    if not tables: return None
    tab = next((t for t in tables if t and len(t) > 1
                and "Cod. Articolo" in " ".join(str(c or "") for c in t[0])), None)
    if not tab: return None

    risultato = []

    for row in tab[1:]:
        # Filtra righe troppo corte o vuote
        if not row or len(row) < 4: continue

        # ── Colonna 0: Cod. Articolo ──────────────────────────────────────────
        raw_codice = str(row[0] or "").strip()
        if not raw_codice: continue

        codice_base, variante_raw = _normalizza_cella_codice(raw_codice)
        if not codice_base: continue

        # ── Colonna 1: Descrizione ────────────────────────────────────────────
        # Multi-riga nella cella → unita con spazio, nessuna altra modifica
        descrizione = re.sub(r'\s+', ' ', str(row[1] or "").replace('\n', ' ')).strip()

        # ── Colonna 2: Peso/Kg ────────────────────────────────────────────────
        try:
            kg = Decimal(str(row[2] or "0").replace(",", ".").strip() or "0")
        except Exception:
            kg = Decimal("0")

        # ── Colonna 3: Quantità ───────────────────────────────────────────────
        quantita_raw = str(row[3] or "").strip()
        quantita = _parse_quantita_da_cella(quantita_raw)

        # Caso speciale 10-GEL: la quantità è in colonna 4 (porzioni)
        if not quantita and "10-GEL" in codice_base:
            porz = str(row[4] or "").strip() if len(row) > 4 else ""
            if porz.isdigit():
                quantita = [(int(porz), "pz")]

        if not quantita: continue

        # ── Colonna 5: Confezionamento (opzionale) ────────────────────────────
        confezionamento = str(row[5] or "").strip() if len(row) > 5 else ""

        risultato.append({
            "codice_base": codice_base,
            "variante_raw": variante_raw,
            "descrizione": descrizione,
            "scadenza": SCAD_RE.search(descrizione).group(1) if SCAD_RE.search(descrizione) else "",
            "kg": kg,
            "quantita": quantita,
            "confezionamento": confezionamento,
        })

    return risultato if risultato else None


def _raccogli_articoli_da_pdf(pdf_path: Path, tipo: str) -> list:
    import pdfplumber
    tutti = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                arts = _estrai_articoli_da_tabella(page)
                tutti.extend(arts or [])
    except Exception as e:
        print(f"    WARN  Errore lettura {pdf_path.name}: {e}")
    return tutti


def _aggrega_articoli(lista: list) -> dict:
    """
    Aggrega articoli con la stessa chiave (codice_base, variante_raw), sommando kg e quantita.
    Questa chiave è stabile perché proviene dalla cella originale, solo normalizzata
    (spazi e trattini), senza alcuna interpretazione semantica.
    """
    agg = {}
    for art in lista:
        # Chiave = (codice_base, variante_raw) — entrambi estratti direttamente dalla cella
        chiave = (art["codice_base"], art["variante_raw"])
        if chiave not in agg:
            agg[chiave] = {
                "codice_base":  art["codice_base"],
                "variante_raw": art["variante_raw"],
                "descrizione":  art["descrizione"],
                "scadenza":     art["scadenza"],
                "kg":           Decimal("0"),
                "quantita":     [],
                "confezionamento": art["confezionamento"],
            }
        agg[chiave]["kg"] += art["kg"]
        agg[chiave]["quantita"].extend(art["quantita"])
        if not agg[chiave]["confezionamento"] and art["confezionamento"]:
            agg[chiave]["confezionamento"] = art["confezionamento"]
    return agg


# ──────────────────────────────────────────────────────────────────────────────
# RICERCA PDF
# ──────────────────────────────────────────────────────────────────────────────

def _carica_rientri(data_attuale: str = None) -> dict:
    """
    Legge rientri_ddt.xlsx e restituisce un dizionario:
        { codice_lower: 'DD-MM-YYYY' }
    Colonna A = codice consegna (es. 'p1745')
    Colonna B = data DDT originale (datetime o stringa)
    Colonna C = stato: se contiene 'allegato' (senza 'lavorazione') -> già consegnato, salta.

    Il PDF fisico va cercato in:
        CONSEGNE_{data}/DDT-ORIGINALI-DIVISI/{FRUTTA o LATTE}/{codice}_{data}.pdf
    """
    if not RIENTRI_XLSX.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(RIENTRI_XLSX, read_only=True, data_only=True)
        ws = wb.active
        rientri = defaultdict(list)
        all_rientri_rows = [] # (r_idx, codice, data_str, status)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            codice_raw = row[0].value
            data_b = row[1].value
            if not codice_raw or not data_b:
                continue

            stato = str(row[2].value or "").strip().lower()
            codice = str(codice_raw).strip().lower()
            
            if hasattr(data_b, 'strftime'):
                data_str = data_b.strftime("%d-%m-%Y")
            else:
                data_str = str(data_b).strip()

            all_rientri_rows.append((r_idx, codice, data_str, stato))

            # Filtro per il caricamento nel dizionario di ricerca PDF
            if "allegato" in stato and "lavorazione" not in stato:
                if not data_attuale or data_attuale.lower() not in stato:
                    continue

            if codice and data_str:
                rientri[codice].append(data_str)
        wb.close()
        return rientri, all_rientri_rows
    except Exception as e:
        print(f"  WARN  Errore lettura rientri_ddt.xlsx: {e}")
        return {}, []

def _aggiorna_stato_rientri_excel(aggiornamenti: list):
    """aggiornamenti: list of (r_idx, nuovo_testo)"""
    if not aggiornamenti: return
    from openpyxl import load_workbook
    try:
        wb = load_workbook(RIENTRI_XLSX)
        ws = wb.active
        for r_idx, testo in aggiornamenti:
            ws.cell(row=r_idx, column=3).value = testo
        wb.save(RIENTRI_XLSX)
        print(f"   Excel Rientri aggiornato ({len(aggiornamenti)} modifiche).")
    except PermissionError:
        print(f"  WARN  ERRORE: Impossibile aggiornare Excel Rientri. Chiudi il file!")
    except Exception as e:
        print(f"  WARN  Errore salvataggio rientri: {e}")



def _trova_pdf(codice: str, data: str, cartella: Path) -> Path | None:
    """
    Cerca il PDF del cliente nella cartella specificata.
    Nome atteso: {codice}_{data}.pdf  (es. p2067_26-03-2026.pdf)

    Supporta cartelle multi-data (es. "30-03-2026_31-03-2026"):
      1. Prova la data esatta composta (es. p1745_30-03-2026_31-03-2026.pdf)
      2. Per cartelle multi-data, prova ogni singola data separata (es. p1745_30-03-2026.pdf)
      3. Fallback glob: qualsiasi file che inizia con {codice}_
    """
    if codice == CODICE_VUOTO or not codice:
        return None

    # 1. Ricerca esatta (funziona per date singole e come primo tentativo)
    p = cartella / f"{codice}_{data}.pdf"
    if p.exists():
        return p

    # 2. Se la data è multi-data (contiene "_" che separa due date DD-MM-YYYY_DD-MM-YYYY),
    #    prova ogni singola data estratta
    #    Formato atteso: "30-03-2026_31-03-2026" -> ['30-03-2026', '31-03-2026']
    parti_data = re.findall(r"\d{2}-\d{2}-\d{4}", data)
    if len(parti_data) > 1:
        for d in parti_data:
            p = cartella / f"{codice}_{d}.pdf"
            if p.exists():
                return p

    # 3. Fallback glob: cerca qualsiasi file che inizia con il codice
    matches = list(cartella.glob(f"{codice}_*.pdf"))
    if matches:
        return matches[0]

    return None


def _trova_cartella(data_arg: str | None) -> Path:
    if data_arg:
        if re.match(r"^\d{2}-\d{2}$", data_arg):
            data_arg = f"{data_arg}-2026"
        p = CONSEGNE_DIR / f"CONSEGNE_{data_arg}"
        if not p.exists():
            raise FileNotFoundError(f"Cartella non trovata: {p}")
        return p
    folders = sorted(
        [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")],
        key=lambda d: d.stat().st_ctime
    )
    if not folders:
        raise FileNotFoundError("Nessuna cartella CONSEGNE_* trovata.")
    return folders[-1]

def _blocco_distinta(viaggio: dict, articoli_viaggio: dict, data_ddt: str, copia: int, styles, colors, mm, Paragraph, Spacer, Table, TableStyle, PageBreak):
    """Restituisce la lista di elementi Flowable per una singola copia della distinta."""
    from reportlab.lib.styles import ParagraphStyle
    st_titolo = ParagraphStyle("titolo", parent=styles["Heading1"], fontSize=14, spaceAfter=3)
    st_sub    = ParagraphStyle("sub",    parent=styles["Normal"],   fontSize=9,  spaceAfter=2)
    st_body   = ParagraphStyle("body",   parent=styles["Normal"],   fontSize=8)
    st_warn   = ParagraphStyle("warn",   parent=styles["Normal"],   fontSize=8, textColor=colors.red)

    nome_giro = viaggio.get("nome_giro", "?")
    zone      = ", ".join(viaggio.get("zone", []))
    n_fermate = viaggio.get("num_fermate", 0)
    label     = f"{'COPIA AUTISTA' if copia == 1 else 'COPIA UFFICIO'}"
    elementi  = []

    # Intestazione
    elementi.append(Paragraph(f"DISTINTA DI CARICO — {nome_giro}  [{label}]", st_titolo))
    elementi.append(Paragraph(f"Zone: {zone}  |  Fermate: {n_fermate}  |  Data: {data_ddt}", st_sub))
    elementi.append(Paragraph("PERICOLO: Caricare nell'ordine inverso: l'ULTIMA fermata va caricata PER PRIMA.", st_warn))
    elementi.append(Spacer(1, 4*mm))

    # ── SEZIONE 1: ARTICOLI ──
    elementi.append(Paragraph("RIEPILOGO ARTICOLI DA CARICARE PER GIRO:", st_body))
    dati_art = [["Codice Articolo", "Descrizione Natura Qualità", "Quantità Consolidata", "Confezionamento"]]
    
    # Ordiniamo per (codice_base, variante_raw) — coerente con la chiave di aggregazione
    for chiave, art in sorted(articoli_viaggio.items(), key=lambda x: (x[0][0], x[0][1])):
        qty_cons, display = _consolida_quantita(art["codice_base"], art["quantita"])
        
        # Codice stampato = base + variante raw (es. "FVNS-03-" + "FOLDER" → "FVNS-03- FOLDER")
        variante = art.get("variante_raw", "")
        codice_stampato = f"{art['codice_base']} {variante}".strip() if variante else art["codice_base"]

        dati_art.append([
            codice_stampato,
            art.get("descrizione", "")[:55], # Aumentato limite caratteri visto lo spazio extra
            display or "—",
            art.get("confezionamento", "")[:30] or "—",
        ])
    ts_art = TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#10b981")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
        ("FONTSIZE",       (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
        ("GRID",           (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TEXTCOLOR",      (4, 1), (4, -1),  colors.red),
        ("LEFTPADDING",    (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 3),
    ])
    t_art = Table(dati_art, colWidths=[35*mm, 75*mm, 35*mm, 35*mm])
    t_art.setStyle(ts_art)
    elementi.append(t_art)
    elementi.append(Spacer(1, 10*mm))

    # ── SEZIONE 2: LISTA CLIENTI ──
    elementi.append(Paragraph("ORDINE DI CARICO (carica dal basso: N.1 = ultima fermata = primo da caricare):", st_body))
    fermate     = viaggio.get("lista_punti", [])
    fermate_inv = list(reversed(fermate))

    # colonne: #, Cod.F, Cod.L, Nome, Indirizzo
    dati_fermate = [["#", "Cod. F", "Cod. L", "Nome", "Indirizzo"]]
    for idx, f in enumerate(fermate_inv, 1):
        cf = f.get("codice_frutta", "") or ""
        cl = f.get("codice_latte",  "") or ""
        dati_fermate.append([
            str(idx),
            cf if cf != CODICE_VUOTO else "—",
            cl if cl != CODICE_VUOTO else "—",
            f.get("nome", "")[:40],
            f.get("indirizzo", "")[:50],
        ])
    ts_fermate = TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1e293b")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
        ("FONTSIZE",       (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID",           (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING",    (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 3),
    ])
    t_fermate = Table(dati_fermate, colWidths=[8*mm, 22*mm, 22*mm, 50*mm, 65*mm])
    t_fermate.setStyle(ts_fermate)
    elementi.append(t_fermate)
    
    return elementi


def _genera_distinta_pdf(viaggio: dict, articoli_viaggio: dict, out_path: Path, data_ddt: str, pdf_ddt: list[Path]):
    """Genera il PDF della distinta di carico per un viaggio in DOPPIA COPIA + DDT allegati x2."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    # --- Genera le due copie della distinta in un PDF temporaneo ---
    import tempfile, os
    tmp = out_path.parent / (out_path.stem + "_TMP.pdf")

    doc = SimpleDocTemplate(
        str(tmp), pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm
    )
    styles = getSampleStyleSheet()

    elementi = []
    # Copia 1 — AUTISTA
    elementi += _blocco_distinta(viaggio, articoli_viaggio, data_ddt, 1, styles, colors, mm, Paragraph, Spacer, Table, TableStyle, PageBreak)
    elementi.append(PageBreak())
    # Copia 2 — UFFICIO
    elementi += _blocco_distinta(viaggio, articoli_viaggio, data_ddt, 2, styles, colors, mm, Paragraph, Spacer, Table, TableStyle, PageBreak)

    doc.build(elementi)

    # --- Assembla fascicoli: [Distinta Copy 1 + DDTs] + [Distinta Copy 2 + DDTs] ---
    # I DDT vengono allegati in ordine INVERSO rispetto al percorso:
    # -> Nel PDF: stop N, stop N-1, ..., stop 1
    # -> Dopo la stampa (i fogli escono impilati): stop 1 è in CIMA alla pila OK
    # Così l'autista trova subito il DDT della prima consegna senza sfogliare.
    pdf_ddt_inv = list(reversed(pdf_ddt))

    try:
        from pypdf import PdfWriter, PdfReader
        writer = PdfWriter()
        reader_tmp = PdfReader(str(tmp))

        # ── FIX DOPPIA COPIA MULTI-PAGINA ────────────────────────────────────
        # Il PDF temporaneo contiene: [Copia1_pag1, Copia1_pagN, PageBreak, Copia2_pag1, Copia2_pagN]
        # Le due copie occupano lo stesso numero di pagine (n_pagine_copia).
        # Per trovare il confine, le pagin totali si dividono in metà.
        n_tot = len(reader_tmp.pages)
        n_per_copia = n_tot // 2  # ogni copia occupa esattamente metà delle pagine totali

        # Copia 1 (AUTISTA): pagine da 0 a n_per_copia-1
        for i in range(n_per_copia):
            writer.add_page(reader_tmp.pages[i])

        # Copia 2 (UFFICIO): pagine da n_per_copia in poi
        for i in range(n_per_copia, n_tot):
            writer.add_page(reader_tmp.pages[i])

        # 2. Ogni DDT viene aggiunto in doppia copia consecutiva
        for pdf in pdf_ddt_inv:
            writer.append(str(pdf)) # Prima copia (autista)
            writer.append(str(pdf)) # Seconda copia (ufficio)

        with open(out_path, "wb") as f:
            writer.write(f)
        tmp.unlink(missing_ok=True)
    except Exception as e:
        # Fallback se pypdf fallisce
        import shutil
        shutil.move(str(tmp), str(out_path))
        print(f"    WARN  Errore assemblaggio pypdf: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# RIEPILOGO ZONE
# ──────────────────────────────────────────────────────────────────────────────

def _genera_pagina_riepilogo_zone(viaggi: list, out_path: Path, data_ddt: str) -> Path | None:
    """
    Genera un PDF a COPIA UNICA con l'elenco di tutte le zone trovate
    nei DDT di tutti i viaggi del giorno. Va come prima pagina del Master PDF.
    Contiene: elenco zone in grande (leggibile a colpo d'occhio) + tabella
    riepilogativa Giro -> Zone.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        return None

    # Raccoglie zone filtrando valori vuoti o privi di senso (es. "B" singola lettera)
    def _zona_valida(z: str) -> bool:
        return len(z.strip()) >= 2

    def _zona_base(z: str) -> str:
        """Rimuove suffissi/prefissi non numerici: '3109_B' → '3109', '!_3109' → '3109'."""
        z = z.strip()
        z = re.sub(r'^[^0-9]+', '', z)  # Rimuove prefissi non numerici
        z = re.sub(r'[^0-9]+$', '', z)  # Rimuove suffissi non numerici
        return z

    tutte_le_zone: set[str] = set()   # set con codici BASE (solo numeri) per elenco grande
    giri_con_zone: list[tuple[str, list[str]]] = []
    for v in viaggi:
        zone_v = [z for z in v.get("zone", []) if _zona_valida(z)]
        nome_v = v.get("nome_giro", "?")
        giri_con_zone.append((nome_v, zone_v))           # tabella: codici completi (con B)
        tutte_le_zone.update(_zona_base(z) for z in zone_v if _zona_base(z))  # elenco grande: pulito

    try:
        doc = SimpleDocTemplate(
            str(out_path), pagesize=A4,
            leftMargin=20*mm, rightMargin=20*mm,
            topMargin=20*mm, bottomMargin=20*mm
        )
        styles = getSampleStyleSheet()
        st_titolo = ParagraphStyle("zt", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
        st_sub    = ParagraphStyle("zs", parent=styles["Normal"],   fontSize=10, spaceAfter=4,
                                   textColor=colors.HexColor("#475569"))
        st_zona   = ParagraphStyle("zz", parent=styles["Normal"],   fontSize=16,
                                   spaceBefore=6, spaceAfter=6,
                                   leading=22,
                                   textColor=colors.HexColor("#1e293b"),
                                   fontName="Helvetica-Bold")

        elementi = []
        elementi.append(Paragraph(f"RIEPILOGO ZONE — {data_ddt}", st_titolo))
        elementi.append(Paragraph("Zone coperte da tutti i giri di oggi:", st_sub))
        elementi.append(Spacer(1, 8*mm))

        # Elenco zone in grande (leggibile a colpo d'occhio)
        for zona in sorted(tutte_le_zone):
            elementi.append(Paragraph(f"&#x25cf;  {zona}", st_zona))

        elementi.append(Spacer(1, 12*mm))
        elementi.append(Paragraph("— Dettaglio per giro:", st_sub))
        elementi.append(Spacer(1, 4*mm))

        # Tabella riepilogativa: Giro | Zone
        dati_tab = [["Giro", "Zone"]]
        for nome_v, zone_v in giri_con_zone:
            zone_display = ", ".join(sorted(zone_v)) if zone_v else "—"
            dati_tab.append([nome_v, zone_display])

        ts = TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1e293b")),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTSIZE",       (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("LEFTPADDING",    (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
        ])
        t = Table(dati_tab, colWidths=[70*mm, 100*mm])
        t.setStyle(ts)
        elementi.append(t)

        doc.build(elementi)
        return out_path
    except Exception as e:
        print(f"  WARN  Errore generazione riepilogo zone: {e}")
        return None


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    try:
        import pdfplumber
        from reportlab.lib.pagesizes import A4
    except ImportError as e:
        print(f"ERR Libreria mancante: {e}  ->  pip install pdfplumber reportlab")
        sys.exit(1)

    data_arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    try:
        cartella = _trova_cartella(data_arg)
    except FileNotFoundError as e:
        print(f"ERR {e}")
        sys.exit(1)

    data_ddt = cartella.name.replace("CONSEGNE_", "")
    json_path = cartella / "viaggi_giornalieri_OTTIMIZZATO.json"

    if not json_path.exists():
        print(f"ERR File non trovato: {json_path.name}")
        print("   Esegui prima: py 8_genera_json_ottimizzato.py")
        sys.exit(1)

    viaggi = json.loads(json_path.read_text(encoding="utf-8"))
    viaggi = [v for v in viaggi if v.get("id_zona", "") != "DDT_DA_INSERIRE"]
    divisi_dir   = cartella / "DDT-ORIGINALI-DIVISI"
    dir_frutta   = divisi_dir / "FRUTTA"
    dir_latte    = divisi_dir / "LATTE"
    out_dir      = cartella / "DISTINTE_VIAGGIO"
    out_dir.mkdir(exist_ok=True)

    print(f"\n{'='*65}")
    print(f"  9_GENERA_DISTINTE - {data_ddt}  ({len(viaggi)} giri)")
    print(f"{'='*65}\n")

    # Carica mappa rientri: { codice: [date...] } e lista righe
    rientri, all_rientri_rows = _carica_rientri(data_ddt)
    if rientri:
        print(f"  [RIENTRI] Caricati da rientri_ddt.xlsx: {len(rientri)} codici mappati")

    # Traccia i PDF usati (per verifica orfani e finalizzazione stati)
    pdf_usati: set[Path] = set()
    rientri_usati: set[tuple] = set() # (codice, data_str)
    pdf_generati: list[Path] = []

    for viaggio in viaggi:
        nome_giro = viaggio["nome_giro"]
        zone      = ", ".join(viaggio.get("zone", []))
        punti     = viaggio.get("lista_punti", [])
        data_v    = viaggio.get("data_ddt", data_ddt)

        print(f"  [GIRO] {nome_giro} (zone: {zone}) - {len(punti)} fermate")

        articoli_giro: list[dict] = []
        pdf_non_trovati: list[str] = []
        pdf_usati_viaggio: list[Path] = []  # PDF di questo viaggio specifico
        zone_punti: set[str] = set()        # Zone raccolte dai singoli punti

        for punto in punti:
            cf   = punto.get("codice_frutta", "") or ""
            cl   = punto.get("codice_latte",  "") or ""
            d_p  = punto.get("data_consegna", data_v) or data_v
            nome = punto.get("nome", "?")[:40]
            z = punto.get("zona", "").strip()
            if z:
                zone_punti.add(z)
                # Deduce la zona equivalente per il latte o frutta in base ai prodotti richiesti
                if cl and cl.lower() != "p00000" and z.startswith("3"):
                    zone_punti.add("4" + z[1:])
                if cf and cf.lower() != "p00000" and z.startswith("4"):
                    zone_punti.add("3" + z[1:])

            # Cerca nella cartella corretta per data (supporta rientri di altre date)
            if d_p != data_v:
                cartella_r   = CONSEGNE_DIR / f"CONSEGNE_{d_p}"
                dir_frutta_r = cartella_r / "DDT-ORIGINALI-DIVISI" / "FRUTTA"
                dir_latte_r  = cartella_r / "DDT-ORIGINALI-DIVISI" / "LATTE"
            else:
                dir_frutta_r = dir_frutta
                dir_latte_r  = dir_latte

            for codice, tipo in [(cf, "FRUTTA"), (cl, "LATTE")]:
                if codice == CODICE_VUOTO or not codice:
                    continue

                # Se si tratta di un punto con data storica (rientro), aggiungi solo
                # se il singolo codice è tra quelli da rientrare su rientri_ddt.xlsx.
                if d_p != data_v and codice.lower() not in rientri:
                    # Questo previene l'aggiunta automatica di codici Latte non necessari
                    # recuperati via Mappatura Master.
                    continue

                pdfs_da_processare = []

                if codice.lower() in rientri:
                    # ── RIENTRO: cerca nei PDF delle date storiche (col. B) ──
                    date_rientro = rientri[codice.lower()]
                    for d_r in date_rientro:
                        cart_storica = CONSEGNE_DIR / f"CONSEGNE_{d_r}" / "DDT-ORIGINALI-DIVISI"
                        for sotto in ["FRUTTA", "LATTE"]:
                            pdf_found = _trova_pdf(codice, d_r, cart_storica / sotto)
                            if pdf_found:
                                pdfs_da_processare.append((pdf_found, sotto, d_r))
                                break
                else:
                    # ── CONSEGNA NORMALE: cerca nella cartella della sessione corrente ──
                    cart_tipo = dir_frutta_r if tipo == "FRUTTA" else dir_latte_r
                    pdf_found = _trova_pdf(codice, d_p, cart_tipo)
                    if pdf_found:
                        pdfs_da_processare.append((pdf_found, tipo, d_p))

                if not pdfs_da_processare:
                    pdf_non_trovati.append(f"{codice} ({tipo})")
                    print(f"       !! {nome:<40} {codice} ({tipo}) -> PDF non trovato")
                else:
                    for pdf_obj, tp, d_r in pdfs_da_processare:
                        pdf_usati.add(pdf_obj)
                        pdf_usati_viaggio.append(pdf_obj)
                        articoli = _raccogli_articoli_da_pdf(pdf_obj, tp)
                        articoli_giro.extend(articoli)
                        n_art = len(articoli)
                        is_rientro = (codice.lower() in rientri) and (d_r != d_p)
                        tag = f" [RIENTRO<-{d_r}]" if is_rientro else ""
                        print(f"       OK {nome:<40} {codice} ({tp}){tag} -> {n_art} art.")
                        if is_rientro:
                            rientri_usati.add((codice.lower(), d_r))

        # Aggrega articoli del viaggio
        articoli_agg = _aggrega_articoli(articoli_giro)
        print(f"       -> Totale articoli distinti: {len(articoli_agg)}")

        # Genera PDF distinta
        sanitized = re.sub(r'[\\/*?:"<>|]', '_', nome_giro)
        # Zone = unione di quelle nel JSON + quelle dai punti reali (garanzia completezza)
        tutte_zone = sorted(set(viaggio.get("zone", [])) | zone_punti)
        zone_str   = "_".join(tutte_zone)
        # Aggiorna anche il dict viaggio in memoria (usato da _blocco_distinta e riepilogo)
        viaggio["zone"] = tutte_zone
        pdf_name  = f"DISTINTA_{sanitized}_Zone_{zone_str}.pdf"
        out_pdf   = out_dir / pdf_name

        try:
            _genera_distinta_pdf(viaggio, articoli_agg, out_pdf, data_ddt, list(pdf_usati_viaggio))
            pdf_generati.append(out_pdf)
            print(f"       DOC Salvato: {pdf_name} (doppia copia + {len(pdf_usati_viaggio)} DDT x2)\n")
        except Exception as e:
            print(f"       ERR Errore PDF: {e}\n")

    # ── Genera Pagina Riepilogo Zone (COPIA UNICA, va in TESTA al Master) ──
    riepilogo_zone_path = out_dir / f"00_RIEPILOGO_ZONE_{data_ddt}.pdf"
    rz = _genera_pagina_riepilogo_zone(viaggi, riepilogo_zone_path, data_ddt)
    if rz:
        print(f"  ZON Riepilogo zone: {riepilogo_zone_path.name}")

    # ── Assembla Master PDF (Riepilogo Zone + tutte le distinte) ──
    if pdf_generati:
        try:
            from pypdf import PdfWriter
            master_path = cartella / f"MASTER_DISTINTE_{data_ddt}.pdf"
            writer = PdfWriter()
            # 1. Prima pagina: riepilogo zone (copia unica)
            if rz and riepilogo_zone_path.exists():
                writer.append(str(riepilogo_zone_path))
            # 2. Poi tutte le distinte viaggio (doppia copia ciascuna)
            for p in pdf_generati:
                writer.append(str(p))
            with open(master_path, "wb") as f:
                writer.write(f)
            print(f"  BIB Master PDF: {master_path.name}")
        except ImportError:
            print("  WARN  pypdf non installato — Master PDF non generato (pip install pypdf)")
        except Exception as e:
            print(f"  WARN  Errore Master PDF: {e}")

    # ── Verifica orfani ──
    print(f"\n  FIN Verifica DDT orfani...")
    tutti_pdf: list[Path] = []
    for d in [dir_frutta, dir_latte]:
        if d.exists():
            tutti_pdf.extend(d.glob("*.pdf"))

    orfani = [p for p in tutti_pdf if p not in pdf_usati]
    if orfani:
        print(f"  !! {len(orfani)} PDF non assegnati a nessun viaggio:")
        for p in sorted(orfani):
            print(f"       - {p.parent.name}/{p.name}")
    else:
        print(f"  OK Tutti i PDF DDT sono stati assegnati a un viaggio.")

    # ── Finalizzazione stati Rientri (Punto 6 workflow) ──
    print(f"\n  FIN Finalizzazione stati in rientri_ddt.xlsx...")
    aggiornamenti_excel = []
    for r_idx, cod, d_str, stato in all_rientri_rows:
        is_usato = (cod, d_str) in rientri_usati
        
        # Se era in lavorazione
        if "lavorazione" in stato:
            if is_usato:
                # Promosso ad allegato
                aggiornamenti_excel.append((r_idx, f"allegato DDT {data_ddt}"))
            else:
                # Rimandato: sbianca la cella
                aggiornamenti_excel.append((r_idx, ""))
        
        # Se non era in lavorazione ma è stato usato comunque oggi (es. abbinamento automatico)
        elif is_usato and "allegato" not in stato:
            aggiornamenti_excel.append((r_idx, f"allegato DDT {data_ddt}"))

    if aggiornamenti_excel:
        _aggiorna_stato_rientri_excel(aggiornamenti_excel)
    else:
        print("    OK Nessuno stato da aggiornare.")

    # ── Riepilogo finale ──
    print(f"\n{'='*65}")
    print(f"  OK COMPLETATO!")
    print(f"     Distinte generate: {len(pdf_generati)}")
    print(f"     Cartella output:   {out_dir.name}")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
