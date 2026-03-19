#!/usr/bin/env python3
import sys
from pathlib import Path
import re
from openpyxl import load_workbook
from crea_tabella_aggiornamento_articoli import (
    _estrai_codice, 
    _descrizione_minima, 
    CONSOLIDAMENTO, 
    _estrai_porzioni_per_unita,
    _raccogli_dati_da_pdf
)

EXCEL_PATH = Path(r"c:\Gestione DDT viaggi\tabella_aggiornamento_articoli.xlsx")

def read_excel_articles():
    if not EXCEL_PATH.exists():
        return {}
    
    articles = {}
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        # Skip header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            cod = str(row[0]).strip()
            articles[cod] = {
                'desc': str(row[1] or "").strip(),
                'conf': str(row[2] or "").strip(),
                'ratio': str(row[6] or "").strip(), # Column G
                'porz_unita': str(row[7] or "").strip(), # Column H
            }
        return articles
    except Exception as e:
        print(f"Errore lettura Excel: {e}")
        return {}

def main():
    print("Confronto articoli estratti dai PDF con articoli in tabella_aggiornamento_articoli.xlsx\n")
    
    # 1. Leggi Excel esistente
    excel_articles = read_excel_articles()
    print(f"Articoli in Excel: {len(excel_articles)}")
    
    # 2. Estrai dai PDF
    dati_pdf = _raccogli_dati_da_pdf()
    pdf_articles = {}
    for cell0, cell1, cell5 in dati_pdf:
        cod = _estrai_codice(cell0)
        if not cod or not re.search(r"[A-Z0-9]{2,}-[A-Z0-9\-]+", cod):
            continue
        desc = _descrizione_minima(cell1)
        conf = (cell5 or "").strip()
        
        if cod not in pdf_articles or len(desc) > len(pdf_articles[cod]['desc']):
            # Calcola porz_unita o ratio
            ratio = ""
            porz_unita = ""
            if cod in CONSOLIDAMENTO:
                up, us, r = CONSOLIDAMENTO[cod]
                ratio = str(r)
            else:
                porz_unita = _estrai_porzioni_per_unita(conf)
                if cod == "10-GEL" and not porz_unita:
                    porz_unita = "1"
            
            pdf_articles[cod] = {
                'desc': desc,
                'conf': conf,
                'ratio': ratio,
                'porz_unita': porz_unita
            }
    
    print(f"Articoli estratti dai PDF: {len(pdf_articles)}\n")
    
    # 3. Confronto
    all_codes = sorted(set(excel_articles.keys()) | set(pdf_articles.keys()))
    
    diffs = []
    nuovi = []
    mancanti = []
    
    for cod in all_codes:
        in_excel = cod in excel_articles
        in_pdf = cod in pdf_articles
        
        if in_excel and in_pdf:
            # Confronto caratteristiche
            ex = excel_articles[cod]
            pd = pdf_articles[cod]
            
            # Normalizziamo per il confronto (es. ratio stringhe vs numeri)
            d_desc = ex['desc'] != pd['desc']
            # Per il ratio/porz_unita, confrontiamo come stringhe pulite
            d_ratio = str(ex['ratio']).replace(".0", "") != str(pd['ratio']).replace(".0", "")
            d_porz = str(ex['porz_unita']).replace(".0", "") != str(pd['porz_unita']).replace(".0", "")
            
            if d_desc or d_ratio or d_porz:
                diffs.append((cod, ex, pd))
        elif in_pdf:
            nuovi.append(cod)
        else:
            mancanti.append(cod)
            
    # Output
    print("-" * 40)
    print(f"NUOVI ARTICOLI (nei PDF ma non in Excel): {len(nuovi)}")
    for c in nuovi:
        print(f"  + {c}: {pdf_articles[c]['desc']}")
        
    print(f"\nARTICOLI MANCANTI (in Excel ma non nei PDF odierni): {len(mancanti)}")
    for c in mancanti:
        print(f"  - {c}: {excel_articles[c]['desc']}")
        
    print(f"\nDIFFERENZE CARATTERISTICHE (per codici esistenti): {len(diffs)}")
    for cod, ex, pd in diffs:
        print(f"  * {cod}:")
        if ex['desc'] != pd['desc']:
            print(f"    Descrizione: '{ex['desc']}' -> '{pd['desc']}'")
        if str(ex['ratio']).replace(".0", "") != str(pd['ratio']).replace(".0", ""):
            print(f"    Ratio: '{ex['ratio']}' -> '{pd['ratio']}'")
        if str(ex['porz_unita']).replace(".0", "") != str(pd['porz_unita']).replace(".0", ""):
            print(f"    Porz/Unità: '{ex['porz_unita']}' -> '{pd['porz_unita']}'")

if __name__ == "__main__":
    main()
