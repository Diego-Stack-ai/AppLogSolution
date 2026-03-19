#!/usr/bin/env python3
"""
Crea la Distinta di carico per il magazziniere.
Per OGNI PDF nella cartella DDT-[data] (es. 3101.pdf, 3102.pdf), estrae gli articoli
e genera una distinta. Le distinte vengono salvate in DDT-[data]/RIEPILOGO/.
Output: lista merce nel formato del modello "da usare", senza elenco bolle.
"""

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from decimal import Decimal

try:
    import pdfplumber
except ImportError:
    raise ImportError("pip install pdfplumber")

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    pass  # usato solo per verifica codici

try:
    from pypdf import PdfWriter, PdfReader
except ImportError:
    PdfWriter = PdfReader = None  # per filtro e unione DDT

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from xml.sax.saxutils import escape as _xml_escape
except ImportError:
    raise ImportError("pip install reportlab")

# Funziona sia come script sia come exe (PyInstaller): cartella principale = Gestione DDT viaggi
BASE_DIR = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent.parent
PROGRAMMA_DIR = BASE_DIR / "Programma"
DDT_FRUTTA_DIR = BASE_DIR / "DDT frutta"
DDT_LATTE_DIR = BASE_DIR / "DDT latte"
GIRI_LAVORATI_DIR = BASE_DIR / "Giri lavorati"
MAPPATURA_XLSX = BASE_DIR / "mappatura_destinazioni.xlsx"
NUOVI_CODICI_XLSX = BASE_DIR / "nuovi_codici_consegna.xlsx"
REPORT_ORARI_MANCANTI_XLSX = BASE_DIR / "report_orari_mancanti.xlsx"
RIENTRI_DDT_XLSX = BASE_DIR / "rientri_ddt.xlsx"
REPORT_RIENTRI_NON_INTEGRABILI_XLSX = BASE_DIR / "report_rientri_non_integrabili.xlsx"
CODICE_CIVETTA = "p00000"


def _crea_struttura_cartelle():
    """Crea la struttura cartelle se non esiste (utile per exe su PC nuovi)."""
    for d in (DDT_FRUTTA_DIR, DDT_LATTE_DIR, GIRI_LAVORATI_DIR, PROGRAMMA_DIR):
        d.mkdir(parents=True, exist_ok=True)

DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)

# Codice LUOGO (Luogo di destinazione): p/P + 4 o 5 cifre (es. p2731, P4848)
LUOGO_RE = re.compile(r'[Ll]uogo [Dd]i [Dd]estinazione:\s*([pP]\d{4,5})')

# Codice CAUSALE (trasporto): lettera MAIUSCOLA + 4 cifre (es. A3101, P3101)
# Con orari: H10 = orario max 10:00; numero 730/800/715 = orario min 07:30/08:00/07:15
# Cercato SOLO dopo "conto di" o "ordine e conto di"
CAUSALE_SEZIONE_MARKER = "CAUSALE DEL TRASPORTO"
CAUSALE_SEZIONE_LUNGHEZZA = 150
CAUSALE_RE = re.compile(
    r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})(?:\s+H(\d{2}))?(?:\s+(\d{3}))?',
    re.I
)
PROVINCIA_RE = re.compile(r'\(([A-Z]{2})\)')

# Colonne mappatura (nuova struttura unificata):
# A=Codice Frutta, B=Codice Latte, C-N=Indirizzo unico (A chi va, Tipologia, Indirizzo, CAP, Città, Provincia, ... Orario min, Orario max)
COL_CODICE_FRUTTA = 1   # A
COL_CODICE_LATTE = 2    # B
COL_ORARIO_MIN = 11     # K
COL_ORARIO_MAX = 12     # L
ORARIO_MAX_DEFAULT = "14:00"  # chiusura scuole se H10 non presente

# Codice DDT -> (codice distinta, descrizione distinta)
CODICE_MAP = {
    "FVNS-03-POSTER": ("10-MANIFESTO", "Manifesto programma"),
}

# Articoli gia noti (con descrizione); se nuovo, viene segnalato in distinta
ARTICOLI_NOTI = frozenset({
    "ME-T-DI-V0-NA", "PE-T-DI-L3-NA", "10-GEL", "10-FLYER", "10-MANIFESTO", "LT-DL-02-LC", "LT-ES-04-LS",
    "LT-ESL-IN-LB", "LT-AQ-04-LV", "YO-BI-MN-04-LB", "YO-DL-02-LC", "AP-SU-PC", "FO-DI-PV-04-LB",
})
# Unita aggiunte di recente: segnalate in distinta per attenzione magazziniere
UNITA_NUOVE = frozenset({"Collo"})  # Fette/Fetta ora considerate unità standard

# Articoli con consolidamento multi-unita: (unita_principale, unita_secondaria, ratio)
# ratio = quante unita_secondarie in 1 unita_principale
CONSOLIDAMENTO = {
    "LT-ES-04-LS": ("Fardelli", "Bottiglie", 10),
    "LT-ESL-IN-LB": ("Fardelli", "Bottiglie", 6),
    "YO-BI-MN-04-LB": ("Cartoni", "Cluster", 10),
    "YO-DL-02-LC": ("Cartoni", "Porzioni", 6),
    "AP-SU-PC": ("Cartoni", "Porzioni", 24),
}


def _estrai_data_e_territori(pdf_paths: list[Path]) -> tuple[str | None, set[str]]:
    """Estrae data e territori unici dai PDF."""
    data_ddt = None
    territori = set()
    for pdf_path in pdf_paths:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if not data_ddt:
                    m = DATA_DDT_RE.search(text)
                    if m:
                        data_ddt = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
                sezione_causale = _estrai_sezione_causale(text)
                for match in CAUSALE_RE.finditer(sezione_causale):
                    territorio = match.group(1)[1:5]
                    territori.add(territorio)
    return data_ddt, territori


# Unita riconosciute nelle quantita (include minuscole per DDT con formato variabile)
UNITA_QTY = r"(Confezioni|Confezione|confezioni|confezione|Colli|Collo|colli|collo|Brick|brick|Fardelli|Fardello|fardelli|fardello|Bottiglie|Bottiglia|bottiglie|bottiglia|Cartoni|Cartone|cartoni|cartone|Cluster|cluster|Porzioni|Porzione|porzioni|porzione|Fascette|Fascetta|fascette|fascetta|Manifesti|Manifesto|manifesti|manifesto|Fette|Fetta|fette|fetta|pz)"

# Confezionamento: X Porzioni/Bottiglie/Cluster/Fetta / Unit
CONF_PATTERN = r"[\d,\.]+\s*(?:Porzioni?|Bottiglie?|Cluster|Fetta?)\s*/\s*\w+"

SCAD_RE = re.compile(r"Scad\.\s*min\.\s*(\d{2}/\d{2}/\d{4})", re.I)


def _parse_quantita_da_cella(cell) -> list[tuple[int, str]]:
    """
    Estrae lista (qty, unita) da cella quantità (Col 3).
    Es: "1 Brick" -> [(1, "Brick")]; "2 Fardelli\\ne 8 Bottiglie" -> [(2, "Fardelli"), (8, "Bottiglie")]
    """
    if not cell or not str(cell).strip():
        return []
    text = str(cell).replace("\n", " ").replace("  ", " ")
    quantita = []
    for m in re.finditer(r"(?:^|e\s+)(\d+)\s+(" + UNITA_QTY + r")", text, re.I):
        quantita.append((int(m.group(1)), _normalizza_unita(m.group(2))))
    if not quantita and re.search(r"^(\d+)\s*$", text.strip()):
        quantita.append((int(text.strip()), "pz"))
    return quantita


def _parse_riga_prodotto(line: str, next_lines: list[str], codice: str) -> dict | None:
    """
    Estrae da una riga prodotto: kg, quantita (lista di (num, unita)), confezionamento.
    next_lines: righe successive (per "e X Unit" nei multi-unit).
    """
    kg = Decimal("0")
    quantita = []
    confezionamento = ""

    # 10-GEL: "0 6 6" - quantita 6, no confezionamento
    gel_match = re.search(r'10-GEL.*?0\s+(\d+)\s+\d+', line)
    if gel_match:
        return {"kg": Decimal("0"), "quantita": [(int(gel_match.group(1)), "pz")], "confezionamento": ""}

    # Cerca "X Unit" + "porzioni_effettive" + confezionamento (flessibile: Porzioni, Bottiglie, Cluster)
    # Es: "0 2 Confezioni 100 50 Porzioni / Confezione"
    # Es: "16,8 2 Fardelli 112 6 Bottiglie / Fardello"
    # Es: "13,5 5 Cartoni 108 10 Cluster / Cartone"
    qty_match = re.search(
        r'(\d+[,\.]?\d*)\s+(\d+)\s+' + UNITA_QTY + r'\s+\d+\s+(' + CONF_PATTERN + r')',
        line
    )
    if qty_match:
        kg = Decimal(qty_match.group(1).replace(",", "."))
        quantita.append((int(qty_match.group(2)), qty_match.group(3)))
        confezionamento = qty_match.group(4).replace(",", ",")

    # Pattern alternativo
    if not quantita:
        alt = re.search(r'(\d+)\s+' + UNITA_QTY + r'\s+\d+\s+(' + CONF_PATTERN + r')', line)
        if alt:
            quantita.append((int(alt.group(1)), alt.group(2)))
            confezionamento = alt.group(3).replace(",", ",")

    # Cerca "e X Unit" nelle righe successive (multi-unit)
    for nl in next_lines[:3]:
        e_match = re.search(r'[eE]\s+(\d+)\s+(\w+)\s+(' + CONF_PATTERN + r')', nl)
        if e_match:
            quantita.append((int(e_match.group(1)), e_match.group(2)))
            if not confezionamento:
                confezionamento = e_match.group(3).replace(",", ",")
            break
        e_porz = re.search(r'[eE]\s+(\d+)\s+porzioni?', nl, re.I)
        if e_porz:
            quantita.append((int(e_porz.group(1)), "Porzioni"))
            break

    if not quantita:
        return None
    return {"kg": kg, "quantita": quantita, "confezionamento": confezionamento}


def _estrai_articoli_da_pagina(lines: list[str], tipo: str) -> list[dict]:
    """Estrae righe articolo da una pagina DDT."""
    articoli = []
    i = 0
    # Salta fino all'header prodotti
    while i < len(lines) and "Cod. Articolo" not in lines[i] and "Confezionamento" not in lines[i]:
        i += 1
    i += 2  # salta header

    code_pattern = re.compile(
        r'^([A-Z0-9]{2,}-[A-Z0-9-]+|FVNS-\d+-)\s*'  # codice
    )

    while i < len(lines):
        line = lines[i]
        if "________________________________________________________________" in line or "Scadenza" in line:
            break

        # Cerca codice prodotto all'inizio riga (include FVNS-03- con dash finale)
        code_m = re.match(r'^([A-Z0-9]{2,}-[A-Z0-9\-]*)', line)
        if not code_m or len(code_m.group(1)) < 4:
            # FVNS-03- può avere POSTER su riga successiva
            if i > 0 and "POSTER" in line and "FVNS" in lines[i-1]:
                i += 1
                continue
            i += 1
            continue

        codice_raw = code_m.group(1).strip()
        # Completa FVNS-03- -> FVNS-03-POSTER
        if codice_raw == "FVNS-03-" or codice_raw.startswith("FVNS-"):
            codice_raw = "FVNS-03-POSTER"

        # Normalizza
        codice, _ = CODICE_MAP.get(codice_raw, (codice_raw, ""))

        # Estrai righe successive (ferma al prossimo codice prodotto inizio riga)
        next_lines = []
        for k in range(i + 1, min(i + 6, len(lines))):
            stripped = lines[k].strip()
            if stripped and re.match(r'^[A-Z0-9]{2,}-[A-Z0-9\-]', stripped):
                break
            next_lines.append(lines[k])
        parsed = _parse_riga_prodotto(line, next_lines, codice)

        if not parsed:
            i += 1
            continue

        # Descrizione completa dal DDT (linea + righe successive: nome, Data distribuzione, quantità)
        # Ferma a "Codice:" (es. Codice: 1-2- 11/03/2026) per evitare di includere colonne successive
        blocco = " ".join([line] + next_lines)
        desc_match = re.search(r'^[A-Z0-9\-]+\s+(.+?)(?=\s+Codice:|\Z)', blocco, re.DOTALL)
        descrizione = desc_match.group(1).strip() if desc_match else ""
        descrizione = " ".join(descrizione.split())  # normalizza spazi
        if not descrizione:
            descrizione = codice

        # Scadenza: Scad. min. DD/MM/YYYY (per raggruppamento: stesso codice + stessa scadenza = somma)
        scad_match = re.search(r'Scad\.\s*min\.\s*(\d{2}/\d{2}/\d{4})', blocco, re.I)
        scadenza = scad_match.group(1) if scad_match else ""

        articoli.append({
            "codice": codice,
            "descrizione": descrizione,
            "scadenza": scadenza,
            "kg": parsed["kg"],
            "quantita": parsed["quantita"],
            "confezionamento": parsed["confezionamento"],
        })
        i += 1

    return articoli


def _raccogli_articoli_da_pdf(pdf_paths: list[Path], tipo: str) -> list[dict]:
    """Raccoglie tutti gli articoli da una lista di PDF. Prova prima extract_tables(), poi fallback a extract_text()."""
    tutti = []
    for pdf_path in pdf_paths:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                arts = _estrai_articoli_da_tabella(page)
                if arts is None:
                    text = page.extract_text() or ""
                    lines = text.split("\n")
                    arts = _estrai_articoli_da_pagina(lines, tipo)
                tutti.extend(arts or [])
    return tutti


def _normalizza_unita(u: str) -> str:
    """Normalizza varianti plurali (Collo->Colli, Fetta->Fette, Bottiglia->Bottiglie, etc.)."""
    u = u.strip().lower()
    mapping = {"bottiglia": "Bottiglie", "bottiglie": "Bottiglie", "fardello": "Fardelli", "fardelli": "Fardelli",
               "cartone": "Cartoni", "cartoni": "Cartoni", "cluster": "Cluster", "porzione": "Porzioni",
               "porzioni": "Porzioni", "collo": "Colli", "colli": "Colli", "fetta": "Fette", "fette": "Fette", "brick": "Brick",
               "confezione": "Confezioni", "confezioni": "Confezioni", "manifesto": "Manifesti",
               "manifesti": "Manifesti", "fascetta": "Fascette", "fascette": "Fascette", "pz": "pz", "confezioni": "Confezioni"}
    return mapping.get(u, u.title() if u else u)


def _estrai_articoli_da_tabella(page) -> list[dict] | None:
    """
    Estrae articoli usando extract_tables(). Colonne separate = descrizione e confezionamento puliti.
    Ritorna lista articoli o None se tabella non trovata.
    """
    tables = page.extract_tables()
    if not tables:
        return None
    tab_articoli = None
    for t in tables:
        if not t or len(t) < 2:
            continue
        header = " ".join(str(c or "") for c in (t[0] or []))
        if "Cod. Articolo" in header or "Cod. Articolo" in str(t[0]):
            tab_articoli = t
            break
    if not tab_articoli:
        return None

    articoli = []
    for row in tab_articoli[1:]:
        if not row or len(row) < 4:
            continue
        cell0 = str(row[0] or "").strip()
        if not cell0:
            continue

        codice_raw = None
        for linea in cell0.split("\n"):
            linea = linea.strip()
            if linea.startswith("Codice:"):
                continue
            if re.match(r"^[A-Z0-9]{2,}-[A-Z0-9\-]+", linea):
                codice_raw = re.match(r"^([A-Z0-9]{2,}-[A-Z0-9\-]+)", linea).group(1).strip()
                break
        if not codice_raw:
            continue
        if codice_raw == "FVNS-03-" or (codice_raw.startswith("FVNS-") and "POSTER" not in codice_raw):
            codice_raw = "FVNS-03-POSTER"
        codice, _ = CODICE_MAP.get(codice_raw, (codice_raw, ""))

        try:
            kg_val = str(row[2] or "0").replace(",", ".").strip()
            kg = Decimal(kg_val) if kg_val else Decimal("0")
        except Exception:
            kg = Decimal("0")

        cell3 = row[3] if len(row) > 3 else ""
        quantita = _parse_quantita_da_cella(cell3)
        if not quantita and codice == "10-GEL":
            porz = str(row[4] or "") if len(row) > 4 else ""
            if porz.isdigit():
                quantita = [(int(porz), "pz")]
        if not quantita:
            continue

        cell1 = str(row[1] or "").strip()
        scad_match = SCAD_RE.search(cell1)
        scadenza = scad_match.group(1) if scad_match else ""
        descrizione = cell1 if cell1 else codice

        confezionamento = str(row[5] or "").strip() if len(row) > 5 else ""

        articoli.append({
            "codice": codice,
            "descrizione": descrizione,
            "scadenza": scadenza,
            "kg": kg,
            "quantita": quantita,
            "confezionamento": confezionamento,
        })
    return articoli if articoli else None


def _consolida_quantita(codice: str, lista_qty: list[tuple[int, str]]) -> tuple[list[tuple[int, str]], str]:
    """
    Consolida le quantita per gli articoli multi-unita.
    Ritorna (quantita_consolidate, formato_quantita_per_display).
    """
    if codice not in CONSOLIDAMENTO:
        # Articolo singolo: somma tutte le quantita con stessa unita (normalizzata)
        by_unit = defaultdict(int)
        for qty, unit in lista_qty:
            nu = _normalizza_unita(unit)
            by_unit[nu] += qty
        result = [(v, k) for k, v in sorted(by_unit.items()) if v > 0]
        parts = [f"{q} {u}" for q, u in result]
        return result, " ".join(parts)

    unit_princ, unit_second, ratio = CONSOLIDAMENTO[codice]
    tot_princ = 0
    tot_second = 0

    for qty, unit in lista_qty:
        unit_lower = unit.lower()
        if unit_princ.lower() in unit_lower or unit_lower in ("fardello", "fardelli"):
            tot_princ += qty
        elif unit_second.lower() in unit_lower or unit_lower in ("bottiglia", "bottiglie", "cluster", "porzioni", "porzione"):
            tot_second += qty
        elif unit_lower in ("cartoni", "cartone"):
            tot_princ += qty
        elif "brick" in unit_lower or "colli" in unit_lower or "confezioni" in unit_lower or "manifesti" in unit_lower or "fascette" in unit_lower:
            tot_princ += qty
        else:
            tot_second += qty

    # Consolida: tot_second possono diventare unit_princ
    extra_princ = tot_second // ratio
    resto_second = tot_second % ratio
    tot_princ += extra_princ

    # Nomi per display (singolare italiano)
    _sing = {"Fardelli": "Fardello", "Bottiglie": "Bottiglia", "Cartoni": "Cartone", "Porzioni": "Porzione",
             "Cluster": "Cluster"}
    nm_princ = _sing.get(unit_princ, unit_princ) if tot_princ == 1 else unit_princ
    nm_second = _sing.get(unit_second, unit_second) if resto_second == 1 else unit_second

    parts = []
    if tot_princ > 0:
        parts.append(f"{tot_princ} {nm_princ}")
    if resto_second > 0:
        parts.append(f"{resto_second} {nm_second}")

    result = []
    if tot_princ > 0:
        result.append((tot_princ, unit_princ))
    if resto_second > 0:
        result.append((resto_second, unit_second))

    return result, " ".join(parts) if parts else "0"


def _aggrega_articoli(articoli: list[dict]) -> list[dict]:
    """Aggrega articoli per (codice, scadenza). Stesso codice + stessa scadenza = somma; scadenze diverse = righe separate."""
    by_key = defaultdict(lambda: {"kg": Decimal("0"), "quantita": [], "descrizione": "", "confezionamento": ""})

    for a in articoli:
        chiave = (a["codice"], a.get("scadenza", ""))
        by_key[chiave]["kg"] += a["kg"]
        by_key[chiave]["quantita"].extend(a["quantita"])
        if not by_key[chiave]["descrizione"]:
            by_key[chiave]["descrizione"] = a["descrizione"]
        if not by_key[chiave]["confezionamento"] and a["confezionamento"]:
            by_key[chiave]["confezionamento"] = a["confezionamento"]

    risultato = []
    for (codice, scadenza) in sorted(by_key.keys()):
        dati = by_key[(codice, scadenza)]
        qty_cons, qty_display = _consolida_quantita(codice, dati["quantita"])
        conf = dati["confezionamento"]
        # Per multi-unit, confezionamento può avere due parti
        if codice in CONSOLIDAMENTO:
            _, _, ratio = CONSOLIDAMENTO[codice]
            if "Fardelli" in codice or "LT-ES" in codice or "LT-ESL" in codice:
                conf = f"{ratio} Bottiglie / Fardello\n6,6 Porzioni / Bottiglia"
            elif "YO-BI" in codice:
                conf = "10 Cluster / Cartone\n2 Porzioni / Cluster"
            elif "YO-DL" in codice:
                conf = "6 Porzioni / Cartone"
            elif "AP-SU" in codice:
                conf = "24 Porzioni / Cartone"
        risultato.append({
            "codice": codice,
            "descrizione": dati["descrizione"],
            "quantita": qty_display,
            "confezionamento": conf or "-",
            "kg": dati["kg"],
        })

    return risultato


NOTA_FORNITORI = "Latte Busche: BL - VE       Latterie Soligo: PD - TV       Latterie Vicentine: VI"


def _genera_pdf(articoli: list[dict], data_ddt: str, num_distinta: str, zone_suffix: str, zone_nums: list[str], output_path: Path, causali_province: dict[str, list[str]] | None = None, rientri_territorio: list[str] | None = None):
    """Genera il PDF della distinta. num_distinta: '01','02'...; zone_suffix: '3107-4107-4120-VI'; zone_nums per riepilogo."""
    causali_province = causali_province or {}
    rientri_territorio = rientri_territorio or []

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    elements = []

    # Titolo: "Distinta 01 Zona -3107-4107-4120-VI"
    titolo = f"Distinta {num_distinta} Zona -{zone_suffix}"
    elements.append(Paragraph(titolo, ParagraphStyle(name="Titolo", fontSize=14, fontName="Helvetica-Bold", spaceAfter=12)))
    elements.append(Spacer(1, 6))

    # Tabella info
    info_data = [
        ["Tipo", "DISTINTA DI CARICO Da TP a giro di consegna."],
        ["Creazione", data_ddt.replace("-", "/") + " 00:00:00"],
        ["Ultima Chiusura", data_ddt.replace("-", "/") + " 00:00:00"],
    ]
    info_table = Table(info_data, colWidths=[35 * mm, 120 * mm])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8E8E8")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # Tabella articoli: Paragraph per descrizione e confezionamento (wrap automatico)
    style_cell = ParagraphStyle(name="Cell", fontSize=9, leading=10, wordWrap="CJK")
    headers = ["CODICE", "DESCRIZIONE", "QUANTITÀ", "CONFEZIONAMENTO"]
    rows = [headers]
    kg_totali = Decimal("0")
    for a in articoli:
        desc = _xml_escape(str(a["descrizione"]))
        conf = _xml_escape(str(a["confezionamento"]).replace("\n", "<br/>"))
        qty = str(a["quantita"] or "")
        if qty:
            qty_formatted = re.sub(r"^(\d+ \S+) (\d+ \S+)$", r"\1<br/>\2", qty)
            if "<br/>" in qty_formatted:
                parts = [_xml_escape(p) for p in qty_formatted.split("<br/>")]
                qty_cell = Paragraph("<br/>".join(parts), style_cell)
            else:
                qty_cell = qty
        else:
            qty_cell = qty
        rows.append([
            a["codice"],
            Paragraph(desc, style_cell),
            qty_cell,
            Paragraph(conf, style_cell),
        ])
        kg_totali += a["kg"]

    table = Table(rows, colWidths=[32 * mm, 75 * mm, 33 * mm, 40 * mm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"KG totali: {kg_totali:.3f}".replace(".", ","), ParagraphStyle(name="Kg", fontSize=10, fontName="Helvetica-Bold")))

    # Segnalazione nuovi articoli o unità (in grande, sotto KG totali)
    nuovi_articoli = [a["codice"] for a in articoli if a["codice"] not in ARTICOLI_NOTI]
    nuove_unita = []
    for a in articoli:
        q = a.get("quantita", "") or ""
        for u in UNITA_NUOVE:
            if u in q and u not in nuove_unita:
                nuove_unita.append(u)
    if nuovi_articoli or nuove_unita:
        elements.append(Spacer(1, 10))
        parti = []
        if nuovi_articoli:
            parti.append("NUOVO ARTICOLO: " + ", ".join(sorted(nuovi_articoli)))
        if nuove_unita:
            parti.append("NUOVA UNITÀ: " + ", ".join(nuove_unita))
        testo_avviso = "   •   ".join(parti)
        elements.append(Paragraph(
            testo_avviso,
            ParagraphStyle(name="Avviso", fontSize=12, fontName="Helvetica-Bold", textColor=colors.HexColor("#B22222"), spaceAfter=6)
        ))

    # Riepilogo zone (senza prefisso DDT) e province di consegna
    zone_str = " ".join(zone_nums) if zone_nums else ""
    if zone_str or causali_province or rientri_territorio:
        elements.append(Spacer(1, 12))
        header_zone = f"Riepilogo zone {zone_str} (altre zone) e province di consegna:" if zone_str else "Riepilogo zone e province di consegna:"
        elements.append(Paragraph(header_zone, ParagraphStyle(name="RiepHeader", fontSize=10, fontName="Helvetica-Bold", spaceAfter=6)))
        for causale in sorted(k for k in causali_province.keys() if k != "_rientri"):
            province = causali_province.get(causale, [])
            if isinstance(province, list):
                riga = f"{causale}: {', '.join(province)}"
                elements.append(Paragraph(riga, ParagraphStyle(name="RiepRow", fontSize=9, leftIndent=10, spaceAfter=2)))
        if rientri_territorio:
            elements.append(Paragraph(
                f"Rientri integrati: {', '.join(sorted(rientri_territorio))}",
                ParagraphStyle(name="Rientri", fontSize=9, leftIndent=10, spaceAfter=2, textColor=colors.HexColor("#006400")),
            ))
        elements.append(Spacer(1, 6))

    def _footer_canvas(canvas, _doc):
        """Nota fornitori in fondo alla pagina: nero, stessa grandezza del titolo."""
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 14)
        canvas.setFillColor(colors.black)
        page_width = A4[0]
        canvas.drawCentredString(page_width / 2, 15 * mm, NOTA_FORNITORI)
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer_canvas, onLaterPages=_footer_canvas)
    print(f"  Creato: {output_path.name}")


# Pattern per estrazione dati consegna (CAP 5 cifre, provincia)
CAP_RE = re.compile(r"\b(\d{5})\b")

# Latte: dopo "Luogo di destinazione" possono esserci P.Iva, CF, Albo (altra casella testo).
# Escludere Dnr Srl / Tel (dati fornitore, non destinatario).
CF_LIKE_RE = re.compile(r"^(Cf|C\.F\.|CF|Partita\s*Iva|P\.?\s*I\.?)\s*[:\s]", re.I)
DNR_FORNITORE_RE = re.compile(r"Dnr\s+Srl|Tel:\s*\d{6,}", re.I)  # da escludere


def _riga_sembra_cf(riga: str) -> bool:
    """True se la riga sembra Codice Fiscale o Partita Iva (layout latte)."""
    if not riga or len(riga) > 60:
        return False
    if CF_LIKE_RE.match(riga.strip()):
        return True
    s = re.sub(r"[\s.]", "", riga)
    return len(s) == 16 and s.isalnum()


def _riga_da_escludere(riga: str) -> bool:
    """Esclude righe con Dnr Srl, Tel (fornitore), o solo P.Iva/CF."""
    if not riga:
        return True
    if DNR_FORNITORE_RE.search(riga):
        return True
    if _riga_sembra_cf(riga):
        return True
    return False


def _splitta_nome_indirizzo(riga: str) -> tuple[str, str]:
    """Se la riga contiene ' Via ' o ' V. ' ecc., splitta in nome + indirizzo (con Via)."""
    for sep in (" Via ", " V. ", " Viale ", " Corso ", " C.so ", " Piazza ", " P.zza "):
        if sep in riga:
            idx = riga.find(sep)
            nome = riga[:idx].strip()
            indirizzo = riga[idx:].strip()  # include Via nell'indirizzo
            if nome and len(nome) > 1 and indirizzo:
                return (nome, indirizzo)
    return (riga, "")


def _estrai_dati_consegna_da_testo(text: str, codice: str, da_frutta: bool = True) -> dict:
    """
    Estrae destinatario, indirizzo, CAP, città, provincia dal testo di una pagina DDT.
    FRUTTA: logica originale invariata - Luogo -> nome (i+1) -> indirizzo (i+2).
    LATTE:  layout diverso, usa candidati ed esclusioni.
    """
    res = {"destinatario": "", "indirizzo": "", "cap": "", "citta": "", "provincia": ""}
    if codice.lower() not in text.lower():
        return res
    idx_l = text.find("Luogo di destinazione")
    if idx_l < 0:
        return res

    if da_frutta:
        blocco = text[idx_l : idx_l + 650]
        lines = [ln.strip() for ln in blocco.split("\n") if ln.strip() and not ln.strip().upper().startswith("RESPONSABILE")]
        # FRUTTA: logica originale, non modificare
        for i, ln in enumerate(lines):
            if LUOGO_RE.search(ln):
                if i + 1 < len(lines):
                    res["destinatario"] = lines[i + 1]
                if i + 2 < len(lines):
                    res["indirizzo"] = lines[i + 2]
                break
    else:
        # LATTE: blocco PRIMA di "CAUSALE DEL TRASPORTO"
        # Regole: nome cliente tra "CF" (e varianti) e "Albo"; indirizzo tra "Albo" e "RESPONSABILE DEL TRASPORTO:"
        # Su una riga ciascuno: riga CF = nome, riga Albo = indirizzo
        idx_causale = text.upper().find("CAUSALE DEL TRASPORTO")
        blocco = text[:idx_causale] if idx_causale > 0 else text[idx_l : idx_l + 900]
        for ln in blocco.split("\n"):
            ln = ln.strip()
            cf_m = re.match(r"^[Cc]\.?[Ff]\.?\s+", ln)
            if cf_m:
                res["destinatario"] = ln[cf_m.end():].strip()
            else:
                albo_m = re.match(r"^[Aa]lbo\s+", ln, re.I)
                if albo_m:
                    res["indirizzo"] = ln[albo_m.end():].strip()

    # CAP, città, provincia: per LATTE cercare solo dopo "RESPONSABILE DEL TRASPORTO"
    # (altrimenti si prende 35030 VEGGIANO dall'header LOG.SOLUTIONS)
    blocco_prov = blocco
    if not da_frutta:
        idx_resp = blocco.upper().find("RESPONSABILE DEL TRASPORTO")
        if idx_resp >= 0:
            blocco_prov = blocco[idx_resp:]

    for prov_m in PROVINCIA_RE.finditer(blocco_prov):
        sigla = prov_m.group(1)
        ctx = blocco_prov[max(0, prov_m.start() - 40) : prov_m.start()]
        if sigla == "MN" and ("Pomponesco" in ctx or "46030" in ctx):
            continue
        res["provincia"] = sigla
        cap_possibili = list(CAP_RE.finditer(blocco_prov[: prov_m.start()]))
        if cap_possibili:
            cap_m = cap_possibili[-1]
            res["cap"] = cap_m.group(1)
            pre = blocco_prov[cap_m.end() : cap_m.end() + 60]
            citta_m = re.search(r"\s*[-]?\s*([A-Za-zÀ-ÿ\s'.]+?)\s*\([A-Z]{2}\)", pre)
            if citta_m:
                res["citta"] = citta_m.group(1).strip()
        break
    return res


def _estrai_dati_consegna_per_codice(pdf_paths: list[Path], codice: str) -> dict:
    """Cerca una pagina DDT con il codice e estrae dati consegna. Preferisce frutta (nome cliente)."""
    codice = codice.lower()
    # Ordina: frutta prima, poi latte (per codici in entrambi, nome da frutta è più pulito)
    def _ordina(p: Path):
        return (0 if "frutta" in str(p).lower() else 1, str(p))
    ordinati = sorted(pdf_paths, key=_ordina)
    for pdf_path in ordinati:
        try:
            da_frutta = "frutta" in str(pdf_path).lower()
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    if f"destinazione: {codice}" not in text.lower() and f"destinazione:{codice}" not in text.lower():
                        continue
                    return _estrai_dati_consegna_da_testo(text, codice, da_frutta=da_frutta)
        except Exception:
            continue
    return {"destinatario": "", "indirizzo": "", "cap": "", "citta": "", "provincia": ""}


def _estrai_codici_luogo_da_pdf(pdf_path: Path) -> set[str]:
    """Estrae tutti i codici luogo (p####) da un PDF."""
    codici = set()
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for m in LUOGO_RE.finditer(text):
                codici.add(m.group(1).lower())
    return codici


def _estrai_codici_per_tipo(pdf_paths: list[Path]) -> set[str]:
    """Estrae codici luogo da una lista di PDF."""
    codici = set()
    for p in pdf_paths:
        codici.update(_estrai_codici_luogo_da_pdf(p))
    return codici


def _leggi_codici_mappatura() -> set[str]:
    """Legge codici da colonne A e I di mappatura_destinazioni (solo lettura)."""
    if not MAPPATURA_XLSX.exists():
        return set()
    wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    codici = set()
    for row in ws.iter_rows(min_row=2, max_col=COL_CODICE_LATTE):
        for col_idx in (COL_CODICE_FRUTTA - 1, COL_CODICE_LATTE - 1):  # 0-based
            val = (row[col_idx].value or "").strip().lower()
            if val:
                codici.add(val)
    wb.close()
    return codici


def _verifica_e_valida_codici() -> bool:
    """
    Verifica: 1) entrambe le cartelle frutta e latte hanno file PDF;
              2) tutti i codici luogo nei DDT sono presenti in mappatura.
    Ritorna True se OK, False se procedura interrotta.
    """
    # 1) Verifica file nelle due cartelle (possibili aggiunte di ordine in attesa)
    pdf_frutta = list(DDT_FRUTTA_DIR.glob("*.pdf")) if DDT_FRUTTA_DIR.exists() else []
    pdf_latte = list(DDT_LATTE_DIR.glob("*.pdf")) if DDT_LATTE_DIR.exists() else []

    if not pdf_frutta:
        print("ERRORE: Nessun PDF nella cartella 'DDT frutta'.")
        print("        Verificare se sono in attesa aggiunte di ordine.")
        return False
    if not pdf_latte:
        print("ERRORE: Nessun PDF nella cartella 'DDT latte'.")
        print("        Verificare se sono in attesa aggiunte di ordine.")
        return False

    # 2) Estrai codici da DDT separati per frutta e latte
    codici_frutta = _estrai_codici_per_tipo(pdf_frutta)
    codici_latte = _estrai_codici_per_tipo(pdf_latte)

    # 3) Confronta con mappatura (colonne A e B)
    codici_mappatura = _leggi_codici_mappatura()
    if not codici_mappatura and MAPPATURA_XLSX.exists():
        print("ATTENZIONE: mappatura_destinazioni.xlsx non contiene codici (verificare foglio 'Mappatura' e colonne A, B).")
    nuovi = (codici_frutta | codici_latte) - codici_mappatura

    if nuovi:
        print("ERRORE: Trovati codici di consegna non presenti in mappatura_destinazioni.xlsx")
        print(f"        Codici nuovi: {', '.join(sorted(nuovi))}")
        print(f"\n        Salvati in: {NUOVI_CODICI_XLSX.name}")
        print("        Completare indirizzi e copiare in mappatura_destinazioni.xlsx")
        # Stesso formato di mappatura (solo lettura, mai modificare mappatura): tutte le colonne incluso Orario
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Nuovi codici"
        headers = []
        if MAPPATURA_XLSX.exists():
            wb_map = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
            ws_map = wb_map["Mappatura"] if "Mappatura" in wb_map.sheetnames else wb_map.active
            headers = [ws_map.cell(1, c).value for c in range(1, ws_map.max_column + 1)]
            wb_map.close()
        if not headers:
            headers = ["Codice Frutta", "Codice Latte", "A chi va consegnato", "Tipologia grado", "Indirizzo", "CAP", "Città", "Provincia",
                      "Tipologia consegna", "Tipologia consegna.1", "Email", "Sito web", "Orario min", "Orario max"]
        ws_out.append(headers)
        n_cols = max(len(headers), 10)
        tutti_pdf = list(pdf_frutta) + list(pdf_latte)
        # Fallback: se cartelle vuote (PDF già spostati), cerca in DDT-ORIGINALI
        if not tutti_pdf and GIRI_LAVORATI_DIR.exists():
            for cartella in sorted([d for d in GIRI_LAVORATI_DIR.iterdir() if d.is_dir()], key=lambda x: x.stat().st_mtime, reverse=True):
                orig = cartella / "DDT-ORIGINALI"
                if orig.exists():
                    tutti_pdf = list(orig.glob("*.pdf"))
                    if tutti_pdf:
                        break
        # Estrai dati per tutti i nuovi codici in un passaggio (cache per velocità)
        print("        Estrazione indirizzi dai DDT...")
        dati_per_codice: dict[str, dict] = {}
        for pdf_path in tutti_pdf:
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        for m in LUOGO_RE.finditer(text):
                            cod = m.group(1).lower()
                            if cod not in nuovi or cod in dati_per_codice:
                                continue
                            da_f = "latte" not in str(pdf_path).lower()
                            dati_per_codice[cod] = _estrai_dati_consegna_da_testo(text, cod, da_frutta=da_f)
            except Exception:
                continue
        def _formatta_titolo(s: str) -> str:
            """Prima lettera di ogni parola maiuscola, resto minuscolo (per nuovi_codici)."""
            return s.strip().title() if s and isinstance(s, str) else (s or "")

        for c in sorted(nuovi):
            dati = dati_per_codice.get(c, {"destinatario": "", "indirizzo": "", "cap": "", "citta": "", "provincia": ""})
            in_f = c in codici_frutta
            in_l = c in codici_latte
            dest = _formatta_titolo(dati["destinatario"])
            indir = _formatta_titolo(dati["indirizzo"])
            citta = _formatta_titolo(dati["citta"])
            prov = (dati["provincia"] or "").strip().upper()[:2]
            cap = (dati["cap"] or "").strip()
            row = [""] * n_cols
            # Nuova struttura: A=Cod Frutta, B=Cod Latte, C-N=indirizzo unico
            if in_f and in_l:
                row[0], row[1] = c, c
                row[2], row[4], row[5], row[6], row[7] = dest, indir, cap, citta, prov
                row[8], row[9] = "Frutta e Latte", "Frutta e Latte"
            elif in_f:
                row[0], row[1] = c, CODICE_CIVETTA
                row[2], row[4], row[5], row[6], row[7] = dest, indir, cap, citta, prov
                row[8], row[9] = "Frutta", ""
            else:
                row[0], row[1] = CODICE_CIVETTA, c
                row[2], row[4], row[5], row[6], row[7] = dest, indir, cap, citta, prov
                row[8], row[9] = "", "Latte"
            ws_out.append(row)
        wb_out.save(NUOVI_CODICI_XLSX)
        print(f"        Salvati {len(nuovi)} codici in: {NUOVI_CODICI_XLSX}")
        return False

    return True


def _orario_min_da_numero(s: str) -> str:
    """Converte 730->07:30, 800->08:00, 715->07:15."""
    if not s or len(s) != 3:
        return ""
    h = int(s[0])
    m = int(s[1:3])
    return f"{h:02d}:{m:02d}"


def _orario_max_da_h(h_digits: str) -> str:
    """Converte H10->10:00, H09->09:00. Se manca ritorna default 14:00."""
    if not h_digits or len(h_digits) != 2:
        return ORARIO_MAX_DEFAULT
    return f"{int(h_digits):02d}:00"


def _estrai_sezione_causale(text: str) -> str:
    """Ritorna solo il blocco di testo dopo 'CAUSALE DEL TRASPORTO', per evitare falsi match (Tel, ecc.)."""
    idx = text.upper().find(CAUSALE_SEZIONE_MARKER.upper())
    if idx < 0:
        return text  # fallback: intera pagina
    start = idx + len(CAUSALE_SEZIONE_MARKER)
    return text[start:start + CAUSALE_SEZIONE_LUNGHEZZA]


def _estrai_luogo_territorio(text: str) -> tuple[str, str] | None:
    """Ritorna (luogo, territorio) da una pagina DDT."""
    luogo_m = LUOGO_RE.search(text)
    if not luogo_m:
        return None
    luogo = luogo_m.group(1).lower()
    sezione_causale = _estrai_sezione_causale(text)
    causale_m = CAUSALE_RE.search(sezione_causale)
    territorio = causale_m.group(1)[1:5] if causale_m else ""
    return (luogo, territorio)


def _estrai_causale_provincia(text: str) -> tuple[str, str] | None:
    """Ritorna (causale, provincia) da una pagina DDT. Es: (A3101, VE).
    Esclude (MN) da '46030 Pomponesco (MN)' che è l'indirizzo DNR, non la destinazione."""
    sezione_causale = _estrai_sezione_causale(text)
    causale_m = CAUSALE_RE.search(sezione_causale)
    if not causale_m:
        return None
    causale = causale_m.group(1)
    idx_luogo = text.find('Luogo di destinazione')
    if idx_luogo < 0:
        return (causale, "")
    rest = text[idx_luogo:idx_luogo + 800]
    for m in PROVINCIA_RE.finditer(rest):
        sigla = m.group(1)
        # Escludi (MN) da indirizzo DNR "46030 Pomponesco (MN)"
        ctx_pre = rest[max(0, m.start() - 35) : m.start()]
        if sigla == "MN" and ("Pomponesco" in ctx_pre or "46030" in ctx_pre):
            continue
        return (causale, sigla)
    return (causale, "")


def _estrai_orari_da_causale(text: str) -> tuple[str, str]:
    """Estrae (orario_min, orario_max) dalla sezione causale. Es: A3101 H10 730 -> (07:30, 10:00)."""
    sezione_causale = _estrai_sezione_causale(text)
    m = CAUSALE_RE.search(sezione_causale)
    if not m:
        return ("", ORARIO_MAX_DEFAULT)
    h_digits = m.group(2)
    num_orario = m.group(3)
    orario_max = _orario_max_da_h(h_digits) if h_digits else ORARIO_MAX_DEFAULT
    orario_min = _orario_min_da_numero(num_orario) if num_orario else ""
    return (orario_min, orario_max)


def _aggiorna_mappatura_orari(row_updates: dict[int, tuple[str, str]]) -> None:
    """
    Aggiorna colonne M (Orario min) e N (Orario max) nella mappatura.
    Scrive SOLO se la cella è vuota - non sovrascrive dati inseriti a mano.
    row_updates: {riga_excel: (orario_min, orario_max)}
    """
    if not MAPPATURA_XLSX.exists() or not row_updates:
        return
    wb = load_workbook(MAPPATURA_XLSX)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    aggiornati = 0
    for row_idx, (om, oM) in row_updates.items():
        cell_om = ws.cell(row=row_idx, column=COL_ORARIO_MIN)
        cell_oM = ws.cell(row=row_idx, column=COL_ORARIO_MAX)
        if cell_om.value is None or str(cell_om.value).strip() == "":
            cell_om.value = om if om and om != "00:00" else "false"  # false = da chiamare
            aggiornati += 1
        if cell_oM.value is None or str(cell_oM.value).strip() == "":
            cell_oM.value = oM if oM and oM != "00:00" else "false"  # false = da chiamare
            aggiornati += 1
    wb.save(MAPPATURA_XLSX)
    if aggiornati > 0:
        print(f"\n  Aggiornati {aggiornati} orari in mappatura_destinazioni.xlsx (colonne M e N)")


def _verifica_orari_e_report(
    territorio_pages: dict,
    page_orari: dict,
    page_luogo: dict,
) -> bool:
    """
    Verifica che tutti i codici consegna abbiano orario min estratto.
    Se manca, crea report_orari_mancanti.xlsx (non blocca: va aggiornato dopo chiamate scuole).
    """
    mancanti: dict[str, tuple[str, str]] = {}
    for territorio, pages in territorio_pages.items():
        for pdf_path, page_idx in pages:
            key = (str(pdf_path), page_idx)
            luogo = page_luogo.get(key)
            if not luogo:
                continue
            om, oM = page_orari.get(key, ("", ORARIO_MAX_DEFAULT))
            if not om and luogo not in mancanti:
                mancanti[luogo] = ("-", oM or ORARIO_MAX_DEFAULT)

    if mancanti:
        print("ATTENZIONE: Alcuni codici hanno orario min non estraibile dalla causale.")
        print(f"            Codici: {', '.join(sorted(mancanti.keys()))}")
        print(f"            Report salvato in: {REPORT_ORARI_MANCANTI_XLSX.name}")
        print("            Aggiornare mappatura dopo le chiamate alle scuole.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Orari mancanti"
        ws.append(["Codice luogo", "Orario min", "Orario max", "Azione richiesta"])
        for cod in sorted(mancanti.keys()):
            om, oM = mancanti[cod]
            ws.append([cod, om, oM, "Causale deve contenere numero 3 cifre (730, 800, 715) dopo H10"])
        wb.save(REPORT_ORARI_MANCANTI_XLSX)
        return False
    return True


def _legge_codice_to_punto_consegna() -> dict[str, int]:
    """
    Mappa ogni codice (p####) al punto di consegna (row index).
    Stessa riga mappatura = stesso punto: frutta p1234 e latte p4567 sulla stessa riga = 1 consegna.
    """
    if not MAPPATURA_XLSX.exists():
        return {}
    wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    codice_to_punto: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        cod_f = (ws.cell(row_idx, COL_CODICE_FRUTTA).value or "").strip().lower()
        cod_l = (ws.cell(row_idx, COL_CODICE_LATTE).value or "").strip().lower()
        codici_riga = [
            c for c in (cod_f, cod_l)
            if c and c != CODICE_CIVETTA
        ]
        for cod in codici_riga:
            codice_to_punto[cod] = row_idx
    wb.close()
    return codice_to_punto


def _legge_abbinamenti_mappatura() -> dict[str, str]:
    """Ritorna {cod_frutta: cod_latte} dalla mappatura (solo lettura)."""
    if not MAPPATURA_XLSX.exists():
        return {}
    wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    abbinamenti = {}
    for row in ws.iter_rows(min_row=2, max_col=COL_CODICE_LATTE):
        cod_f = (row[COL_CODICE_FRUTTA - 1].value or "").strip().lower()
        cod_l = (row[COL_CODICE_LATTE - 1].value or "").strip().lower()
        if cod_f and cod_l and cod_f != CODICE_CIVETTA and cod_l != CODICE_CIVETTA:
            abbinamenti[cod_f] = cod_l
    wb.close()
    return abbinamenti


def _costruisci_nome_zona(territorio: str, territorio_causali_province: dict) -> str:
    """Costruisce il nome zona (es. 3107-4107-4120-VI) per il report rientri."""
    cp = territorio_causali_province.get(territorio, {})
    zone_nums = sorted(set(c[1:5] for c in cp.keys() if len(c) >= 5))
    main = territorio
    others = [z for z in zone_nums if z != main]
    zone_ordered = [main] + others
    province = sorted(set(p for provs in cp.values() for p in provs if p))
    return "-".join(zone_ordered) + ("-" + "-".join(province) if province else "")


def _scrivi_report_rientri(
    rientri_esito: list[tuple[str, str, str, str]],
) -> None:
    """
    Salva report_rientri_non_integrabili.xlsx con esito di TUTTI i rientri.
    rientri_esito: [(codice, data_ddt, stato, zona_o_motivo), ...]
    stato: Integrato | Non integrabile | Non trovato
    zona_o_motivo: nome zona se integrato, altrimenti motivo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Esito rientri"
    ws.append(["Codice consegna", "Data DDT", "Stato", "Zona / Motivo"])
    for codice, data_ddt, stato, zona_motivo in rientri_esito:
        ws.append([codice, data_ddt, stato, zona_motivo])
    integrati = sum(1 for r in rientri_esito if r[2] == "Integrato")
    non_int = sum(1 for r in rientri_esito if r[2] == "Non integrabile")
    non_trov = sum(1 for r in rientri_esito if r[2] == "Non trovato")
    ignorati = sum(1 for r in rientri_esito if r[2] == "Ignorato")
    ws_summary = wb.create_sheet("Riepilogo", 0)
    ws_summary.append(["Integrati", integrati])
    ws_summary.append(["Non integrabili", non_int])
    ws_summary.append(["Non trovati", non_trov])
    if ignorati:
        ws_summary.append(["Ignorati (data futura)", ignorati])
    wb.save(REPORT_RIENTRI_NON_INTEGRABILI_XLSX)
    print(f"            Report salvato in: {REPORT_RIENTRI_NON_INTEGRABILI_XLSX.name}")


def _aggiorna_rientri_allegato(codici_integrati: set[str], data_ddt: str) -> None:
    """
    Aggiorna rientri_ddt.xlsx colonna C con "Allegato con DDT (data)" per i rientri integrati.
    data_ddt: data del giro in cui il rientro è stato allegato (es. 12-03-2026).
    """
    if not codici_integrati or not RIENTRI_DDT_XLSX.exists():
        return
    try:
        wb = load_workbook(RIENTRI_DDT_XLSX)
        ws = wb["Rientri"] if "Rientri" in wb.sheetnames else wb.active
        data_legible = data_ddt.replace("-", "/")
        valore = f"Allegato con DDT ({data_legible})"
        # Header colonna C se non presente
        if ws.cell(row=1, column=3).value is None or ws.cell(row=1, column=3).value == "":
            ws.cell(row=1, column=3, value="Allegato con DDT (data)")
        for row_idx in range(2, ws.max_row + 1):
            cod = (ws.cell(row=row_idx, column=1).value or "").strip().lower()
            if cod and cod in codici_integrati:
                ws.cell(row=row_idx, column=3, value=valore)
        wb.save(RIENTRI_DDT_XLSX)
        print(f"            Aggiornato {RIENTRI_DDT_XLSX.name} colonna C per {len(codici_integrati)} rientri allegati")
    except Exception as e:
        print(f"            Avviso: impossibile aggiornare rientri_ddt.xlsx: {e}")


def _carica_rientri() -> list[tuple[str, str]]:
    """
    Legge rientri_ddt.xlsx (colonne: Codice consegna, Data DDT, Stato/Allegato).
    La data indica la cartella DDT-{data} dove cercare il DDT non consegnato.
    Esclude righe con colonna Stato già popolata (es. "Allegato con DDT (data)").
    Ritorna lista di (codice, data_str) in formato dd-mm-yyyy.
    """
    if not RIENTRI_DDT_XLSX.exists():
        return []
    wb = load_workbook(RIENTRI_DDT_XLSX, read_only=True, data_only=True)
    ws = wb["Rientri"] if "Rientri" in wb.sheetnames else wb.active
    righe = []
    for row in ws.iter_rows(min_row=2, max_col=3):
        cod = (row[0].value or "").strip().lower()
        data_val = row[1].value if len(row) > 1 else None
        stato = (row[2].value or "").strip() if len(row) > 2 else ""
        if not cod or cod == "codice consegna":
            continue
        if stato and "allegato" in stato.lower():
            continue
        data_str = _normalizza_data(data_val) if data_val else ""
        righe.append((cod, data_str))
    wb.close()
    return righe


def _normalizza_data(val) -> str | None:
    """Converte vari formati data in dd-mm-yyyy."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    s = str(val).strip()
    if not s:
        return None
    for sep in ("-", "/", "."):
        if sep in s:
            parts = re.split(r"[-/.]", s, maxsplit=2)
            if len(parts) >= 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return f"{d:02d}-{m:02d}-{y:04d}"
                except (ValueError, TypeError):
                    pass
    return None


def _trova_pagina_rientro(codice: str, data_str: str) -> tuple[Path, int, str, str, str] | None:
    """
    Cerca il DDT del rientro in Giri lavorati/DDT-{data}/DDT-ORIGINALI/.
    La data indica la cartella specifica (DDT non consegnato in quella data).
    Se data_str vuota: fallback su tutte le cartelle DDT-* (più recenti prima).
    Ritorna (pdf_path, page_idx, territorio, causale, provincia) o None se non trovato.
    """
    codice = codice.lower()

    def _cerca_in_cartella(cartella: Path) -> tuple[Path, int, str, str, str] | None:
        if not cartella.exists():
            return None
        # File singolo *_{codice}.pdf (dopo crea_ddt_originali)
        for singolo in cartella.glob(f"*_{codice}.pdf"):
            try:
                with pdfplumber.open(singolo) as pdf:
                    if pdf.pages:
                        text = pdf.pages[0].extract_text() or ""
                        cp = _estrai_causale_provincia(text)
                        territorio = cp[0][1:5] if cp else ""
                        causale = cp[0] if cp else ""
                        provincia = cp[1] if cp else ""
                        return (singolo, 0, territorio, causale, provincia)
            except Exception:
                pass
        # Cerca in tutti i PDF multi-pagina
        for pdf_path in sorted(cartella.glob("*.pdf")):
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text = page.extract_text() or ""
                        if f"destinazione: {codice}" not in text.lower() and f"destinazione:{codice}" not in text.lower():
                            continue
                        cp = _estrai_causale_provincia(text)
                        territorio = cp[0][1:5] if cp else ""
                        causale = cp[0] if cp else ""
                        provincia = cp[1] if cp else ""
                        return (pdf_path, i, territorio, causale, provincia)
            except Exception:
                continue
        return None

    if data_str:
        cartella = GIRI_LAVORATI_DIR / f"DDT-{data_str}" / "DDT-ORIGINALI"
        return _cerca_in_cartella(cartella)
    # Senza data: cerca in tutte le cartelle DDT-* (ordine: più recenti prima)
    if not GIRI_LAVORATI_DIR.exists():
        return None
    subdirs = sorted([d for d in GIRI_LAVORATI_DIR.iterdir() if d.is_dir() and d.name.startswith("DDT-") and d.name != "DDT-ORIGINALI"], key=lambda x: x.name, reverse=True)
    for subdir in subdirs:
        cartella = subdir / "DDT-ORIGINALI"
        result = _cerca_in_cartella(cartella)
        if result:
            return result
    return None


def _filtro_e_unione_ddt() -> str | None:
    """
    Filtra e unisce i DDT frutta+latte per territorio. Crea i PDF in DDT-[data]/.

    Ordine operazioni (IMPORTANTE):
    1) Lavora DDT frutta
    2) Lavora DDT latte e li integra con frutta (abbinamenti)
    3) Crea le zone dai codici (territorio_pages)
    4) Merge allegati (4101->3101, ecc.)
    5) SOLO DOPO: controlla rientri e li allega ai file già creati (mai nuovi giri)
    """
    if not PdfWriter or not PdfReader:
        print("ERRORE: pypdf necessario per filtro e unione DDT. pip install pypdf")
        return None

    pdf_frutta = sorted(DDT_FRUTTA_DIR.glob("*.pdf")) if DDT_FRUTTA_DIR.exists() else []
    pdf_latte = sorted(DDT_LATTE_DIR.glob("*.pdf")) if DDT_LATTE_DIR.exists() else []
    if not pdf_frutta or not pdf_latte:
        return None

    # Data dal primo PDF
    data_ddt = None
    with pdfplumber.open(pdf_frutta[0]) as pdf:
        if pdf.pages:
            text = pdf.pages[0].extract_text() or ""
            m = DATA_DDT_RE.search(text)
            if m:
                data_ddt = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    if not data_ddt:
        print("ERRORE: Data non trovata nei DDT.")
        return None

    output_dir = GIRI_LAVORATI_DIR / f"DDT-{data_ddt}"
    if output_dir.exists():
        shutil.rmtree(output_dir)
        print(f"\nRimossa cartella precedente {output_dir.name}")
    output_dir.mkdir(parents=True, exist_ok=True)
    ddt_zona_dir = output_dir / "DDT-ZONA"
    ddt_zona_dir.mkdir(parents=True, exist_ok=True)
    print(f"Filtro e unione DDT -> {output_dir.name}")

    luogo_to_page: dict[str, tuple[Path, int]] = {}
    luogo_territorio: dict[str, str] = {}
    territorio_pages: dict[str, list[tuple[Path, int]]] = defaultdict(list)
    territorio_luogo_gia_aggiunto: dict[str, set[str]] = defaultdict(set)  # evita DDT duplicati
    territorio_causali_province: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    page_causale_provincia: dict[tuple[str, int], tuple[str, str]] = {}
    page_orari: dict[tuple[str, int], tuple[str, str]] = {}
    page_luogo: dict[tuple[str, int], str] = {}

    def add_page_to_territorio(territorio: str, pdf_path: Path, page_idx: int, luogo: str | None = None):
        # Se abbiamo un luogo e l'abbiamo già aggiunto per questo territorio, salta (evita copie duplicate)
        if luogo and luogo in territorio_luogo_gia_aggiunto[territorio]:
            return
        if luogo:
            territorio_luogo_gia_aggiunto[territorio].add(luogo)
        key = (str(pdf_path), page_idx)
        territorio_pages[territorio].append((pdf_path, page_idx))
        if key in page_causale_provincia:
            causale, provincia = page_causale_provincia[key]
            territorio_causali_province[territorio][causale]  # assicura chiave per zone_nums anche se provincia vuota
            if provincia:
                territorio_causali_province[territorio][causale].add(provincia)

    def process_pdf(pdf_path: Path, skip_duplicati_frutta: bool = False):
        """skip_duplicati_frutta: nel PDF frutta ogni DDT è duplicato (2 pagine = stesso DDT), prendi solo la prima."""
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                if skip_duplicati_frutta and i % 2 == 1:
                    continue  # salta la copia duplicata (pagine 1,3,5,...)
                text = page.extract_text()
                if not text:
                    continue
                res = _estrai_luogo_territorio(text)
                cp = _estrai_causale_provincia(text)
                om, oM = _estrai_orari_da_causale(text)
                key = (str(pdf_path), i)
                if cp:
                    page_causale_provincia[key] = cp
                page_orari[key] = (om, oM)
                if res:
                    luogo, territorio = res
                    page_luogo[key] = luogo
                    if not territorio:
                        continue
                    luogo_territorio[luogo] = territorio
                    if luogo not in luogo_to_page:
                        luogo_to_page[luogo] = (pdf_path, i)
                    add_page_to_territorio(territorio, pdf_path, i, luogo)

    for p in pdf_frutta:
        process_pdf(p, skip_duplicati_frutta=True)
    for p in pdf_latte:
        process_pdf(p)

    abbinamenti = _legge_abbinamenti_mappatura()
    # Latte abbinate: aggiungi al territorio frutta e rimuovi dal territorio latte (evita doppi)
    # I DDT latte con corrispondenza frutta vanno solo nel giro frutta, non più nei file latte
    for luogo_f, luogo_l in abbinamenti.items():
        if luogo_f in luogo_territorio and luogo_l in luogo_to_page:
            territorio_frutta = luogo_territorio[luogo_f]
            pdf_path, page_idx = luogo_to_page[luogo_l]
            add_page_to_territorio(territorio_frutta, pdf_path, page_idx, luogo_l)
            key = (str(pdf_path), page_idx)
            if key in page_causale_provincia:
                territorio_latte = page_causale_provincia[key][0][1:5]
                territorio_pages[territorio_latte] = [
                    (p, i) for (p, i) in territorio_pages[territorio_latte]
                    if (str(p), i) != key
                ]

    # Verifica orari: se orario min non estratto, crea report (non blocca: va aggiornato dopo chiamate scuole)
    _verifica_orari_e_report(territorio_pages, page_orari, page_luogo)

    # Territori "allegati" (frutta) che non devono avere file propri: unisci al principale
    # Da Esempio reale: 4101, 4104, 4124 compaiono solo come DDT ALLEGATI, mai come file principale
    ALLEGATO_A_PRINCIPALE = {"4101": "3101", "4104": "3104", "4124": "3208"}
    for allegato, principale in ALLEGATO_A_PRINCIPALE.items():
        if allegato not in territorio_pages:
            continue
        if principale not in territorio_pages:
            territorio_pages[principale] = []
        for pdf_path, page_idx in territorio_pages[allegato]:
            add_page_to_territorio(principale, pdf_path, page_idx, page_luogo.get((str(pdf_path), page_idx)))
        for causale, provs in territorio_causali_province.get(allegato, {}).items():
            territorio_causali_province[principale][causale].update(provs)
        del territorio_pages[allegato]
        if allegato in territorio_causali_province:
            del territorio_causali_province[allegato]

    # Rientri DDT: SOLO allegare a zone già create, mai creare nuovi giri (dopo frutta, latte, abbinamenti, allegati)
    rientri = _carica_rientri()
    territorio_rientri: dict[str, list[str]] = defaultdict(list)
    rientri_esito: list[tuple[str, str, str, str]] = []  # (codice, data, stato, zona_o_motivo)
    codici_oggi = set(page_luogo.values())
    luogo_to_territorio_output: dict[str, str] = {}
    for territorio, pages in territorio_pages.items():
        for pdf_path, page_idx in pages:
            luogo = page_luogo.get((str(pdf_path), page_idx))
            if luogo:
                luogo_to_territorio_output[luogo] = territorio
    for codice, data_rientro in rientri:
        result = _trova_pagina_rientro(codice, data_rientro)
        if not result:
            rientri_esito.append((codice, data_rientro, "Non trovato", "DDT non presente in Giri lavorati"))
            continue
        pdf_path, page_idx, territorio, causale, provincia = result
        if codice not in codici_oggi:
            rientri_esito.append((codice, data_rientro, "Non integrabile", "Codice non presente nel giro del giorno"))
            continue
        territorio_oggi = luogo_to_territorio_output.get(codice)
        if not territorio_oggi:
            rientri_esito.append((codice, data_rientro, "Non integrabile", "Territorio non determinato"))
            continue
        key = (str(pdf_path), page_idx)
        territorio_pages[territorio_oggi].append((pdf_path, page_idx))
        page_luogo[key] = codice
        page_causale_provincia[key] = (causale, provincia)
        if provincia:
            territorio_causali_province[territorio_oggi][causale].add(provincia)
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[page_idx].extract_text() or ""
        page_orari[key] = _estrai_orari_da_causale(text)
        territorio_rientri[territorio_oggi].append(codice)
        zona_str = _costruisci_nome_zona(territorio_oggi, territorio_causali_province)
        rientri_esito.append((codice, data_rientro, "Integrato", zona_str))

    if rientri:
        integrati = sum(len(v) for v in territorio_rientri.values())
        non_int = sum(1 for r in rientri_esito if r[2] == "Non integrabile")
        non_trov = sum(1 for r in rientri_esito if r[2] == "Non trovato")
        print(f"  Rientri: {integrati} integrati, {non_int} non integrabili, {non_trov} non trovati")
        _scrivi_report_rientri(rientri_esito)
        codici_integrati = {r[0] for r in rientri_esito if r[2] == "Integrato"}
        if codici_integrati:
            _aggiorna_rientri_allegato(codici_integrati, data_ddt)

    metadata: dict[str, dict[str, list[str]]] = {}
    distinta_luoghi: dict[str, list[str]] = {}
    territori_lista = sorted(territorio_pages.keys())
    print(f"  Territori trovati: {len(territori_lista)} (uno per file output)")
    if len(territori_lista) > 20:
        print(f"  Esempi: {', '.join(territori_lista[:5])} ... {', '.join(territori_lista[-3:])}")
    creati = 0
    for territorio in territori_lista:
        pages = territorio_pages[territorio]
        if not pages:
            continue
        cp = territorio_causali_province[territorio]
        zone_nums = sorted(set(c[1:5] for c in cp.keys() if len(c) >= 5))
        main = territorio
        others = [z for z in zone_nums if z != main]
        zone_ordered = [main] + others
        province = sorted(set(p for provs in cp.values() for p in provs if p))
        filename_base = f"DDT-{'-'.join(zone_ordered)}-{'-'.join(province)}" if province else f"DDT-{'-'.join(zone_ordered)}"

        luoghi_distinta = sorted({page_luogo.get((str(p), i)) for p, i in pages if page_luogo.get((str(p), i))})
        distinta_luoghi[filename_base] = luoghi_distinta

        writer = PdfWriter()
        seen = set()
        for pdf_path, page_idx in pages:
            key = (str(pdf_path), page_idx)
            if key in seen:
                continue
            seen.add(key)
            try:
                reader = PdfReader(pdf_path)
                writer.add_page(reader.pages[page_idx])
            except Exception as e:
                print(f"  Errore {pdf_path.name} p{page_idx}: {e}")
        out_file = ddt_zona_dir / f"{filename_base}.pdf"
        writer.write(out_file)
        writer.close()
        print(f"  Creato: {out_file.name} ({len(seen)} pagine)")
        creati += 1
        meta = {c: sorted(provs) for c, provs in cp.items()}
        meta["_rientri"] = territorio_rientri.get(territorio, [])
        metadata[filename_base] = meta

    # Aggiorna orari in mappatura (colonne M e N): solo se celle vuote
    row_updates = _costruisci_aggiornamenti_orari(
        territorio_pages, territorio_luogo_gia_aggiunto, page_orari, page_luogo
    )
    if row_updates:
        _aggiorna_mappatura_orari(row_updates)

    return (data_ddt, metadata, distinta_luoghi) if creati > 0 else None


def _genera_report_consegne(cartella: Path, data_ddt: str, distinta_luoghi: dict[str, list[str]], pdf_files: list[Path]) -> Path | None:
    """
    Report punti di consegna (non DDT): stessa riga mappatura = 1 consegna.
    Total e parziali per distinta. Genera PDF, ritorna path per inserimento in DDT assemblato.
    """
    codice_to_punto = _legge_codice_to_punto_consegna()

    def _punti_unici(codici: list[str]) -> set:
        punti = set()
        for cod in codici:
            punto = codice_to_punto.get(cod)
            if punto is not None:
                punti.add(("r", punto))
            else:
                punti.add(("n", cod))
        return punti

    righe_report: list[tuple[str, int]] = []
    tutti_punti: set = set()
    for idx, pdf_path in enumerate(pdf_files):
        filename_base = pdf_path.stem
        luoghi = distinta_luoghi.get(filename_base, [])
        punti = _punti_unici(luoghi)
        n = len(punti)
        num_distinta = f"{idx + 1:02d}"
        zone_suffix = filename_base.replace("DDT-", "", 1) if filename_base.startswith("DDT-") else filename_base
        nome_distinta = f"Distinta {num_distinta} Zona -{zone_suffix}"
        righe_report.append((nome_distinta, n))
        tutti_punti.update(punti)

    totale = len(tutti_punti)
    report_path = cartella / "report_consegne.pdf"

    doc = SimpleDocTemplate(
        str(report_path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    elements = []
    elements.append(Paragraph("Report punti di consegna", ParagraphStyle(name="RpTitolo", fontSize=14, fontName="Helvetica-Bold", spaceAfter=12)))
    elements.append(Paragraph(f"Data DDT: {data_ddt.replace('-', '/')}", ParagraphStyle(name="RpData", fontSize=10, spaceAfter=8)))
    elements.append(Paragraph(f"Totale consegne (punti unici): {totale}", ParagraphStyle(name="RpTot", fontSize=12, fontName="Helvetica-Bold", spaceAfter=12)))
    elements.append(Spacer(1, 6))
    table_data = [["Distinta", "Consegne"]] + [[nome, str(n)] for nome, n in righe_report]
    tbl = Table(table_data, colWidths=[130 * mm, 40 * mm])
    tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)

    def _footer_canvas(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 14)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(A4[0] / 2, 15 * mm, NOTA_FORNITORI)
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer_canvas, onLaterPages=_footer_canvas)
    print(f"\nReport consegne (totale {totale} punti) -> inserito in DDT-{data_ddt}.pdf")
    return report_path


def _costruisci_aggiornamenti_orari(
    territorio_pages: dict,
    territorio_luogo_gia_aggiunto: dict,
    page_orari: dict,
    page_luogo: dict,
) -> dict[int, tuple[str, str]]:
    """
    Per ogni riga mappatura, aggrega orari dai DDT del viaggio.
    Logica: orario_min = più vincolante (earliest), orario_max = min (10:00 se H10, altrimenti 14:00).
    Se due destinazioni nella stessa riga hanno orari diversi: si prende il più vincolante.
    """
    if not MAPPATURA_XLSX.exists():
        return {}
    wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    row_updates = {}
    for row_idx in range(2, ws.max_row + 1):
        cod_f = (ws.cell(row_idx, COL_CODICE_FRUTTA).value or "").strip().lower()
        cod_l = (ws.cell(row_idx, COL_CODICE_LATTE).value or "").strip().lower()
        if not cod_f and not cod_l:
            continue
        codici_riga = {c for c in (cod_f, cod_l) if c and c != CODICE_CIVETTA}
        if not codici_riga:
            continue
        collected = []
        for territorio, luoghi in territorio_luogo_gia_aggiunto.items():
            if not (codici_riga & luoghi):
                continue
            for pdf_path, page_idx in territorio_pages.get(territorio, []):
                key = (str(pdf_path), page_idx)
                luogo = page_luogo.get(key)
                if luogo and luogo in codici_riga:
                    om, oM = page_orari.get(key, ("", ORARIO_MAX_DEFAULT))
                    if om or oM:
                        collected.append((om or "00:00", oM or ORARIO_MAX_DEFAULT))
        if collected:
            orari_min = [x[0] for x in collected if x[0]]
            orari_max = [x[1] for x in collected if x[1]]
            om_final = min(orari_min) if orari_min else ""
            oM_final = min(orari_max) if orari_max else ORARIO_MAX_DEFAULT
            row_updates[row_idx] = (om_final, oM_final)
    wb.close()
    return row_updates


def _assembla_file_unico(cartella: Path, data_ddt: str, pdf_files: list[Path], riepilogo_dir: Path, report_pdf_path: Path | None = None) -> None:
    """
    Crea un unico PDF DDT-(data).pdf nella root della cartella.
    Se presente: report consegne (1 copia) come prima pagina.
    Per ogni viaggio: distinta (2 copie) + pagine DDT zona (2 copie ciascuna).
    """
    out_path = cartella / f"DDT-{data_ddt}.pdf"
    writer = PdfWriter()
    if report_pdf_path and report_pdf_path.exists():
        for page in PdfReader(report_pdf_path).pages:
            writer.add_page(page)
    for idx, pdf_path in enumerate(pdf_files):
        filename_base = pdf_path.stem
        zone_suffix = filename_base.replace("DDT-", "", 1) if filename_base.startswith("DDT-") else filename_base
        num_distinta = f"{idx + 1:02d}"
        distinta_path = riepilogo_dir / f"Distinta {num_distinta} Zona -{zone_suffix}.pdf"
        if distinta_path.exists():
            for page in PdfReader(distinta_path).pages:
                writer.add_page(page)
                writer.add_page(page)
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)
            writer.add_page(page)
    writer.write(out_path)
    writer.close()
    print(f"\nCreato file unico: {out_path.name}")


def main():
    print("Creazione Distinte magazzino...")

    # 0) Crea struttura cartelle se mancante (utile quando si usa l'exe su PC nuovi)
    _crea_struttura_cartelle()

    # 1) Verifica preliminare: cartelle piene e codici in mappatura
    if not _verifica_e_valida_codici():
        return

    # 2) Filtro e unione DDT (crea PDF per territorio in DDT-[data]/)
    risultato = _filtro_e_unione_ddt()
    if not risultato:
        print("ERRORE: Filtro e unione DDT non completato.")
        return
    data_ddt, metadata, distinta_luoghi = risultato

    # 3) Creazione distinte
    cartella = GIRI_LAVORATI_DIR / f"DDT-{data_ddt}"
    ddt_zona_dir = cartella / "DDT-ZONA"
    creati = 0
    pdf_files = [f for f in sorted(ddt_zona_dir.glob("DDT-*.pdf")) if f.parent == ddt_zona_dir]
    if not pdf_files:
        print(f"\n{cartella.name}: nessun PDF territorio, skip distinte.")
    report_pdf_path: Path | None = None
    if pdf_files:
        riepilogo_dir = cartella / "RIEPILOGO"
        riepilogo_dir.mkdir(parents=True, exist_ok=True)
        print(f"\nCreazione distinte ({len(pdf_files)} PDF)")
        for idx, pdf_path in enumerate(pdf_files):
            filename_base = pdf_path.stem
            num_distinta = f"{idx + 1:02d}"
            zone_suffix = filename_base.replace("DDT-", "", 1) if filename_base.startswith("DDT-") else filename_base
            zone_nums = re.findall(r"\d{4}", filename_base)
            articoli = _raccogli_articoli_da_pdf([pdf_path], "misto")
            if not articoli:
                print(f"  {pdf_path.name}: nessun articolo, skip.")
                continue
            aggregati = _aggrega_articoli(articoli)
            nome_file = f"Distinta {num_distinta} Zona -{zone_suffix}.pdf"
            output_path = riepilogo_dir / nome_file
            meta = metadata.get(filename_base, {})
            causali_province = {k: v for k, v in meta.items() if k != "_rientri"}
            rientri_territorio = meta.get("_rientri", []) if isinstance(meta.get("_rientri"), list) else []
            _genera_pdf(aggregati, data_ddt, num_distinta, zone_suffix, zone_nums, output_path, causali_province, rientri_territorio)
            creati += 1

        # Report punti di consegna (PDF, total e parziali per distinta)
        report_pdf_path = _genera_report_consegne(cartella, data_ddt, distinta_luoghi, pdf_files)

    # 4) Crea cartella DDT-ORIGINALI e sposta i PDF frutta e latte (rimuovendoli dalle cartelle sorgente)
    originali_dir = cartella / "DDT-ORIGINALI"
    originali_dir.mkdir(parents=True, exist_ok=True)
    pdf_frutta = list(DDT_FRUTTA_DIR.glob("*.pdf")) if DDT_FRUTTA_DIR.exists() else []
    pdf_latte = list(DDT_LATTE_DIR.glob("*.pdf")) if DDT_LATTE_DIR.exists() else []
    spostati = 0
    for p in pdf_frutta + pdf_latte:
        try:
            shutil.move(str(p), str(originali_dir / p.name))
            spostati += 1
        except Exception as e:
            print(f"  Errore spostamento {p.name}: {e}")
    if spostati > 0:
        print(f"\nSpostati {spostati} PDF originali in {originali_dir.name}/")

    # 5) Assemblaggio file unico DDT-(data).pdf nella root della cartella
    if pdf_files and PdfWriter and PdfReader:
        _assembla_file_unico(cartella, data_ddt, pdf_files, riepilogo_dir, report_pdf_path)
        if report_pdf_path and report_pdf_path.exists():
            try:
                report_pdf_path.unlink()
            except OSError:
                pass

    print(f"\nCompletato: {creati} distinte create in RIEPILOGO.")


if __name__ == "__main__":
    main()
