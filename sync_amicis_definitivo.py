import json
from openpyxl import load_workbook
import string

f2 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026_copia/punti_consegna_unificati.json'
with open(f2, encoding='utf-8') as f: j2 = json.load(f)

punti_veri = []
for p in j2.get('punti', []):
    if p.get('lat') and 'amicis' in p['nome'].lower():
        punti_veri.append(p)

map_file = 'G:/Il mio Drive/App/AppLogSolution/dati/PROGRAMMA/mappatura_destinazioni.xlsx'
wb = load_workbook(map_file)
ws = wb.active
headers = [str(c.value).strip().lower() for c in ws[1]]
col_nome = next(i for i, h in enumerate(headers) if h in ['a chi va consegnato', 'nome'])
col_ind = next(i for i, h in enumerate(headers) if h in ['indirizzo', 'via'])
col_lat = next(i for i, h in enumerate(headers) if h == 'latitudine')
col_lon = next(i for i, h in enumerate(headers) if h == 'longitudine')

saved = 0
for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    n = str(row[col_nome].value).strip().lower()
    i = str(row[col_ind].value).strip().lower()
    
    if n == 'edmondo de amicis':
        # exact coordinate assignment based on string fragments
        lat, lon = None, None
        if 'vittorio' in i:
            lat, lon = 45.42695237766134, 12.077898433043623
        elif 'caltana' in i:
            lat, lon = 45.46527050981497, 12.11488789812301
        elif 'battisti' in i:
            lat, lon = 45.39191463806938, 12.022580558576914
        
        if lat and lon:
            ws.cell(row=r_idx, column=col_lat+1, value=lat)
            ws.cell(row=r_idx, column=col_lon+1, value=lon)
            print(f'Sistemato {i}: {lat}, {lon}')
            saved += 1

wb.save(map_file)
print(f'\\nFATTO DEFINITIVO! {saved}')
