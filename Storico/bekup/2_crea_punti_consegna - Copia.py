#!/usr/bin/env python3
"""
Crea un unico punti_consegna.xlsx nella cartella
CONSEGNE/CONSEGNE_{data}/, aggregando FRUTTA e LATTE.

Evita duplicati: se lo stesso cliente ha DDT in entrambe le cartelle,
aggiorna le colonne Codice Frutta e Codice Latte nella stessa riga.

Uso: py crea_punti_consegna.py <data>   (es. 16-03-2026)
"""

import re
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
MAPPATURA = BASE_DIR / "mappatura_destinazioni.xlsx"

DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'[Ll]uogo [Dd]i [Dd]estinazione:\s*([pP]\d{4,5})')
CAUSALE_SEZIONE_MARKER = "CAUSALE DEL TRASPORTO"
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})', re.I)



def _val(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s


def _estrai_data_luogo_zona(text: str) -> tuple[str | None, str | None, str | None]:
    """Estrae (data, luogo, zona) da testo pagina. data formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    
    idx = text.upper().find(CAUSALE_SEZIONE_MARKER.upper())
    zona = ""
    if idx >= 0:
        sezione = text[idx:idx+200]
        m_z = CAUSALE_RE.search(sezione)
        if m_z: zona = m_z.group(1)[1:5]
        
    return (data, luogo, zona)


def _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov):
    ind = _val(vals[col_ind]) if col_ind < len(vals) else ""
    cap_val = vals[col_cap] if col_cap < len(vals) else None
    cap = str(int(cap_val)) if cap_val is not None and isinstance(cap_val, (int, float)) and not (isinstance(cap_val, float) and str(cap_val) == "nan") else (_val(cap_val) or "")
    citta = _val(vals[col_citta]) if col_citta < len(vals) else ""
    prov = _val(vals[col_prov]) if col_prov < len(vals) else ""
    parts = []
    if ind:
        parts.append(ind)
    if cap or citta:
        loc = f"{cap} {citta}".strip()
        if prov:
            loc = f"{loc} ({prov})"
        parts.append(loc)
    return ", ".join(parts) if parts else ""


def _carica_mappatura():
    """Carica mappatura: codice -> (row_idx, dato con cod_f, cod_l, nome, indirizzo, lat, lon, orari)."""
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    col_cod_f = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Frutta"), 0)
    col_cod_l = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Latte"), 1)
    col_nome = next((i for i, h in enumerate(headers) if "chi va" in str(h or "").lower()), 2)
    col_ind = next((i for i, h in enumerate(headers) if h == "Indirizzo"), 4)
    col_cap = next((i for i, h in enumerate(headers) if h == "CAP"), 5)
    col_citta = next((i for i, h in enumerate(headers) if h == "Città"), 6)
    col_prov = next((i for i, h in enumerate(headers) if h == "Provincia"), 7)
    col_om = next((i for i, h in enumerate(headers) if h == "Orario min"), 10)
    col_oM = next((i for i, h in enumerate(headers) if h == "Orario max"), 11)
    col_lat = next((i for i, h in enumerate(headers) if h == "Latitudine"), 12)
    col_lon = next((i for i, h in enumerate(headers) if h == "Longitudine"), 13)

    map_codice = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row]
        try:
            lat = float(vals[col_lat]) if col_lat < len(vals) and vals[col_lat] is not None else None
            lon = float(vals[col_lon]) if col_lon < len(vals) and vals[col_lon] is not None else None
        except (TypeError, ValueError, IndexError):
            lat, lon = None, None
        cod_f = _val(vals[col_cod_f]) if col_cod_f < len(vals) else ""
        cod_l = _val(vals[col_cod_l]) if col_cod_l < len(vals) else ""
        dato = {
            "codice_frutta": cod_f.lower() if cod_f else "",
            "codice_latte": cod_l.lower() if cod_l else "",
            "nome": _val(vals[col_nome]) if col_nome < len(vals) else "",
            "indirizzo": _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov),
            "orario_min": _val(vals[col_om]) if col_om < len(vals) else "",
            "orario_max": _val(vals[col_oM]) if col_oM < len(vals) else "",
            "lat": lat,
            "lon": lon,
        }
        for cod in (cod_f, cod_l):
            if cod:
                c = cod.lower()
                if c not in map_codice:
                    map_codice[c] = (row_idx, dato)
    wb.close()
    return map_codice


def _elabora_cartella(cartella_input: Path, map_codice: dict) -> list[dict]:
    """Estrae i DDT dalla cartella e ritorna lista di punti trovati."""
    import pdfplumber
    punti_per_riga: dict[int, dict] = {}

    for pdf_path in sorted(cartella_input.glob("*.pdf")):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    _, luogo, zona = _estrai_data_luogo_zona(text)
                    if not luogo:
                        continue
                    if luogo not in map_codice:
                        continue
                    row_idx, dato = map_codice[luogo]
                    if row_idx not in punti_per_riga:
                        punti_per_riga[row_idx] = {
                            **dato,
                            "codici_ddt_trovati": [luogo],
                            "zona": zona
                        }
                    else:
                        punti_per_riga[row_idx]["codici_ddt_trovati"].append(luogo)
                        if zona: punti_per_riga[row_idx]["zona"] = zona
                        # aggiorna i codici Frutta / Latte se presenti
                        if dato["codice_frutta"]:
                            punti_per_riga[row_idx]["codice_frutta"] = dato["codice_frutta"]
                        if dato["codice_latte"]:
                            punti_per_riga[row_idx]["codice_latte"] = dato["codice_latte"]
        except Exception as e:
            print(f"  Errore {pdf_path.name}: {e}")

    # unisci codici DDT trovati senza duplicati
    for pt in punti_per_riga.values():
        pt["codici_ddt_trovati"] = ", ".join(sorted(set(pt["codici_ddt_trovati"])))

    return list(punti_per_riga.values())


def _salva_excel(punti: list[dict], out_path: Path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Punti consegna"
    headers = [
        "Codice Frutta", "Codice Latte", "Codici DDT trovati", "Zona",
        "Nome", "Indirizzo", "Orario min", "Orario max", "Latitudine", "Longitudine"
    ]
    ws.append(headers)
    for pt in punti:
        ws.append([
            pt.get("codice_frutta", ""),
            pt.get("codice_latte", ""),
            pt.get("codici_ddt_trovati", ""),
            pt.get("zona", ""),
            pt.get("nome", ""),
            pt.get("indirizzo", ""),
            pt.get("orario_min", ""),
            pt.get("orario_max", ""),
            pt.get("lat") if pt.get("lat") is not None else "",
            pt.get("lon") if pt.get("lon") is not None else "",
        ])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Salvato: {out_path}")


def main():
    if len(sys.argv) < 2:
        # prendi l'ultima cartella
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py crea_punti_consegna.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name.split("_")[1]
        print(f"Nessuna data specificata. Uso l'ultima cartella: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data):
        data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    input_root = output_base / "DDT-ORIGINALI-DIVISI" if (output_base / "DDT-ORIGINALI-DIVISI").exists() else output_base
    input_frutta = input_root / "FRUTTA"
    input_latte = input_root / "LATTE"
    output_file = output_base / "punti_consegna.xlsx"

    print(f"\n--- Creazione punti consegna unificati ---\nCartella: {output_base}\n")

    if not MAPPATURA.exists():
        print(f"Mappatura non trovata: {MAPPATURA}")
        return 1

    map_codice = _carica_mappatura()
    print(f"Mappatura caricata: {len(map_codice)} codici\n")

    punti_totali = []
    for cartella in [input_frutta, input_latte]:
        if cartella.exists():
            punti_totali.extend(_elabora_cartella(cartella, map_codice))
        else:
            print(f"  Cartella non trovata: {cartella}")

    # elimina duplicati basati su nome + indirizzo
    punti_unici: dict[tuple[str, str], dict] = {}
    for pt in punti_totali:
        chiave = (pt["nome"], pt["indirizzo"])
        if chiave in punti_unici:
            # aggiorna codici e DDT trovati
            esistente = punti_unici[chiave]
            if pt.get("codice_frutta"):
                esistente["codice_frutta"] = pt["codice_frutta"]
            if pt.get("codice_latte"):
                esistente["codice_latte"] = pt["codice_latte"]
            esistente["codici_ddt_trovati"] = ", ".join(
                sorted(set(esistente["codici_ddt_trovati"].split(", ") + pt["codici_ddt_trovati"].split(", ")))
            )
        else:
            punti_unici[chiave] = pt

    _salva_excel(list(punti_unici.values()), output_file)
    print("\n--- Completato ---\n")
    return 0


if __name__ == "__main__":
    exit(main())