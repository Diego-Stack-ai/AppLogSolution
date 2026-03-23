#!/usr/bin/env python3
"""
3_crea_lista_unificata.py (versione aggiornata)
Legge un unico file punti_consegna.xlsx e gestisce i rientri da rientri_ddt.xlsx.
Produce punti_consegna_unificati.json.
"""

import json
import re
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"

def _val(x):
    if x is None: return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

def _carica_excel(path: Path):
    if not path.exists(): return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = []
    headers = [str(c.value or "").split("-")[-1].strip().lower().replace(" ", "_") for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        d = {}
        for i, h in enumerate(headers):
            if i < len(vals): d[h] = vals[i]
        data.append(d)
    wb.close()
    return data

def _carica_rientri(map_codice: dict):
    path = BASE_DIR / "rientri_ddt.xlsx"
    if not path.exists(): return {}
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rientri_per_riga = {}
    for row in ws.iter_rows(min_row=2):
        cod_r = _val(row[0].value)
        if not cod_r: continue
        stato = _val(row[2].value) if len(row) > 2 else ""
        if stato: continue
        c = cod_r.lower()
        if c in map_codice:
            row_idx, _ = map_codice[c]
            if row_idx not in rientri_per_riga: rientri_per_riga[row_idx] = []
            if c not in rientri_per_riga[row_idx]:
                rientri_per_riga[row_idx].append(c)
    wb.close()
    return rientri_per_riga

def _carica_mappatura_veloce():
    path = BASE_DIR / "mappatura_destinazioni.xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    res = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        c_f = _val(row[0].value).lower()
        c_l = _val(row[1].value).lower()
        if c_f: res[c_f] = (r_idx, "F")
        if c_l: res[c_l] = (r_idx, "L")
    wb.close()
    return res

def main():
    if len(sys.argv) < 2:
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py 3_crea_lista_unificata.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name.split("_")[1]
        print(f"Nessuna data specificata. Uso l'ultima cartella trovata: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data): data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    file_punti = output_base / "punti_consegna.xlsx"
    out_json = output_base / "punti_consegna_unificati.json"

    print(f"\n--- Unificazione punti consegna e RIENTRI ({data}) ---")

    punti = _carica_excel(file_punti)
    map_codice = _carica_mappatura_veloce()
    rientri_globale = _carica_rientri(map_codice)

    unificati = [] # Cambiato in lista per evitare raggruppamenti
    # Mappa per agganciare velocemente i rientri ai punti (idx_mappa -> lista di riferimenti ai punti)
    map_idx_mappa_to_punti = {}

    def add_punto(p, row_idx=None):
        punto = {
            "nome": _val(p.get("nome")),
            "indirizzo": _val(p.get("indirizzo")),
            "codice_frutta": _val(p.get("codice_frutta")),
            "codice_latte": _val(p.get("codice_latte")),
            "data_consegna": _val(p.get("data_consegna")),
            "orario_min": _val(p.get("orario_min")),
            "orario_max": _val(p.get("orario_max")),
            "lat": p.get("latitudine"),
            "lon": p.get("longitudine"),
            "codici_ddt_frutta": [],
            "codici_ddt_latte": [],
            "rientri_alert": [],
            "row_idx_mappatura": row_idx
        }
        
        ddt_str = _val(p.get("codici_ddt_trovati"))
        ddts = [x.strip().lower() for x in ddt_str.split(",") if x.strip()]
        for d in ddts:
            if punto.get("codice_frutta"):
                punto["codici_ddt_frutta"].append(d)
            if punto.get("codice_latte"):
                punto["codici_ddt_latte"].append(d)
        
        unificati.append(punto)
        
        if row_idx is not None:
            if row_idx not in map_idx_mappa_to_punti:
                map_idx_mappa_to_punti[row_idx] = []
            map_idx_mappa_to_punti[row_idx].append(punto)

    # Carica punti da Excel
    for p in punti:
        c = _val(p.get("codice_frutta")).lower()
        idx = map_codice[c][0] if c in map_codice else None
        add_punto(p, idx)

    # Gestione rientri
    for idx_mappa, codici in rientri_globale.items():
        if idx_mappa in map_idx_mappa_to_punti:
            # Aggancia il rientro a TUTTI i punti che hanno questo indice di mappatura
            for up in map_idx_mappa_to_punti[idx_mappa]:
                current_codes = (up.get("codici_ddt_frutta") or []) + (up.get("codici_ddt_latte") or [])
                for cr in codici:
                    status = "yellow" if cr in current_codes else "red"
                    # Evita duplicati di alert
                    if not any(a["codice"] == cr for a in up["rientri_alert"]):
                        up["rientri_alert"].append({"codice": cr, "status": status})
        else:
            # Nuovo punto SOLO per rientro (non presente in consegne)
            path_m = BASE_DIR / "mappatura_destinazioni.xlsx"
            from openpyxl import load_workbook
            wb_m = load_workbook(path_m, read_only=True, data_only=True)
            ws_m = wb_m.active
            row = list(ws_m.iter_rows(min_row=idx_mappa, max_row=idx_mappa))[0]
            nome_r = _val(row[2].value) or f"Rientro {codici[0]}"
            ind_r = _val(row[4].value)
            punto_r = {
                "nome": nome_r,
                "indirizzo": ind_r,
                "codice_frutta": _val(row[0].value),
                "codice_latte": _val(row[1].value),
                "data_consegna": data,
                "orario_min": _val(row[10].value),
                "orario_max": _val(row[11].value),
                "lat": row[12].value,
                "lon": row[13].value,
                "codici_ddt_frutta": [],
                "codici_ddt_latte": [],
                "rientri_alert": [{"codice": cr, "status": "red"} for cr in codici],
                "row_idx_mappatura": idx_mappa
            }
            unificati.append(punto_r)
            wb_m.close()

    lista_finale = unificati
    mappati = sum(1 for p in lista_finale if p.get("row_idx_mappatura") is not None)
    fallback = len(lista_finale) - mappati

    for p in lista_finale:
        p["geo_query_nome_indirizzo"] = f"{p['nome']} {p['indirizzo']}"
        p["geo_query_indirizzo"] = p["indirizzo"]

    out_json.write_text(json.dumps({"data": data, "punti": lista_finale}, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Punti totali: {len(lista_finale)} (Mappati: {mappati}, Fuori mappa: {fallback})")
    print(f"  Rientri agganciati: {sum(len(p['rientri_alert']) for p in lista_finale)}")
    print(f"  Salvato: {out_json.name}")
    print("--- Completato ---\n")
    return 0

if __name__ == "__main__":
    exit(main())