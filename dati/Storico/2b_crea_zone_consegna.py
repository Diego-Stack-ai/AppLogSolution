#!/usr/bin/env python3
"""
2b_crea_zone_consegna.py (Versione Fruit-First)
Logica:
1. Carica i DDT Frutta -> Estraggono la Causale Trasporto (ID Zona).
2. Carica i DDT Latte -> Si abbinano alle zone frutta per destinazione (p####).
3. Residuo Latte -> Creano zone standalone solo se non c'è frutta quel giorno.
Output: 2b_crea_zone_consegna.json
"""

import json
import re
import sys
from pathlib import Path
from collections import defaultdict

BASE_DIR = Path(__file__).resolve().parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
MAPPATURA_XLSX = BASE_DIR / "mappatura_destinazioni.xlsx"

# Regex per estrarre Luogo e Causale (Zona) dai DDT
LUOGO_RE = re.compile(r'[Ll]uogo [Dd]i [Dd]estinazione:\s*([pP]\d{4,5})')
CAUSALE_SEZIONE_MARKER = "CAUSALE DEL TRASPORTO"
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})', re.I)

def _estrai_info_da_pdf(pdf_path):
    import pdfplumber
    risultati = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                luogo_m = LUOGO_RE.search(text)
                if not luogo_m: continue
                luogo = luogo_m.group(1).lower()
                
                idx = text.upper().find(CAUSALE_SEZIONE_MARKER.upper())
                zona = ""
                if idx >= 0:
                    sezione = text[idx:idx+200]
                    m = CAUSALE_RE.search(sezione)
                    if m: zona = m.group(1)[1:5]
                
                risultati.append((luogo, zona))
    except Exception as e:
        print(f"  Errore nel leggere {pdf_path.name}: {e}")
    return risultati

def _carica_abbinamenti():
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
    ws = wb.active
    latte_to_frutta = {}
    for row in ws.iter_rows(min_row=2):
        cf = str(row[0].value or "").strip().lower()
        cl = str(row[1].value or "").strip().lower()
        if cf and cl: latte_to_frutta[cl] = cf
    wb.close()
    return latte_to_frutta

def main():
    if len(sys.argv) < 2:
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders: return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name.split("_")[1]
    else:
        data = sys.argv[1].strip()
    
    if re.match(r"^\d{2}-\d{2}$", data): data = f"{data}-2026"
    
    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    input_root = output_base / "DDT-ORIGINALI-DIVISI" if (output_base / "DDT-ORIGINALI-DIVISI").exists() else output_base
    input_frutta = input_root / "FRUTTA"
    input_latte = input_root / "LATTE"
    
    if not input_frutta.exists() or not input_latte.exists():
        print(f"Cartelle non trovate in {output_base}")
        return 1

    print(f"\n--- Creazione ZONE Consegna (Fruit-First) [{data}] ---")
    
    latte_to_frutta = _carica_abbinamenti()
    
    # 1. Frutta Master
    zone_dict = defaultdict(lambda: {"punti": set()})
    punto_to_zona = {}
    
    for p in input_frutta.glob("*.pdf"):
        for luogo, zona in _estrai_info_da_pdf(p):
            if not zona: continue
            zone_dict[zona]["punti"].add(luogo)
            punto_to_zona[luogo] = zona

    # 2. Latte Integration
    latte_standalone = defaultdict(list)
    for p in input_latte.glob("*.pdf"):
        for luogo, zona_latte in _estrai_info_da_pdf(p):
            target_f = latte_to_frutta.get(luogo)
            # Verifica se il punto latte (o il suo equivalente frutta) ha una zona frutta oggi
            zona_f = punto_to_zona.get(luogo) or (punto_to_zona.get(target_f) if target_f else None)
            
            if zona_f:
                # Il latte segue la frutta
                zone_dict[zona_f]["punti"].add(luogo)
            else:
                # Il latte crea la sua zona standalone
                latte_standalone[zona_latte].append(luogo)

    # 3. Finalize
    final_output = []
    # Ordiniamo per ID zona
    for zid in sorted(zone_dict.keys()):
        final_output.append({
            "id_zona": zid,
            "nome_zona": f"Zona {zid}",
            "codici_luogo": sorted(list(zone_dict[zid]["punti"])),
            "tipologia": "mista/frutta"
        })
    for zid in sorted(latte_standalone.keys()):
        final_output.append({
            "id_zona": zid,
            "nome_zona": f"Zona {zid} (Solo Latte)",
            "codici_luogo": sorted(list(set(latte_standalone[zid]))),
            "tipologia": "solo_latte"
        })

    out_file = output_base / "2b_crea_zone_consegna.json"
    out_file.write_text(json.dumps(final_output, indent=2, ensure_ascii=False), encoding="utf-8")
    
    print(f"Analizzate {len(zone_dict)} zone frutta/miste.")
    print(f"Analizzate {len(latte_standalone)} zone residue solo latte.")
    print(f"Salvato: {out_file.name}")
    return 0

if __name__ == "__main__":
    exit(main())
