"""
Aggiunge le 2 nuove colonne orario latte a mappatura_destinazioni.xlsx
e rinomina quelle esistenti da "Orario min/max" a "Orario min/max Frutta".
"""
from pathlib import Path
from openpyxl import load_workbook

MAPPATURA = Path(__file__).resolve().parent / "mappatura_destinazioni.xlsx"

wb = load_workbook(MAPPATURA)
ws = wb.active

# Trova posizioni attuali per nome
headers = [str(c.value or "").strip() for c in ws[1]]
print(f"Colonne attuali: {headers}")

col_om = next((i + 1 for i, h in enumerate(headers) if h == "Orario min"), None)
col_oM = next((i + 1 for i, h in enumerate(headers) if h == "Orario max"), None)

if col_om is None or col_oM is None:
    print("ERRORE: Colonne 'Orario min' o 'Orario max' non trovate.")
    exit(1)

print(f"  'Orario min' trovata in colonna {col_om}")
print(f"  'Orario max' trovata in colonna {col_oM}")

# 1. Rinomina le colonne esistenti aggiungendo " Frutta"
ws.cell(row=1, column=col_om).value = "Orario min Frutta"
ws.cell(row=1, column=col_oM).value = "Orario max Frutta"
print("  Rinominate: 'Orario min Frutta' e 'Orario max Frutta'")

# 2. Inserisce 2 nuove colonne DOPO "Orario max Frutta"
#    Le colonne a destra si spostano automaticamente
insert_after = col_oM + 1
ws.insert_cols(insert_after, 2)
ws.cell(row=1, column=insert_after).value     = "Orario min Latte"
ws.cell(row=1, column=insert_after + 1).value = "Orario max Latte"
print(f"  Inserite: 'Orario min Latte' (col {insert_after}) e 'Orario max Latte' (col {insert_after+1})")

# 3. Verifica finale
headers_nuovi = [str(c.value or "").strip() for c in ws[1]]
print(f"\nColonne finali:")
for i, h in enumerate(headers_nuovi, 1):
    tag = " ← NUOVA" if h in ("Orario min Latte", "Orario max Latte") else \
          " ← rinominata" if h in ("Orario min Frutta", "Orario max Frutta") else ""
    if h: print(f"  Col {i:>2}: {h}{tag}")

wb.save(MAPPATURA)
print(f"\nSalvato: {MAPPATURA.name}")
