from openpyxl import load_workbook
from pathlib import Path

PROG_DIR = Path(r'g:\Il mio Drive\App\AppLogSolutionLocale\dati\PROGRAMMA')
backup  = PROG_DIR / 'mappatura_destinazioni - Copia.xlsx'
attuale = PROG_DIR / 'mappatura_destinazioni.xlsx'

CLIENTI_DA_RIPRISTINARE = [
    'LORENZO MILANI - FARRA CAP.',
    'G.ANCILLOTTO - SOLIGO',
    'Gianni Rodari - Col San Martino',
    'D. ALIGHIERI',
    'LA NOSTRA FAMIGLIA',
    '"GIANNI RODARI" - PARE\'',
    'CAMPOLONGO'
]

print("Lettura coordinate dal backup...")
wb_b = load_workbook(backup, read_only=True, data_only=True)
ws_b = wb_b.active

coord_backup = {}
for row in ws_b.iter_rows(min_row=2):
    vals = [c.value for c in row]
    nome = str(vals[2] or '').strip()
    if any(c in nome for c in CLIENTI_DA_RIPRISTINARE):
        lat = vals[12] if len(vals) > 12 else None
        lon = vals[13] if len(vals) > 13 else None
        if lat and lon:
            coord_backup[nome] = (lat, lon)
wb_b.close()

print(f"Trovate {len(coord_backup)} coordinate nel backup.")

print("\nRipristino su mappatura attuale...")
wb_a = load_workbook(attuale)
ws_a = wb_a.active

headers = [str(c.value or '').strip() for c in ws_a[1]]
col_lat = next((i + 1 for i, h in enumerate(headers) if h == 'Latitudine'), 15)
col_lon = next((i + 1 for i, h in enumerate(headers) if h == 'Longitudine'), 16)

ripristinati = 0
for row_idx, row in enumerate(ws_a.iter_rows(min_row=2), start=2):
    vals = [c.value for c in row]
    nome = str(vals[2] or '').strip()
    if nome in coord_backup:
        lat, lon = coord_backup[nome]
        ws_a.cell(row=row_idx, column=col_lat, value=lat)
        ws_a.cell(row=row_idx, column=col_lon, value=lon)
        print(f"  OK Ripristinato: {nome:<35} -> {lat}, {lon}")
        ripristinati += 1

wb_a.save(attuale)
print(f"\nOperazione completata. {ripristinati} clienti ripristinati.")
