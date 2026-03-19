#!/usr/bin/env python3
"""
Crea punti_consegna_frutta.xlsx e punti_consegna_latte.xlsx nella cartella
CONSEGNE/CONSEGNE_{data}/.

Legge i DDT da CONSEGNE/CONSEGNE_{data}/DDT-ORIGINALI-DIVISI/FRUTTA e LATTE.
Per ogni DDT: estrae Codice luogo e data. Lookup in mappatura: stessa riga = stesso punto.
Output: entrambi i codici (A, B) per uso futuro.

Uso: py crea_punti_consegna.py <data>   (es. 16-03-2026)
"""

import re
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
MAPPATURA = BASE_DIR / "mappatura_destinazioni.xlsx"

DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'[Ll]uogo [Dd]i [Dd]estinazione:\s*([pP]\d{4,5})')


def _val(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s


def _estrai_data_luogo(text: str) -> tuple[str | None, str | None]:
    """Estrae (data, luogo) da testo pagina. data formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    return (data, luogo)


def _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov):
    ind = _val(vals[col_ind]) if col_ind < len(vals) else ""
    cap_val = vals[col_cap] if col_cap < len(vals) else None
    cap = str(int(cap_val)) if cap_val is not None and isinstance(cap_val, (int, float)) and not (isinstance(cap_val, float) and str(cap_val) == "nan") else (_val(cap_val) or "")
    citta = _val(vals[col_citta]) if col_citta < len(vals) else ""
    prov = _val(vals[col_prov]) if col_prov < len(vals) else ""
    parts = []
    if ind:
        parts.append(ind)
    if cap or citta:
        loc = f"{cap} {citta}".strip()
        if prov:
            loc = f"{loc} ({prov})"
        parts.append(loc)
    return ", ".join(parts) if parts else ""


def _carica_mappatura():
    """Carica mappatura: codice -> (row_idx, dato con cod_f, cod_l, nome, indirizzo, lat, lon, orari)."""
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    col_cod_f = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Frutta"), 0)
    col_cod_l = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Latte"), 1)
    col_nome = next((i for i, h in enumerate(headers) if "chi va" in str(h or "").lower()), 2)
    col_ind = next((i for i, h in enumerate(headers) if h == "Indirizzo"), 4)
    col_cap = next((i for i, h in enumerate(headers) if h == "CAP"), 5)
    col_citta = next((i for i, h in enumerate(headers) if h == "Città"), 6)
    col_prov = next((i for i, h in enumerate(headers) if h == "Provincia"), 7)
    col_om = next((i for i, h in enumerate(headers) if h == "Orario min"), 10)
    col_oM = next((i for i, h in enumerate(headers) if h == "Orario max"), 11)
    col_lat = next((i for i, h in enumerate(headers) if h == "Latitudine"), 12)
    col_lon = next((i for i, h in enumerate(headers) if h == "Longitudine"), 13)

    map_codice = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row]
        try:
            lat = float(vals[col_lat]) if col_lat < len(vals) and vals[col_lat] is not None else None
            lon = float(vals[col_lon]) if col_lon < len(vals) and vals[col_lon] is not None else None
        except (TypeError, ValueError, IndexError):
            lat, lon = None, None
        cod_f = _val(vals[col_cod_f]) if col_cod_f < len(vals) else ""
        cod_l = _val(vals[col_cod_l]) if col_cod_l < len(vals) else ""
        dato = {
            "codice_frutta": cod_f.lower() if cod_f else "",
            "codice_latte": cod_l.lower() if cod_l else "",
            "nome": _val(vals[col_nome]) if col_nome < len(vals) else "",
            "indirizzo": _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov),
            "orario_min": _val(vals[col_om]) if col_om < len(vals) else "",
            "orario_max": _val(vals[col_oM]) if col_oM < len(vals) else "",
            "lat": lat,
            "lon": lon,
        }
        for cod in (cod_f, cod_l):
            if cod:
                c = cod.lower()
                if c not in map_codice:
                    map_codice[c] = (row_idx, dato)
    wb.close()
    return map_codice


def _elabora_cartella(cartella_input: Path, map_codice: dict) -> tuple[list[dict], int, int]:
    """
    Per ogni PDF nella cartella: estrae (codice, data), lookup mappatura.
    Raggruppa per riga mappatura (stesso punto). Ritorna (lista punti, mancanti, totale_ddt).
    """
    import pdfplumber
    punti_per_riga: dict[int, dict] = {}
    codici_mancanti = []
    totale_ddt = 0

    for pdf_path in sorted(cartella_input.glob("*.pdf")):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    data, luogo = _estrai_data_luogo(text)
                    if not luogo:
                        continue
                    totale_ddt += 1
                    if luogo not in map_codice:
                        codici_mancanti.append(luogo)
                        continue
                    row_idx, dato = map_codice[luogo]
                    if row_idx not in punti_per_riga:
                        punti_per_riga[row_idx] = {
                            **dato,
                            "data_consegna": data or "",
                            "codici_ddt_trovati": [],
                        }
                    punti_per_riga[row_idx]["codici_ddt_trovati"].append(luogo)
                    if data and not punti_per_riga[row_idx]["data_consegna"]:
                        punti_per_riga[row_idx]["data_consegna"] = data
        except Exception as e:
            print(f"  Errore {pdf_path.name}: {e}")

    for pt in punti_per_riga.values():
        pt["codici_ddt_trovati"] = ", ".join(sorted(set(pt["codici_ddt_trovati"])))

    return (list(punti_per_riga.values()), codici_mancanti, totale_ddt)


def _salva_excel(punti: list[dict], out_path: Path, etichetta: str):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Punti consegna"
    col_refs = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    headers = [
        "Codice Frutta", "Codice Latte", "Data Consegna", "Codici DDT trovati",
        "Nome", "Indirizzo", "Orario min", "Orario max", "Latitudine", "Longitudine"
    ]
    ws.append([f"{col_refs[i]} - {h}" for i, h in enumerate(headers)])
    for pt in punti:
        ws.append([
            pt.get("codice_frutta", ""),
            pt.get("codice_latte", ""),
            pt.get("data_consegna", ""),
            pt.get("codici_ddt_trovati", ""),
            pt.get("nome", ""),
            pt.get("indirizzo", ""),
            pt.get("orario_min", ""),
            pt.get("orario_max", ""),
            pt.get("lat") if pt.get("lat") is not None else "",
            pt.get("lon") if pt.get("lon") is not None else "",
        ])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Salvato: {out_path}")


def main():
    if len(sys.argv) < 2:
        # Cerca l'ultima cartella creata
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py 2_crea_punti_consegna.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name.split("_")[1]
        print(f"Nessuna data specificata. Uso l'ultima cartella trovata: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data):
        data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    # Gestione intelligente percorsi: cerca in DDT-ORIGINALI-DIVISI se presente (per archivio storico)
    sub_base = output_base / "DDT-ORIGINALI-DIVISI"
    input_root = sub_base if sub_base.exists() else output_base
    
    input_frutta = input_root / "FRUTTA"
    input_latte = input_root / "LATTE"
    output_frutta = output_base / "punti_consegna_frutta.xlsx"
    output_latte = output_base / "punti_consegna_latte.xlsx"

    print("\n--- Creazione punti consegna (FRUTTA e LATTE) ---\n")
    print(f"Cartella: {output_base}\n")

    if not MAPPATURA.exists():
        print(f"Mappatura non trovata: {MAPPATURA}")
        return 1

    map_codice = _carica_mappatura()
    print(f"Mappatura caricata: {len(map_codice)} codici\n")

    # FRUTTA
    print("FRUTTA:")
    if input_frutta.exists():
        punti_f, manc_f, tot_f = _elabora_cartella(input_frutta, map_codice)
        print(f"  DDT letti: {tot_f} | Punti unici: {len(punti_f)}")
        if manc_f:
            print(f"  Codici non in mappatura: {', '.join(sorted(set(manc_f))[:10])}{'...' if len(set(manc_f)) > 10 else ''}")
        _salva_excel(punti_f, output_frutta, "FRUTTA")
    else:
        print(f"  Cartella non trovata: {input_frutta}")

    print()

    # LATTE
    print("LATTE:")
    if input_latte.exists():
        punti_l, manc_l, tot_l = _elabora_cartella(input_latte, map_codice)
        print(f"  DDT letti: {tot_l} | Punti unici: {len(punti_l)}")
        if manc_l:
            print(f"  Codici non in mappatura: {', '.join(sorted(set(manc_l))[:10])}{'...' if len(set(manc_l)) > 10 else ''}")
        _salva_excel(punti_l, output_latte, "LATTE")
    else:
        print(f"  Cartella non trovata: {input_latte}")

    print("\n--- Completato ---\n")
    return 0


if __name__ == "__main__":
    exit(main())
