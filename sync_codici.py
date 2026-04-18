import json
from openpyxl import load_workbook

f2 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026_copia/punti_consegna_unificati.json'
with open(f2, encoding='utf-8') as f: j2 = json.load(f)

# Prep dictionary based ONLY on Codice Frutta and Codice Latte
punti_veri = {}
for p in j2.get('punti', []):
    if p.get('lat') and p.get('lon'):
        cf = str(p.get('codice_frutta', '')).strip().lower()
        cl = str(p.get('codice_latte', '')).strip().lower()
        if cf == 'p00000': cf = ''
        if cl == 'p00000': cl = ''
        if cf or cl:
            punti_veri[(cf, cl)] = (p['lat'], p['lon'])

map_file = 'G:/Il mio Drive/App/AppLogSolution/dati/PROGRAMMA/mappatura_destinazioni.xlsx'
wb = load_workbook(map_file)
ws = wb.active
headers = [str(c.value).strip().lower() for c in ws[1]]

col_f = next((i for i, h in enumerate(headers) if 'frutta' in h or 'cod. fr' in h), 0)
col_l = next((i for i, h in enumerate(headers) if 'latte' in h or 'cod. la' in h), 1)
col_lat = next((i for i, h in enumerate(headers) if h == 'latitudine'), -1)
col_lon = next((i for i, h in enumerate(headers) if h == 'longitudine'), -1)

saved = 0
for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    rf = str(row[col_f].value or "").strip().lower()
    rl = str(row[col_l].value or "").strip().lower()
    
    # Check if either exact pair or single code matches
    match = punti_veri.get((rf, rl))
    if not match and rf:
        match = punti_veri.get((rf, ''))
    if not match and rl:
        match = punti_veri.get(('', rl))
        
    if match:
        old_lat = row[col_lat].value
        old_lon = row[col_lon].value
        
        # se sono diverse
        if old_lat != match[0] or old_lon != match[1]:
            ws.cell(row=r_idx, column=col_lat+1, value=match[0])
            ws.cell(row=r_idx, column=col_lon+1, value=match[1])
            print(f"Corretto da COPIA tramite codici: {rf}/{rl} -> {match[0]}, {match[1]}")
            saved += 1

wb.save(map_file)
print(f"\\nTOTALE SISTEMATI TRAMITE CODICE: {saved}")
