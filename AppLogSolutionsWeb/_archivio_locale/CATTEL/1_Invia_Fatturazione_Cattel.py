import os
import pandas as pd
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment

# MODALITÀ DI TEST: Se True, il file Excel non verrà salvato. Verranno solo stampate le discrepanze.
DRY_RUN = False

KPI_FILE = r"G:\Il mio Drive\Fatturazione\CATTEL\KPI Report.xlsx"
FATT_FILE = r"G:\Il mio Drive\Fatturazione\FATTURAZIONE 2026.xlsx"

mesi_map = {1: 'GENNAIO', 2: 'FEBBRAIO', 3: 'MARZO', 4: 'APRILE', 5: 'MAGGIO', 6: 'GIUGNO',
            7: 'LUGLIO', 8: 'AGOSTO', 9: 'SETTEMBRE', 10: 'OTTOBRE', 11: 'NOVEMBRE', 12: 'DICEMBRE'}

def clean_name(name):
    return str(name).strip().upper() if pd.notna(name) else ""

def get_macro_zona(zona_str):
    zona_str = str(zona_str).upper()
    has_bs = "BS" in zona_str
    has_la = "LA" in zona_str
    if has_bs and has_la: return "BS ESTERNO"
    if has_la: return "LAGO GARDA"
    if has_bs: return "BS CENTRO"
    return None

def main():
    print("================================================")
    print("--- AUTOMAZIONE FATTURAZIONE CATTEL ---")
    print("Modalità DRY RUN: ON" if DRY_RUN else "Modalità DRY RUN: OFF")
    print("Nessuna modifica verrà scritta sul file." if DRY_RUN else "ATTENZIONE: File verrà modificato.")
    print("================================================\n")

    # 1. Lettura KPI
    print(f"Leggendo {KPI_FILE}...")
    try:
        df_kpi = pd.read_excel(KPI_FILE)
    except Exception as e:
        print(f"Errore lettura KPI: {e}")
        return

    # Convertiamo la data
    df_kpi['Data'] = pd.to_datetime(df_kpi['Data'], format='%d/%m/%Y', errors='coerce')
    df_kpi = df_kpi.dropna(subset=['Data'])
    
    if len(df_kpi) == 0:
        print("Nessun dato valido nel KPI.")
        return

    mese = df_kpi['Data'].dt.month.iloc[0]
    sheet_name = mesi_map.get(mese)
    print(f"Mese rilevato: {sheet_name} ({mese})")

    # Estrazione autisti univoci dal KPI
    unique_drivers = df_kpi['Autista'].dropna().unique()
    unique_drivers = [clean_name(d) for d in unique_drivers]
    print(f"Autisti univoci nel KPI ({len(unique_drivers)}): {unique_drivers}\n")

    # 2. Caricamento Fatturazione
    print(f"Caricamento {FATT_FILE} (richiede qualche secondo)...")
    wb = openpyxl.load_workbook(FATT_FILE)
    if sheet_name not in wb.sheetnames:
        print(f"ERRORE: Foglio {sheet_name} non trovato in FATTURAZIONE 2026.xlsx!")
        return
    ws = wb[sheet_name]

    # Trova blocco CATTEL
    cattel_start_row = None
    colli_row = None
    navetta_row = None
    macro_zones_rows = {}
    
    # Esploriamo colonna A e B
    for r in range(1, 300):
        val_a = clean_name(ws.cell(row=r, column=1).value)
        val_b = clean_name(ws.cell(row=r, column=2).value)
        
        if val_a == "CATTEL":
            cattel_start_row = r
        elif cattel_start_row and val_a in ["LAGO GARDA", "BS CENTRO", "BS ESTERNO"]:
            macro_zones_rows[val_a] = r
        elif cattel_start_row and val_a == "NAVETTA":
            navetta_row = r
        elif cattel_start_row and (val_a == "COLLI" or val_b == "COLLI"):
            colli_row = r
            break
            
    print(f"-> CATTEL Inizia alla riga {cattel_start_row}")
    print(f"-> Zone Trovate: {macro_zones_rows}")
    print(f"-> Riga NAVETTA trovata alla riga {navetta_row}")
    print(f"-> Riga COLLI trovata alla riga {colli_row}\n")

    if not DRY_RUN and cattel_start_row and colli_row:
        print("--- PULIZIA DATI ESISTENTI E AGGIORNAMENTO AUTISTI ---")
        # Pulisce le celle dei giorni (da colonna 4 a 34) per tutte le righe del blocco Cattel
        for r in range(cattel_start_row + 1, colli_row + 1):
            for c in range(4, 35):
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).comment = None
                
        # Scrive i nuovi autisti univoci nelle 3 zone (Colonna B = 2)
        for zona, start_r in macro_zones_rows.items():
            end_r = colli_row
            for r in range(start_r + 1, colli_row + 1):
                val_a = clean_name(ws.cell(row=r, column=1).value)
                if val_a and val_a in ["LAGO GARDA", "BS CENTRO", "BS ESTERNO", "NAVETTA", "COLLI"]:
                    end_r = r
                    break
            
            idx_driver = 0
            for r in range(start_r, end_r):
                if idx_driver < len(unique_drivers):
                    ws.cell(row=r, column=2).value = unique_drivers[idx_driver]
                    idx_driver += 1
                else:
                    ws.cell(row=r, column=2).value = None 

    colli_giornalieri = {}
    targhe_c = ['EK832AW', 'EN201DB', 'EN364DB']

    # Oggetto per l'allineamento centrato
    center_aligned_text = Alignment(horizontal='center', vertical='center')

    print("--- ELABORAZIONE DATI KPI ---")
    # 3. Elaborazione Righe KPI
    for idx, row in df_kpi.iterrows():
        giorno = row['Data'].day
        autista = clean_name(row['Autista'])
        targa = str(row['Codice mezzo']).strip().upper()
        zona = get_macro_zona(row['Zona'])
        destinazioni = row['Destinazioni']
        km = row['Km']
        quantita = row['Quantità']
        peso = row['Peso']
        
        if not zona: continue
        
        # Assegnazione Patente
        patente = "C" if targa in targhe_c else "B"
        
        if pd.notna(quantita):
            try:
                colli_giornalieri[giorno] = colli_giornalieri.get(giorno, 0) + float(quantita)
            except:
                pass
            
        start_r = macro_zones_rows.get(zona)
        if not start_r:
            continue
            
        autista_row = None
        limit_r = colli_row if colli_row else start_r + 50
        for r in range(start_r, limit_r):
            val_a = clean_name(ws.cell(row=r, column=1).value)
            if val_a and val_a != zona and val_a in ["LAGO GARDA", "BS CENTRO", "BS ESTERNO", "NAVETTA", "COLLI"]:
                break
                
            val_b = clean_name(ws.cell(row=r, column=2).value)
            val_b_clean = " ".join(val_b.split())
            aut_clean = " ".join(autista.split())
            
            if val_b_clean == aut_clean or aut_clean in val_b_clean or val_b_clean in aut_clean:
                autista_row = r
                break
                
        if not autista_row:
            continue
            
        col_giorno = 3 + giorno
        cella = ws.cell(row=autista_row, column=col_giorno)
        
        nota_attesa = f"Targa: {targa}\nDestinazioni: {destinazioni}\nKm: {km}\nPeso: {peso}"
        
        if not DRY_RUN:
            cella.value = patente
            cella.comment = Comment(nota_attesa, "Script")
            cella.alignment = center_aligned_text
            

    # Scrittura Colli
    for g, totale_colli in sorted(colli_giornalieri.items()):
        col_giorno = 3 + g
        if not DRY_RUN and colli_row:
            cella_colli = ws.cell(row=colli_row, column=col_giorno)
            cella_colli.value = totale_colli
            cella_colli.alignment = center_aligned_text

    if not DRY_RUN:
        print(f"\nSalvataggio modifiche in corso...")
        wb.save(FATT_FILE)
        print("Modifiche salvate con successo!")
    else:
        print("\n================================================")
        print("FINE DRY RUN: Analisi completata senza modificare il file.")
        print("================================================")

if __name__ == "__main__":
    main()
