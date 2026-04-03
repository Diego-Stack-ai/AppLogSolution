import os
import openpyxl
import re

xlsx_path = r"G:\Il mio Drive\AppLogSolutions\Fatturazione\Anagrafica_Clienti_Master.xlsx"

indirizzi_da_trovare = [
    'PIAZZA SEN. A. ALBERTI4, LAZISE VR',
    'VIA LARGO ARMANDO DIAZ 2, MEL - BORGO VALBELLUNA- BL',
    'VIA TORRE NR.1 SAN MARTINO DELLA BATTAGLIA, DESENZANO DEL GARDA BS',
    'VIA CADOLA 19, PONTE NELLE ALPI BL',
    'VIA PAGNANGHE 5, TREGNAGO VR',
    'VIALE DELLA STAZIONE 1, PONTE NELLE ALPI BL',
    'LOC. SAN VEROLO 1, COSTERMANO SUL GARDA VR',
    'VIA RAVENNA 297, BORCA DI CADORE BL',
    'VIA STRAIBAN 17, VODO CADORE BL',
    'LOC. VIA DA DEL SALE 15, SOMMACAMPAGNA VR',
    'VAL DE LA BRUSSA 26, ERTO E CASSO PN',
    'VIA CORTE GIRARDI 1, ZEVIO VR',
    'VIA BACH 24, SAPPADA BL',
    'BORGATA KRATTEN 8, SAPPADA UD',
    'BORGATA KRATTEN 11, SAPPADA UD',
    'GALLERIA PELLICIANI 12, VERONA VR',
    'STRADA PROVINCIALE 1, CASALOLDO MN',
    'LOCALITA  PONTAROLA 12, MARANO DI VALPOLICELLA VR',
    'PIAZZA S. ANASTASIA 4, VERONA VR',
    'LOC CIMA GOGNA, AURONZO DI CADORE BL',
    'VIA VOLTA 9, POZZOLO MN',
    'LOC. LAGAZUOI, CORTINA D AMPEZZO BL',
    'VIA ELENOIRE ROOSEVELT 1, AURONZO DI CADORE BL',
    'PIAN DEI BUOI SNC, LOZZO DI CADORE BL',
    'VIA MEZZAVALLE 9, TAIBON AGORDINO BL'
]

# Normalizziamo gli indirizzi
def norm(addr):
    return re.sub(r'[^a-zA-Z0-9]', '', addr.lower())

indirizzi_norm = {norm(addr): addr for addr in indirizzi_da_trovare}
risultati = {addr: "Ragione Sociale Non Trovata" for addr in indirizzi_da_trovare}

try:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # Cerchiamo le colonne
    header = [str(x).lower().strip() if x else "" for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx_rs = -1
    idx_indirizzo = -1
    idx_loc = -1
    idx_pr = -1
    
    for i, col in enumerate(header):
        if 'ragione sociale' in col or 'cliente' in col:
            idx_rs = i
        elif 'indirizzo' in col:
            idx_indirizzo = i
        elif 'localit' in col:
            idx_loc = i
        elif col == 'pr' or col == 'pr.' or col == 'provincia':
            idx_pr = i
            
    if idx_rs == -1: idx_rs = 3
    if idx_indirizzo == -1: idx_indirizzo = 4
    if idx_loc == -1: idx_loc = 5
    if idx_pr == -1: idx_pr = 6
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_rs] or not row[idx_indirizzo]: continue
        
        rs = str(row[idx_rs]).strip()
        ind = str(row[idx_indirizzo]).strip()
        loc = str(row[idx_loc]).strip() if idx_loc != -1 and idx_loc < len(row) and row[idx_loc] else ""
        pr = str(row[idx_pr]).strip() if idx_pr != -1 and idx_pr < len(row) and row[idx_pr] else ""
        
        full_a = f"{ind}, {loc} {pr}".strip(", ")
        n_full = norm(full_a)
        
        for k_norm, orig_addr in indirizzi_norm.items():
            if k_norm in n_full or n_full in k_norm:
                risultati[orig_addr] = rs
            elif norm(ind) in k_norm and norm(loc) in k_norm:
                risultati[orig_addr] = rs
                
except Exception as e:
    print(f"Errore: {e}")

for i, addr in enumerate(indirizzi_da_trovare, 1):
    print(f"{i}. {risultati[addr]} - '{addr}'")
