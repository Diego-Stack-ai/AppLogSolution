import openpyxl, requests, urllib.parse, time

API_KEY = 'AIzaSyAHQ3HjuEEIS8bn5KMh6N3UoM6kZ2MYGL4'

wb = openpyxl.load_workbook(
    r'G:\Il mio Drive\AppLogSolutions\Progetto Scuole\PROGRAMMA\mappatura_destinazioni.xlsx',
    data_only=True
)
ws = wb.active
headers = [c.value for c in ws[1]]
print('Colonne:', headers[:15])
print()

senza = []
for idx_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row[0] and not row[1]:
        continue
    lat = row[12] if len(row) > 12 else None
    lon = row[13] if len(row) > 13 else None
    if not lat or not lon:
        nome = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        ind  = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        cap  = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        cap  = str(int(float(cap))) if cap.replace('.','').isdigit() else cap
        cit  = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        prov = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        senza.append({
            'idx_row': idx_row,
            'nome': nome,
            'ind': ind,
            'cap': cap,
            'citta': cit,
            'prov': prov,
        })

print(f'Scuole senza coordinate: {len(senza)}')
print()
for i, s in enumerate(senza, 1):
    addr = f"{s['ind']}, {s['cap']} {s['citta']} ({s['prov']})".strip(', ')
    print(f"{i:2}. {s['nome'][:45]:<45} | {addr}")

print()
print('--- AVVIO GEOCODING GOOGLE ---')
print()

risultati = []
for s in senza:
    addr = f"{s['ind']}, {s['cap']} {s['citta']} ({s['prov']}), Italia".strip(', ')
    query = f"{s['nome']} {addr}" if s['nome'] else addr
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={urllib.parse.quote(query)}&key={API_KEY}"
    try:
        r = requests.get(url, timeout=5).json()
        if r['status'] == 'OK':
            loc = r['results'][0]['geometry']['location']
            loc_type = r['results'][0]['geometry']['location_type']
            formatted = r['results'][0].get('formatted_address', '')
            stato = 'OK' if loc_type != 'APPROXIMATE' else 'APPROX'
            risultati.append({**s, 'lat': loc['lat'], 'lon': loc['lng'], 'stato': stato, 'formatted': formatted})
            print(f"  {'✅' if stato=='OK' else '⚠️ '} {s['nome'][:40]:<40} -> {loc['lat']:.5f},{loc['lng']:.5f} [{loc_type}]")
        else:
            risultati.append({**s, 'lat': None, 'lon': None, 'stato': 'ERRORE', 'formatted': r['status']})
            print(f"  ❌ {s['nome'][:40]:<40} -> {r['status']}")
    except Exception as e:
        risultati.append({**s, 'lat': None, 'lon': None, 'stato': 'ERRORE', 'formatted': str(e)})
        print(f"  ❌ {s['nome'][:40]:<40} -> ERRORE: {e}")
    time.sleep(0.2)

print()
ok   = sum(1 for r in risultati if r['stato'] == 'OK')
appr = sum(1 for r in risultati if r['stato'] == 'APPROX')
err  = sum(1 for r in risultati if r['stato'] == 'ERRORE')
print(f'Trovati precisi: {ok} | Approssimativi: {appr} | Errori: {err}')
