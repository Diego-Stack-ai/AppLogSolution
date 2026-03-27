#!/usr/bin/env python3
"""
3_crea_lista_unificata.py (versione aggiornata)
Legge un unico file punti_consegna.xlsx e gestisce i rientri da rientri_ddt.xlsx.
Produce punti_consegna_unificati.json.
Geocoding automatico per punti senza coordinate:
  - Tentativo 1: "Nome Indirizzo"
  - Tentativo 2: solo "Indirizzo"
I risultati vengono cachati in geocode_cache.json e salvati su mappatura_destinazioni.xlsx.
"""

import json
import re
import sys
import time
from pathlib import Path

PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
GEOCODE_CACHE_PATH = PROG_DIR / "geocode_cache.json"

# --- GEOCODING ---

def _carica_geocache() -> dict:
    if GEOCODE_CACHE_PATH.exists():
        try:
            return json.loads(GEOCODE_CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _salva_geocache(cache: dict):
    GEOCODE_CACHE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def _pulisci_indirizzo(addr: str) -> str:
    """Normalizza abbreviazioni comuni negli indirizzi italiani."""
    s = addr.strip()
    abbrev = {
        r'\bV\.LE\b': 'Viale', r'\bV\.?\s*LE\b': 'Viale',
        r'\bVIA\b': 'Via', r'\bP\.ZZA\b': 'Piazza', r'\bPZA\b': 'Piazza',
        r'\bP\.ZA\b': 'Piazza', r'\bFRAZ\.\b': 'Frazione',
        r'\bLOC\.\b': 'Localita', r'\bS\.\b': 'San',
        r'\bVLE\b': 'Viale', r'\bL\.GO\b': 'Largo',
        r'\bC\.SO\b': 'Corso', r'\bP\.LE\b': 'Piazzale',
    }
    for pat, repl in abbrev.items():
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)
    # Rimuove contenuti tra parentesi (es. "(VICENZA)")
    s = re.sub(r'\([^)]*\)', '', s).strip()
    # Normalizza spazi multipli
    s = re.sub(r'\s+', ' ', s)
    return s

def _genera_varianti_query(indirizzo: str) -> list[str]:
    """Produce varianti progressive di ricerca (da specifica a generica)."""
    addr = _pulisci_indirizzo(indirizzo)
    varianti = [addr]
    # Rimuove numero civico (es. "Via Roma 5" -> "Via Roma")
    senza_civico = re.sub(r'\s+\d+[/\w]*($|,)', ',', addr)
    senza_civico = re.sub(r',\s*,', ',', senza_civico).strip().rstrip(',')
    if senza_civico != addr:
        varianti.append(senza_civico)
    # Estrae solo il CAP + città se presente
    m = re.search(r'(\d{5})\s+([A-Z][^,]+?)(?:\s*,|\s*$)', addr, re.IGNORECASE)
    if m:
        varianti.append(f"{m.group(1)} {m.group(2).strip()}, Italia")
    return list(dict.fromkeys(varianti))  # dedup preservando ordine

def _geocodifica_query(query: str, cache: dict) -> tuple[float, float] | None:
    """Cerca lat/lon per la query. Prima in cache, poi via Nominatim."""
    key = query.strip().lower()
    if key in cache:
        entry = cache[key]
        if entry.get("lat") and entry.get("lon"):
            return (entry["lat"], entry["lon"])
        return None  # già cercato e non trovato

    # Chiamata Nominatim con varianti progressive
    try:
        from geopy.geocoders import Nominatim
        geolocator = Nominatim(user_agent="applog_logistics_v2", timeout=10)
        varianti = _genera_varianti_query(query)

        for variante in varianti:
            time.sleep(1.1)  # rispetta rate limit Nominatim (1 req/sec)
            location = geolocator.geocode(
                variante, exactly_one=True, language="it", country_codes="it"
            )
            if location:
                lat, lon = round(location.latitude, 7), round(location.longitude, 7)
                cache[key] = {"lat": lat, "lon": lon, "status": "found", "query_ok": variante}
                return (lat, lon)

        cache[key] = {"lat": None, "lon": None, "status": "not_found"}
        return None
    except Exception as e:
        print(f"    ⚠️  Geocoding error per '{query[:60]}': {e}")
        return None

def _geocodifica_punto(punto: dict, cache: dict) -> tuple[float, float] | None:
    """
    Prova nome+indirizzo, poi solo indirizzo.
    Ritorna (lat, lon) oppure None.
    """
    q1 = punto.get("geo_query_nome_indirizzo", "").strip()
    q2 = punto.get("geo_query_indirizzo", "").strip()
    if q1:
        result = _geocodifica_query(q1, cache)
        if result:
            return result
    if q2 and q2 != q1:
        return _geocodifica_query(q2, cache)
    return None

def _aggiorna_mappatura_coord(row_idx: int, lat: float, lon: float):
    """Salva le coordinate trovate su mappatura_destinazioni.xlsx (permanente)."""
    path = PROG_DIR / "mappatura_destinazioni.xlsx"
    if not path.exists() or row_idx is None:
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        # Colonne Latitudine=13 (M), Longitudine=14 (N) — adatta se diverse
        row = ws[row_idx]
        ws.cell(row=row_idx, column=13, value=lat)
        ws.cell(row=row_idx, column=14, value=lon)
        wb.save(path)
    except PermissionError:
        print(f"    ⚠️  mappatura_destinazioni.xlsx è aperto — coordinate non salvate sull'Excel.")
    except Exception as e:
        print(f"    ⚠️  Errore salvataggio Excel: {e}")

# --- UTILITY ---


def _val(x):
    if x is None: return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

def _carica_excel(path: Path):
    if not path.exists(): return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = []
    headers = [str(c.value or "").split("-")[-1].strip().lower().replace(" ", "_") for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        d = {}
        for i, h in enumerate(headers):
            if i < len(vals): d[h] = vals[i]
        data.append(d)
    wb.close()
    return data

def _carica_rientri(map_codice: dict):
    """Legge rientri_ddt.xlsx.
    Col A = codice, Col B = data DDT originale (→ cartella CONSEGNE_), Col C = stato.
    Restituisce:
      rientri_per_riga  : {row_idx_mappa: [(codice, data_originale), ...]}
      righe_da_lavorare : [(codice, riga_excel, data_originale), ...]
    """
    path = BASE_DIR / "rientri_ddt.xlsx"
    if not path.exists(): return {}, []
    from openpyxl import load_workbook
    from datetime import datetime as _dt
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rientri_per_riga = {}
    righe_da_lavorare = []  # (codice, riga_excel, data_originale)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cod_r = _val(row[0].value)
        if not cod_r: continue
        stato = _val(row[2].value) if len(row) > 2 else ""

        # Saltiamo solo DDT già assegnati (contengono 'allegato' senza 'lavorazione')
        stato_lower = stato.lower()
        if 'allegato' in stato_lower and 'lavorazione' not in stato_lower:
            continue

        # ── Leggi colonna B: Data DDT originale ──────────────────────────────
        val_b = row[1].value if len(row) > 1 else None
        if val_b:
            if hasattr(val_b, 'strftime'):               # datetime object
                data_orig = val_b.strftime("%d-%m-%Y")
            else:                                         # stringa "YYYY-MM-DD ..."
                try:
                    data_orig = _dt.strptime(str(val_b)[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
                except Exception:
                    data_orig = ""
        else:
            data_orig = ""

        c = cod_r.lower()
        if c in map_codice:
            row_idx_mappa, _ = map_codice[c]
            if row_idx_mappa not in rientri_per_riga:
                rientri_per_riga[row_idx_mappa] = []
            if not any(x[0] == c for x in rientri_per_riga[row_idx_mappa]):
                rientri_per_riga[row_idx_mappa].append((c, data_orig))
                righe_da_lavorare.append((c, r_idx, data_orig))
    wb.close()
    return rientri_per_riga, righe_da_lavorare

def _aggiorna_stato_rientri_excel(righe_validate: list[int], testo: str):
    """Scrive `testo` nella col C per le righe specificate di rientri_ddt.xlsx."""
    if not righe_validate: return
    path = BASE_DIR / "rientri_ddt.xlsx"
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path)
        ws = wb.active
        for r_idx in righe_validate:
            ws.cell(row=r_idx, column=3).value = testo
        wb.save(path)
        print(f"  💾 Excel Rientri: {len(righe_validate)} righe → '{testo}'")
    except PermissionError:
        print(f"  ⚠️  ERRORE: Impossibile aggiornare Excel Rientri. Chiudi il file!")
    except Exception as e:
        print(f"  ⚠️  Errore salvataggio rientri: {e}")

def _carica_mappatura_veloce():
    path = PROG_DIR / "mappatura_destinazioni.xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    res = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        c_f = _val(row[0].value).lower()
        c_l = _val(row[1].value).lower()
        if c_f: res[c_f] = (r_idx, "F")
        if c_l: res[c_l] = (r_idx, "L")
    wb.close()
    return res

def main():
    if len(sys.argv) < 2:
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py 3_crea_lista_unificata.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name[len("CONSEGNE_"):]  # es. "30-03-2026" o "30-03-2026_31-03-2026"
        print(f"Nessuna data specificata. Uso l'ultima cartella trovata: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data): data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    file_punti = output_base / "punti_consegna.xlsx"
    out_json = output_base / "punti_consegna_unificati.json"

    print(f"\n--- Unificazione punti consegna e RIENTRI ({data}) ---")

    punti = _carica_excel(file_punti)
    map_codice = _carica_mappatura_veloce()
    rientri_globale, righe_excel_rientri = _carica_rientri(map_codice)
    
    righe_rientri_da_marcare = []       # matched → scriverò data consegna
    righe_rientri_in_lavorazione = []   # unmatched → scriverò 'In lavorazione'

    unificati = [] # Cambiato in lista per evitare raggruppamenti
    # Mappa per agganciare velocemente i rientri ai punti (idx_mappa -> lista di riferimenti ai punti)
    map_idx_mappa_to_punti = {}

    def add_punto(p, row_idx=None):
        c_f = _val(p.get("codice_frutta")) or "p00000"
        c_l = _val(p.get("codice_latte")) or "p00000"
        # Legge il codice univoco composito dalla colonna C (formato: codice_frutta_codice_latte)
        cod_univoco = _val(p.get("codici_ddt_trovati")) or f"{c_f}_{c_l}"

        punto = {
            "nome": _val(p.get("nome")),
            "indirizzo": _val(p.get("indirizzo")),
            "codice_frutta": c_f,
            "codice_latte": c_l,
            "codice_univoco": cod_univoco,  # chiave univoca stabile per il /save
            "zona": _val(p.get("zona")) or "0000",
            "data_consegna": _val(p.get("data_consegna")),
            "orario_min": _val(p.get("orario_min")),
            "orario_max": _val(p.get("orario_max")),
            "lat": p.get("latitudine"),
            "lon": p.get("longitudine"),
            "codici_ddt_frutta": [],
            "codici_ddt_latte": [],
            "rientri_alert": [],
            "row_idx_mappatura": row_idx
        }

        # Popola le liste DDT dai componenti del codice univoco
        parti = [x.strip() for x in cod_univoco.split("_")]
        c_f_ddt = parti[0] if len(parti) > 0 else ""
        c_l_ddt = parti[1] if len(parti) > 1 else ""
        if c_f_ddt:
            punto["codici_ddt_frutta"].append(c_f_ddt)
        if c_l_ddt:
            punto["codici_ddt_latte"].append(c_l_ddt)
        
        unificati.append(punto)
        
        if row_idx is not None:
            if row_idx not in map_idx_mappa_to_punti:
                map_idx_mappa_to_punti[row_idx] = []
            map_idx_mappa_to_punti[row_idx].append(punto)

    # Carica punti da Excel
    for p in punti:
        c = _val(p.get("codice_frutta")).lower()
        idx = map_codice[c][0] if c in map_codice else None
        add_punto(p, idx)

    # Gestione rientri
    # rientri_globale = {idx_mappa: [(codice, data_originale), ...]}
    for idx_mappa, codici_con_date in rientri_globale.items():
        if idx_mappa in map_idx_mappa_to_punti:
            # Abbina il rientro ai punti di consegna esistenti (cliente presente oggi)
            for up in map_idx_mappa_to_punti[idx_mappa]:
                current_codes = (up.get("codici_ddt_frutta") or []) + (up.get("codici_ddt_latte") or [])
                for cr, data_orig in codici_con_date:
                    status = "yellow" if cr in current_codes else "red"
                    if not any(a["codice"] == cr for a in up["rientri_alert"]):
                        up["rientri_alert"].append({"codice": cr, "status": status, "data_ddt": data_orig})
                        r_excel = next((r for c, r, d in righe_excel_rientri if c == cr), None)
                        if r_excel: righe_rientri_da_marcare.append(r_excel)
        else:
            # DDT non abbinato ad alcuna consegna oggi → zona speciale sulla mappa
            path_m = PROG_DIR / "mappatura_destinazioni.xlsx"
            from openpyxl import load_workbook
            wb_m = load_workbook(path_m, read_only=True, data_only=True)
            ws_m = wb_m.active
            row = list(ws_m.iter_rows(min_row=idx_mappa, max_row=idx_mappa))[0]
            # Usa il primo codice per il nome fallback
            primo_codice, primo_data_orig = codici_con_date[0]
            nome_r = _val(row[2].value) or f"Rientro {primo_codice}"
            ind_r  = _val(row[4].value)
            lat_r  = row[12].value
            lon_r  = row[13].value
            punto_r = {
                "nome": nome_r,
                "indirizzo": ind_r,
                "codice_frutta": _val(row[0].value),
                "codice_latte":  _val(row[1].value),
                # ← data originale del DDT (colonna B): indica la cartella CONSEGNE_ dove
                #   si trova fisicamente il PDF. Usata da 9_genera_distinte per trovarlo.
                "data_consegna": primo_data_orig or data,
                "orario_min": _val(row[10].value),
                "orario_max": _val(row[11].value),
                "lat": lat_r,
                "lon": lon_r,
                "zona": "DDT_DA_INSERIRE",
                "codici_ddt_frutta": [],
                "codici_ddt_latte":  [],
                "rientri_alert": [{"codice": cr, "status": "red", "data_ddt": d} for cr, d in codici_con_date],
                "row_idx_mappatura": idx_mappa,
                "_is_rientro_speciale": True
            }
            for cr, data_orig in codici_con_date:
                tipo = _val(row[0].value).lower()
                if tipo:
                    punto_r["codici_ddt_frutta"].append(cr)
                else:
                    punto_r["codici_ddt_latte"].append(cr)
            unificati.append(punto_r)
            wb_m.close()
            # Marca come 'In lavorazione'
            for cr, _ in codici_con_date:
                r_excel = next((r for c, r, d in righe_excel_rientri if c == cr), None)
                if r_excel: righe_rientri_in_lavorazione.append(r_excel)

    lista_finale = unificati
    mappati = sum(1 for p in lista_finale if p.get("row_idx_mappatura") is not None)
    fallback = len(lista_finale) - mappati

    for p in lista_finale:
        p["geo_query_nome_indirizzo"] = f"{p['nome']} {p['indirizzo']}"
        p["geo_query_indirizzo"] = p["indirizzo"]

    # --- GEOCODING AUTOMATICO per i punti senza coordinate ---
    senza_coord = [p for p in lista_finale if not p.get("lat") or not p.get("lon")]
    if senza_coord:
        print(f"\n  🌍 Geocoding automatico: {len(senza_coord)} punti senza coordinate...")
        geo_cache = _carica_geocache()
        trovati = 0
        for p in senza_coord:
            nome = p.get('nome', '')
            coord = _geocodifica_punto(p, geo_cache)
            if coord:
                p["lat"], p["lon"] = coord
                row_idx = p.get("row_idx_mappatura")
                _aggiorna_mappatura_coord(row_idx, coord[0], coord[1])
                trovati += 1
                print(f"    ✅ {nome[:55]:<55} → {coord[0]:.5f}, {coord[1]:.5f}")
            else:
                print(f"    🟡 {nome[:55]:<55} → non trovato (marker giallo)")
        _salva_geocache(geo_cache)
        rimasti = len(senza_coord) - trovati
        print(f"  Geocoding completato: {trovati} trovati, {rimasti} ancora da posizionare manualmente.\n")
    else:
        print("  ✅ Tutti i punti hanno già le coordinate.")

    out_json.write_text(json.dumps({"data": data, "punti": lista_finale}, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Punti totali: {len(lista_finale)} (Mappati: {mappati}, Fuori mappa: {fallback})")
    print(f"  Rientri agganciati: {sum(len(p['rientri_alert']) for p in lista_finale)}")
    print(f"  Salvato: {out_json.name}")

    # ── AGGIORNAMENTO EXCEL RIENTRI ──
    if righe_rientri_da_marcare:
        _aggiorna_stato_rientri_excel(list(set(righe_rientri_da_marcare)), f"allegato DDT {data}")
    if righe_rientri_in_lavorazione:
        _aggiorna_stato_rientri_excel(list(set(righe_rientri_in_lavorazione)), "In lavorazione")

    print("--- Completato ---\n")
    return 0

if __name__ == "__main__":
    exit(main())