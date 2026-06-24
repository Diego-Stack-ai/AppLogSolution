import json
from openpyxl import load_workbook
import string

f2 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026_copia/punti_consegna_unificati.json'
with open(f2, encoding='utf-8') as f: j2 = json.load(f)

# Prep dictionary with key (nome, indirizzo)
punti_veri = {}
for p in j2.get('punti', []):
    if p.get('lat'):
        n = p['nome'].strip().lower()
        i = p.get('indirizzo', '').strip().lower()
        # Normalizing address a bit: remove commas, numbers, "via", "piazza" etc to make fuzzy matching if needed
        # Or better: just keep it simple, since we can match on exact string if they are identical
        punti_veri[(n, i)] = p

map_file = 'G:/Il mio Drive/App/AppLogSolution/dati/PROGRAMMA/mappatura_destinazioni.xlsx'
wb = load_workbook(map_file)
ws = wb.active
headers = [str(c.value).strip().lower() for c in ws[1]]
col_nome = next(i for i, h in enumerate(headers) if h in ['a chi va consegnato', 'nome'])
col_ind = next(i for i, h in enumerate(headers) if h in ['indirizzo', 'via'])
col_lat = next(i for i, h in enumerate(headers) if h == 'latitudine')
col_lon = next(i for i, h in enumerate(headers) if h == 'longitudine')

saved = 0
for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    n = str(row[col_nome].value).strip().lower()
    i = str(row[col_ind].value).strip().lower()
    
    # Try exact match first
    match = punti_veri.get((n, i))
    
    # Try partial address match
    if not match:
        for (pn, pi), p in punti_veri.items():
            if pn == n:
                # if address is partially contained
                if i in pi or pi in i or (i.split()[0] in pi):
                    match = p
                    break
                    
    if match and n == 'edmondo de amicis':
        ws.cell(row=r_idx, column=col_lat+1, value=match['lat'])
        ws.cell(row=r_idx, column=col_lon+1, value=match['lon'])
        print(f'Corretto: {n} | {i}  -> {match["lat"]}, {match["lon"]}')
        saved += 1

wb.save(map_file)
print(f'\\nRipristinati {saved} punti (Edmondo De Amicis disambiguati) nel master Excel!')
