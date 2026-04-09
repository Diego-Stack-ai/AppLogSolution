#!/usr/bin/env python3
"""
3_crea_lista_unificata.py (Firebase Version)
Legge un unico file punti_consegna.xlsx e gestisce i rientri da rientri_ddt.xlsx.
Produce punti_consegna_unificati.json.
Ora dipende interamente da Firebase per tutte le anagrafiche e coordinate.
Nessun file Excel locale 'mappatura_destinazioni.xlsx' viene usato.
"""

import json
import re
import sys
from pathlib import Path
import firebase_admin
from firebase_admin import credentials, firestore

PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"

def _val(x):
    if x is None: return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

def _carica_excel(path: Path):
    if not path.exists(): return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = []
    headers = [str(c.value or "").split("-")[-1].strip().lower().replace(" ", "_") for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        d = {}
        for i, h in enumerate(headers):
            if i < len(vals): d[h] = vals[i]
        data.append(d)
    wb.close()
    return data

def init_firebase():
    import glob
    import os
    cred_files = glob.glob(os.path.join(os.path.dirname(__file__), '..', '..', 'backend', 'config', 'log-solution-*-firebase-adminsdk-*.json'))
    if not firebase_admin._apps and cred_files:
        cred = credentials.Certificate(cred_files[0])
        firebase_admin.initialize_app(cred)
    elif not cred_files:
        print("❌ Credenziali Firebase non trovate in backend/config/")
        return None
    return firestore.client()

def _carica_firebase_veloce():
    db = init_firebase()
    if not db: return {}, {}
    docs = db.collection("customers").document("DNR").collection("clienti").stream()
    
    res = {}
    full_data = {}
    
    row_idx = 2
    for doc in docs:
        d = doc.to_dict()
        c_f = str(d.get("codiceFrutta", "")).strip().lower()
        c_l = str(d.get("codiceLatte", "")).strip().lower()
        if c_f == "nan": c_f = ""
        if c_l == "nan": c_l = ""
        
        if c_f: res[c_f] = (row_idx, "F")
        if c_l: res[c_l] = (row_idx, "L")
        
        # Salviamo l'intero dato del database per ripescarlo se serve come RIENTRO orfano
        d["row_idx"] = row_idx
        full_data[row_idx] = d
        row_idx += 1
        
    return res, full_data

def _carica_rientri(map_codice: dict):
    path = BASE_DIR / "rientri_ddt.xlsx"
    if not path.exists(): return {}, []
    from openpyxl import load_workbook
    from datetime import datetime as _dt
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rientri_per_riga = {}
    righe_da_lavorare = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cod_r = _val(row[0].value)
        if not cod_r: continue
        stato = _val(row[2].value) if len(row) > 2 else ""

        stato_lower = stato.lower()
        if 'allegato' in stato_lower and 'lavorazione' not in stato_lower:
            continue

        val_b = row[1].value if len(row) > 1 else None
        if val_b:
            if hasattr(val_b, 'strftime'):
                data_orig = val_b.strftime("%d-%m-%Y")
            else:
                try:
                    data_orig = _dt.strptime(str(val_b)[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
                except Exception:
                    data_orig = ""
        else:
            data_orig = ""

        c = cod_r.lower()
        if c in map_codice:
            row_idx_mappa, _ = map_codice[c]
            if row_idx_mappa not in rientri_per_riga:
                rientri_per_riga[row_idx_mappa] = []
            if not any(x[0] == c for x in rientri_per_riga[row_idx_mappa]):
                rientri_per_riga[row_idx_mappa].append((c, data_orig))
                righe_da_lavorare.append((c, r_idx, data_orig))
    wb.close()
    return rientri_per_riga, righe_da_lavorare

def _aggiorna_stato_rientri_excel(righe_validate: list[int], testo: str):
    if not righe_validate: return
    path = BASE_DIR / "rientri_ddt.xlsx"
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path)
        ws = wb.active
        for r_idx in righe_validate:
            ws.cell(row=r_idx, column=3).value = testo
        wb.save(path)
        print(f"  💾 Excel Rientri: {len(righe_validate)} righe → '{testo}'")
    except PermissionError:
        print(f"  ⚠️  ERRORE: Impossibile aggiornare Excel Rientri. Chiudi il file!")
    except Exception as e:
        print(f"  ⚠️  Errore salvataggio rientri: {e}")

def main():
    if len(sys.argv) < 2:
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py 3_crea_lista_unificata.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name[len("CONSEGNE_"):]
        print(f"Nessuna data specificata. Uso l'ultima cartella trovata: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data): data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    file_punti = output_base / "punti_consegna.xlsx"
    out_json = output_base / "punti_consegna_unificati.json"

    print(f"\n--- Unificazione punti consegna e RIENTRI Firebase ({data}) ---")

    punti = _carica_excel(file_punti)
    
    print("📡 Caricamento dati da Firebase...")
    map_codice, raw_firebase_data = _carica_firebase_veloce()
    
    rientri_globale, righe_excel_rientri = _carica_rientri(map_codice)
    
    righe_rientri_da_marcare = []
    righe_rientri_in_lavorazione = []

    unificati = []
    map_idx_mappa_to_punti = {}

    def add_punto(p, row_idx=None):
        c_f = _val(p.get("codice_frutta")) or "p00000"
        c_l = _val(p.get("codice_latte")) or "p00000"
        cod_univoco = _val(p.get("codici_ddt_trovati")) or f"{c_f}_{c_l}"

        punto = {
            "nome": _val(p.get("nome")),
            "indirizzo": _val(p.get("indirizzo")),
            "codice_frutta": c_f,
            "codice_latte": c_l,
            "codice_univoco": cod_univoco,
            "zona": _val(p.get("zona")) or "0000",
            "data_consegna": _val(p.get("data_consegna")),
            "orario_min": _val(p.get("orario_min")),
            "orario_max": _val(p.get("orario_max")),
            "lat": p.get("latitudine"),
            "lon": p.get("longitudine"),
            "codici_ddt_frutta": [],
            "codici_ddt_latte": [],
            "rientri_alert": [],
            "row_idx_mappatura": row_idx
        }

        parti = [x.strip() for x in cod_univoco.split("_")]
        c_f_ddt = parti[0] if len(parti) > 0 else ""
        c_l_ddt = parti[1] if len(parti) > 1 else ""
        if c_f_ddt: punto["codici_ddt_frutta"].append(c_f_ddt)
        if c_l_ddt: punto["codici_ddt_latte"].append(c_l_ddt)
        
        unificati.append(punto)
        
        if row_idx is not None:
            if row_idx not in map_idx_mappa_to_punti:
                map_idx_mappa_to_punti[row_idx] = []
            map_idx_mappa_to_punti[row_idx].append(punto)

    for p in punti:
        c = _val(p.get("codice_frutta")).lower()
        idx = map_codice[c][0] if c in map_codice else None
        add_punto(p, idx)

    for idx_mappa, codici_con_date in rientri_globale.items():
        if idx_mappa in map_idx_mappa_to_punti:
            for up in map_idx_mappa_to_punti[idx_mappa]:
                current_codes = (up.get("codici_ddt_frutta") or []) + (up.get("codici_ddt_latte") or [])
                for cr, data_orig in codici_con_date:
                    status = "yellow" if cr in current_codes else "red"
                    if not any(a["codice"] == cr for a in up["rientri_alert"]):
                        up["rientri_alert"].append({"codice": cr, "status": status, "data_ddt": data_orig})
                        r_excel = next((r for c, r, d in righe_excel_rientri if c == cr), None)
                        if r_excel: righe_rientri_da_marcare.append(r_excel)
        else:
            fb_row = raw_firebase_data.get(idx_mappa, {})
            primo_codice, primo_data_orig = codici_con_date[0]
            
            nome_r = fb_row.get("nome", f"Rientro {primo_codice}")
            ind_r  = fb_row.get("indirizzo", "")
            lat_r  = fb_row.get("lat")
            lon_r  = fb_row.get("lon") or fb_row.get("lng")
            
            punto_r = {
                "nome": nome_r,
                "indirizzo": ind_r,
                "codice_frutta": fb_row.get("codiceFrutta", ""),
                "codice_latte":  fb_row.get("codiceLatte", ""),
                "data_consegna": primo_data_orig or data,
                "orario_min": fb_row.get("orarioMin", ""),
                "orario_max": fb_row.get("orarioMax", ""),
                "lat": lat_r,
                "lon": lon_r,
                "zona": "DDT_DA_INSERIRE",
                "codici_ddt_frutta": [],
                "codici_ddt_latte":  [],
                "rientri_alert": [{"codice": cr, "status": "red", "data_ddt": d} for cr, d in codici_con_date],
                "row_idx_mappatura": idx_mappa,
                "_is_rientro_speciale": True
            }
            
            for cr, data_orig in codici_con_date:
                # euristica molto semplice
                if cr.startswith('p') or cr.startswith('l'):
                    punto_r["codici_ddt_latte" if cr.startswith('l') else "codici_ddt_frutta"].append(cr)
            
            unificati.append(punto_r)
            for cr, _ in codici_con_date:
                r_excel = next((r for c, r, d in righe_excel_rientri if c == cr), None)
                if r_excel: righe_rientri_in_lavorazione.append(r_excel)

    lista_finale = unificati
    mappati = sum(1 for p in lista_finale if p.get("row_idx_mappatura") is not None)
    fallback = len(lista_finale) - mappati

    # Nessun geocoding automatico qui. Firebase PWA web interface handles geocoding.
    senza_coord = sum(1 for p in lista_finale if not p.get("lat") or not p.get("lon"))
    if senza_coord > 0:
        print(f"⚠️ Ci sono {senza_coord} punti senza coordinate.")
        print(f"   Aggiorna la loro posizione direttamente dalla Dashboard Web PWA (Clienti)!")

    out_json.write_text(json.dumps({"data": data, "punti": lista_finale}, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Punti totali: {len(lista_finale)} (Mappati: {mappati}, Fuori mappa: {fallback})")
    print(f"  Salvato: {out_json.name}")

    if righe_rientri_da_marcare:
        _aggiorna_stato_rientri_excel(list(set(righe_rientri_da_marcare)), f"allegato DDT {data}")
    if righe_rientri_in_lavorazione:
        _aggiorna_stato_rientri_excel(list(set(righe_rientri_in_lavorazione)), "In lavorazione")

    print("--- Completato ---\n")
    return 0

if __name__ == "__main__":
    exit(main())