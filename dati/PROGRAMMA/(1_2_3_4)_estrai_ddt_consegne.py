#!/usr/bin/env python3
"""
(1_2_3_4)_estrai_ddt_consegne.py
════════════════════════════════
Estrae tutti i DDT dai PDF in CONSEGNE/DDT-ORIGINALI/FRUTTA e LATTE.
Identifica nuovi clienti ed estrae automaticamente i loro dati (Indirizzo, CAP, Orari, ecc.).

Uso: py (1_2_3_4)_estrai_ddt_consegne.py [data]
"""

import re
import sys
import subprocess
import time
import json
from pathlib import Path
from datetime import datetime

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    print("pip install pypdf")
    sys.exit(1)

# --- CONFIGURAZIONE ---
PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
INPUT_FRUTTA = BASE_DIR / "FRUTTA"
INPUT_LATTE = BASE_DIR / "LATTE"
MAPPATURA_XLSX = PROG_DIR / "mappatura_destinazioni.xlsx"

# Regex per estrazione dati
DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'(?:[Ll]uogo [Dd]i [Dd]estinazione|[Cc]odice [Dd]estinazione):\s*([pP]\d{4,5})')
CAP_RE = re.compile(r"\b(\d{5})\b")
PROVINCIA_RE = re.compile(r"\(([A-Z]{2})\)")
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})(?:\s+H(\d{2}))?(?:\s+(\d{3}))?', re.I)


def _estrai_data_luogo(text: str) -> tuple[str | None, str | None]:
    """Estrae (data, luogo) da una pagina DDT. data in formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    return (data, luogo)


def _estrai_dati_consegna_da_testo(text: str, codice: str, da_frutta: bool) -> dict:
    """Estrae destinatario, indirizzo, CAP, città, provincia e orari dal testo di una pagina DDT."""
    res = {"dest": "", "ind": "", "cap": "", "cit": "", "prov": "", "om": "", "oM": "14:00"}
    if codice.lower() not in text.lower():
        return res
    
    idx_l = text.find("Luogo di destinazione")
    if idx_l < 0: return res

    # 1. Nome e Indirizzo
    if da_frutta:
        blocco = text[idx_l : idx_l + 650]
        lines = [ln.strip() for ln in blocco.split("\n") if ln.strip()]
        for i, ln in enumerate(lines):
            if LUOGO_RE.search(ln):
                if i + 1 < len(lines): res["dest"] = lines[i + 1].strip().title()
                if i + 2 < len(lines): res["ind"] = lines[i + 2].strip().title()
                break
    else:
        idx_causale = text.upper().find("CAUSALE DEL TRASPORTO")
        blocco = text[:idx_causale] if idx_causale > 0 else text[idx_l : idx_l + 900]
        for ln in blocco.split("\n"):
            ln = ln.strip()
            cf_m = re.match(r"^[Cc]\.?[Ff]\.?\s+", ln)
            if cf_m: res["dest"] = ln[cf_m.end():].strip().title()
            else:
                albo_m = re.match(r"^[Aa]lbo\s+", ln, re.I)
                if albo_m: res["ind"] = ln[albo_m.end():].strip().title()

    # 2. CAP, Provincia, Città
    idx_resp = text.upper().find("RESPONSABILE DEL TRASPORTO")
    blocco_prov = text[idx_resp:] if idx_resp >= 0 else text
    
    for prov_m in PROVINCIA_RE.finditer(blocco_prov):
        sigla = prov_m.group(1)
        if sigla == "MN" and ("Pomponesco" in blocco_prov[max(0, prov_m.start()-40):prov_m.start()] or "46030" in blocco_prov):
            continue
        res["prov"] = sigla
        caps = list(CAP_RE.finditer(blocco_prov[:prov_m.start()]))
        if caps:
            res["cap"] = caps[-1].group(1)
            pre = blocco_prov[caps[-1].end() : caps[-1].end() + 60]
            citta_m = re.search(r"\s*[-]?\s*([A-Za-zÀ-ÿ\s'.]+?)\s*\([A-Z]{2}\)", pre)
            if citta_m: res["cit"] = citta_m.group(1).strip().title()
        break
        
    # 3. Orari
    idx_c = text.upper().find("CAUSALE DEL TRASPORTO")
    if idx_c >= 0:
        sezione = text[idx_c:idx_c+150]
        m = CAUSALE_RE.search(sezione)
        if m:
            if m.group(2): res["oM"] = f"{int(m.group(2)):02d}:00"
            if m.group(3):
                s = m.group(3)
                if len(s) == 3: res["om"] = f"{int(s[0]):02d}:{int(s[1:3]):02d}"
    return res


def _ricava_data_da_pdf() -> str | None:
    import pdfplumber
    for cart in (INPUT_FRUTTA, INPUT_LATTE):
        if not cart.exists(): continue
        for pdf_path in sorted(cart.glob("*.pdf")):
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        data, _ = _estrai_data_luogo(text)
                        if data: return data
            except: pass
    return None


def _leggi_codici_mappatura():
    if not MAPPATURA_XLSX.exists(): return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
        ws = wb.active
        codici = set()
        for row in ws.iter_rows(min_row=2, max_col=2):
            for cell in row:
                val = str(cell.value or "").strip().lower()
                if val and val != "p00000": codici.add(val)
        wb.close()
        return codici
    except Exception as e:
        print(f"⚠️ Errore mappatura: {e}")
        return set()


def _verifica_nuovi_clienti(dati_nuovi: dict):
    if not dati_nuovi: return True
    print("\n" + "!"*60)
    print(f"🛑 RILEVATI {len(dati_nuovi)} NUOVI CODICI CLIENTE:")
    for cod, info in sorted(dati_nuovi.items()):
        print(f"   - {cod} ({info['tipo']}): {info['dest']} - {info['cit']}")
    print("!"*60 + "\n")
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nuovi Codici"
        headers = ["Codice Frutta", "Codice Latte", "A chi va consegnato", "Tipologia grado", "Indirizzo", "CAP", "Città", "Provincia",
                  "Tipologia consegna", "Tipologia consegna.1", "Email", "Sito web", "Orario min", "Orario max"]
        ws.append(headers)
        for cod, info in sorted(dati_nuovi.items()):
            row = [""] * 14
            if info["tipo"] == "FRUTTA": row[0] = cod
            else:                        row[1] = cod
            row[2], row[4], row[5], row[6], row[7], row[12], row[13] = info["dest"], info["ind"], info["cap"], info["cit"], info["prov"], info["om"], info["oM"]
            ws.append(row)
        hp = BASE_DIR / "nuovi_codici_consegna.xlsx"
        wb.save(hp)
        print(f"📄 File generato: {hp.name}\n💡 Copia i dati in mappatura_destinazioni.xlsx per proseguire.\n")
    except Exception as e: print(f"⚠️ Errore excel: {e}")
    return False


def _estrai_da_cartella(cart_in: Path, cart_out: Path, etichetta: str, data_v: str, mappati: set, *, duplicata: bool = False):
    nuovi_dati = {}
    if not cart_in.exists(): return 0, 0, nuovi_dati
    cart_out.mkdir(parents=True, exist_ok=True)
    pdf_files = list(cart_in.glob("*.pdf"))
    if not pdf_files: return 0, 0, nuovi_dati
    import pdfplumber
    creati = 0
    visti = {}
    for pdf_path in pdf_files:
        try:
            reader = PdfReader(pdf_path)
            with pdfplumber.open(pdf_path) as pdf:
                for i in range(0, len(pdf.pages), 2 if duplicata else 1):
                    text = pdf.pages[i].extract_text() or ""
                    d, l = _estrai_data_luogo(text)
                    if not d or not l or d != data_v: continue
                    if l not in mappati and l not in nuovi_dati:
                        nuovi_dati[l] = _estrai_dati_consegna_da_testo(text, l, duplicata)
                        nuovi_dati[l]["tipo"] = etichetta
                    chiave = (d, l)
                    cnt = visti.get(chiave, 0) + 1
                    visti[chiave] = cnt
                    fname = f"{l}_{d}_{cnt}.pdf" if cnt > 1 else f"{l}_{d}.pdf"
                    writer = PdfWriter()
                    writer.add_page(reader.pages[i])
                    with open(cart_out / fname, "wb") as f: writer.write(f)
                    creati += 1
                    if creati <= 3 or creati % 50 == 0: print(f"    {fname}")
        except Exception as e: print(f"  Errore {pdf_path.name}: {e}")
    return creati, sum(1 for c in visti.values() if c > 1), nuovi_dati


def _pulisci_sorgenti(cart_in: Path, data_v: str):
    import pdfplumber
    rimossi = 0
    for p in cart_in.glob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                if any(_estrai_data_luogo(pg.extract_text() or "")[0] == data_v for pg in pdf.pages):
                    pdf.close(); p.unlink(); rimossi += 1
        except: pass
    return rimossi


def _pulisci_output(base: Path, data_v: str):
    for f in [base/"punti_consegna.xlsx", base/"punti_consegna_unificati.json", base/"4_mappa_zone_google.html", base/f"zone_google_{data_v.replace('-', '_')}.kml"]:
        if f.exists():
            try: f.unlink()
            except: pass


# --- ARTICOLI NOTI ---
ARTICOLI_NOTI = {"ME-T-DI-V0-NA", "PE-T-DI-L3-NA", "10-GEL", "10-FLYER", "10-MANIFESTO", "LT-DL-02-LC", "LT-ES-04-LS",
                 "LT-ESL-IN-LB", "LT-AQ-04-LV", "YO-BI-MN-04-LB", "YO-DL-02-LC", "AP-SU-PC", "FO-DI-PV-04-LB",
                 "CA-Z-BI-L3-NA", "FO-DI-GP-01-NI", "FVNS-03-GADGET", "KI-S-BI-L3-NA", "FVNS-03-POSTER",
                 # Aggiunti 26/03/2026
                 "FI-Z-BI-L3-NA",   # Finocchio biologico da porzionare in classe
                 "ME-S-BI-L3-NA",   # Mela biologica del territorio per estratto
                 }

def _verifica_nuovi_articoli(base):
    print("Verifica articoli..."); import pdfplumber; trovati = set()
    # Regex: cattura codici standard (es. 10-FLYER) e codici data flyer (es. --300326)
    art_re = re.compile(r'^([A-Z0-9]{2,}-[A-Z0-9\-]+|FVNS-\d+-|--\d{6})', re.M)
    divisi = base / "DDT-ORIGINALI-DIVISI"
    if not divisi.exists(): return True
    for p in divisi.rglob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                for pg in pdf.pages:
                    for m in art_re.finditer(pg.extract_text() or ""):
                        c = m.group(1).strip()
                        if c == "FVNS-03-": c = "FVNS-03-POSTER"
                        # Normalizza codici data flyer (--NNNNNN) → 10-FLYER
                        if re.match(r'^--\d{6}$', c): c = "10-FLYER"
                        trovati.add(c)
        except: continue
    nuovi = trovati - ARTICOLI_NOTI
    if nuovi:
        print("\n" + "!"*60 + f"\n🛑 NUOVI ARTICOLI: {', '.join(sorted(nuovi))}\n" + "!"*60 + "\n")
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Nuovi Articoli"; ws.append(["Codice Articolo", "Data", "Note"])
            for c in sorted(nuovi): ws.append([c, datetime.now().strftime("%d/%m/%Y"), "Aggiungi a ARTICOLI_NOTI"])
            rp = BASE_DIR / "nuovi_articoli_rilevati.xlsx"; wb.save(rp); print(f"📄 Report: {rp.name}")
        except Exception as e: print(f"⚠️ Errore: {e}")
        return False
    print("✓ Articoli OK.\n"); return True


def main():
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    if arg and re.match(r"^\d{2}-\d{2}$", arg): arg = f"{arg}-2026"
    data = arg or _ricava_data_da_pdf()
    if not data: return print("❌ Nessun PDF.")
    base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    if base.exists() and not arg: return print(f"⚠️ Cartella {base.name} esiste.")
    _pulisci_output(base, data)
    print(f"\n--- Estrazione DDT ({data}) ---\n")
    out_f, out_l = base/"DDT-ORIGINALI-DIVISI"/"FRUTTA", base/"DDT-ORIGINALI-DIVISI"/"LATTE"
    mappati = _leggi_codici_mappatura()
    res_f = _estrai_da_cartella(INPUT_FRUTTA, out_f, "FRUTTA", data, mappati, duplicata=True)
    res_l = _estrai_da_cartella(INPUT_LATTE, out_l, "LATTE", data, mappati)
    if not _verifica_nuovi_clienti({**res_f[2], **res_l[2]}): sys.exit(1)
    if not _verifica_nuovi_articoli(base): sys.exit(1)
    print("Pulizia e avvio pipeline..."); _pulisci_sorgenti(INPUT_FRUTTA, data); _pulisci_sorgenti(INPUT_LATTE, data); time.sleep(1)
    for s in ["2_crea_punti_consegna.py", "3_crea_lista_unificata.py", "4_mappa_zone_google.py"]:
        p = PROG_DIR / s
        if p.exists(): print(f"⚙️ {s}..."); subprocess.run([sys.executable, str(p), data], cwd=BASE_DIR)
    print(f"\n✅ COMPLETATO ({data})!")

if __name__ == "__main__": main()
