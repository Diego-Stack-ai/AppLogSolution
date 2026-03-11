#!/usr/bin/env python3
"""
Allinea i nomi in mappatura_destinazioni - da aggiustare.xlsx.
Righe 2-455: target con indirizzi vecchi da confrontare.
Righe 456+: sorgenti con indirizzi nuovi da copiare nelle righe 2-455.
Modifica solo colonne C-J (A chi va, Tipologia, Indirizzo, CAP, Città, Provincia, Email, Home Page).

Struttura: A=Cod Frutta, B=Cod Latte, C=A chi va, D=Tipologia, E=Indirizzo, F=CAP, G=Città, H=Provincia, I=Email, J=Home Page, K-P=non modificate
Ripresa: allinea_nomi_completati.txt salva le righe già fatte (se interrompi con q, al prossimo avvio riprendi da lì).
Eseguire: py -3 Programma/allinea_nomi_mappatura.py
"""
import difflib
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)
try:
    from colorama import init, Fore, Style
    init()
    COLORE_ORIGINALE = Fore.YELLOW
    COLORE_ALTERNATIVA = Fore.CYAN
    COLORE_RESET = Style.RESET_ALL
except ImportError:
    COLORE_ORIGINALE = COLORE_ALTERNATIVA = COLORE_RESET = ""

# Cartella principale (parent di Programma/)
BASE = Path(__file__).resolve().parent.parent
MAPPATURA = BASE / "mappatura_destinazioni - da aggiustare.xlsx"
COMPLETATI_FILE = Path(__file__).parent / "allinea_nomi_completati.txt"  # in Programma

# Colonne da modificare: solo C-J (3-10)
# A=Cod Frutta, B=Cod Latte, C=A chi va, D=Tipologia, E=Indirizzo, F=CAP, G=Città, H=Provincia, I=Email, J=Home Page
COL_NOME, COL_TIPOLOGIA, COL_INDIRIZZO, COL_CAP, COL_CITTA, COL_PROVINCIA = 3, 4, 5, 6, 7, 8
COLS_COPIA = list(range(3, 11))  # C:J (colonne 3-10)

FILL_DA_SOSTITUIRE = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # giallo chiaro
FILL_COMPLETATO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde chiaro

TARGET_INF, TARGET_SUP = 2, 455
SOURCE_START = 456


def _val(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _norm(s):
    """Normalizza per confronto: minuscole, spazi, apostrofi/virgolette trattati come spazio."""
    if not s:
        return ""
    t = str(s).lower()
    for c in "'\"`ʼ′‛‚''":
        t = t.replace(c, " ")
    return " ".join(t.split())


def _carica_completati() -> set:
    """Carica i numeri di riga già completati (per riprendere dopo interruzione)."""
    if not COMPLETATI_FILE.exists():
        return set()
    righe = set()
    for ln in COMPLETATI_FILE.read_text(encoding="utf-8").splitlines():
        ln = ln.strip()
        if ln.isdigit():
            righe.add(int(ln))
    return righe


def _salva_completato(riga: int) -> None:
    """Salva il numero di riga completata."""
    esistente = COMPLETATI_FILE.read_text(encoding="utf-8") if COMPLETATI_FILE.exists() else ""
    COMPLETATI_FILE.write_text(esistente + str(riga) + "\n", encoding="utf-8")


def _match_indirizzo_cap(d1, e1, d2, e2):
    """Indirizzo e CAP uguali (normalizzati)."""
    return _norm(d1) == _norm(d2) and _norm(e1) == _norm(e2)


def _norm_cap(v):
    """CAP: normalizza 35040.0 -> 35040 per confronto."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return _norm(s)


def _match_indirizzo_cap_citta(ind1, cap1, citta1, ind2, cap2, citta2):
    """Indirizzo, CAP e Città tutti uguali (100% match)."""
    return (_norm(ind1) == _norm(ind2) and
            _norm_cap(cap1) == _norm_cap(cap2) and
            _norm(citta1) == _norm(citta2))


def _fuzzy_score(target_b: str, source_bc: str) -> float:
    """Score tra target B e sorgente (B+C). Se target è contenuto in sorgente → alto score."""
    nt, ns = _norm(target_b), _norm(source_bc)
    if not nt:
        return 0.0
    base = difflib.SequenceMatcher(None, nt, ns).ratio()
    # Se target è contenuto in sorgente (es. "marco polo" in "marco polo scuola primaria") → 95%
    if len(nt) >= 4 and nt in ns:
        return max(base, 0.95)
    return base


def main():
    if not MAPPATURA.exists():
        print(f"File non trovato: {MAPPATURA}")
        print("Verifica che lo script sia nella stessa cartella del file Excel.")
        return

    wb = load_workbook(MAPPATURA)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active

    print(f"File: {MAPPATURA.name} | Righe totali: {ws.max_row} | Sorgenti (da {SOURCE_START}): {max(0, ws.max_row - TARGET_SUP)}")
    print("(Chiudi Excel prima di procedere)")
    if ws.max_row < SOURCE_START:
        print(f"ATTENZIONE: nessuna riga {SOURCE_START}+. Il file deve contenere i dati corretti dalla riga {SOURCE_START}.")
        input("Invio per uscire.")
        return

    completati = _carica_completati()
    if completati:
        print(f"Ripresa: {len(completati)} righe già completate (saltate). File: {COMPLETATI_FILE.name}")
    fatti_oggi = 0

    # Righe target da processare (dal basso)
    target_sup = min(TARGET_SUP, ws.max_row)
    target_rows = list(range(target_sup, TARGET_INF - 1, -1))

    # --- FASE 1: AUTO-APPLY ---
    # Criteri: Indirizzo+CAP+Città 100% uguali, Nome fuzzy >= 90% → sostituisci automaticamente
    SOGLIA_AUTO = 0.90
    auto_applies = []
    used_sources = set()

    for target_row in target_rows:
        if target_row in completati:
            continue
        cod = _val(ws.cell(target_row, 1)) or _val(ws.cell(target_row, 2))  # A=Cod Frutta, B=Cod Latte
        if not cod:
            continue
        nome_t = _val(ws.cell(target_row, COL_NOME))
        ind_t = _val(ws.cell(target_row, COL_INDIRIZZO))
        cap_t = _val(ws.cell(target_row, COL_CAP))
        citta_t = _val(ws.cell(target_row, COL_CITTA))
        if not ind_t and not cap_t:
            continue

        best_src, best_score = None, 0.0
        for r in range(SOURCE_START, ws.max_row + 1):
            if r in used_sources:
                continue
            ind_s = _val(ws.cell(r, COL_INDIRIZZO))
            cap_s = _val(ws.cell(r, COL_CAP))
            citta_s = _val(ws.cell(r, COL_CITTA))
            if not _match_indirizzo_cap_citta(ind_t, cap_t, citta_t, ind_s, cap_s, citta_s):
                continue
            nome_s = _val(ws.cell(r, COL_NOME))
            tip_s = _val(ws.cell(r, COL_TIPOLOGIA))
            nome_completo = f"{nome_s} {tip_s}".strip()
            score = _fuzzy_score(nome_t, nome_completo)
            if score >= SOGLIA_AUTO and score > best_score:
                best_src, best_score = r, score

        if best_src is not None:
            auto_applies.append((target_row, best_src))
            used_sources.add(best_src)
            completati.add(target_row)

    # Applica auto: ordina per riga sorgente decrescente (così le eliminazioni non spostano)
    auto_applies.sort(key=lambda x: -x[1])
    for target_row, src_row in auto_applies:
        for col in COLS_COPIA:
            val = ws.cell(src_row, col).value
            cell = ws.cell(target_row, col)
            cell.value = val
            cell.fill = FILL_COMPLETATO
        ws.delete_rows(src_row, 1)
        _salva_completato(target_row)
        fatti_oggi += 1

    if auto_applies:
        wb.save(MAPPATURA)
        print(f"\n{COLORE_ORIGINALE}AUTO-APPLY: {len(auto_applies)} righe sostituite automaticamente (Ind+CAP+Città 100%, Nome >= 90%){COLORE_RESET}\n")

    # --- FASE 2: INTERATTIVO (righe rimanenti) ---
    for target_row in target_rows:
        if target_row in completati:
            continue
        cod = _val(ws.cell(target_row, 1)) or _val(ws.cell(target_row, 2))  # A=Cod Frutta, B=Cod Latte
        if not cod:
            print(f"\n--- Riga {target_row}: nessun codice, skip ---")
            continue
        nome_t = _val(ws.cell(target_row, COL_NOME))
        ind_t = _val(ws.cell(target_row, COL_INDIRIZZO))
        cap_t = _val(ws.cell(target_row, COL_CAP))
        citta_t = _val(ws.cell(target_row, COL_CITTA))

        print(f"\n{'='*60}")
        print(f"RIGA TARGET {target_row}: {cod}")
        print(f"  {COLORE_ORIGINALE}ORIGINALE (da sostituire):  Nome: {nome_t} | Indirizzo: {ind_t} | CAP: {cap_t} | Città: {citta_t}{COLORE_RESET}")

        # Cerca in TUTTE le sorgenti (455 fino a max_row): fuzzy su target_B vs source(B+C)
        if ws.max_row < SOURCE_START:
            print(f"  Nessuna riga sorgente (max_row={ws.max_row}). Righe {SOURCE_START}+ assenti.")
            input("[Invio per continuare]")
            continue

        sources = []
        for r in range(SOURCE_START, ws.max_row + 1):
            nome_s = _val(ws.cell(r, COL_NOME))
            tip_s = _val(ws.cell(r, COL_TIPOLOGIA))
            if not nome_s and not tip_s:
                continue
            nome_completo = f"{nome_s} {tip_s}".strip()
            score = _fuzzy_score(nome_t, nome_completo)
            ind_s = _val(ws.cell(r, COL_INDIRIZZO))
            cap_s = _val(ws.cell(r, COL_CAP))
            citta_s = _val(ws.cell(r, COL_CITTA))
            if _match_indirizzo_cap(ind_t, cap_t, ind_s, cap_s):
                score = max(score + 0.2, 0.95)
            elif _norm(cap_t) == _norm(cap_s) or _norm(citta_t) == _norm(citta_s):
                score = max(score + 0.1, 0.9)
            score = min(score, 1.0)
            sources.append((r, f"{nome_s} | {tip_s}".strip(" |"), nome_s, tip_s, ind_s, cap_s, citta_s, score))

        sources.sort(key=lambda x: -x[7])
        candidates = sources[:4]

        if not candidates:
            print(f"  Nessun match (scansite {len(sources)} righe sorgente). [Invio]")
            input()
            continue

        # Evidenzia in Excel le celle C:J da sostituire (giallo)
        try:
            for col in COLS_COPIA:
                ws.cell(target_row, col).fill = FILL_DA_SOSTITUIRE
            wb.save(MAPPATURA)
            print("  (Riga evidenziata in giallo nell'Excel)")
        except PermissionError:
            print("  ATTENZIONE: Chiudi Excel per permettere il salvataggio.")

        print(f"\n  Trovate {len(sources)} sorgenti. Top 4:")
        for i, c in enumerate(candidates, 1):
            nome = c[1] or "-"
            indirizzo = c[4] or "-"
            cap = c[5] or "-"
            citta = c[6] or "-"
            print(f"  {COLORE_ALTERNATIVA}[{i}] (riga {c[0]}, {c[7]:.0%}) Nome: {nome} | Indirizzo: {indirizzo} | CAP: {cap} | Città: {citta}{COLORE_RESET}")

        scelta = input("\n  [1-4] Sostituisci  [s] Salta  [q] Esci: ").strip().lower()

        if scelta == "q":
            wb.save(MAPPATURA)
            print("Salvato. Uscita.")
            return
        if scelta == "s":
            # Togli evidenziazione gialla (riga saltata)
            for col in COLS_COPIA:
                ws.cell(target_row, col).fill = PatternFill()
            wb.save(MAPPATURA)
            continue

        try:
            idx = int(scelta)
            if 1 <= idx <= len(candidates):
                src_row = candidates[idx - 1][0]
                for col in COLS_COPIA:
                    val = ws.cell(src_row, col).value
                    cell = ws.cell(target_row, col)
                    cell.value = val
                    cell.fill = FILL_COMPLETATO  # verde = completato
                ws.delete_rows(src_row, 1)
                wb.save(MAPPATURA)
                _salva_completato(target_row)
                completati.add(target_row)
                fatti_oggi += 1
                print(f"  OK: copiato da riga {src_row}, riga eliminata.")
        except (ValueError, IndexError):
            print("  Input non valido.")

    print(f"\nFine. Sostituiti in questa sessione: {fatti_oggi}")
    if completati:
        print(f"Totale righe completate (salvate in {COMPLETATI_FILE.name}): {len(completati)}")
    wb.close()


if __name__ == "__main__":
    try:
        main()
    except PermissionError:
        print("\n" + "="*50)
        print("ERRORE: Il file Excel è aperto (es. in Microsoft Excel).")
        print("        Chiudi completamente Excel e riprova.")
        print("="*50)
    except Exception as e:
        print(f"\nERRORE: {e}")
        import traceback
        traceback.print_exc()
    input("\nPremi Invio per chiudere...")
