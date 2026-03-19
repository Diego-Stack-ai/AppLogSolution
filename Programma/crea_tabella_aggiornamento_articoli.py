#!/usr/bin/env python3
"""
Crea tabella_aggiornamento_articoli.xlsx.
Estrae articoli direttamente dai PDF (DDT frutta, DDT latte, DDT-ORIGINALI).
Se nessun PDF trovato: ferma e indica di aggiungere i file da elaborare.
"""
import re
import sys
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("pip install pdfplumber")
    sys.exit(1)

from openpyxl import Workbook
from openpyxl.styles import Alignment

BASE_DIR = Path(__file__).resolve().parent.parent
DDT_FRUTTA = BASE_DIR / "DDT frutta"
DDT_LATTE = BASE_DIR / "DDT latte"
GIRI = BASE_DIR / "Giri lavorati"
OUTPUT = BASE_DIR / "tabella_aggiornamento_articoli.xlsx"

# Consolidamento multi-unita: (unita_principale, unita_secondaria, ratio)
# LT-AQ-04-LV: 6 Bottiglie/Fardello come LT-ESL-IN-LB
CONSOLIDAMENTO = {
    "LT-ES-04-LS": ("Fardelli", "Bottiglie", 10),
    "LT-ESL-IN-LB": ("Fardelli", "Bottiglie", 6),
    "LT-AQ-04-LV": ("Fardelli", "Bottiglie", 6),
    "YO-BI-MN-04-LB": ("Cartoni", "Cluster", 10),
    "YO-DL-02-LC": ("Cartoni", "Porzioni", 6),
    "AP-SU-PC": ("Cartoni", "Porzioni", 24),
}


def _estrai_porzioni_per_unita(confezionamento: str) -> str:
    """Estrae ratio numerico per articoli a unità singola (es. '20 Porzioni / Fascetta' -> 20)."""
    if not confezionamento:
        return ""
    # Prima riga: "20 Porzioni / Fascetta" o "3,3 Porzioni / Brick" o "17,85 Porzioni / Fetta"
    m = re.search(r"^([\d,\.]+)\s+(?:Porzioni|Porzione|Bottiglie?|Cluster|Fetta?)\s*/\s*\w+", str(confezionamento).strip(), re.I)
    if m:
        return m.group(1).replace(",", ".")
    return ""


def _estrai_codice(cell0: str) -> str:
    """Estrae il codice articolo base (es. 10-FLYER da '10-FLYER\\nCodice: 3025')."""
    if not cell0 or not str(cell0).strip():
        return ""
    parts = str(cell0).split("Codice:")
    base = parts[0].strip().replace("\n", "").replace("\r", "").strip()
    return base


def _normalizza_descrizione(desc: str) -> str:
    """Sostituisce le date con (Data) nella descrizione."""
    if not desc or not isinstance(desc, str):
        return desc or ""
    d = desc
    d = re.sub(r"\s*-\s*Data\s*\n?\s*distribuzione\s*:\s*\n?\s*\d{1,2}/\d{1,2}/\d{4}", " (Data)", d, flags=re.IGNORECASE)
    d = re.sub(r"\s*Data\s*\n?\s*distribuzione\s*:\s*\n?\s*\d{1,2}/\d{1,2}/\d{4}", " (Data)", d, flags=re.IGNORECASE)
    d = re.sub(r"\s*-\s*Scad\.\s*min\.\s*\n?\s*\d{1,2}/\d{1,2}/\d{4}", "", d)
    d = re.sub(r"\s*Scad\.\s*min\.\s*\n?\s*\d{1,2}/\d{1,2}/\d{4}", "", d)
    d = re.sub(r"\d{1,2}/\d{1,2}/\d{4}", "(Data)", d)
    d = re.sub(r"\(\s*Data\s*\)\s*(\(\s*Data\s*\)\s*)+", "(Data) ", d)
    d = re.sub(r"\s+", " ", d).strip()
    return d


def _descrizione_minima(desc: str) -> str:
    """Rimuove (Data) e optional SPECIALE 1 per descrizione comparativa."""
    if not desc:
        return ""
    d = _normalizza_descrizione(desc)
    if d.startswith("SPECIALE 1 "):
        d = d[11:]
    return d


def _raccogli_dati_da_pdf() -> list[tuple[str, str, str]]:
    """Estrae (codice_raw, descrizione, confezionamento) dai PDF. Ritorna lista vuota se nessun PDF."""
    pdfs = []
    if DDT_FRUTTA.exists():
        pdfs.extend(DDT_FRUTTA.glob("*.pdf"))
    if DDT_LATTE.exists():
        pdfs.extend(DDT_LATTE.glob("*.pdf"))
    if not pdfs and GIRI.exists():
        for d in sorted(GIRI.iterdir(), key=lambda x: x.name, reverse=True)[:5]:
            if not d.is_dir():
                continue
            orig = d / "DDT-ORIGINALI"
            if orig.exists():
                pdfs = list(orig.glob("*.pdf"))
                if pdfs:
                    break

    if not pdfs:
        return []

    righe_viste = set()
    dati = []
    for pdf_path in pdfs:
        try:
            with pdfplumber.open(pdf_path) as doc:
                for page in doc.pages:
                    for t in page.extract_tables() or []:
                        if not t or len(t) < 2 or "Cod. Articolo" not in str(t[0]):
                            continue
                        for row in t[1:]:
                            if not row or len(row) < 4:
                                continue
                            cell0 = str(row[0] or "").strip()
                            cell1 = str(row[1] or "").strip()
                            cell5 = str(row[5] or "").strip() if len(row) > 5 else ""
                            if not cell0 or not re.search(r"[A-Z0-9]{2,}-[A-Z0-9\-]+", cell0):
                                continue
                            chiave = (cell0, cell1, cell5)
                            if chiave not in righe_viste:
                                righe_viste.add(chiave)
                                dati.append((cell0, cell1, cell5))
        except Exception as e:
            print(f"Errore {pdf_path.name}: {e}")
    return dati


from openpyxl import Workbook, load_workbook

def main():
    # 1. Raccogli dati dai PDF
    dati_pdf = _raccogli_dati_da_pdf()
    if not dati_pdf:
        print("Nessun PDF trovato da elaborare.")
        return 1

    # 2. Carica articoli esistenti (se presenti)
    articoli_esistenti = {}
    if OUTPUT.exists():
        try:
            wb_old = load_workbook(OUTPUT, data_only=True)
            ws_old = wb_old.active
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                cod = str(row[0]).strip()
                articoli_esistenti[cod] = {
                    "cod": cod,
                    "desc": str(row[1] or ""),
                    "conf": str(row[2] or ""),
                    "unit_princ": str(row[3] or ""),
                    "per": str(row[4] or ""),
                    "unit_sec": str(row[5] or ""),
                    "ratio": str(row[6] or ""),
                    "porz_unita": str(row[7] or ""),
                }
        except Exception as e:
            print(f"Nota: Impossibile leggere il file esistente ({e}). Creazione nuovo file.")

    # 3. Elabora dati dai PDF
    by_codice = {}
    for cell0, cell1, cell5 in dati_pdf:
        cod = _estrai_codice(cell0)
        if not cod or not re.search(r"[A-Z0-9]{2,}-[A-Z0-9\-]+", cod):
            continue
        desc = _descrizione_minima(cell1)
        conf = (cell5 or "").strip()
        
        # Se abbiamo più descrizioni per lo stesso codice, prendiamo la più lunga
        if cod not in by_codice or len(desc) > len(by_codice[cod]['desc']):
            by_codice[cod] = {'desc': desc, 'conf': conf}

    # 4. Unisci: i dati in PDF hanno la priorità per desc/conf, ma preserviamo mappature esistenti
    for cod, info in by_codice.items():
        desc = info['desc']
        conf = info['conf']
        
        if cod in articoli_esistenti:
            # Aggiorna solo desc e conf se necessario, tieni il resto
            articoli_esistenti[cod]["desc"] = desc
            articoli_esistenti[cod]["conf"] = conf
        else:
            # Crea nuovo record
            unit_princ = ""
            per = ""
            unit_sec = ""
            ratio = ""
            porz_unita = ""
            
            if cod in CONSOLIDAMENTO:
                up, us, r = CONSOLIDAMENTO[cod]
                unit_princ = up
                per = 1
                unit_sec = us
                ratio = r
            else:
                porz_unita = _estrai_porzioni_per_unita(conf)
                if cod == "10-GEL" and not porz_unita:
                    porz_unita = "1"
            
            articoli_esistenti[cod] = {
                "cod": cod,
                "desc": desc,
                "conf": conf,
                "unit_princ": unit_princ,
                "per": per,
                "unit_sec": unit_sec,
                "ratio": ratio,
                "porz_unita": porz_unita,
            }

    # 5. Salva nuovo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Articoli"
    ws.append([
        "Codice",
        "Descrizione",
        "Confezionamento",
        "Unità principale",
        "Per",
        "Unità secondaria",
        "Ratio",
        "Porzioni/Unità",
    ])

    for cod in sorted(articoli_esistenti.keys()):
        a = articoli_esistenti[cod]
        ws.append([
            a["cod"], a["desc"], a["conf"], 
            a["unit_princ"], a["per"], a["unit_sec"], 
            a["ratio"], a["porz_unita"]
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 14

    try:
        wb.save(OUTPUT)
        print(f"Aggiornato (Merge): {OUTPUT.name} ({len(articoli_esistenti)} articoli totali)")
    except PermissionError:
        temp_out = OUTPUT.parent / f"temp_{OUTPUT.name}"
        wb.save(temp_out)
        print(f"ERRORE: File bloccato. Salvata una copia temporanea in: {temp_out}")
    
    return 0


if __name__ == "__main__":
    exit(main())
