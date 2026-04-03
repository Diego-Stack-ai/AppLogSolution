#!/usr/bin/env python3
"""
Crea un unico punti_consegna.xlsx nella cartella
CONSEGNE/CONSEGNE_{data}/, aggregando FRUTTA e LATTE.

Evita duplicati: se lo stesso cliente ha DDT in entrambe le cartelle,
aggiorna le colonne Codice Frutta e Codice Latte nella stessa riga.

Uso: py crea_punti_consegna.py <data>   (es. 16-03-2026)
"""

import re
import sys
from pathlib import Path

PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
MAPPATURA = PROG_DIR / "mappatura_destinazioni.xlsx"

DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'[Ll]uogo [Dd]i [Dd]estinazione:\s*([pP]\d{4,5})')
CAUSALE_SEZIONE_MARKER = "CAUSALE DEL TRASPORTO"
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})', re.I)



def _val(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s


def _estrai_data_luogo_zona(text: str) -> tuple[str | None, str | None, str | None]:
    """Estrae (data, luogo, zona) da testo pagina. data formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    
    idx = text.upper().find(CAUSALE_SEZIONE_MARKER.upper())
    zona = ""
    if idx >= 0:
        sezione = text[idx:idx+200]
        m_z = CAUSALE_RE.search(sezione)
        if m_z: zona = m_z.group(1)[1:5]
        
    return (data, luogo, zona)


def _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov):
    ind = _val(vals[col_ind]) if col_ind < len(vals) else ""
    cap_val = vals[col_cap] if col_cap < len(vals) else None
    cap = str(int(cap_val)) if cap_val is not None and isinstance(cap_val, (int, float)) and not (isinstance(cap_val, float) and str(cap_val) == "nan") else (_val(cap_val) or "")
    citta = _val(vals[col_citta]) if col_citta < len(vals) else ""
    prov = _val(vals[col_prov]) if col_prov < len(vals) else ""
    parts = []
    if ind:
        parts.append(ind)
    if cap or citta:
        loc = f"{cap} {citta}".strip()
        if prov:
            loc = f"{loc} ({prov})"
        parts.append(loc)
    return ", ".join(parts) if parts else ""


def _carica_mappatura():
    """Carica mappatura: codice -> (row_idx, dato con cod_f, cod_l, nome, indirizzo, lat, lon, orari)."""
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA, read_only=True, data_only=True)
    ws = wb["Mappatura"] if "Mappatura" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    col_cod_f = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Frutta"), 0)
    col_cod_l = next((i for i, h in enumerate(headers) if str(h or "").strip() == "Codice Latte"), 1)
    col_nome = next((i for i, h in enumerate(headers) if "chi va" in str(h or "").lower()), 2)
    col_ind = next((i for i, h in enumerate(headers) if h == "Indirizzo"), 4)
    col_cap = next((i for i, h in enumerate(headers) if h == "CAP"), 5)
    col_citta = next((i for i, h in enumerate(headers) if h == "Città"), 6)
    col_prov = next((i for i, h in enumerate(headers) if h == "Provincia"), 7)
    col_om = next((i for i, h in enumerate(headers) if h == "Orario min"), 10)
    col_oM = next((i for i, h in enumerate(headers) if h == "Orario max"), 11)
    col_lat = next((i for i, h in enumerate(headers) if h == "Latitudine"), 12)
    col_lon = next((i for i, h in enumerate(headers) if h == "Longitudine"), 13)

    map_codice = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row]
        try:
            lat = float(vals[col_lat]) if col_lat < len(vals) and vals[col_lat] is not None else None
            lon = float(vals[col_lon]) if col_lon < len(vals) and vals[col_lon] is not None else None
        except (TypeError, ValueError, IndexError):
            lat, lon = None, None
        cod_f = _val(vals[col_cod_f]) if col_cod_f < len(vals) else ""
        cod_l = _val(vals[col_cod_l]) if col_cod_l < len(vals) else ""
        dato = {
            "codice_frutta": cod_f.lower() if cod_f else "",
            "codice_latte": cod_l.lower() if cod_l else "",
            "nome": _val(vals[col_nome]) if col_nome < len(vals) else "",
            "indirizzo": _build_indirizzo(vals, col_ind, col_cap, col_citta, col_prov),
            "orario_min": _val(vals[col_om]) if col_om < len(vals) else "",
            "orario_max": _val(vals[col_oM]) if col_oM < len(vals) else "",
            "lat": lat,
            "lon": lon,
        }
        for cod in (cod_f, cod_l):
            if cod:
                c = cod.lower()
                if c not in map_codice:
                    map_codice[c] = (row_idx, dato)
    wb.close()
    return map_codice


def _elabora_cartella(cartella_input: Path, map_codice: dict,
                      pdf_frutta_presenti: set, pdf_latte_presenti: set) -> list[dict]:
    """Estrae i DDT dalla cartella e ritorna lista di punti trovati.

    codice_frutta / codice_latte vengono impostati SOLO se il PDF
    corrispondente e' fisicamente presente nella rispettiva cartella.
    Questo evita i falsi ⚠️ 'PDF non trovato' in 9_genera_distinte
    per scuole che in mappatura hanno entrambi i codici ma quel giorno
    ricevono solo frutta (o solo latte).
    """
    import pdfplumber
    punti_per_riga: dict[int, dict] = {}

    for pdf_path in sorted(cartella_input.glob("*.pdf")):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    _, luogo, zona = _estrai_data_luogo_zona(text)
                    if not luogo:
                        continue
                    if luogo not in map_codice:
                        continue
                    row_idx, dato = map_codice[luogo]
                    if row_idx not in punti_per_riga:
                        punti_per_riga[row_idx] = {
                            **dato,
                            "codice_frutta": "p00000",   # di default: nessuno
                            "codice_latte":  "p00000",   # di default: nessuno
                            "codici_ddt_trovati": [luogo],
                            "zona": zona
                        }
                    else:
                        punti_per_riga[row_idx]["codici_ddt_trovati"].append(luogo)
                        if zona: punti_per_riga[row_idx]["zona"] = zona

                    # Imposta codice_frutta SOLO se il PDF frutta e' presente
                    cf = dato.get("codice_frutta", "")
                    if cf and cf.lower() not in ("p00000", "") and cf.lower() in pdf_frutta_presenti:
                        punti_per_riga[row_idx]["codice_frutta"] = cf

                    # Imposta codice_latte SOLO se il PDF latte e' presente
                    cl = dato.get("codice_latte", "")
                    if cl and cl.lower() not in ("p00000", "") and cl.lower() in pdf_latte_presenti:
                        punti_per_riga[row_idx]["codice_latte"] = cl

        except Exception as e:
            print(f"  Errore {pdf_path.name}: {e}")

    # unisci codici DDT trovati senza duplicati
    for pt in punti_per_riga.values():
        pt["codici_ddt_trovati"] = ", ".join(sorted(set(pt["codici_ddt_trovati"])))

    return list(punti_per_riga.values())


def _salva_excel(punti: list[dict], out_path: Path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Punti consegna"
    headers = [
        "Codice Frutta", "Codice Latte", "Codici DDT trovati", "Zona",
        "Nome", "Indirizzo", "Orario min", "Orario max", "Latitudine", "Longitudine"
    ]
    ws.append(headers)
    for pt in punti:
        ws.append([
            pt.get("codice_frutta", ""),
            pt.get("codice_latte", ""),
            f"{pt.get('codice_frutta') or 'p00000'}_{pt.get('codice_latte') or 'p00000'}",  # codice univoco composito
            pt.get("zona", ""),
            pt.get("nome", ""),
            pt.get("indirizzo", ""),
            pt.get("orario_min", ""),
            pt.get("orario_max", ""),
            pt.get("lat") if pt.get("lat") is not None else "",
            pt.get("lon") if pt.get("lon") is not None else "",
        ])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Salvato: {out_path}")


def main():
    if len(sys.argv) < 2:
        # prendi l'ultima cartella
        folders = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not folders:
            print("Uso: py crea_punti_consegna.py <data>")
            return 1
        folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = folders[0].name[len("CONSEGNE_"):]  # es. "30-03-2026" o "30-03-2026_31-03-2026"
        print(f"Nessuna data specificata. Uso l'ultima cartella: {data}")
    else:
        data = sys.argv[1].strip()
    if re.match(r"^\d{2}-\d{2}$", data):
        data = f"{data}-2026"

    output_base = CONSEGNE_DIR / f"CONSEGNE_{data}"
    input_root = output_base / "DDT-ORIGINALI-DIVISI" if (output_base / "DDT-ORIGINALI-DIVISI").exists() else output_base
    input_frutta = input_root / "FRUTTA"
    input_latte = input_root / "LATTE"
    output_file = output_base / "punti_consegna.xlsx"

    print(f"\n--- Creazione punti consegna unificati ---\nCartella: {output_base}\n")

    if not MAPPATURA.exists():
        print(f"Mappatura non trovata: {MAPPATURA}")
        return 1

    map_codice = _carica_mappatura()
    print(f"Mappatura caricata: {len(map_codice)} codici\n")

    # ── Costruisce i set di codici PDF fisicamente presenti ──────────────────
    # Legge i nomi file nelle cartelle FRUTTA e LATTE ed estrae il codice
    # (es. "p2063_30-03-2026.pdf" → "p2063")
    def _codici_in_cartella(cart: Path) -> set:
        if not cart.exists():
            return set()
        codici = set()
        for f in cart.glob("*.pdf"):
            # nome atteso: {codice}_{data}.pdf  oppure {codice}_{data}_N.pdf
            parti = f.stem.split("_")
            if parti:
                codici.add(parti[0].lower())
        return codici

    pdf_frutta_presenti = _codici_in_cartella(input_frutta)
    pdf_latte_presenti  = _codici_in_cartella(input_latte)
    print(f"  PDF presenti → FRUTTA: {len(pdf_frutta_presenti)}  LATTE: {len(pdf_latte_presenti)}")

    punti_totali = []
    for cartella in [input_frutta, input_latte]:
        if cartella.exists():
            punti_totali.extend(_elabora_cartella(cartella, map_codice,
                                                  pdf_frutta_presenti, pdf_latte_presenti))
        else:
            print(f"  Cartella non trovata: {cartella}")

    # elimina duplicati basati su nome + indirizzo
    punti_unici: dict[tuple[str, str], dict] = {}
    for pt in punti_totali:
        chiave = (pt["nome"], pt["indirizzo"])
        if chiave in punti_unici:
            # aggiorna codici e DDT trovati
            # Aggiorna solo se il nuovo valore è un codice reale (non p00000, non vuoto)
            esistente = punti_unici[chiave]
            cf_new = pt.get("codice_frutta", "")
            cl_new = pt.get("codice_latte",  "")
            if cf_new and cf_new.lower() not in ("p00000", ""):
                esistente["codice_frutta"] = cf_new
            if cl_new and cl_new.lower() not in ("p00000", ""):
                esistente["codice_latte"] = cl_new
            esistente["codici_ddt_trovati"] = ", ".join(
                sorted(set(esistente["codici_ddt_trovati"].split(", ") + pt["codici_ddt_trovati"].split(", ")))
            )
        else:
            punti_unici[chiave] = pt

    _salva_excel(list(punti_unici.values()), output_file)
    print("\n--- Completato ---\n")
    return 0


if __name__ == "__main__":
    exit(main())