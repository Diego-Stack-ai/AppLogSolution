"""
migra_clienti_firebase.py
=========================
Script di migrazione UNA TANTUM che popola la struttura "a cartelle":
  customers/{PARTNER}/clienti/{ID_CLIENTE}

I 4 Partner principali (DNR, GRAN CHEF, BAUER, CATTEL)
vengono inizializzati automaticamente.

Dipendenze:
    pip install firebase-admin openpyxl pandas python-dotenv

Uso:
    python migra_clienti_firebase.py
"""

import os
import sys
from pathlib import Path

# ── Configurazioni di Percorso ──────────────────────────────────────────────
ROOT_DIR = Path(__file__).resolve().parent

# ── Percorsi file Excel sorgente ─────────────────────────────────────────────
EXCEL_DNR       = ROOT_DIR / "Progetto Scuole" / "PROGRAMMA" / "mappatura_destinazioni.xlsx"
EXCEL_GRAN_CHEF = ROOT_DIR / "Fatturazione" / "Anagrafica_Clienti_Master.xlsx"

# ── SDK Firebase ─────────────────────────────────────────────────────────────
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except ImportError:
    print("❌ Dipendenza mancante. Esegui: pip install firebase-admin")
    sys.exit(1)

try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("❌ Dipendenza mancante. Esegui: pip install openpyxl pandas")
    sys.exit(1)


def init_firebase():
    """Inizializza Firebase Admin SDK con Service Account."""
    if not firebase_admin._apps:
        candidates = list((ROOT_DIR / "backend" / "config").glob("*firebase-adminsdk*.json")) + \
                     [ROOT_DIR / "firebase-service-account.json"]

        sa_path = next((p for p in candidates if p.exists()), None)

        if sa_path:
            cred = credentials.Certificate(str(sa_path))
            firebase_admin.initialize_app(cred)
            print(f"✅ Firebase inizializzato con: {sa_path.name}")
        else:
            print("❌ Nessun Service Account trovato.")
            sys.exit(1)
    return firestore.client()


def cancella_collezione(db, percorso_coll: str, batch_size: int = 400):
    """Cancella tutti i documenti di una collezione al percorso dato in batch."""
    coll_ref = db.collection(percorso_coll)
    deleted = 0
    while True:
        docs = list(coll_ref.limit(batch_size).stream())
        if not docs:
            break
        batch = db.batch()
        for doc in docs:
            batch.delete(doc.reference)
        batch.commit()
        deleted += len(docs)
    if deleted > 0:
        print(f"   🗑️  Puliti {deleted} vecchi documenti da '{percorso_coll}'.")


def setup_macroaree(db):
    """Crea la struttura di base e pulisce eventuali dati vecchi vecchi."""
    print("\n🏗️  Setup della Struttura Architetturale (Macroaree)...")
    
    # 1. Pulisce eventuali vecchi dati flat caricati nell'ultimo test
    cancella_collezione(db, "customers")
    
    # 2. Crea / Inizializza i documenti Partner e pulisce le loro sotto-collezioni
    partners = ["DNR", "GRAN CHEF", "BAUER", "CATTEL"]
    for p in partners:
        # A. Salva info del partner 
        db.collection("customers").document(p).set({
            "nome_partner": p,
            "tipo": "macroarea_trasporti",
            "attivo": True
        })
        
        # B. Svuota la sub-collection 'clienti' del partner nel caso di riesecuzione
        cancella_collezione(db, f"customers/{p}/clienti")
        
    print("✅ Macroaree (DNR, GRAN CHEF, BAUER, CATTEL) pronte e pulite.")


def carica_dnr(db) -> int:
    """Legge mappatura_destinazioni.xlsx e inserisce in customers/DNR/clienti/"""
    print(f"\n📂 Caricamento clienti in customers/DNR/clienti/...")
    if not EXCEL_DNR.exists():
        print(f"❌ File non trovato: {EXCEL_DNR}")
        return 0

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_DNR, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(nome):
        return next((i for i, h in enumerate(headers) if str(h or "").strip() == nome), None)

    i_cod_f, i_cod_l, i_nome = col("Codice Frutta"), col("Codice Latte"), col("A chi va consegnato")
    i_ind, i_cap, i_citta, i_prov = col("Indirizzo"), col("CAP"), col("Città"), col("Provincia")
    i_lat, i_lon, i_gps = col("Latitudine"), col("Longitudine"), col("COORDINATE_REALI_GPS")
    
    i_tipo = col("Tipologia grado")
    i_email = col("Email")
    i_sito = col("Sito web")
    i_omin = col("Orario min")
    i_omax = col("Orario max")

    def v(row, idx):
        if idx is None or idx >= len(row): return None
        val = row[idx]
        if val is None: return None
        s = str(val).strip()
        return None if s.lower() in ("nan", "") else s

    coll = db.collection("customers").document("DNR").collection("clienti")
    batch = db.batch()
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_f, cod_l, nome = v(row, i_cod_f), v(row, i_cod_l), v(row, i_nome)

        if not cod_f and not cod_l and not nome:
            continue

        # Converti CAP come stringa fissa se numerico
        cap_raw = row[i_cap] if i_cap is not None and i_cap < len(row) else None
        try: cap_str = str(int(float(cap_raw))) if cap_raw is not None else None
        except: cap_str = str(cap_raw).strip() if cap_raw else None

        lat = lon = None
        try:
            lat = float(row[i_lat]) if i_lat is not None and i_lat < len(row) and row[i_lat] else None
            lon = float(row[i_lon]) if i_lon is not None and i_lon < len(row) and row[i_lon] else None
        except: pass

        doc = {
            "codice_frutta":         cod_f or "",
            "codice_latte":          cod_l or "",
            "nome":                  nome or "",
            "indirizzo":             v(row, i_ind) or "",
            "cap":                   cap_str or "",
            "citta":                 v(row, i_citta) or "",
            "provincia":             v(row, i_prov) or "",
            "lat":                   lat,
            "lon":                   lon,
            "coordinate_reali_gps":  v(row, i_gps) or "",
            "tipologiaGrado":        v(row, i_tipo) or "-",
            "email":                 v(row, i_email) or "",
            "homePage":              v(row, i_sito) or "",
            "orarioMin":             v(row, i_omin) or "",
            "orarioMax":             v(row, i_omax) or ""
        }

        doc_id = f"DNR_{cod_f or 'X'}_{cod_l or 'X'}".replace(" ", "_")
        batch.set(coll.document(doc_id), doc)
        count += 1

        if count % 400 == 0:
            batch.commit()
            batch = db.batch()

    if count % 400 != 0:
        batch.commit()

    wb.close()
    print(f"✅ DNR: {count} clienti caricati correttamente nella cartella DNR.")
    return count


def carica_gran_chef(db) -> int:
    """Legge l'Anagrafica Gran Chef e inserisce in customers/GRAN CHEF/clienti/"""
    print(f"\n📂 Caricamento clienti in customers/GRAN CHEF/clienti/...")
    if not EXCEL_GRAN_CHEF.exists():
        print(f"❌ File non trovato: {EXCEL_GRAN_CHEF}")
        return 0

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_GRAN_CHEF, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(nome):
        return next((i for i, h in enumerate(headers) if str(h or "").strip() == nome), None)

    i_cod, i_nome, i_ind = col("Codice Cliente"), col("Ragione Sociale"), col("Indirizzo")
    i_loc, i_prov = col("Località"), col("Provincia")

    def v(row, idx):
        if idx is None or idx >= len(row): return None
        val = row[idx]
        if val is None: return None
        s = str(val).strip()
        return None if s.lower() in ("nan", "") else s

    coll = db.collection("customers").document("GRAN CHEF").collection("clienti")
    batch = db.batch()
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        codice, nome = v(row, i_cod), v(row, i_nome)
        if not codice and not nome: continue

        doc = {
            "codice_cliente":       codice or "",
            "nome":                 nome or "",
            "indirizzo":            v(row, i_ind) or "",
            "cap":                  "",
            "localita":             v(row, i_loc) or "",
            "provincia":            v(row, i_prov) or "",
            "lat":                  None,
            "lon":                  None,
            "coordinate_reali_gps": "",
        }

        cod_clean = str(codice).replace(" ", "_") if codice else "SENZA_CODICE"
        doc_id = f"GC_{cod_clean}"
        batch.set(coll.document(doc_id), doc)
        count += 1

        if count % 400 == 0:
            batch.commit()
            batch = db.batch()

    if count % 400 != 0:
        batch.commit()

    wb.close()
    print(f"✅ GRAN CHEF: {count} clienti caricati correttamente nella cartella GRAN CHEF.")
    return count


def main():
    print("=" * 60)
    print("  MIGRAZIONE STRUTTURA A MACROAREE FIREBASE FIRESTORE")
    print("=" * 60)

    db = init_firebase()

    risposta = input("\n⚠️  Vuoi SOVRASCRIVERE E RIORGANIZZARE tutti i dati su Firebase con la nuova architettura a cartelle? (s/N): ").strip().lower()
    if risposta != "s":
        print("Operazione annullata.")
        return

    # Inizializza macroaree e pulisce i dati correnti
    setup_macroaree(db)

    # Scrittura sub-collections
    tot_dnr = carica_dnr(db)
    tot_gc = carica_gran_chef(db)

    print("\n" + "=" * 60)
    print(f"  ✅ Riorganizzazione Firebase conclusa con successo!")
    print(f"     -> customers/DNR/clienti/       ({tot_dnr} doc)")
    print(f"     -> customers/GRAN CHEF/clienti/ ({tot_gc} doc)")
    print(f"     -> customers/BAUER              (Inizializzato vuoto)")
    print(f"     -> customers/CATTEL             (Inizializzato vuoto)")
    print("=" * 60)


if __name__ == "__main__":
    main()
