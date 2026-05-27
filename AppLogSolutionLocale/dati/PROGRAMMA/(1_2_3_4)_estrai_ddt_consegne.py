#!/usr/bin/env python3
"""
(1_2_3_4)_estrai_ddt_consegne.py
════════════════════════════════
Estrae tutti i DDT dai PDF in CONSEGNE/DDT-ORIGINALI/FRUTTA e LATTE.
Identifica nuovi clienti ed estrae automaticamente i loro dati (Indirizzo, CAP, Orari, ecc.).

Uso: py (1_2_3_4)_estrai_ddt_consegne.py [data]
"""

import re
import sys
import subprocess
import time
import json
from pathlib import Path
from datetime import datetime

# Forza la codifica UTF-8 per il terminale Windows (evita errori con le emoji)
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    print("pip install pypdf")
    sys.exit(1)

# --- CONFIGURAZIONE ---
PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
INPUT_FRUTTA = BASE_DIR / "FRUTTA"
INPUT_LATTE = BASE_DIR / "LATTE"
MAPPATURA_XLSX = PROG_DIR / "mappatura_destinazioni.xlsx"

# Regex per estrazione dati
DATA_DDT_RE = re.compile(r'del\s+(\d{2})/(\d{2})/(\d{4})', re.I)
LUOGO_RE = re.compile(r'(?:[Ll]uogo [Dd]i [Dd]estinazione|[Cc]odice [Dd]estinazione):\s*([pP]\d{4,5})')
CAP_RE = re.compile(r"\b(\d{5})\b")
PROVINCIA_RE = re.compile(r"\(([A-Z]{2})\)")
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4})(?:\s+H(\d{2}))?(?:\s+(\d{3}))?', re.I)
NUM_DDT_RE = re.compile(r'DDT\s*[Nn][°º\.\s]*([A-Za-z0-9/-]+)', re.I)


def _estrai_data_luogo(text: str) -> tuple[str | None, str | None, str | None]:
    """Estrae (data, luogo, num_ddt) da una pagina DDT. data in formato DD-MM-YYYY."""
    data = None
    m = DATA_DDT_RE.search(text)
    if m:
        data = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    luogo_m = LUOGO_RE.search(text)
    luogo = luogo_m.group(1).lower() if luogo_m else None
    
    num_m = NUM_DDT_RE.search(text)
    num_ddt = num_m.group(1).replace("/", "-") if num_m else "UNK"
    
    return (data, luogo, num_ddt)


def _estrai_dati_consegna_da_testo(text: str, codice: str, da_frutta: bool) -> dict:
    """Estrae destinatario, indirizzo, CAP, città, provincia e orari dal testo di una pagina DDT."""
    res = {"dest": "", "ind": "", "cap": "", "cit": "", "prov": "", "om": "", "oM": "14:00"}
    if codice.lower() not in text.lower():
        return res
    
    idx_l = text.find("Luogo di destinazione")
    if idx_l < 0: return res

    # 1. Nome e Indirizzo
    if da_frutta:
        blocco = text[idx_l : idx_l + 650]
        lines = [ln.strip() for ln in blocco.split("\n") if ln.strip()]
        for i, ln in enumerate(lines):
            if LUOGO_RE.search(ln):
                if i + 1 < len(lines): res["dest"] = lines[i + 1].strip().title()
                if i + 2 < len(lines): res["ind"] = lines[i + 2].strip().title()
                break
    else:
        idx_causale = text.upper().find("CAUSALE DEL TRASPORTO")
        blocco = text[:idx_causale] if idx_causale > 0 else text[idx_l : idx_l + 900]
        for ln in blocco.split("\n"):
            ln = ln.strip()
            cf_m = re.match(r"^[Cc]\.?[Ff]\.?\s+", ln)
            if cf_m: res["dest"] = ln[cf_m.end():].strip().title()
            else:
                albo_m = re.match(r"^[Aa]lbo\s+", ln, re.I)
                if albo_m: res["ind"] = ln[albo_m.end():].strip().title()

    # 2. CAP, Provincia, Città
    idx_resp = text.upper().find("RESPONSABILE DEL TRASPORTO")
    blocco_prov = text[idx_resp:] if idx_resp >= 0 else text
    
    for prov_m in PROVINCIA_RE.finditer(blocco_prov):
        sigla = prov_m.group(1)
        if sigla == "MN" and ("Pomponesco" in blocco_prov[max(0, prov_m.start()-40):prov_m.start()] or "46030" in blocco_prov):
            continue
        res["prov"] = sigla
        caps = list(CAP_RE.finditer(blocco_prov[:prov_m.start()]))
        if caps:
            res["cap"] = caps[-1].group(1)
            pre = blocco_prov[caps[-1].end() : caps[-1].end() + 60]
            citta_m = re.search(r"\s*[-]?\s*([A-Za-zÀ-ÿ\s'.]+?)\s*\([A-Z]{2}\)", pre)
            if citta_m: res["cit"] = citta_m.group(1).strip().title()
        break
        
    # 3. Orari
    idx_c = text.upper().find("CAUSALE DEL TRASPORTO")
    if idx_c >= 0:
        sezione = text[idx_c:idx_c+150]
        m = CAUSALE_RE.search(sezione)
        if m:
            if m.group(2): res["oM"] = f"{int(m.group(2)):02d}:00"
            if m.group(3):
                s = m.group(3)
                if len(s) == 3: res["om"] = f"{int(s[0]):02d}:{int(s[1:3]):02d}"
    return res


def _ricava_date_da_pdf() -> list[str]:
    """Raccoglie TUTTE le date distinte trovate nei PDF di FRUTTA e LATTE."""
    import pdfplumber
    date_trovate: set[str] = set()
    for cart in (INPUT_FRUTTA, INPUT_LATTE):
        if not cart.exists(): continue
        for pdf_path in sorted(cart.glob("*.pdf")):
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        data, _, _ = _estrai_data_luogo(text)
                        if data: date_trovate.add(data)
            except: pass
    return sorted(date_trovate)


def _leggi_codici_mappatura() -> dict:
    """Restituisce {codice: (row_idx, om_frutta, oM_frutta, om_latte, oM_latte)}."""
    if not MAPPATURA_XLSX.exists(): return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(MAPPATURA_XLSX, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        col_om_f = next((i for i, h in enumerate(headers) if h == "Orario min Frutta"), 10)
        col_oM_f = next((i for i, h in enumerate(headers) if h == "Orario max Frutta"), 11)
        col_om_l = next((i for i, h in enumerate(headers) if h == "Orario min Latte"),  12)
        col_oM_l = next((i for i, h in enumerate(headers) if h == "Orario max Latte"),  13)
        codici = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = [c.value for c in row]
            def _v(x): return str(x).strip() if x is not None else ""
            c_f = _v(vals[0]).lower()
            c_l = _v(vals[1]).lower() if len(vals) > 1 else ""
            om_f = _v(vals[col_om_f]) if col_om_f < len(vals) else ""
            oM_f = _v(vals[col_oM_f]) if col_oM_f < len(vals) else ""
            om_l = _v(vals[col_om_l]) if col_om_l < len(vals) else ""
            oM_l = _v(vals[col_oM_l]) if col_oM_l < len(vals) else ""
            if c_f and c_f != "p00000": codici[c_f] = (row_idx, om_f, oM_f, om_l, oM_l)
            if c_l and c_l != "p00000": codici[c_l] = (row_idx, om_f, oM_f, om_l, oM_l)
        wb.close()
        return codici
    except Exception as e:
        print(f"⚠️ Errore mappatura: {e}")
        return {}


def _aggiorna_orari_mappatura(row_idx: int, orario_min: str, orario_max: str, tipo: str):
    """Scrive Orario min/max sulla colonna FRUTTA o LATTE in mappatura_destinazioni.xlsx."""
    if not MAPPATURA_XLSX.exists() or not row_idx: return
    tipo_cap = tipo.capitalize()  # "Frutta" o "Latte"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(MAPPATURA_XLSX)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        col_om = next((i + 1 for i, h in enumerate(headers) if h == f"Orario min {tipo_cap}"), None)
        col_oM = next((i + 1 for i, h in enumerate(headers) if h == f"Orario max {tipo_cap}"), None)
        if col_om is None or col_oM is None:
            print(f"    ⚠️  Colonne orario {tipo_cap} non trovate in mappatura.")
            return
        if orario_min: ws.cell(row=row_idx, column=col_om, value=orario_min)
        if orario_max: ws.cell(row=row_idx, column=col_oM, value=orario_max)
        wb.save(MAPPATURA_XLSX)
        print(f"    ✅ Mappatura [{tipo_cap}] riga {row_idx}: min={orario_min or '—'}  max={orario_max or '—'}")
    except PermissionError:
        print(f"    ⚠️  mappatura_destinazioni.xlsx e' aperto — orari non aggiornati.")
    except Exception as e:
        print(f"    ⚠️  Errore aggiornamento orari: {e}")


def _verifica_nuovi_clienti(dati_nuovi: dict):
    if not dati_nuovi: return True
    print("\n" + "!"*60)
    print(f"🛑 RILEVATI {len(dati_nuovi)} NUOVI CODICI CLIENTE:")
    for cod, info in sorted(dati_nuovi.items()):
        print(f"   - {cod} ({info['tipo']}): {info['dest']} - {info['cit']}")
    print("!"*60 + "\n")
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nuovi Codici"
        headers = ["Codice Frutta", "Codice Latte", "A chi va consegnato", "Tipologia grado", "Indirizzo", "CAP", "Città", "Provincia",
                  "Tipologia consegna", "Tipologia consegna.1", "Email", "Sito web", "Orario min", "Orario max"]
        ws.append(headers)
        for cod, info in sorted(dati_nuovi.items()):
            row = [""] * 14
            t = info.get("tipo", "")
            if t == "FRUTTA" or t == "GRAND CHEF":
                row[0] = cod
            else:
                row[1] = cod
            row[2], row[3], row[4], row[5], row[6], row[7], row[12], row[13] = info["dest"], t, info["ind"], info["cap"], info["cit"], info["prov"], info["om"], info["oM"]
            ws.append(row)
        hp = BASE_DIR / "nuovi_codici_consegna.xlsx"
        wb.save(hp)
        print(f"📄 File generato: {hp.name}\n💡 Copia i dati in mappatura_destinazioni.xlsx per proseguire.\n")
    except Exception as e: print(f"⚠️ Errore excel: {e}")
    return False


def _estrai_solo_articoli_per_confronto(text: str) -> str:
    idx = text.find("Cod. Articolo")
    if idx == -1: idx = text.find("Descrizione Natura")
    if idx == -1: return text
    return text[idx:]

def _estrai_da_cartella(cart_in: Path, cart_out: Path, etichetta: str, date_valide: set[str], mappati: dict, *, duplicata: bool = False):
    """Estrae DDT raggruppandoli per (Punto di Consegna, Data, Numero DDT). Per FRUTTA decodifica l'ordine di stampa."""
    nuovi_dati = {}
    aggiornati_orari: set[int] = set()
    if not cart_in.exists(): return 0, 0, nuovi_dati
    cart_out.mkdir(parents=True, exist_ok=True)
    pdf_files = list(cart_in.glob("*.pdf"))
    if not pdf_files: return 0, 0, nuovi_dati
    
    import pdfplumber
    creati = 0
    visti = {}
    
    blocchi = {}  # { (l, d, num_ddt) : [(text, pypdf_page)] }
    readers_open = []  # Manteniamo aperti i reader fino alla scrittura
    
    # 1. LETTURA GLOBALE E RAGGRUPPAMENTO
    for pdf_path in pdf_files:
        try:
            reader = PdfReader(pdf_path)
            readers_open.append(reader)
            with pdfplumber.open(pdf_path) as pdf:
                for i in range(len(pdf.pages)):
                    text = pdf.pages[i].extract_text() or ""
                    d, l, num_ddt = _estrai_data_luogo(text)
                    if not d or not l or d not in date_valide: continue
                    
                    # Gestione nuovi clienti
                    if l not in mappati and l not in nuovi_dati:
                        nuovi_dati[l] = _estrai_dati_consegna_da_testo(text, l, duplicata)
                        nuovi_dati[l]["tipo"] = etichetta
                    elif l in mappati:
                        row_idx_m, om_f, oM_f, om_l, oM_l = mappati[l]
                        om_mappa = om_f if etichetta == "FRUTTA" else om_l
                        oM_mappa = oM_f if etichetta == "FRUTTA" else oM_l
                        if row_idx_m not in aggiornati_orari:
                            idx_c = text.upper().find("CAUSALE DEL TRASPORTO")
                            if idx_c >= 0:
                                m_c = CAUSALE_RE.search(text[idx_c:idx_c+150])
                                if m_c:
                                    oM_ddt = f"{int(m_c.group(2)):02d}:00" if m_c.group(2) else ""
                                    om_ddt = ""
                                    if m_c.group(3):
                                        s = m_c.group(3)
                                        if len(s) == 3: om_ddt = f"{int(s[0]):02d}:{int(s[1:3]):02d}"
                                        elif len(s) == 4: om_ddt = f"{int(s[:2]):02d}:{int(s[2:]):02d}"
                                    needs_update = ((oM_ddt and oM_ddt != oM_mappa) or (om_ddt and om_ddt != om_mappa))
                                    if needs_update:
                                        print(f"    [ORARIO {etichetta}] {l}: DDT({om_ddt or '-'}/{oM_ddt or '-'}) vs Mappatura({om_mappa or '-'}/{oM_mappa or '-'})")
                                        _aggiorna_orari_mappatura(row_idx_m, om_ddt or None, oM_ddt or None, etichetta)
                                        if etichetta == "FRUTTA": mappati[l] = (row_idx_m, om_ddt or om_f, oM_ddt or oM_f, om_l, oM_l)
                                        else: mappati[l] = (row_idx_m, om_f, oM_f, om_ddt or om_l, oM_ddt or oM_l)
                                        aggiornati_orari.add(row_idx_m)
                                        
                    chiave = (l, d, num_ddt)
                    if chiave not in blocchi:
                        blocchi[chiave] = []
                    blocchi[chiave].append((text, reader.pages[i]))
        except Exception as e: print(f"  Errore {pdf_path.name}: {e}")

    # 2. ANALISI DEL BLOCCO CAMPIONE (Solo per la Frutta)
    is_fascicolato = False
    '''
    if duplicata and blocchi:
        max_chiave = max(blocchi.keys(), key=lambda k: len(blocchi[k]))
        max_pages = blocchi[max_chiave]
        n_pages = len(max_pages)
        
        if n_pages > 2:
            # Confronto gli articoli tra la prima e la seconda pagina del blocco
            art_p1 = _estrai_solo_articoli_per_confronto(max_pages[0][0])
            art_p2 = _estrai_solo_articoli_per_confronto(max_pages[1][0])
            if art_p1 == art_p2:
                is_fascicolato = False
                print(f"    [PATTERN {etichetta}] Non Fascicolato (1,1,2,2) rilevato sul blocco {max_chiave[0]} ({n_pages} pag.)")
            else:
                is_fascicolato = True
                print(f"    [PATTERN {etichetta}] Fascicolato (1,2,1,2) rilevato sul blocco {max_chiave[0]} ({n_pages} pag.)")
    '''

    # 3. TAGLIO FOTOCOPIE E SCRITTURA
    doppioni_totali = 0
    for chiave, lista_pagine in blocchi.items():
        writer = PdfWriter()
        l, d, num_ddt = chiave
        n = len(lista_pagine)
        pagine_da_salvare = []
        
        '''
        if duplicata:
            doppioni_totali += 1 if n > 1 else 0
            if is_fascicolato:
                # Tieni la prima metà
                half = n // 2
                pagine_da_salvare = [p[1] for p in lista_pagine[:half]]
            else:
                # Tieni una pagina sì e una no (indici pari)
                pagine_da_salvare = [lista_pagine[i][1] for i in range(0, n, 2)]
        else:
        '''
        # LATTE o copia singola: salva tutto il blocco intatto
        doppioni_totali += 1 if n > 1 else 0
        pagine_da_salvare = [p[1] for p in lista_pagine]
            
        for pg in pagine_da_salvare:
            writer.add_page(pg)
            
        cnt = visti.get(chiave, 0) + 1
        visti[chiave] = cnt
        fname = f"{l}_{d}_{num_ddt}_{cnt}.pdf" if cnt > 1 else f"{l}_{d}_{num_ddt}.pdf"
        with open(cart_out / fname, "wb") as f:
            writer.write(f)
        creati += 1
        if creati <= 3 or creati % 50 == 0: print(f"    {fname}")

    return creati, doppioni_totali, nuovi_dati


def _pulisci_sorgenti(cart_in: Path, date_valide: set[str]):
    """Elimina i PDF sorgenti che contengono pagine di qualsiasi data elaborata."""
    import pdfplumber
    rimossi = 0
    for p in cart_in.glob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                if any(_estrai_data_luogo(pg.extract_text() or "")[0] in date_valide for pg in pdf.pages):
                    pdf.close(); p.unlink(); rimossi += 1
        except: pass
    return rimossi


def _pulisci_output(base: Path, data_v: str):
    for f in [base/"punti_consegna.xlsx", base/"punti_consegna_unificati.json", base/"4_mappa_zone_google.html", base/f"zone_google_{data_v.replace('-', '_')}.kml"]:
        if f.exists():
            try: f.unlink()
            except: pass


# --- ARTICOLI NOTI (SORGENTE DI VERITÀ) ---
ARTICOLI_NOTI = {
    "10-FLYER", "10-GEL", "10-MANIFESTO", "10-AT-01", "10-BICC", "10-CUCCH", "10-PIATTO",
    "AP-SU-PC", "FO-DI-PV-04-LB", "FO-DI-GP-01-NI", "FVNS-03", "FVNS-03-", 
    "LT-AQ-04-LV", "LT-AQ-04-LB", "LT-AQ-04-LS", "LT-DL-02-LC", "LT-ES-04-LS", "LT-ESL-IN-LB", 
    "MA-T-LI-L3-NA", "ME-T-DI-V0-NA", "ME-S-BI-L3-NA", "PE-T-DI-L3-NA",
    "YO-BI-MN-04-LB", "YO-DL-02-LC", "FI-Z-BI-L3-NA", "FR-M-BI-L3-NI",
    "LNS-04-GADGET", "LNS-04-", "CA-Z-BI-L3-NA", "KI-S-BI-L3-NA", "ME-S-DI-L3-NA", "FO-DI-AS-04-LV",
    "AL-M-BI-L3-NI", "SUCCO-REC", "PF-T-LI-L3-NA", "SU-M-BI-L3-NI", "YO-CN-MN-04-",
    "AL-T-LI-NA", "NE-M-BI-L3-NI",
}

def _is_primary_code(text: str) -> bool:
    """Verifica se una stringa corrisponde a un codice base noto (Sorgente di Verità)."""
    if not text: return False
    t = text.strip().upper()
    if t in ARTICOLI_NOTI: return True
    for prefix in ARTICOLI_NOTI:
        if prefix.endswith('-') and t.startswith(prefix.upper()):
            return True
    return False


def _normalizza_cella_codice_base(raw: str) -> str:
    """
    Dato il contenuto grezzo di una cella Cod. Articolo,
    restituisce SOLO il codice_base (prima riga significativa, filtro metadati).
    Uguale alla logica in 9_genera_distinte_da_viaggi.py.
    """
    righe = [l.strip() for l in raw.split('\n')
             if l.strip() and not l.strip().startswith("Codice:")]
    if not righe: return ""
    
    codice_base = righe[0]
    
    # ── OPZIONE A: Ricomposizione per codici troncati ──
    # Se la prima riga finisce col trattino, uniamo la prima parola della riga successiva
    if len(righe) > 1 and codice_base.endswith('-'):
        pezzi = righe[1].split()
        if pezzi:
            codice_base += pezzi[0]
            
    return codice_base


def _verifica_nuovi_articoli(base):
    """
    Verifica la presenza di codici articolo non presenti in ARTICOLI_NOTI.
    Usa logica column-based: estrae la colonna Cod. Articolo dalle tabelle PDF
    e confronta il codice_base (NON la variante) con la Sorgente di Verita'.
    Questo evita falsi positivi per varianti note (es. FVNS-03-FOLDER e' noto
    perche' il suo codice_base FVNS-03- e' in ARTICOLI_NOTI).
    """
    print("Verifica articoli (logica column-based)...")
    import pdfplumber

    codici_base_trovati = set()
    divisi = base / "DDT-ORIGINALI-DIVISI"
    if not divisi.exists(): return True

    for p in divisi.rglob("*.pdf"):
        try:
            with pdfplumber.open(p) as pdf:
                for pg in pdf.pages:
                    tables = pg.extract_tables()
                    if not tables: continue
                    # Cerca la tabella con "Cod. Articolo"
                    tab = next((t for t in tables if t and len(t) > 1
                                and "Cod. Articolo" in " ".join(str(c or "") for c in t[0])), None)
                    if not tab: continue
                    for row in tab[1:]:
                        if not row or not row[0]: continue
                        codice_base = _normalizza_cella_codice_base(str(row[0]))
                        if codice_base:
                            codici_base_trovati.add(codice_base)
        except: continue

    # Confronta i codici_base trovati con ARTICOLI_NOTI
    nuovi = {c for c in codici_base_trovati if not _is_primary_code(c)}

    if nuovi:
        print("\n" + "!"*60 + f"\n[!] NUOVI ARTICOLI RILEVATI: {', '.join(sorted(nuovi))}\n" + "!"*60 + "\n")
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Nuovi Articoli"
            ws.append(["Codice Articolo", "Data Rilevamento", "Azione"])
            for c in sorted(nuovi):
                ws.append([c, datetime.now().strftime("%d/%m/%Y"), "Aggiungere a ARTICOLI_NOTI"])
            rp = BASE_DIR / "nuovi_articoli_rilevati.xlsx"
            wb.save(rp)
            print(f"Report salvato: {rp.name}")
        except Exception as e:
            print(f"Errore salvataggio report: {e}")
        return False

    print(f"Articoli OK. ({len(codici_base_trovati)} codici base verificati)\n")
    return True


def clean_client_code(code_val):
    if code_val is None or (hasattr(code_val, "isna") and code_val.isna()):
        return ""
    code_str = str(code_val).strip()
    if code_str.endswith(".0"):
        code_str = code_str[:-2]
    return code_str

def parse_fascia_oraria(val):
    if val is None or (hasattr(val, "isna") and val.isna()) or val == "":
        return "", ""
    val_str = str(val).strip()
    match_range = re.findall(r'(\d{2}:\d{2})', val_str)
    if len(match_range) == 2:
        return match_range[0], match_range[1]
    match_dopo = re.search(r'(?:Dopo le|dopo le)\s*(\d{2}:\d{2})', val_str)
    if match_dopo:
        return match_dopo.group(1), ""
    match_entro = re.search(r'(?:Entro le|entro le)\s*(\d{2}:\d{2})', val_str)
    if match_entro:
        return "", match_entro.group(1)
    return "", ""

def _genera_pdf_placeholder_grand_chef(path_out: Path, codice: str, nome: str, ind: str, cit: str, prov: str, note: str, om: str, oM: str, data: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    doc = SimpleDocTemplate(str(path_out), pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('gc_title', parent=styles['Heading1'], fontSize=16, leading=20, textColor=colors.HexColor('#0f172a'), spaceAfter=15)
    body_style = ParagraphStyle('gc_body', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#334155'))
    label_style = ParagraphStyle('gc_label', parent=styles['Normal'], fontSize=10, leading=14, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'))
    
    elements = []
    elements.append(Paragraph(f"SCHEDA DI CONSEGNA - CANALE GRAND CHEF", title_style))
    elements.append(Spacer(1, 10))
    
    data_table = [
        [Paragraph("Codice Cliente:", label_style), Paragraph(codice, body_style)],
        [Paragraph("Destinatario:", label_style), Paragraph(nome, body_style)],
        [Paragraph("Indirizzo:", label_style), Paragraph(ind, body_style)],
        [Paragraph("Città:", label_style), Paragraph(f"{cit} ({prov})", body_style)],
        [Paragraph("Data Consegna:", label_style), Paragraph(data, body_style)],
        [Paragraph("Fascia Oraria:", label_style), Paragraph(f"Da {om or '—'} A {oM or '14:00'}", body_style)],
        [Paragraph("Note Consegna:", label_style), Paragraph(note or "Nessuna nota", body_style)]
    ]
    
    t = Table(data_table, colWidths=[120, 380])
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f8fafc')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 40))
    
    elements.append(Paragraph("<b>FIRMA PER RICEVUTA</b>", label_style))
    elements.append(Spacer(1, 15))
    sig_table = [
        [Paragraph("Data: ____________________", body_style), Paragraph("Firma Leggibile: ___________________________", body_style)]
    ]
    t_sig = Table(sig_table, colWidths=[200, 300])
    t_sig.setStyle(TableStyle([
        ('PADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(t_sig)
    
    doc.build(elements)

def _estrai_grand_chef(base_dir: Path, date_valide: set[str], mappati: dict) -> dict:
    import pandas as pd
    grand_chef_dir = base_dir.parent.parent / "Grand Chef"
    out_gc_dir = base_dir / "DDT-ORIGINALI-DIVISI" / "FRUTTA"
    out_gc_dir.mkdir(parents=True, exist_ok=True)
    
    nuovi_dati = {}
    if not grand_chef_dir.exists():
        print(f"  ⚠️  Cartella Grand Chef non trovata in: {grand_chef_dir}")
        return nuovi_dati
        
    files = list(grand_chef_dir.glob("*.xlsx"))
    if not files:
        print("  ⚠️  Nessun file Excel trovato in Grand Chef.")
        return nuovi_dati
        
    print(f"  📊 Elaborazione {len(files)} file Excel Grand Chef in Grand Chef...")
    
    data_label = base_dir.name.replace("CONSEGNE_", "")
    parti_data = re.findall(r"\d{2}-\d{2}-\d{4}", data_label)
    data_consegna = parti_data[0] if parti_data else data_label
    
    for f in files:
        try:
            df = pd.read_excel(f, sheet_name=0)
            df_clean = df.dropna(how='all')
            
            header_row_idx = None
            for idx, row in df_clean.iterrows():
                row_vals = [str(val).strip().lower() for val in row.values if pd.notna(val)]
                if any('ragione sociale' in rv for rv in row_vals) or any('codice' in rv for rv in row_vals):
                    header_row_idx = idx
                    break
                    
            if header_row_idx is None:
                continue
                
            df_data = df_clean.loc[header_row_idx + 1:]
            for _, row in df_data.iterrows():
                if str(row.iloc[0]).lower().strip() == 'totale':
                    continue
                    
                codice = clean_client_code(row.iloc[0])
                if not codice:
                    continue
                    
                ragione_sociale = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
                indirizzo = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""
                localita = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ""
                provincia = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ""
                note = str(row.iloc[14]).strip() if len(row) > 14 and pd.notna(row.iloc[14]) else ""
                fascia = str(row.iloc[15]).strip() if len(row) > 15 and pd.notna(row.iloc[15]) else ""
                
                orario_min, orario_max = parse_fascia_oraria(fascia)
                if not orario_min and not orario_max and note:
                    orario_min, orario_max = parse_fascia_oraria(note)
                    
                if not orario_max:
                    orario_max = "14:00"
                
                if codice not in mappati:
                    nuovi_dati[codice] = {
                        "dest": ragione_sociale,
                        "ind": indirizzo,
                        "cap": "",
                        "cit": localita,
                        "prov": provincia,
                        "om": orario_min,
                        "oM": orario_max,
                        "tipo": "GRAND CHEF"
                    }
                else:
                    fname = f"{codice}_{data_consegna}.pdf"
                    pdf_path = out_gc_dir / fname
                    if not pdf_path.exists():
                        print(f"    📄 Generazione PDF Grand Chef per {codice}: {ragione_sociale[:35]}")
                        _genera_pdf_placeholder_grand_chef(
                            pdf_path, codice, ragione_sociale, indirizzo, 
                            localita, provincia, note, orario_min, orario_max, data_consegna
                        )
                        
        except Exception as e:
            print(f"  ⚠️  Errore lettura file Grand Chef {f.name}: {e}")
            
    return nuovi_dati


def _ricava_date_da_grand_chef() -> list[str]:
    grand_chef_dir = BASE_DIR / "Grand Chef"
    date_found = set()
    if grand_chef_dir.exists():
        for f in grand_chef_dir.glob("*.xlsx"):
            m = re.search(r"(\d{4})-(\d{2})-(\d{2})", f.name)
            if m:
                date_found.add(f"{m.group(3)}-{m.group(2)}-{m.group(1)}")
    return sorted(list(date_found))


def main():
    # ── Determina le date valide ────────────────────────────────────────────
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else None
    if arg:
        # Supporta: "27-03", "27-03-2026", "27-03-2026_28-03-2026"
        date_valide: set[str] = set()
        for parte in arg.split("_"):
            if re.match(r"^\d{2}-\d{2}$", parte): parte = f"{parte}-2026"
            date_valide.add(parte)
    else:
        date_list = _ricava_date_da_pdf()
        if not date_list:
            date_list = _ricava_date_da_grand_chef()
        if not date_list:
            return print("❌ Nessun PDF o file Grand Chef trovato.")
        date_valide = set(date_list)

    # ── Nome cartella: singola o doppia data ────────────────────────────────
    data_label = "_".join(sorted(date_valide))   # es. "30-03-2026" o "30-03-2026_31-03-2026"
    base = CONSEGNE_DIR / f"CONSEGNE_{data_label}"
    if base.exists() and not arg:
        return print(f"⚠️ Cartella {base.name} esiste già. Passa la data come argomento per forzare.")

    _pulisci_output(base, data_label)
    nd = len(date_valide)
    print(f"\n--- Estrazione DDT ({data_label}) — {nd} data{'e' if nd > 1 else ''} ---\n")
    if nd > 1:
        print(f"  ℹ️  Date rilevate: {', '.join(sorted(date_valide))} → elaborazione accorpata\n")

    out_f = base / "DDT-ORIGINALI-DIVISI" / "FRUTTA"
    out_l = base / "DDT-ORIGINALI-DIVISI" / "LATTE"
    mappati = _leggi_codici_mappatura()
    res_f = _estrai_da_cartella(INPUT_FRUTTA, out_f, "FRUTTA", date_valide, mappati, duplicata=True)
    res_l = _estrai_da_cartella(INPUT_LATTE,  out_l, "LATTE",  date_valide, mappati)

    # Elabora Grand Chef
    res_gc = _estrai_grand_chef(base, date_valide, mappati)

    if not _verifica_nuovi_clienti({**res_f[2], **res_l[2], **res_gc}): sys.exit(1)
    if not _verifica_nuovi_articoli(base): sys.exit(1)

    print("Pulizia sorgenti e avvio pipeline...")
    _pulisci_sorgenti(INPUT_FRUTTA, date_valide)
    _pulisci_sorgenti(INPUT_LATTE,  date_valide)
    time.sleep(1)

    for s in ["2_crea_punti_consegna.py", "3_crea_lista_unificata.py"]:
        p = PROG_DIR / s
        if p.exists():
            print(f"⚙️ {s}...")
            subprocess.run([sys.executable, str(p), data_label], cwd=BASE_DIR)

    p_mappa = PROG_DIR / "4_mappa_zone_google.py"
    if p_mappa.exists():
        print(f"⚙️ 4_mappa_zone_google.py (generazione file, server disabilitato)...")
        subprocess.run([sys.executable, str(p_mappa), data_label, "--no-serve"], cwd=BASE_DIR)

    print(f"\n✅ COMPLETATO ({data_label})!")

if __name__ == "__main__": main()

