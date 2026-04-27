
import os
from pathlib import Path
from collections import defaultdict
import re

RIENTRI_XLSX = Path(r"g:\Il mio Drive\App\AppLogSolutionLocale\dati\rientri_ddt.xlsx")

def _carica_rientri(data_attuale: str = None) -> dict:
    if not RIENTRI_XLSX.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(RIENTRI_XLSX, read_only=True, data_only=True)
        ws = wb.active
        rientri = defaultdict(list)
        all_rientri_rows = [] 
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            codice_raw = row[0].value
            data_b = row[1].value
            if not codice_raw or not data_b:
                continue

            stato = str(row[2].value or "").strip().lower()
            codice = str(codice_raw).strip().lower()
            
            if hasattr(data_b, 'strftime'):
                data_str = data_b.strftime("%d-%m-%Y")
            else:
                data_str = str(data_b).strip()

            all_rientri_rows.append((r_idx, codice, data_str, stato))

            if "allegato" in stato and "lavorazione" not in stato:
                if not data_attuale or data_attuale.lower() not in stato:
                    continue

            if codice and data_str:
                rientri[codice].append(data_str)
        wb.close()
        return rientri, all_rientri_rows
    except Exception as e:
        print(f"  WARN  Errore lettura rientri_ddt.xlsx: {e}")
        return {}, []

rientri, rows = _carica_rientri("28-04-2026")
print(f"Rientri for 28-04-2026: {rientri}")
print(f"Total rows: {len(rows)}")
for r in rows[:20]:
    print(r)
