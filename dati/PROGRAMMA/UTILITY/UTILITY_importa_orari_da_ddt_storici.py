#!/usr/bin/env python3
"""
UTILITY_importa_orari_da_ddt_storici.py
════════════════════════════════════════
1. AZZERA le 4 colonne orario in mappatura_destinazioni.xlsx
2. Scansiona tutte le sorgenti DDT storiche:
   - DNR/Marzo 2026/CONSEGNE_*/DDT-ORIGINALI/       (singoli file per codice)
   - DNR/CONSEGNE_*/DDT-ORIGINALI-DIVISI/FRUTTA+LATTE/  (se presenti)
   - DNR/*.pdf                                       (file singoli nella root)
   - CONSEGNE_*/DDT-ORIGINALI-DIVISI/FRUTTA+LATTE/  (sistema attuale)
   - dati/*.pdf                                      (pdf non ancora elaborati)
3. Popola Orario min Frutta/max Frutta e Orario min Latte/max Latte.
   Il TIPO (FRUTTA/LATTE) è determinato dal codice in mappatura.
"""

import re
from pathlib import Path
from collections import defaultdict

PROG_DIR     = Path(__file__).resolve().parent
BASE_DIR     = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
DNR_DIR      = BASE_DIR.parent / "DNR"   # g:\Il mio Drive\App\DNR
MAPPATURA    = PROG_DIR / "mappatura_destinazioni.xlsx"

LUOGO_RE   = re.compile(r'(?:Luogo di destinazione|Codice destinazione):\s*([pP]\d{4,5})', re.I)
CAUSALE_RE = re.compile(r'(?:conto di|ordine e conto di)\s+([A-Z]\d{4,5})(?:\s+H(\d{2}))?(?:\s+(\d{3,4}))?', re.I)


def _converti_orario(s: str) -> str:
    s = s.strip()
    if len(s) == 3:
        return f"{int(s[0]):02d}:{s[1:]}"
    elif len(s) == 4:
        return f"{s[:2]}:{s[2:]}"
    return ""


# ─── CARICA MAPPATURA ────────────────────────────────────────────────────────

def _carica_mappatura_completa() -> tuple[dict, dict]:
    """
    Restituisce:
      mappa_dati  : {row_idx: [om_f, oM_f, om_l, oM_l]}
      codice_tipo : {codice_p: (row_idx, 'FRUTTA'|'LATTE')}
    """
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_om_f = next((i for i, h in enumerate(headers) if h == "Orario min Frutta"), 10)
    col_oM_f = next((i for i, h in enumerate(headers) if h == "Orario max Frutta"), 11)
    col_om_l = next((i for i, h in enumerate(headers) if h == "Orario min Latte"),  12)
    col_oM_l = next((i for i, h in enumerate(headers) if h == "Orario max Latte"),  13)

    mappa_dati  = {}
    codice_tipo = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row]
        def _v(x): return str(x).strip() if x is not None else ""
        c_f = _v(vals[0]).lower()
        c_l = _v(vals[1]).lower() if len(vals) > 1 else ""
        om_f = _v(vals[col_om_f]) if col_om_f < len(vals) else ""
        oM_f = _v(vals[col_oM_f]) if col_oM_f < len(vals) else ""
        om_l = _v(vals[col_om_l]) if col_om_l < len(vals) else ""
        oM_l = _v(vals[col_oM_l]) if col_oM_l < len(vals) else ""
        mappa_dati[row_idx] = [om_f, oM_f, om_l, oM_l]
        if c_f and c_f != "p00000": codice_tipo[c_f] = (row_idx, "FRUTTA")
        if c_l and c_l != "p00000": codice_tipo[c_l] = (row_idx, "LATTE")
    wb.close()
    return mappa_dati, codice_tipo


# ─── AZZERAMENTO 4 COLONNE ───────────────────────────────────────────────────

def _azzera_colonne_orario():
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    cols_da_azzerare = [
        "Orario min Frutta", "Orario max Frutta",
        "Orario min Latte",  "Orario max Latte"
    ]
    col_indices = []
    for nome in cols_da_azzerare:
        idx = next((i + 1 for i, h in enumerate(headers) if h == nome), None)
        if idx: col_indices.append(idx)

    righe_azzerate = 0
    for row in ws.iter_rows(min_row=2):
        for col_idx in col_indices:
            ws.cell(row=row[0].row, column=col_idx, value=None)
        righe_azzerate += 1
    wb.save(MAPPATURA)
    print(f"  Azzerate {len(col_indices)} colonne x {righe_azzerate} righe.")


# ─── ESTRAZIONE ORARI DA UN PDF ───────────────────────────────────────────────

def _estrai_orari_da_pdf(pdf_path: Path) -> dict:
    """
    Restituisce {codice_p: (om, oM)} leggendo la CAUSALE DEL TRASPORTO.
    oM e' valorizzato solo se trovato il prefisso H nel DDT.
    """
    import pdfplumber
    risultati = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                m_luogo = LUOGO_RE.search(text)
                if not m_luogo:
                    continue
                codice = m_luogo.group(1).lower()

                om_ddt, oM_ddt = "", ""
                idx_c = text.upper().find("CAUSALE DEL TRASPORTO")
                if idx_c >= 0:
                    m_c = CAUSALE_RE.search(text[idx_c:idx_c + 200])
                    if m_c:
                        if m_c.group(2):            # H10, H08, ecc.
                            oM_ddt = f"{int(m_c.group(2)):02d}:00"
                        if m_c.group(3):            # 800, 745, ecc.
                            om_ddt = _converti_orario(m_c.group(3))

                if om_ddt or oM_ddt:
                    if codice not in risultati:
                        risultati[codice] = (om_ddt, oM_ddt)
    except Exception as e:
        print(f"  ERRORE {pdf_path.name}: {e}")
    return risultati


# ─── AGGIORNAMENTO EXCEL ──────────────────────────────────────────────────────

def _aggiorna_excel(nuovi_orari: dict):
    """
    nuovi_orari = {row_idx: [om_f, oM_f, om_l, oM_l]}
    Sovrascrive i valori nelle 4 colonne.
    """
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_map = {
        0: next((i + 1 for i, h in enumerate(headers) if h == "Orario min Frutta"), None),
        1: next((i + 1 for i, h in enumerate(headers) if h == "Orario max Frutta"), None),
        2: next((i + 1 for i, h in enumerate(headers) if h == "Orario min Latte"),  None),
        3: next((i + 1 for i, h in enumerate(headers) if h == "Orario max Latte"),  None),
    }
    scritti = 0
    for row_idx, valori in nuovi_orari.items():
        for pos, valore in enumerate(valori):
            col = col_map.get(pos)
            if col and valore:
                ws.cell(row=row_idx, column=col, value=valore)
                scritti += 1
    wb.save(MAPPATURA)
    return scritti


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*65)
    print("  UTILITY — Importa orari H10 da DDT storici (SOVRASCRITTURA)")
    print("="*65)

    if not MAPPATURA.exists():
        print("ERRORE: mappatura_destinazioni.xlsx non trovata.")
        return

    # 1. Carica mappatura
    mappa_dati, codice_tipo = _carica_mappatura_completa()
    print(f"  Mappatura caricata: {len(codice_tipo)} codici, {len(mappa_dati)} righe")

    # 2. Azzera le 4 colonne
    print("\n  [1/3] Azzeramento colonne orario...")
    _azzera_colonne_orario()

    # 3. Definisci tutte le sorgenti PDF da scansionare
    sorgenti = []  # (pdf_path, tipo_forzato)
    #    tipo_forzato = "FRUTTA"/"LATTE" se noto, None se da determinare da mappatura

    # CONSEGNE correnti (DDT-ORIGINALI-DIVISI/FRUTTA e LATTE)
    for cart in CONSEGNE_DIR.iterdir():
        if not cart.is_dir() or not cart.name.startswith("CONSEGNE_"): continue
        for tipo, subdir in [("FRUTTA", "DDT-ORIGINALI-DIVISI/FRUTTA"),
                             ("LATTE",  "DDT-ORIGINALI-DIVISI/LATTE")]:
            d = cart / subdir
            if d.exists():
                for p in d.glob("*.pdf"):
                    sorgenti.append((p, tipo))

    # DNR — struttura con DDT-ORIGINALI (non separati)
    if DNR_DIR.exists():
        for p in DNR_DIR.rglob("*.pdf"):
            if "DDT-ORIGINALI" in str(p) and "DIVISI" not in str(p):
                sorgenti.append((p, None))    # tipo determinato da mappatura
            elif "DDT-ORIGINALI-DIVISI" in str(p):
                tipo = "FRUTTA" if "FRUTTA" in str(p).upper() else "LATTE"
                sorgenti.append((p, tipo))

        # File singoli nella root DNR (pXXXXX_data.pdf)
        for p in DNR_DIR.glob("*.pdf"):
            sorgenti.append((p, None))

    # PDF non ancora elaborati nella cartella dati principale
    for p in BASE_DIR.glob("*.pdf"):
        tipo = None
        if "frutta" in p.name.lower(): tipo = "FRUTTA"
        elif "latte" in p.name.lower(): tipo = "LATTE"
        sorgenti.append((p, tipo))

    print(f"\n  [2/3] Scansione di {len(sorgenti)} file PDF...")

    # Accumula: {row_idx: [om_f, oM_f, om_l, oM_l]}
    nuovi_orari = defaultdict(lambda: ["", "", "", ""])
    rilevamenti = 0

    for pdf_path, tipo_forzato in sorgenti:
        orari = _estrai_orari_da_pdf(pdf_path)
        for codice, (om, oM) in orari.items():
            if codice not in codice_tipo:
                continue
            row_idx, tipo_mappa = codice_tipo[codice]
            tipo = tipo_forzato or tipo_mappa

            valori = nuovi_orari[row_idx]
            if tipo == "FRUTTA":
                if om:  valori[0] = om
                if oM:  valori[1] = oM
            else:  # LATTE
                if om:  valori[2] = om
                if oM:  valori[3] = oM
            rilevamenti += 1

    # 4. Scrivi su Excel
    print(f"\n  [3/3] Scrittura {len(nuovi_orari)} righe su mappatura...")
    try:
        scritti = _aggiorna_excel(nuovi_orari)
        print(f"  Celle scritte: {scritti}")
    except PermissionError:
        print("  ERRORE: mappatura_destinazioni.xlsx e' aperto. Chiudi e riprova.")
        return

    # 5. Report finale
    print(f"\n{'='*65}")
    print(f"  REPORT — {rilevamenti} rilevamenti, {len(nuovi_orari)} clienti aggiornati")
    print(f"{'='*65}")
    h10_frutta = sum(1 for v in nuovi_orari.values() if v[1] == "10:00")
    h10_latte  = sum(1 for v in nuovi_orari.values() if v[3] == "10:00")
    print(f"  Clienti H10 Frutta (Orario max = 10:00): {h10_frutta}")
    print(f"  Clienti H10 Latte  (Orario max = 10:00): {h10_latte}")
    print(f"  Righe mappatura aggiornate: {len(nuovi_orari)}")
    print("="*65 + "\n")


if __name__ == "__main__":
    main()
