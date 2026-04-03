import os
import glob
import pandas as pd
import openpyxl

DRIVE_PATH = r"G:\Il mio Drive\Fatturazione"
INPUT_DIR = os.path.join(DRIVE_PATH, "Riepiloghi_Giornalieri")
OUTPUT_FILE = os.path.join(DRIVE_PATH, "Anagrafica_Clienti_Master.xlsx")

def estrai_anagrafica():
    print(f"Scansione file in {INPUT_DIR}...")
    file_list = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    
    anagrafica = {} # Dizionario per evitare duplicati, chiave: codice
    
    for file_path in file_list:
        file_name = os.path.basename(file_path)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if "Riepilogo_Giornaliero" in wb.sheetnames:
                ws = wb["Riepilogo_Giornaliero"]
            else:
                ws = wb.active
                
            col_map = {}
            for row in ws.iter_rows(values_only=True):
                # Cerchiamo la riga delle intestazioni
                row_strs = [str(c).lower().strip() for c in row if c]
                if 'codice' in row_strs and 'ragione sociale' in row_strs:
                    col_map = {}
                    for i, val in enumerate(row):
                        if val and isinstance(val, str):
                            col_map[val.lower().strip()] = i
                    continue
                
                # Se abbiamo trovato le intestazioni, leggiamo i dati
                if col_map:
                    if not any(row):  # fine della tabella del viaggio
                        col_map = {}
                        continue
                        
                    idx_cod = col_map.get('codice')
                    idx_rs = col_map.get('ragione sociale')
                    idx_ind = col_map.get('indirizzo')
                    idx_loc = col_map.get('località', col_map.get('localita'))
                    idx_pr = col_map.get('pr.', col_map.get('provincia'))
                    
                    if idx_cod is not None and idx_cod < len(row) and row[idx_cod]:
                        codice = str(row[idx_cod]).strip().upper()
                        # Salta righe di totale, spurie, o numerazioni strane
                        if not codice or codice == "NONE" or "TOTALE" in codice or codice == "CODICE":
                            continue
                            
                        # Estrazione valori
                        rs = str(row[idx_rs]).strip() if idx_rs is not None and idx_rs < len(row) and row[idx_rs] else ""
                        ind = str(row[idx_ind]).strip() if idx_ind is not None and idx_ind < len(row) and row[idx_ind] else ""
                        loc = str(row[idx_loc]).strip() if idx_loc is not None and idx_loc < len(row) and row[idx_loc] else ""
                        pr = str(row[idx_pr]).strip() if idx_pr is not None and idx_pr < len(row) and row[idx_pr] else ""
                        
                        # Aggiorniamo o creiamo l'anagrafica (il codice cliente assicura niente doppioni)
                        if codice not in anagrafica:
                            anagrafica[codice] = {
                                "Codice Cliente": codice,
                                "Ragione Sociale": rs.upper(),
                                "Indirizzo": ind.upper(),
                                "Località": loc.upper(),
                                "Provincia": pr.upper()
                            }
        except Exception as e:
            print(f"Errore nella lettura di {file_name}: {e}")
            
    # Convertiamo in DataFrame e salviamo
    df_output = pd.DataFrame(list(anagrafica.values()))
    
    # Ordiniamo per Ragione Sociale
    if not df_output.empty:
        df_output = df_output.sort_values(by="Ragione Sociale")
        df_output.to_excel(OUTPUT_FILE, index=False)
        print(f"\n✅ Anagrafica completata! Estratti {len(anagrafica)} CLIENTI UNICI.")
        print(f"File salvato in: {OUTPUT_FILE}")
    else:
        print("Nessun dato anagrafico trovato.")

if __name__ == "__main__":
    estrai_anagrafica()
