#!/usr/bin/env python3
"""
UTILITY_patch_json_orari.py
═══════════════════════════
Aggiorna orario_min e orario_max nei file viaggi_giornalieri*.json
senza ri-eseguire la pipeline completa.

Legge Orario min/max Frutta e Latte dalla mappatura aggiornata
e calcola l'orario piu' restrittivo per ogni punto di consegna.
"""

import json
from pathlib import Path

PROG_DIR     = Path(__file__).resolve().parent
BASE_DIR     = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"
MAPPATURA    = PROG_DIR / "mappatura_destinazioni.xlsx"


def _carica_orari_mappatura() -> dict:
    """Restituisce {codice_p: (om_f, oM_f, om_l, oM_l)}."""
    from openpyxl import load_workbook
    wb = load_workbook(MAPPATURA, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_om_f = next((i for i, h in enumerate(headers) if h == "Orario min Frutta"), 10)
    col_oM_f = next((i for i, h in enumerate(headers) if h == "Orario max Frutta"), 11)
    col_om_l = next((i for i, h in enumerate(headers) if h == "Orario min Latte"),  12)
    col_oM_l = next((i for i, h in enumerate(headers) if h == "Orario max Latte"),  13)
    orari = {}
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        def _v(x): return str(x).strip() if x is not None else ""
        c_f = _v(vals[0]).lower()
        c_l = _v(vals[1]).lower() if len(vals) > 1 else ""
        om_f = _v(vals[col_om_f]) if col_om_f < len(vals) else ""
        oM_f = _v(vals[col_oM_f]) if col_oM_f < len(vals) else ""
        om_l = _v(vals[col_om_l]) if col_om_l < len(vals) else ""
        oM_l = _v(vals[col_oM_l]) if col_oM_l < len(vals) else ""
        entry = (om_f, oM_f, om_l, oM_l)
        if c_f and c_f != "p00000": orari[c_f] = entry
        if c_l and c_l != "p00000": orari[c_l] = entry
    wb.close()
    return orari


def _orario_piu_restrittivo(a: str, b: str) -> str:
    """Restituisce il piu' piccolo tra due orari HH:MM non vuoti."""
    if a and b: return a if a < b else b
    return a or b


def _patch_punto(punto: dict, orari_mappa: dict) -> bool:
    """Aggiorna orario_min e orario_max nel punto. Ritorna True se modificato."""
    cf = str(punto.get("codice_frutta", "") or "").lower()
    cl = str(punto.get("codice_latte",  "") or "").lower()

    om_f, oM_f, om_l, oM_l = "", "", "", ""
    if cf and cf != "p00000" and cf in orari_mappa:
        om_f, oM_f, _, _ = orari_mappa[cf]
    if cl and cl != "p00000" and cl in orari_mappa:
        _, _, om_l, oM_l = orari_mappa[cl]

    # Orario generale = piu' restrittivo tra frutta e latte
    om_new = _orario_piu_restrittivo(om_f, om_l)
    oM_new = _orario_piu_restrittivo(oM_f, oM_l)

    changed = False
    if om_new and om_new != punto.get("orario_min", ""):
        punto["orario_min"] = om_new
        changed = True
    if oM_new and oM_new != punto.get("orario_max", ""):
        punto["orario_max"] = oM_new
        changed = True

    # Aggiunge anche i 4 campi specifici per uso futuro
    punto["orario_min_frutta"] = om_f
    punto["orario_max_frutta"] = oM_f
    punto["orario_min_latte"]  = om_l
    punto["orario_max_latte"]  = oM_l

    return changed


def _patch_json(json_path: Path, orari_mappa: dict) -> tuple[int, int]:
    """Patcha un file JSON. Ritorna (totale_punti, punti_modificati)."""
    data = json.loads(json_path.read_text(encoding="utf-8"))

    # Supporta sia lista di viaggi che dict con key 'viaggi'
    if isinstance(data, list):
        viaggi = data
    elif isinstance(data, dict):
        viaggi = data.get("viaggi", [])
    else:
        return 0, 0

    totale, modificati = 0, 0
    for viaggio in viaggi:
        punti_key = "lista_punti" if "lista_punti" in viaggio else "punti"
        for punto in viaggio.get(punti_key, []):
            totale += 1
            if _patch_punto(punto, orari_mappa):
                modificati += 1

    json_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return totale, modificati


def main():
    print("\n" + "="*60)
    print("  UTILITY — Patch JSON orari da mappatura aggiornata")
    print("="*60)

    orari_mappa = _carica_orari_mappatura()
    print(f"  Mappatura caricata: {len(orari_mappa)} codici con orari\n")

    # Trova tutte le cartelle CONSEGNE con JSON da patchare
    cartelle = sorted([
        d for d in CONSEGNE_DIR.iterdir()
        if d.is_dir() and d.name.startswith("CONSEGNE_")
    ])

    totale_file = 0
    for cart in cartelle:
        json_files = list(cart.glob("viaggi_giornalieri*.json"))
        if not json_files:
            continue
        print(f"  {cart.name}:")
        for jp in json_files:
            tot, mod = _patch_json(jp, orari_mappa)
            print(f"    {jp.name:<45} {tot} punti, {mod} aggiornati")
            totale_file += 1

    print(f"\n  File JSON patchati: {totale_file}")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
