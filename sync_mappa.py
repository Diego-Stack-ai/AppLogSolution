import json
from openpyxl import load_workbook

f2 = 'G:/Il mio Drive/App/AppLogSolution/dati/CONSEGNE/CONSEGNE_20-04-2026_copia/punti_consegna_unificati.json'
with open(f2, encoding='utf-8') as f: j2 = json.load(f)
punti_veri = {p['nome'].strip().lower(): p for p in j2.get('punti', []) if p.get('lat')}

map_file = 'G:/Il mio Drive/App/AppLogSolution/dati/PROGRAMMA/mappatura_destinazioni.xlsx'
wb = load_workbook(map_file)
ws = wb.active
headers = [str(c.value).strip().lower() for c in ws[1]]
col_nome = next(i for i, h in enumerate(headers) if h in ['a chi va consegnato', 'nome'])
col_lat = next(i for i, h in enumerate(headers) if h == 'latitudine')
col_lon = next(i for i, h in enumerate(headers) if h == 'longitudine')

saved = 0
for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    n = str(row[col_nome].value).strip().lower()
    if n in punti_veri:
        vero = punti_veri[n]
        if row[col_lat].value != vero['lat'] or row[col_lon].value != vero['lon']:
            ws.cell(row=r_idx, column=col_lat+1, value=vero['lat'])
            ws.cell(row=r_idx, column=col_lon+1, value=vero['lon'])
            print(f'Aggiornato {n}: {vero["lat"]}, {vero["lon"]}')
            saved += 1

wb.save(map_file)
print(f'\nAggiornati {saved} punti in mappatura da unificati.json di COPIA!')
