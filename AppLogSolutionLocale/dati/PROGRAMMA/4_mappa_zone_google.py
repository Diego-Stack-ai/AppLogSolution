import json, sys, re, webbrowser, threading, time, logging
import html as html_module
import xml.etree.ElementTree as ET
from pathlib import Path

# Flask e Flask-CORS verranno importati solo se serve avviare il server
try:
    from flask import Flask, render_template_string, request, jsonify
    from flask_cors import CORS
    HAS_FLASK = True
except ImportError:
    HAS_FLASK = False
    Flask = lambda *args, **kwargs: None
    render_template_string = request = jsonify = None

# --- CONFIGURAZIONE ---
PROG_DIR = Path(__file__).resolve().parent
BASE_DIR = PROG_DIR.parent
CONSEGNE_DIR = BASE_DIR / "CONSEGNE"

# INSERISCI QUI LA TUA API KEY DI GOOGLE MAPS
import os as _os
def _carica_env():
    _p = Path(__file__).resolve().parent / '.env'
    if _p.exists():
        for _l in _p.read_text(encoding='utf-8').splitlines():
            _l = _l.strip()
            if _l and not _l.startswith('#') and '=' in _l:
                _k, _v = _l.split('=', 1)
                _os.environ.setdefault(_k.strip(), _v.strip())
_carica_env()
GOOGLE_MAPS_API_KEY = _os.environ.get('GOOGLE_MAPS_API_KEY', '')
if HAS_FLASK:
    app = Flask(__name__)
    CORS(app, resources={r"/*": {"origins": "*"}})
else:
    app = None

# Configurazione Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Variabili globali per tracciare i file in uso
TARGET_FILE_UNIFICATO = None
TARGET_FILE_VIAGGI = None
ZONE_LIST_CACHE = []
DATA_GIORNO = ""

def _get_color(idx):
    palette = ["#4f46e5", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6", "#ec4899", "#06b6d4", "#f97316", "#14b8a6", "#6366f1", "#a855f7", "#3b82f6", "#22c55e", "#d946ef", "#84cc16"]
    return palette[idx % len(palette)]

excel_lock = threading.Lock()

def _aggiorna_entrambi_excel(cod_f, cod_l, lat, lon, nome=None):
    """Sincronizzazione Atomica su Excel Master e Excel Giornaliero tramite Codice Frutta e Codice Latte."""
    try:
        from openpyxl import load_workbook
        output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
        targets = [PROG_DIR / "mappatura_destinazioni.xlsx", output_base / "punti_consegna.xlsx"]
        
        success = True
        for file_path in targets:
            if not file_path.exists(): continue
            with excel_lock:
                try:
                    wb = load_workbook(file_path)
                    ws = wb.active
                    headers = [str(c.value or "").strip().lower() for c in ws[1]]
                    
                    col_f_idx = -1
                    col_l_idx = -1
                    for i, h in enumerate(headers):
                        if "frutta" in h or "cod. fr" in h: col_f_idx = i
                        elif "latte" in h or "cod. la" in h: col_l_idx = i
                    if col_f_idx == -1 and col_l_idx == -1:
                        col_f_idx, col_l_idx = 0, 1
                    
                    # Cerca o crea colonne lat e lon
                    col_lat_idx = -1
                    col_lon_idx = -1
                    for i, h in enumerate(headers):
                        if h == "latitudine": col_lat_idx = i
                        elif h == "longitudine": col_lon_idx = i
                        
                    if col_lat_idx == -1:
                        col_lat_idx = len(headers)
                        ws.cell(row=1, column=col_lat_idx+1, value="Latitudine")
                        headers.append("latitudine")
                    if col_lon_idx == -1:
                        col_lon_idx = len(headers)
                        ws.cell(row=1, column=col_lon_idx+1, value="Longitudine")
                        headers.append("longitudine")

                    c_f = str(cod_f).strip().lower() if cod_f and str(cod_f)!="p00000" else ""
                    c_l = str(cod_l).strip().lower() if cod_l and str(cod_l)!="p00000" else ""
                    if not c_f and not c_l: return True
                    trovato = False

                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        rf = str(row[col_f_idx].value or "").strip().lower() if col_f_idx>=0 else ""
                        rl = str(row[col_l_idx].value or "").strip().lower() if col_l_idx>=0 else ""
                        match_f = (c_f and rf == c_f)
                        match_l = (c_l and rl == c_l)
                        
                        if match_f or match_l:
                            ws.cell(row=row_idx, column=col_lat_idx+1, value=lat)
                            ws.cell(row=row_idx, column=col_lon_idx+1, value=lon)
                            trovato = True
                    
                    if trovato:
                        wb.save(file_path)
                        logger.info(f"Aggiornato Excel ({file_path.name}) per {nome}")
                    else:
                        logger.warning(f"Punto non trovato in {file_path.name}: {nome}")
                    wb.close()
                except Exception as e:
                    logger.exception(f"Errore scrittura {file_path.name}")
                    success = False
        return success
    except Exception as e:
        logger.exception("Errore globale Excel update")
        return False

if HAS_FLASK:
    @app.after_request
    def add_header(r):
        r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        r.headers["Pragma"] = "no-cache"
        r.headers["Expires"] = "0"
        return r

    @app.route('/')
    def index():
        is_locked_val = True
        if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
            try:
                unif = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                is_locked_val = unif.get("is_locked", True)
            except: pass
        return render_template_string(HTML_TEMPLATE, DATA_GIORNO=DATA_GIORNO, JSON_ZONE=json.dumps(ZONE_LIST_CACHE, ensure_ascii=False), GOOGLE_MAPS_API_KEY=GOOGLE_MAPS_API_KEY, IS_LOCKED_JS="true" if is_locked_val else "false")

    @app.route('/save', methods=['POST'])
    def save():
        """Salvataggio globale: Zone + Coordinate su tutti i file + aggiornamento rientri DDT."""
        global ZONE_LIST_CACHE
        try:
            data = request.json
            ZONE_LIST_CACHE = data
            output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"

            # 1. Aggiorna Excel (Bulk)
            from openpyxl import load_workbook
            for ex_p in [PROG_DIR / "mappatura_destinazioni.xlsx", output_base / "punti_consegna.xlsx"]:
                if ex_p.exists():
                    with excel_lock:
                        try:
                            wb = load_workbook(ex_p)
                            ws = wb.active
                            headers = [str(c.value or "").strip().lower() for c in ws[1]]
                            
                            col_f_idx = -1
                            col_l_idx = -1
                            for i, h in enumerate(headers):
                                if "frutta" in h or "cod. fr" in h:
                                    col_f_idx = i
                                elif "latte" in h or "cod. la" in h:
                                    col_l_idx = i
                                    
                            if col_f_idx == -1 and col_l_idx == -1:
                                col_f_idx, col_l_idx = 0, 1
                                    
                            col_lat_idx, col_lon_idx = -1, -1
                            for i, h in enumerate(headers):
                                if h == "latitudine": col_lat_idx = i
                                elif h == "longitudine": col_lon_idx = i
                                
                            if col_lat_idx == -1:
                                col_lat_idx = len(headers)
                                ws.cell(row=1, column=col_lat_idx+1, value="Latitudine")
                                headers.append("latitudine")
                            if col_lon_idx == -1:
                                col_lon_idx = len(headers)
                                ws.cell(row=1, column=col_lon_idx+1, value="Longitudine")
                                headers.append("longitudine")
                            
                            mappa_punti = {}
                            for z in data:
                                for p in z.get("lista_punti", []):
                                    c_f = str(p.get('codice_frutta', '')).strip().lower()
                                    c_l = str(p.get('codice_latte', '')).strip().lower()
                                    if c_f == "p00000": c_f = ""
                                    if c_l == "p00000": c_l = ""
                                    if c_f or c_l:
                                        mappa_punti[(c_f, c_l)] = (p['lat'], p['lon'])
                                    
                            changed = False
                            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                                # read file values
                                rf = str(row[col_f_idx].value or "").strip().lower() if col_f_idx>=0 else ""
                                rl = str(row[col_l_idx].value or "").strip().lower() if col_l_idx>=0 else ""
                                
                                for (cf, cl), (lt, ln) in mappa_punti.items():
                                    if (cf and cf == rf) or (cl and cl == rl):
                                        ws.cell(row=row_idx, column=col_lat_idx+1, value=lt)
                                        ws.cell(row=row_idx, column=col_lon_idx+1, value=ln)
                                        changed = True
                                        break
                                    
                            if changed:
                                wb.save(ex_p)
                            wb.close()
                        except Exception as e:
                            logger.exception(f"Errore scrittura bulk su {ex_p.name}")

            # 2. Aggiorna JSON Unificato
            if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
                unif = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                p_map = { (p.get("codice_univoco") or f"{p.get('codice_frutta')}_{p.get('codice_latte')}"): (z["id_zona"], p['lat'], p['lon']) for z in data for p in z["lista_punti"] }
                for p in unif.get("punti", []):
                    pid = p.get("codice_univoco") or f"{p.get('codice_frutta')}_{p.get('codice_latte')}"
                    if pid in p_map:
                        p["zona"], lat, lon = p_map[pid]
                        if lat: p["lat"] = lat
                        if lon: p["lon"] = lon
                TARGET_FILE_UNIFICATO.write_text(json.dumps(unif, indent=2, ensure_ascii=False), encoding="utf-8")

            # 3. Aggiorna JSON Viaggi
            data_f = [z for z in data if len(z.get("lista_punti", [])) > 0]
            if TARGET_FILE_VIAGGI:
                TARGET_FILE_VIAGGI.write_text(json.dumps(data_f, indent=2, ensure_ascii=False), encoding="utf-8")

            # 4. Aggiorna rientri_ddt.xlsx in base alla zona finale dei punti speciali
            _aggiorna_rientri_dopo_salvataggio(data, DATA_GIORNO)

            return jsonify({"status": "ok"})
        except Exception as e:
            logger.exception("Errore save")
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route('/save_coord', methods=['POST'])
    def save_coord():
        """Salvataggio atomico al movimento di un singolo punto."""
        try:
            p = request.json
            cod_f = p.get('cod_f', '')
            cod_l = p.get('cod_l', '')
            nome = p.get('nome', '')
            lat = p['lat']
            lon = p['lon']
            cod_univoco = p.get('codice_univoco', '')

            # IGNORA p00000 per il salvataggio Excel basato sui vecchi codici
            if cod_f.strip().lower() == "p00000": cod_f = ""
            if cod_l.strip().lower() == "p00000": cod_l = ""

            res = _aggiorna_entrambi_excel(cod_f, cod_l, lat, lon, nome=nome)
            if res is True:
                global ZONE_LIST_CACHE
                for z in ZONE_LIST_CACHE:
                    for pt in z.get("lista_punti", []):
                        pt_univoco = pt.get("codice_univoco") or f"{pt.get('codice_frutta', '')}_{pt.get('codice_latte', '')}"
                        if cod_univoco and pt_univoco == cod_univoco:
                            pt['lat'], pt['lon'] = lat, lon

                # Aggiorna JSON Unificato e Viaggi al volo tramite codice univoco
                for f_path, is_unif in [(TARGET_FILE_UNIFICATO, True), (TARGET_FILE_VIAGGI, False)]:
                    if f_path and f_path.exists():
                        d = json.loads(f_path.read_text(encoding='utf-8'))
                        lista = d.get('punti', []) if is_unif else [x for zz in d for x in zz.get('lista_punti', [])]
                        for pt in lista:
                            pt_univoco = pt.get("codice_univoco") or f"{pt.get('codice_frutta', '')}_{pt.get('codice_latte', '')}"
                            if cod_univoco and pt_univoco == cod_univoco:
                                pt['lat'], pt['lon'] = lat, lon
                        f_path.write_text(json.dumps(d, indent=2, ensure_ascii=False), encoding='utf-8')
                return jsonify({"status": "ok", "msg": "Sincronizzazione completata!"})
            return jsonify({"status": "error", "msg": str(res)}), 500
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500

    @app.route('/toggle_lock', methods=['POST'])
    def toggle_lock():
        try:
            req = request.json
            action = req.get('action')
            if TARGET_FILE_UNIFICATO and TARGET_FILE_UNIFICATO.exists():
                unificato = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
                is_locked = (action == 'lock')
                unificato['is_locked'] = is_locked
                TARGET_FILE_UNIFICATO.write_text(json.dumps(unificato, indent=2, ensure_ascii=False), encoding="utf-8")
                return jsonify({"status": "ok", "is_locked": is_locked})
            return jsonify({"status": "error", "message": "File non trovato"}), 404
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route('/preview_percorsi', methods=['GET'])
    def preview_percorsi():
        """Lancia BAT 3 completo (OR-Tools + Google Directions), poi legge il risultato
        dall'OTTIMIZZATO.json per mostrarlo nel drawer. Nessun side-effect sulla UI."""
        import subprocess as _sp
        try:
            if not TARGET_FILE_VIAGGI or not TARGET_FILE_VIAGGI.exists():
                return jsonify({"status": "error", "msg": "Salva prima i giri con SALVA TUTTO"}), 400

            script_bat3 = PROG_DIR / "6_genera_percorsi_veggiano.py"
            if not script_bat3.exists():
                return jsonify({"status": "error", "msg": "Script 6_genera_percorsi_veggiano.py non trovato"}), 500

            # Lancia BAT 3 completo (con OR-Tools)
            result = _sp.run(
                [sys.executable, str(script_bat3)],
                cwd=str(BASE_DIR),
                capture_output=True, text=True, timeout=360
            )
            if result.returncode != 0:
                logger.error(f"BAT3 stderr: {result.stderr[:800]}")
                return jsonify({"status": "error", "msg": result.stderr[-400:]}), 500

            # Legge OTTIMIZZATO.json per restituire la sequenza ottimizzata con timing
            output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
            ottimizzato_path = output_base / "viaggi_giornalieri_OTTIMIZZATO.json"
            riepilogo_path   = output_base / "PERCORSI_VEGGIANO" / "RIEPILOGO_GIRI.html"

            giri = []
            if ottimizzato_path.exists():
                try:
                    ottimizzato = json.loads(ottimizzato_path.read_text(encoding="utf-8"))
                    for z in ottimizzato:
                        punti = z.get("lista_punti", [])
                        if not punti: continue
                        is_gc = any(
                            "GRAND" in str(p.get("tipologia_grado") or "").upper() or
                            "CHEF"  in str(p.get("tipologia_grado") or "").upper() or
                            "GRANCHEF" in str(p.get("zona") or "").upper()
                            for p in punti
                        )
                        fermate = [{
                            "codice_univoco": p.get("codice_univoco", ""),
                            "nome":          p.get("nome", ""),
                            "indirizzo":     p.get("indirizzo", ""),
                            "orario_min":    p.get("orario_min", ""),
                            "orario_max":    p.get("orario_max", ""),
                            "ora_arrivo":    p.get("ora_arrivo", ""),
                            "ora_ripartenza": p.get("ora_ripartenza", ""),
                            "is_late":       p.get("ritardo", False),
                            "lat": p.get("lat"), "lon": p.get("lon"),
                        } for p in punti]
                        n_late = sum(1 for f in fermate if f["is_late"])
                        giri.append({
                            "id_zona":   z.get("id_zona"),
                            "nome_giro": z.get("nome_giro") or z.get("id_zona"),
                            "is_gc":     is_gc,
                            "n_fermate": len(fermate),
                            "n_late":    n_late,
                            "fermate":   fermate,
                        })
                except Exception as e:
                    logger.warning(f"Errore lettura OTTIMIZZATO.json: {e}")

            return jsonify({
                "status": "ok",
                "giri": giri,
                "riepilogo_disponibile": riepilogo_path.exists(),
                "log": result.stdout[-300:]
            })
        except Exception as e:
            logger.exception("Errore preview_percorsi")
            return jsonify({"status": "error", "msg": str(e)}), 500

    @app.route('/serve_riepilogo', methods=['GET'])
    def serve_riepilogo():
        """Serve RIEPILOGO_GIRI.html con i link alle mappe singole riscritti per Flask."""
        from flask import Response
        output_base    = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
        riepilogo_path = output_base / "PERCORSI_VEGGIANO" / "RIEPILOGO_GIRI.html"
        if not riepilogo_path.exists():
            return Response("<h2>Riepilogo non ancora generato. Premi prima Anteprima Percorsi.</h2>", mimetype="text/html")
        html = riepilogo_path.read_text(encoding="utf-8")
        # Riscrivi i link relativi alle mappe singole → /percorsi/<filename>
        # I link nel riepilogo sono tipo: href="GranChef_V01_Zone_123.html"
        html = re.sub(r'href="([^"]+\.html)"', r'href="/percorsi/\1" target="_blank"', html)
        return Response(html, mimetype="text/html")

    @app.route('/percorsi/<path:filename>', methods=['GET'])
    def serve_percorso(filename):
        """Serve i file HTML delle singole mappe giro generate da BAT 3."""
        from flask import Response, abort
        output_base   = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
        percorsi_dir  = output_base / "PERCORSI_VEGGIANO"
        target        = percorsi_dir / filename
        # Sicurezza: rimane dentro la cartella PERCORSI_VEGGIANO
        try:
            target.resolve().relative_to(percorsi_dir.resolve())
        except ValueError:
            abort(403)
        if not target.exists() or not target.suffix == ".html":
            abort(404)
        return Response(target.read_text(encoding="utf-8"), mimetype="text/html")


    @app.route('/conferma_percorsi', methods=['POST'])
    def conferma_percorsi():
        """Gestisce due casi:
        - manuale=False: BAT 3 è già girato (anteprima), non serve rilancarlo.
        - manuale=True:  l'utente ha spostato fermate nel drawer → salva il nuovo
          ordine e rilancia BAT 3 con --usa-ordine-attuale.
        """
        import subprocess as _sp
        try:
            req = request.json
            ordini  = req.get("ordini", {})   # {id_zona: [codice_univoco, ...]}
            manuale = req.get("manuale", False)

            if manuale and ordini:
                # 1. Salva il nuovo ordine manuale nel JSON
                if TARGET_FILE_VIAGGI and TARGET_FILE_VIAGGI.exists():
                    viaggi = json.loads(TARGET_FILE_VIAGGI.read_text(encoding="utf-8"))
                    for z in viaggi:
                        zid = z.get("id_zona")
                        if zid not in ordini: continue
                        nuovo_ordine = ordini[zid]
                        mappa = {p.get("codice_univoco", ""): p for p in z["lista_punti"]}
                        z["lista_punti"] = [mappa[cu] for cu in nuovo_ordine if cu in mappa]
                    TARGET_FILE_VIAGGI.write_text(json.dumps(viaggi, indent=2, ensure_ascii=False), encoding="utf-8")
                    ZONE_LIST_CACHE.clear()
                    ZONE_LIST_CACHE.extend(viaggi)

                # 2. Rilancia BAT 3 con --usa-ordine-attuale
                script_bat3 = PROG_DIR / "6_genera_percorsi_veggiano.py"
                result = _sp.run(
                    [sys.executable, str(script_bat3), "--usa-ordine-attuale"],
                    cwd=str(BASE_DIR),
                    capture_output=True, text=True, timeout=300
                )
                if result.returncode != 0:
                    return jsonify({"status": "error", "msg": result.stderr[-300:]}), 500
                return jsonify({"status": "ok", "msg": "Ordine manuale salvato e mappe rigenerate!"})
            else:
                # BAT 3 è già girato con l'anteprima — nessuna azione necessaria
                return jsonify({"status": "ok", "msg": "Percorsi confermati (nessuna modifica manuale)"})

        except Exception as e:
            logger.exception("Errore conferma_percorsi")
            return jsonify({"status": "error", "msg": str(e)}), 500

def _aggiorna_rientri_dopo_salvataggio(zone_data: list, data_giorno: str):
    """Aggiorna colonna C di rientri_ddt.xlsx in base alla zona finale di OGNI punto."""
    rientri_path = BASE_DIR / "rientri_ddt.xlsx"
    if not rientri_path.exists(): return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(rientri_path)
        ws = wb.active
        
        # Mappatura Codice -> Stato finale desiderato
        status_map = {} # {codice_ddt: "in lavorazione..." o ""}
        
        for z in zone_data:
            zid = z.get("id_zona", "")
            for p in z.get("lista_punti", []):
                # Raccogliamo tutti i possibili codici DDT del punto (da liste o campi singoli)
                codici = []
                codici.extend(p.get("codici_ddt_frutta") or [])
                codici.extend(p.get("codici_ddt_latte") or [])
                if not codici:
                    # Fallback se le liste sono vuote
                    c_f, c_l = p.get("codice_frutta"), p.get("codice_latte")
                    if c_f and c_f != "p00000": codici.append(c_f)
                    if c_l and c_l != "p00000": codici.append(c_l)
                
                # Determiniamo lo stato: "in lavorazione" quando assegnato in mappa.
                # Lo step 9 (distinte) modifichera' "in lavorazione" in "allegato ...".
                final_status = "" if zid == "DDT_DA_INSERIRE" else f"in lavorazione"
                
                for c in codici:
                    status_map[str(c).strip().lower()] = final_status

        if not status_map: return
        
        modifiche = 0
        for row in ws.iter_rows(min_row=2):
            cod_excel = str(row[0].value or '').strip().lower()
            stato_attuale = str(row[2].value or '').strip().lower()
            
            if cod_excel in status_map:
                # SE LA RIGA E' GIA' IN STATO "ALLEGATO" DEFINITIVO (DI GIORNI PASSATI),
                # NON DOBBIAMO ASSOLUTAMENTE TOCCARLA NE' SOVRASCRIVERLA!
                if "allegato" in stato_attuale and "lavorazione" not in stato_attuale:
                    continue
                
                # Applica il nuovo stato solo su righe vuote o "in lavorazione"
                if row[2].value != status_map[cod_excel]:
                    row[2].value = status_map[cod_excel]
                    modifiche += 1
        
        if modifiche:
            wb.save(rientri_path)
            logger.info(f"📂 Excel Rientri aggiornato: {modifiche} righe modificate (saltate quelle gia' allegate).")
    except Exception as e:
        logger.exception("Errore aggiornamento rientri_ddt.xlsx")

def _libera_porta_5000():
    import subprocess as _sp
    try:
        res = _sp.run(["netstat", "-ano"], capture_output=True, text=True)
        for d in res.stdout.splitlines():
            if ":5000 " in d and "LISTENING" in d:
                pid = d.strip().split()[-1]
                if pid.isdigit():
                    _sp.run(["taskkill", "/PID", pid, "/F"], capture_output=True)
                    print(f"🔪 Porta 5000 liberata (PID {pid})")
        time.sleep(0.5)
    except: pass

def _salva_kml(punti, path, data):
    root = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
    doc = ET.SubElement(root, "Document")
    ET.SubElement(doc, "name").text = f"Zone Google {data}"
    for i, p in enumerate(punti, 1):
        if not p.get('lat') or not p.get('lon'): continue
        pm = ET.SubElement(doc, "Placemark")
        ET.SubElement(pm, "name").text = f"{i}. {p.get('nome')[:80]}"
        desc = [f"<b>{p.get('nome')}</b>", f"Indirizzo: {p.get('indirizzo', '')}", f"Zona: {p.get('zona', '')}"]
        ET.SubElement(pm, "description").text = "<br>".join(desc)
        pt = ET.SubElement(pm, "Point")
        ET.SubElement(pt, "coordinates").text = f"{p['lon']},{p['lat']},0"
    tree = ET.ElementTree(root)
    if hasattr(ET, "indent"): ET.indent(tree, space="  ")
    tree.write(path, encoding="utf-8", xml_declaration=True, method="xml")

def _salva_html_fisico(output_base, template_html, data, zone_json, api_key, is_locked):
    is_locked_js = 'true' if is_locked else 'false'
    html_content = template_html.replace("{{DATA_GIORNO}}", data).replace("{{JSON_ZONE | safe}}", zone_json).replace("{{GOOGLE_MAPS_API_KEY}}", api_key).replace("{{IS_LOCKED_JS}}", is_locked_js)
    file_path = output_base / "4_mappa_zone_google.html"
    file_path.write_text(html_content, encoding="utf-8")

# --- TEMPLATE HTML RIPRISTINATO INTEGRALMENTE ---
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="utf-8">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://fonts.googleapis.com/icon?family=Material+Icons+Round" rel="stylesheet">
    <script>
        (g=>{var h,a,k,p="The Google Maps JavaScript API",c="google",l="importLibrary",q="__ib__",m=document,b=window;b=b[c]||(b[c]={});var d=b.maps||(b.maps={}),r=new Set,e=new URLSearchParams,u=()=>h||(h=new Promise(async(f,n)=>{await (a=m.createElement("script"));e.set("libraries",[...r]+"");for(k in g)e.set(k.replace(/[A-Z]/g,t=>"_"+t[0].toLowerCase()),g[k]);e.set("callback",c+".maps."+q);a.src=`https://maps.${c}apis.com/maps/api/js?`+e;d[q]=f;a.onerror=()=>h=n(Error(p+" could not load."));a.nonce=m.querySelector("script[nonce]")?.nonce||"";m.head.append(a)}));d[l]?console.warn(p+" only loads once. See https://goo.gle/js-api-loading-troubleshooting"):d[l]=(f,...n)=>r.add(f)&&u().then(()=>d[l](f,...n))})({
            key: "{{GOOGLE_MAPS_API_KEY}}",
            v: "beta"
        });
    </script>
    <style>
        :root { 
            --primary: #6366f1; --primary-dark: #4f46e5; --bg: #0f172a; 
            --sidebar-bg: #ffffff; --text-main: #1e293b; --text-muted: #64748b;
            --accent: #10b981; --card-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1);
        }
        * { box-sizing: border-box; }
        body { margin: 0; font-family: 'Inter', sans-serif; height: 100vh; display: flex; background: var(--bg); overflow: hidden; }
        #sidebar { width: 400px; height: 100%; background: var(--sidebar-bg); border-right: 1px solid #e2e8f0; display: flex; flex-direction: column; z-index: 1000; box-shadow: 10px 0 30px rgba(0,0,0,0.05); }
        #header { padding: 25px; background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%); color: white; }
        #zone-list { flex: 1; overflow-y: auto; padding: 15px; background: #f8fafc; }
        .zone-card { background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 12px; margin-bottom: 12px; cursor: pointer; transition: 0.2s; box-shadow: var(--card-shadow); }
        .zone-card.selected { border: 2px solid var(--primary); background: #f5f3ff; }
        .zone-header { display: flex; align-items: center; gap: 10px; }
        .color-pill { width: 12px; height: 12px; border-radius: 4px; }
        .zone-title { font-weight: 700; flex: 1; font-size: 0.9rem; }
        #map { flex: 1; }
        .btn { padding: 6px 10px; font-size: 0.7rem; font-weight: 700; cursor: pointer; border-radius: 6px; border: 1px solid #ddd; background: white; }
        .btn-confirm { background: #22c55e; color: white; border: none; }
        .btn-cancel { background: #ef4444; color: white; border: none; }
        .btn-lock { background: #f1f5f9; min-width: 100px; }
        .btn-lock.locked { background: #ef4444; color: white; }
        #save-status { position: fixed; top: 20px; left: 50%; transform: translateX(-50%); padding: 10px 20px; background: #10b981; color: white; border-radius: 30px; font-weight: 700; display: none; z-index: 5000; box-shadow: 0 10px 20px rgba(0,0,0,0.2); }
        .custom-marker { position: absolute; width: 32px; height: 32px; display: flex; align-items: center; justify-content: center; color: white; font-weight: 800; font-size: 11px; box-shadow: 0 4px 10px rgba(0,0,0,0.3); border: 2px solid white; transition: 0.2s; transform: translate(-50%, -100%); }
        .point-card { background: #f8fafc; padding: 7px 9px; margin-bottom: 5px; border-radius: 8px; border: 1px solid #e2e8f0; }
        .point-num { display: inline-flex; align-items: center; justify-content: center; width: 20px; height: 20px; border-radius: 50%; color: white; font-size: 0.65rem; font-weight: 800; margin-right: 6px; flex-shrink: 0; vertical-align: middle; }
        .point-name { font-size: 0.75rem; font-weight: 700; color: #1e293b; }
        .point-addr { font-size: 0.65rem; color: #64748b; margin-top: 2px; padding-left: 26px; }
        .zone-chip { display:inline-block; padding:3px 9px; border-radius:20px; border:2px solid #cbd5e1; font-size:0.6rem; font-weight:700; cursor:pointer; margin:2px 2px 0 0; color:#475569; background:white; transition:0.15s; white-space:nowrap; }
        .zone-chip.sel { color:white; border-color:transparent; }
        .chk-box { display:none; }
        .chk-lbl { display:inline-flex; width:22px; height:22px; border-radius:5px; border:2px solid #cbd5e1; align-items:center; justify-content:center; cursor:pointer; background:white; font-size:13px; color:transparent; flex-shrink:0; transition:0.15s; }
        .chk-box:checked + .chk-lbl { background:#6366f1; border-color:#6366f1; color:white; }
        .m-tonda { border-radius: 50%; }
        .m-goccia { border-radius: 50% 50% 50% 0; transform: rotate(-45deg); }
        .m-goccia span { transform: rotate(45deg); }
        .m-quadrato { border-radius: 4px; }
        .m-triangolo { background: transparent !important; border-left: 18px solid transparent; border-right: 18px solid transparent; border-bottom: 36px solid #ccc; width: 0 !important; height: 0 !important; border-radius:0; }
        .m-triangolo span { position: absolute; top: 16px; left: -8px; width: 16px; }
        /* Zona speciale DDT da inserire */
        .zone-card.speciale { border: 2px dashed #f59e0b; background: #fffbeb; }
        .zone-card.speciale .zone-title span:first-child { color: #92400e; }
        .zone-card.speciale .zone-title span:last-child { color: #b45309; }
        .badge-speciale { display:inline-block; background:#f59e0b; color:white; font-size:0.55rem; font-weight:800; padding:1px 6px; border-radius:10px; margin-left:6px; vertical-align:middle; letter-spacing:0.03em; }
        
        /* Modal rinomina giro */
        #rename-modal-overlay {
            display:none; position:fixed; inset:0; background:rgba(15,23,42,0.6);
            z-index:9999; align-items:center; justify-content:center;
            backdrop-filter: blur(4px);
        }
        #rename-modal-overlay.open { display:flex; }
        #rename-modal {
            background:white; border-radius:20px; padding:28px 24px; width:340px;
            box-shadow:0 25px 60px rgba(0,0,0,0.3); animation: modalIn 0.2s ease;
        }
        @keyframes modalIn { from{opacity:0; transform:scale(0.92)} to{opacity:1; transform:scale(1)} }
        #rename-modal h3 { margin:0 0 6px; font-size:1rem; font-weight:800; color:#1e293b; }
        #rename-modal p  { margin:0 0 16px; font-size:0.72rem; color:#64748b; }
        #rename-select {
            width:100%; padding:10px 12px; border-radius:10px; border:2px solid #e2e8f0;
            font-size:0.82rem; font-family:'Inter',sans-serif; font-weight:600;
            color:#1e293b; background:#f8fafc; cursor:pointer; margin-bottom:12px;
            appearance:none; background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%236366f1' stroke-width='2' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");
            background-repeat:no-repeat; background-position:right 12px center;
            padding-right:32px; transition:border-color 0.2s;
        }
        #rename-select:focus { outline:none; border-color:#6366f1; }
        #rename-custom-row { margin-bottom:16px; }
        #rename-custom-row label { font-size:0.7rem; font-weight:700; color:#64748b; display:block; margin-bottom:4px; }
        #rename-custom {
            width:100%; padding:9px 12px; border-radius:10px; border:2px solid #e2e8f0;
            font-size:0.82rem; font-family:'Inter',sans-serif; color:#1e293b;
            background:#f8fafc; transition:border-color 0.2s;
        }
        #rename-custom:focus { outline:none; border-color:#6366f1; }
        #rename-modal-btns { display:flex; gap:8px; margin-top:4px; }
        #rename-modal-btns button { flex:1; padding:10px; border-radius:10px; border:none;
            font-size:0.8rem; font-weight:800; cursor:pointer; font-family:'Inter',sans-serif; transition:0.15s; }
        #rename-btn-ok { background:#6366f1; color:white; }
        #rename-btn-ok:hover { background:#4f46e5; }
        #rename-btn-cancel { background:#f1f5f9; color:#475569; }
        #rename-btn-cancel:hover { background:#e2e8f0; }
    </style>
</head>
<body>

<!-- Modal rinomina giro -->
<div id="rename-modal-overlay">
    <div id="rename-modal">
        <h3 id="rename-modal-title">✏️ Rinomina viaggio</h3>
        <p id="rename-modal-subtitle">Seleziona un nome dalla lista o scrivi uno personalizzato</p>
        <select id="rename-select">
            <option value="">— Seleziona nome —</option>
        </select>
        <div id="rename-custom-row">
            <label for="rename-custom">Oppure scrivi nome personalizzato:</label>
            <input id="rename-custom" type="text" placeholder="Es. LAGO BS 1 speciale...">
        </div>
        <div id="rename-modal-btns">
            <button id="rename-btn-cancel" onclick="chiudiRenameModal()">Annulla</button>
            <button id="rename-btn-ok" onclick="confermaRenameModal()">✓ Conferma</button>
        </div>
    </div>
</div>

    <div id="sidebar">
        <div id="header">
            <h1 style="margin:0; font-size:1.2rem;">Gestione Zone <span id="tot-points" style="font-size:0.7rem; opacity:0.7;">0 Punti</span></h1>
            <div style="margin-top:10px; display:flex; justify-content: space-between; align-items:center; gap:8px;">
                <button onclick="saveAllToServer()" style="background:var(--accent); color:white; border:none; padding:8px 15px; border-radius:8px; font-weight:800; cursor:pointer; flex: 1;">SALVA TUTTO</button>
                <button onclick="toggleVerificaViaggi()" id="btn-verifica" class="btn btn-verifica" style="padding:8px 10px; height: 35px;">VERIFICA VIAGGI</button>
                <button onclick="apriAnteprima()" id="btn-anteprima" title="Anteprima timing percorsi" style="background:#6366f1; color:white; border:none; padding:8px 10px; border-radius:8px; font-weight:800; cursor:pointer; font-size:1rem; height:35px;">👁</button>
            </div>
            <div style="display:flex; gap:5px; margin-top:10px;">
                <button onclick="toggleLockMap()" id="btn-lock-map" class="btn btn-lock">CARICAMENTO...</button>
                <button id="btn-toggle-drag" onclick="toggleDragging()" class="btn" style="flex:1;">🔒 SPOSTA PUNTI (OFF)</button>
            </div>
            <div style="text-align:right; margin-top:5px;"><small style="opacity:0.6;">{{DATA_GIORNO}}</small></div>
        </div>
        <div id="zone-list"></div>
    </div>
    <div id="map"></div>
    <div id="save-status">💾 AGGIORNAMENTO COMPLETATO!</div>

    <!-- ═══════════════════════════════════════════════════════ -->
    <!-- DRAWER ANTEPRIMA PERCORSI                              -->
    <!-- ═══════════════════════════════════════════════════════ -->
    <div id="drawer-overlay" onclick="chiudiAnteprima()" style="display:none; position:fixed; inset:0; background:rgba(0,0,0,0.4); z-index:2000;"></div>
    <div id="drawer-anteprima" style="display:none; position:fixed; top:0; right:0; width:500px; max-width:95vw; height:100vh; background:white; z-index:2001; box-shadow:-10px 0 40px rgba(0,0,0,0.2); display:flex; flex-direction:column; font-family:'Inter',sans-serif;">
        <!-- Header drawer -->
        <div style="padding:20px 24px; background:linear-gradient(135deg,#1e293b,#0f172a); color:white; flex-shrink:0;">
            <div style="display:flex; align-items:center; justify-content:space-between;">
                <div>
                    <div style="font-size:0.7rem; font-weight:800; opacity:0.7; text-transform:uppercase; letter-spacing:0.05em;">Anteprima</div>
                    <div style="font-size:1.2rem; font-weight:900; margin-top:2px;">👁 Percorsi Stimati</div>
                    <div id="drawer-subtitle" style="font-size:0.72rem; opacity:0.6; margin-top:3px;">Calcolo con distanze Haversine</div>
                </div>
                <button onclick="chiudiAnteprima()" style="background:rgba(255,255,255,0.15); border:none; color:white; width:36px; height:36px; border-radius:8px; font-size:1.2rem; cursor:pointer; display:flex; align-items:center; justify-content:center;">✕</button>
            </div>
        </div>
        <!-- Corpo drawer (scrollabile) -->
        <div id="drawer-body" style="flex:1; overflow-y:auto; padding:16px; background:#f8fafc;"></div>
        <!-- Footer drawer -->
        <div style="padding:16px; border-top:1px solid #e2e8f0; background:white; flex-shrink:0; display:flex; gap:10px;">
            <button onclick="chiudiAnteprima()" style="flex:1; padding:12px; background:#f1f5f9; border:1px solid #e2e8f0; border-radius:10px; font-weight:700; cursor:pointer; color:#475569;">✕ Chiudi</button>
            <button onclick="confermaPercorsi()" id="btn-conferma" style="flex:2; padding:12px; background:#10b981; border:none; border-radius:10px; font-weight:800; cursor:pointer; color:white; font-size:0.9rem;">💾 Salva e Conferma → Lancia BAT 3</button>
        </div>
    </div>
    <!-- Fine drawer -->

    <script>
        let DATA_ZONE = {{JSON_ZONE | safe}};
        let map, gMarkers = [], DRAGGING_ENABLED = false, isLockedGlobal = {{IS_LOCKED_JS}};
        let activeExpandedZid = null, activeAction = null, activeSourceZid = null;
        let _hasUnsavedChanges = false, VIEW_MODE_PULITA = false;

        function toggleVerificaViaggi() {
            VIEW_MODE_PULITA = !VIEW_MODE_PULITA;
            const btn = document.getElementById('btn-verifica');
            if (VIEW_MODE_PULITA) {
                btn.classList.add('active');
                btn.textContent = 'VISTA PULITA ON';
            } else {
                btn.classList.remove('active');
                btn.textContent = 'VERIFICA VIAGGI';
            }
            renderMarkers();
            renderSidebar();
            updateTotals();
        }

        window.addEventListener('beforeunload', function(e) {
            if (_hasUnsavedChanges) {
                e.preventDefault();
                e.returnValue = 'Hai modifiche non salvate. Premi SALVA TUTTO prima di chiudere.';
                return e.returnValue;
            }
        });

        async function initMap() {
            const { Map } = await google.maps.importLibrary("maps");
            map = new Map(document.getElementById("map"), { center: { lat: 45.5, lng: 11.8 }, zoom: 9, mapTypeId: 'hybrid', mapId: "MY_MAP_ID" });
            _updateLockUI();
            updateTotals();
            await renderMarkers();
            renderSidebar();
            if(gMarkers.length) { 
                let b = new google.maps.LatLngBounds(); 
                gMarkers.forEach(m => b.extend(m.position)); 
                map.fitBounds(b); 
            }
        }

        function _updateLockUI() {
            const btn = document.getElementById('btn-lock-map');
            btn.innerHTML = isLockedGlobal ? 'BLOCCATA 🔒' : 'SBLOCCATA 🔓';
            btn.className = 'btn btn-lock' + (isLockedGlobal ? ' locked' : '');
        }

        async function toggleLockMap() {
            isLockedGlobal = !isLockedGlobal;
            fetch('/toggle_lock', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({action: isLockedGlobal?'lock':'unlock'}) })
            .then(() => { _updateLockUI(); renderMarkers(); renderSidebar(); });
        }

        async function toggleDragging() {
            DRAGGING_ENABLED = !DRAGGING_ENABLED;
            const btn = document.getElementById('btn-toggle-drag');
            btn.textContent = DRAGGING_ENABLED ? "🔓 SPOSTA PUNTI (ON)" : "🔒 SPOSTA PUNTI (OFF)";
            btn.style.background = DRAGGING_ENABLED ? "#10b981" : "#4b5563";
            btn.style.color = "white";
            await renderMarkers();
        }

        async function renderMarkers() {
            gMarkers.forEach(m => m.setMap(null)); gMarkers = [];
            const { AdvancedMarkerElement } = await google.maps.importLibrary("marker");
            const coordCounts = {};
            DATA_ZONE.forEach(z => {
                const isSpeciale = (z.id_zona === 'DDT_DA_INSERIRE');
                if (VIEW_MODE_PULITA && isSpeciale) return;

                z.lista_punti.forEach((p, idx) => {
                    if (!p.lat) return;
                    const key = `${p.lat.toFixed(6)}_${p.lon.toFixed(6)}`;
                    coordCounts[key] = (coordCounts[key] || 0) + 1;
                    const layer = coordCounts[key];
                    let shape = 'm-goccia';
                    if (layer === 2) shape = 'm-quadrato'; else if (layer === 3) shape = 'm-triangolo'; else if (layer > 3) shape = 'm-tonda';
                    
                    const t_g = (p.tipologia_grado || '').toUpperCase();
                    const isGrandChef = (p.codice_frutta && (p.codice_frutta.startsWith('100') || p.codice_frutta.length > 6)) || t_g.includes('GRAND CHEF') || t_g.includes('GRAN CHEF');
                    const isCattell = t_g.includes('CATTELL') || t_g.includes('CATTEL') || (p.codice_frutta && p.codice_frutta.toUpperCase().includes('CATTELL'));
                    const isBauer = t_g.includes('BAUER') || (p.codice_frutta && p.codice_frutta.toUpperCase().includes('BAUER'));
                    
                    const el = document.createElement("div");
                    el.className = `custom-marker`;
                    el.style.transform = "translate(-50%, -100%)";
                    el.style.display = "flex";
                    el.style.alignItems = "center";
                    el.style.justifyContent = "center";
                    
                    if (isGrandChef) {
                        el.className = `custom-marker m-chef`;
                        el.style.backgroundColor = z.color;
                        el.style.borderRadius = "50%";
                        el.style.width = "34px";
                        el.style.height = "34px";
                        el.style.border = "2.5px solid white";
                        el.style.boxShadow = "0 4px 10px rgba(0,0,0,0.35)";
                        el.style.display = "flex";
                        el.style.flexDirection = "column";
                        el.style.alignItems = "center";
                        el.style.justifyContent = "center";
                        el.style.color = "white";
                        el.style.transform = "translate(-50%, -100%)";
                        el.innerHTML = `
                            <span style="font-size: 13px; line-height: 1.1; margin-top: -1px; pointer-events: none; display: block;">👨‍🍳</span>
                            <span style="font-size: 8px; font-weight: 900; pointer-events: none; margin-top: -2px; display: block;">${idx+1}</span>
                        `;
                    } else if (isCattell) {
                        const svgCattell = `
                        <svg width="38" height="38" viewBox="0 0 24 24" fill="white" style="filter: drop-shadow(0 4px 6px rgba(0,0,0,0.3));">
                            <circle cx="12" cy="12" r="11" fill="${z.color}" stroke="white" stroke-width="1.5"/>
                            <g transform="translate(3, 3) scale(0.75)">
                                <path d="M3.4 20.6c-.4-.4-.4-1 0-1.4L15.6 7c1-1 2.5-1.2 3.5-.2s.8 2.5-.2 3.5L6.8 22.6c-.4.4-1 .4-1.4 0L3.4 20.6zm17.2 0c.4-.4.4-1 0-1.4L9.8 8.4l-.7.7 2.1 2.1-1.4 1.4-2.1-2.1-.7.7L19.2 22.6c.4.4 1 .4 1.4 0l0 0zm-11.5-12.9L7 5.6c-.8-.8-.8-2 0-2.8s2-.8 2.8 0l2.1 2.1-2.8 2.8z" fill="white"/>
                            </g>
                        </svg>`;
                        el.innerHTML = `
                            <div style="position: relative; width: 38px; height: 38px; display: flex; align-items: center; justify-content: center;">
                                ${svgCattell}
                                <span style="position: absolute; top: 12px; width: 100%; text-align: center; font-size: 8.5px; font-weight: 900; color: white; text-shadow: 1px 1px 2px black; pointer-events: none;">${idx+1}</span>
                            </div>`;
                    } else if (isBauer) {
                        const svgBauer = `
                        <svg width="38" height="38" viewBox="0 0 24 24" fill="white" style="filter: drop-shadow(0 4px 6px rgba(0,0,0,0.3));">
                            <circle cx="12" cy="12" r="11" fill="${z.color}" stroke="white" stroke-width="1.5"/>
                            <g transform="translate(3, 3) scale(0.75)">
                                <path d="M19 8c-.6 0-1.1.2-1.5.6C16.8 5.7 14.6 4 12 4s-4.8 1.7-5.5 4.6c-.4-.4-.9-.6-1.5-.6-1.1 0-2 .9-2 2 0 .7.4 1.4 1 1.7v4.3c0 2.2 1.8 4 4 4h8c2.2 0 4-1.8 4-4v-4.3c.6-.3 1-.9 1-1.7 0-1.1-.9-2-2-2zm-12 8c-.6 0-1-.4-1-1s.4-1 1-1 1 .4 1 1-.4 1-1 1zm10 0c-.6 0-1-.4-1-1s.4-1 1-1 1 .4 1 1-.4 1-1 1z" fill="white"/>
                            </g>
                        </svg>`;
                        el.innerHTML = `
                            <div style="position: relative; width: 38px; height: 38px; display: flex; align-items: center; justify-content: center;">
                                ${svgBauer}
                                <span style="position: absolute; top: 12px; width: 100%; text-align: center; font-size: 8.5px; font-weight: 900; color: white; text-shadow: 1px 1px 2px black; pointer-events: none;">${idx+1}</span>
                            </div>`;
                    } else {
                        el.className = `custom-marker ${shape}`;
                        if (shape === 'm-goccia') {
                            el.style.transform = "translate(-50%, -100%) rotate(-45deg)";
                        } else {
                            el.style.transform = "translate(-50%, -100%)";
                        }
                        
                        let isParziale = false;
                        if (isSpeciale && p.rientri_alert) {
                            isParziale = p.rientri_alert.some(r => r.is_parziale === true);
                        }
                        
                        if (isParziale) {
                            if (shape === 'm-triangolo') {
                                el.style.borderBottomColor = "#f59e0b";
                                el.style.backgroundImage = "repeating-linear-gradient(45deg, black, black 4px, transparent 4px, transparent 8px)";
                            } else {
                                el.style.backgroundImage = "repeating-linear-gradient(45deg, #000, #000 4px, #f59e0b 4px, #f59e0b 8px)";
                                el.style.color = "white";
                                el.style.textShadow = "1px 1px 2px black, -1px -1px 2px black, 0px 0px 3px black";
                                el.style.border = "2px solid black";
                            }
                        } else {
                            if (shape === 'm-triangolo') el.style.borderBottomColor = z.color; else el.style.backgroundColor = z.color;
                        }
                        
                        el.innerHTML = `<span>${idx+1}</span>`;
                    }
                    
                    const m = new AdvancedMarkerElement({ position: {lat: p.lat, lng: p.lon}, map: map, title: p.nome, content: el, gmpDraggable: DRAGGING_ENABLED && !isLockedGlobal });
                    m.addListener("dragend", () => {
                        p.lat = m.position.lat; p.lon = m.position.lng;
                        _hasUnsavedChanges = true;
                        let c_u = p.codice_univoco || (p.codice_frutta + '_' + p.codice_latte);
                        _salvaSingolo(p.codice_frutta || '', p.codice_latte || '', p.nome || '', p.lat, p.lon, c_u);
                    });
                    m.addListener("gmp-click", () => {
                        new google.maps.InfoWindow({ content: `<div style="padding:10px;"><b>${p.nome}</b><br>${p.indirizzo}<br><br><button onclick="window.open('https://www.google.com/maps/search/?api=1&query=${p.lat},${p.lon}')" style="width:100%; cursor:pointer;">VEDI SU GOOGLE</button></div>` }).open(map, m);
                    });
                    gMarkers.push(m);
                });
            });
        }

        function _salvaSingolo(cod_f, cod_l, nome, lat, lon, codice_univoco) {
            fetch('/save_coord', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({cod_f, cod_l, nome, lat, lon, codice_univoco}) })
            .then(r => r.json())
            .then(d => { 
                let st = document.getElementById('save-status');
                if (d.status === 'ok') {
                    _hasUnsavedChanges = false;
                    st.style.background = '#10b981';
                    st.textContent = '\u2705 Salvato: ' + nome;
                } else {
                    st.style.background = '#ef4444';
                    st.textContent = '\u274c Errore: ' + (d.msg || 'sconosciuto');
                }
                st.style.display = 'block'; 
                setTimeout(()=>{st.style.display='none';}, 3000); 
            })
            .catch(err => {
                let st = document.getElementById('save-status');
                st.style.background = '#ef4444';
                st.textContent = '\u274c Connessione persa!';
                st.style.display = 'block';
                setTimeout(()=>{st.style.display='none';}, 3000);
            });
        }

        function renderSidebar() {
            const list = document.getElementById('zone-list');
            list.innerHTML = DATA_ZONE.filter(z => z.lista_punti.length > 0).map(z => {
                const isSpeciale = (z.id_zona === 'DDT_DA_INSERIRE');
                if (VIEW_MODE_PULITA && isSpeciale) return '';
                
                const isSelected = (activeExpandedZid === z.id_zona) || (activeSourceZid === z.id_zona);
                return `
                <div class="zone-card ${isSelected ? 'selected' : ''} ${isSpeciale ? 'speciale' : ''}" onclick="focusZone('${z.id_zona}')">
                    <div class="zone-header">
                        <div class="color-pill" style="background: ${z.color}"></div>
                        <div class="zone-title">
                            <span style="display:block; font-size:0.95rem; font-weight:800;">${z.nome_giro || z.id_zona}${isSpeciale ? '<span class="badge-speciale">RIENTRI</span>' : ''}${!isSpeciale && !isLockedGlobal ? `<button title="Rinomina viaggio" onclick="event.stopPropagation(); renameZona('${z.id_zona}')" style="margin-left:6px; background:none; border:none; cursor:pointer; font-size:0.75rem; opacity:0.5; padding:0; vertical-align:middle;" onmouseover="this.style.opacity=1" onmouseout="this.style.opacity=0.5">✏️</button>` : ''}</span>
                            <span style="display:block; font-size:0.65rem; font-weight:600; color:#94a3b8;">${isSpeciale ? 'Assegna a un viaggio' : 'Zona ' + z.id_zona}</span>
                        </div>
                        <div style="text-align:right; flex-shrink:0;">
                            <div style="font-size:0.72rem; font-weight:800; background:${isSpeciale?'#fde68a':'#e2e8f0'}; padding:2px 7px; border-radius:20px; color:${isSpeciale?'#92400e':'#475569'}; white-space:nowrap;">${z.lista_punti.length} pt</div>
                            ${(() => { const isDNR = !z.id_zona.startsWith('GranChef') && z.id_zona !== 'DDT_DA_INSERIRE' && z.id_zona !== 'SENZA_ZONA'; if (!isDNR) return ''; const s = calcolaValoreZona(z); return s.tot_ddt > 0 ? `<div style="font-size:0.65rem; font-weight:700; color:#10b981; margin-top:2px; white-space:nowrap;">${s.tot_ddt} DDT · € ${s.valore}</div>` : ''; })()}
                        </div>
                    </div>
                    ${isSelected ? `
                        <div style="margin-top:10px; border-top:1px solid #e2e8f0; padding-top:10px; max-height:320px; overflow-y:auto;">
                            ${z.lista_punti.map((p, idx) => {
                                const isParzList = isSpeciale && p.rientri_alert && p.rientri_alert.some(r => r.is_parziale);
                                const bgStyle = isParzList ? 'background: repeating-linear-gradient(45deg, #000, #000 3px, #f59e0b 3px, #f59e0b 6px); color: white; text-shadow: 1px 1px 2px black; border: 1px solid black;' : `background: ${z.color};`;
                                const pid = (p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_');
                                const parts = (p.indirizzo || '').split(',');
                                const via = parts[0] || '';
                                const comune = parts.slice(1).join(',').trim();
                                const codiceDDT = [...(p.codici_ddt_frutta||[]), ...(p.codici_ddt_latte||[])].join(', ');
                                let ctrl = '';
                                if (activeAction === 'dividi' && !isSpeciale) ctrl = `<input type="checkbox" id="chk-${pid}" class="chk-box dividi-chk" value="${pid}" onclick="event.stopPropagation()"><label for="chk-${pid}" class="chk-lbl" onclick="event.stopPropagation()">✓</label>`;
                                else if (activeAction === 'sposta') ctrl = `<div class="zone-chips" data-pid="${pid}" style="padding-left:26px; margin-top:4px;">${DATA_ZONE.filter(o=>o.id_zona!==z.id_zona && o.id_zona!=='DDT_DA_INSERIRE').map(o=>`<span class="zone-chip" data-zone="${o.id_zona}" style="" onclick="event.stopPropagation(); _selChip(this,'${o.color}')">${o.nome_giro||o.id_zona}</span>`).join('')}</div>`;
                                return `<div class="point-card">
                                    <div style="display:flex; align-items:center; justify-content:space-between; gap:4px;">
                                        <div style="display:flex; align-items:center; flex:1; min-width:0; gap:0;">
                                            <span class="point-num" style="${bgStyle}">${idx+1}</span>
                                            <span class="point-name" style="white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${p.nome}</span>
                                        </div>
                                         ${activeAction === 'dividi' ? ctrl : ''}
                                    </div>
                                    ${codiceDDT ? `<div style="font-size:0.6rem; color:#f59e0b; font-weight:700; padding-left:26px; margin-top:2px;">DDT: ${codiceDDT}</div>` : ''}
                                     <div class="point-addr">
                                         <span style="display:block;">${via}</span>
                                         ${comune ? `<span style="font-weight:700; color:#334155;">${comune}</span>` : ''}
                                     </div>
                                     ${(p.orario_min || p.orario_max) ? `<div style="font-size:0.62rem; color:#4f46e5; font-weight:700; padding-left:26px; margin-top:2px;">🕒 ${p.orario_min && p.orario_max ? p.orario_min + ' - ' + p.orario_max : p.orario_min ? 'Dalle ' + p.orario_min : 'Entro le ' + p.orario_max}</div>` : ''}
                                     ${p.note ? `<div style="font-size:0.62rem; color:#d97706; font-weight:600; padding-left:26px; margin-top:3px; background:#fffbeb; border-radius:4px; padding:2px 6px 2px 26px; border:1px solid #fde68a;">📝 ${p.note}</div>` : ''}
                                     ${activeAction === 'sposta' ? ctrl : ''}
                                 </div>`;
                            }).join('')}
                        </div>
                        <div style="margin-top:10px; display:flex; gap:4px;">
                            ${activeAction ? `
                                <button class="btn btn-cancel" onclick="event.stopPropagation(); cancelAction()">✕ Annulla</button>
                                <button class="btn btn-confirm" style="flex:1;" onclick="event.stopPropagation(); ${activeAction==='dividi'?'executeDividi':'executeSposta'}('${z.id_zona}')">✓ Conferma</button>
                            ` : `
                                ${!isSpeciale ? `<button class="btn" style="flex:1;" ${isLockedGlobal?'disabled':''} onclick="event.stopPropagation(); startAction('dividi', '${z.id_zona}')">DIVIDI</button>` : ''}
                                <button class="btn" style="flex:1; ${isSpeciale?'background:#f59e0b;color:white;border:none;':''}" ${isLockedGlobal?'disabled':''} onclick="event.stopPropagation(); startAction('sposta', '${z.id_zona}')">${isSpeciale ? '📦 ASSEGNA A VIAGGIO' : 'SPOSTA'}</button>
                            `}
                        </div>
                    ` : ''}
                </div>`;
            }).join('');
        }

        function focusZone(zid) { if(activeAction) return; activeExpandedZid = (activeExpandedZid === zid) ? null : zid; renderSidebar(); }
        function startAction(type, zid) { activeAction = type; activeSourceZid = zid; renderSidebar(); }
        function cancelAction() { activeAction = null; activeSourceZid = null; renderSidebar(); }

        // ── Nomi predefiniti per il menù a tendina ──────────────────────────────
        const NOMI_DNR = [
            'BS', 'FUORI BS',
            'LAGO BS 1', 'LAGO BS 2', 'LAGO BS 3',
            'VR', 'FUORI VR',
            'MN', 'VR MN',
            'LAGO VR 1', 'LAGO VR 2', 'LAGO VR 3',
        ];
        const NOMI_GRANCHEF = NOMI_DNR.map(n => 'GranChef ' + n);
        // ────────────────────────────────────────────────────────────────────────

        let _renameTargetZid = null;

        function renameZona(zid) {
            const z = DATA_ZONE.find(x => x.id_zona === zid);
            if (!z) return;
            _renameTargetZid = zid;

            const isGC = zid.startsWith('GranChef');
            const nomi = isGC ? NOMI_GRANCHEF : NOMI_DNR;

            // Popola il select
            const sel = document.getElementById('rename-select');
            sel.innerHTML = '<option value="">— Seleziona nome —</option>';
            nomi.forEach(n => {
                const opt = document.createElement('option');
                opt.value = n;
                opt.textContent = n;
                if (n === z.nome_giro) opt.selected = true;
                sel.appendChild(opt);
            });

            // Campo testo: precompila se il nome attuale NON è in lista
            const custom = document.getElementById('rename-custom');
            custom.value = nomi.includes(z.nome_giro) ? '' : (z.nome_giro || '');

            // Titolo modale
            document.getElementById('rename-modal-title').textContent =
                (isGC ? '🍽️ Rinomina giro GranChef' : '🚚 Rinomina giro DNR');
            document.getElementById('rename-modal-subtitle').textContent =
                `Giro attuale: "${z.nome_giro || z.id_zona}"`;

            // Sincronizza select ↔ custom: quando scegli dal menu, svuota il testo libero
            sel.onchange = () => { if (sel.value) custom.value = ''; };
            custom.oninput = () => { if (custom.value.trim()) sel.value = ''; };

            document.getElementById('rename-modal-overlay').classList.add('open');
            setTimeout(() => custom.value ? custom.focus() : sel.focus(), 100);
        }

        function chiudiRenameModal() {
            document.getElementById('rename-modal-overlay').classList.remove('open');
            _renameTargetZid = null;
        }

        function confermaRenameModal() {
            const sel    = document.getElementById('rename-select').value.trim();
            const custom = document.getElementById('rename-custom').value.trim();
            const nuovoNome = custom || sel;
            if (!nuovoNome) { alert('Seleziona un nome o scrivi un nome personalizzato.'); return; }
            const z = DATA_ZONE.find(x => x.id_zona === _renameTargetZid);
            if (z) { z.nome_giro = nuovoNome; renderSidebar(); }
            chiudiRenameModal();
        }

        // Chiudi modal premendo Escape
        document.addEventListener('keydown', e => {
            if (e.key === 'Escape') chiudiRenameModal();
        });

        function _selChip(el, color) {
            el.closest('.zone-chips').querySelectorAll('.zone-chip').forEach(c => { c.classList.remove('sel'); c.style.background='white'; c.style.color='#475569'; });
            el.classList.add('sel'); el.style.background = color; el.style.color = 'white';
        }

        function executeDividi(sourceZid) {
            const z = DATA_ZONE.find(x => x.id_zona === sourceZid);
            const chks = document.querySelectorAll('.dividi-chk:checked');
            if(!chks.length) return alert("Seleziona punti!");
            const ids = Array.from(chks).map(c => c.value);
            const toMove = z.lista_punti.filter(p => ids.includes((p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_')));

            // Rimuovi suffisso _X esistente per trovare il vero id base
            // così splitttare "V03_B" genera "V03_C", non "V03_B_B"
            const baseId = sourceZid.replace(/_[A-Z]$/, '');

            // Trova la prossima lettera libera (B, C, D, ...)
            const LETTERS = 'BCDEFGHIJKLMNOPQRSTUVWXYZ';
            let nextLetter = 'B';
            for (const letter of LETTERS) {
                if (!DATA_ZONE.some(x => x.id_zona === `${baseId}_${letter}`)) {
                    nextLetter = letter;
                    break;
                }
            }

            const subId = `${baseId}_${nextLetter}`;
            // Rimuovi suffisso /X dal nome visualizzato (es. "V01/B" → base "V01")
            const baseNome = (z.nome_giro || z.id_zona || 'Viaggio').replace(/\\/[A-Z]$/, '');
            const newZ = { id_zona: subId, nome_giro: `${baseNome}/${nextLetter}`, color: "#"+Math.floor(Math.random()*16777215).toString(16).padStart(6,'0'), lista_punti: toMove };
            z.lista_punti = z.lista_punti.filter(p => !toMove.includes(p));
            DATA_ZONE.push(newZ);
            cancelAction(); updateTotals(); renderMarkers(); renderSidebar();
        }


        function executeSposta(sourceZid) {
            const z = DATA_ZONE.find(x => x.id_zona === sourceZid);
            document.querySelectorAll('.zone-chips').forEach(chips => {
                const sel = chips.querySelector('.zone-chip.sel');
                if(sel) {
                    const target = DATA_ZONE.find(x => x.id_zona === sel.dataset.zone);
                    const pid = chips.dataset.pid;
                    const pIdx = z.lista_punti.findIndex(p => (p.codice_frutta + "_" + p.nome).replace(/[^a-zA-Z0-9]/g, '_') === pid);
                    if(pIdx >= 0 && target) { target.lista_punti.push(z.lista_punti.splice(pIdx, 1)[0]); }
                }
            });
            cancelAction(); updateTotals(); renderMarkers(); renderSidebar();
        }

        function updateTotals() {
            let totalPunti = 0;
            let totalDdt = 0;
            DATA_ZONE.forEach(z => {
                const isSpeciale = (z.id_zona === 'DDT_DA_INSERIRE');
                if (VIEW_MODE_PULITA && isSpeciale) return;
                totalPunti += z.lista_punti.length;
                totalDdt += calcolaValoreZona(z).tot_ddt;
            });
            document.getElementById('tot-points').textContent = `${totalPunti} Punti - ${totalDdt} DDT`;
        }

        const VALORE_DDT = 16.50;
        function calcolaValoreZona(z) {
            let totDdt = 0;
            z.lista_punti.forEach(p => {
                totDdt += (p.codici_ddt_frutta || []).filter(c => c && c !== 'p00000').length;
                totDdt += (p.codici_ddt_latte  || []).filter(c => c && c !== 'p00000').length;
                // fallback: se le liste non ci sono, conta 1 DDT solo per codici DNR (iniziano con 'p')
                if (!(p.codici_ddt_frutta) && !(p.codici_ddt_latte)) {
                    if (p.codice_frutta && p.codice_frutta.startsWith('p') && p.codice_frutta !== 'p00000') totDdt++;
                    if (p.codice_latte  && p.codice_latte.startsWith('p')  && p.codice_latte  !== 'p00000') totDdt++;
                }
            });
            return { tot_ddt: totDdt, valore: (totDdt * VALORE_DDT).toFixed(2) };
        }

        function saveAllToServer() {
            const btn = event.target; btn.disabled = true; btn.textContent = "...";
            fetch('/save', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(DATA_ZONE) })
            .then(() => { 
                _hasUnsavedChanges = false;
                const st = document.getElementById('save-status');
                st.style.background = '#10b981';
                st.textContent = '✅ SALVATAGGIO COMPLETATO!';
                st.style.display = 'block'; 
                setTimeout(()=>{ st.style.display='none'; }, 3000); 
            }).finally(() => { btn.disabled = false; btn.textContent = "SALVA TUTTO"; });
        }

        window.onload = initMap;

        // ═══════════════════════════════════════════════════
        // ANTEPRIMA PERCORSI
        // ═══════════════════════════════════════════════════

        let _anteprimaGiri = [];  // Dati giri caricati dall'API
        let _sortableInstances = [];  // Istanze SortableJS per cleanup
        let _hasManualChanges = false;  // true se l'utente ha fatto drag&drop nel drawer

        async function apriAnteprima() {
            const drawer  = document.getElementById('drawer-anteprima');
            const overlay = document.getElementById('drawer-overlay');
            const body    = document.getElementById('drawer-body');
            const btn     = document.getElementById('btn-anteprima');
            const btnConf = document.getElementById('btn-conferma');

            // Reset stato manuale
            _hasManualChanges = false;
            btnConf.textContent = '✅ Conferma (percorsi già ottimizzati)';
            btnConf.style.background = '#10b981';
            btnConf.disabled = false;

            // Mostra drawer con spinner
            drawer.style.display = 'flex';
            overlay.style.display = 'block';
            btn.textContent = '⏳';
            document.getElementById('drawer-subtitle').textContent = 'BAT 3 in esecuzione…';
            body.innerHTML = `<div style="text-align:center; padding:60px 20px; color:#64748b;">
                <div style="font-size:2.5rem; margin-bottom:16px; animation:spin 1.5s linear infinite; display:inline-block;">&#9696;</div>
                <div style="font-weight:700; font-size:1rem;">OR-Tools + Google Directions in corso…</div>
                <div style="font-size:0.8rem; margin-top:8px; opacity:0.7;">Potrebbe richiedere 30-60 secondi</div>
            </div>
            <style>@keyframes spin{to{transform:rotate(360deg)}}</style>`;

            try {
                const resp = await fetch('/preview_percorsi');
                const data = await resp.json();
                btn.textContent = '👁';

                if (data.status !== 'ok') {
                    body.innerHTML = `<div style="padding:30px; color:#ef4444; font-weight:700;">⚠️ ${data.msg}</div>`;
                    return;
                }

                _anteprimaGiri = data.giri;
                _renderDrawerBody(data.giri);

                const nTot = data.giri.reduce((s, g) => s + g.n_late, 0);
                document.getElementById('drawer-subtitle').textContent =
                    nTot === 0
                    ? `✅ ${data.giri.length} giri ottimizzati — nessun ritardo`
                    : `⚠️ ${data.giri.length} giri — ${nTot} ritardi rilevati`;

                // Apri riepilogo BAT 3 in nuova tab
                if (data.riepilogo_disponibile) {
                    window.open('/serve_riepilogo', '_blank');
                }

            } catch (err) {
                btn.textContent = '👁';
                body.innerHTML = `<div style="padding:30px; color:#ef4444;">Errore connessione: ${err.message}</div>`;
            }
        }

        function _renderDrawerBody(giri) {
            const body = document.getElementById('drawer-body');
            _sortableInstances.forEach(s => s.destroy());
            _sortableInstances = [];

            body.innerHTML = giri.map((g, gi) => {
                const hasLate    = g.n_late > 0;
                const borderClr  = hasLate ? '#fecaca' : '#bbf7d0';
                const headerBg   = hasLate ? '#fef2f2' : '#f0fdf4';
                const statusClr  = hasLate ? '#ef4444' : '#10b981';
                const statusTxt  = hasLate ? `⚠️ ${g.n_late} ritard${g.n_late===1?'o':'i'}` : '✅ In orario';
                const chevronId  = `chv-${gi}`;
                const bodyId     = `sortable-${gi}`;
                const isOpen     = hasLate;
                const gcBadge    = g.is_gc ? `<span style="background:#f59e0b;color:white;font-size:0.5rem;font-weight:800;padding:2px 5px;border-radius:6px;margin-left:6px;vertical-align:middle;">GC</span>` : '';

                const ferrateHtml = g.fermate.map((f, fi) => {
                    const lateStyle  = f.is_late ? 'background:#fef2f2; border-color:#fecaca;' : '';
                    const lateBadge  = f.is_late ? `<span style="background:#ef4444;color:white;font-size:0.5rem;font-weight:800;padding:1px 4px;border-radius:4px;margin-left:4px;">RITARDO</span>` : '';
                    const orario     = f.orario_min && f.orario_max ? `🕒 ${f.orario_min}–${f.orario_max}` :
                                       f.orario_max ? `🕒 Entro ${f.orario_max}` :
                                       f.orario_min ? `🕒 Dalle ${f.orario_min}` : '';
                    const arrivo     = f.ora_arrivo ? `<div style="font-weight:800; color:#10b981;">▶ ${f.ora_arrivo}</div>` : '';
                    const ripartenza = f.ora_ripartenza ? `<div style="color:#94a3b8;">◀ ${f.ora_ripartenza}</div>` : '';
                    return `
                    <div class="prv-fermata" data-uid="${f.codice_univoco}" style="background:white; border:1px solid #e2e8f0; border-radius:10px; padding:9px 10px; margin-bottom:5px; display:flex; align-items:center; gap:8px; ${lateStyle}">
                        <div style="width:22px; height:22px; background:#6366f1; color:white; border-radius:50%; display:flex; align-items:center; justify-content:center; font-size:0.65rem; font-weight:800; flex-shrink:0;">${fi+1}</div>
                        <div style="flex:1; min-width:0;">
                            <div style="font-size:0.78rem; font-weight:700; color:#1e293b; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${f.nome}${lateBadge}</div>
                            <div style="font-size:0.6rem; color:#94a3b8; margin-top:1px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">${f.indirizzo}</div>
                            ${orario ? `<div style="font-size:0.6rem; color:#4f46e5; font-weight:700; margin-top:2px;">${orario}</div>` : ''}
                        </div>
                        <div style="text-align:right; flex-shrink:0; font-size:0.65rem;">${arrivo}${ripartenza}</div>
                        <div style="color:#cbd5e1; font-size:1rem; cursor:grab; flex-shrink:0;">☰</div>
                    </div>`;
                }).join('');

                return `
                <div style="border-radius:12px; border:1.5px solid ${borderClr}; margin-bottom:10px; overflow:hidden; box-shadow:0 2px 6px rgba(0,0,0,0.04);">
                    <div onclick="_toggleCard(${gi})" style="padding:11px 14px; display:flex; align-items:center; justify-content:space-between; background:${headerBg}; cursor:pointer; user-select:none;">
                        <div style="flex:1; min-width:0;">
                            <div style="font-size:0.78rem; font-weight:800; color:#1e293b;">${g.nome_giro}${gcBadge}</div>
                            <div style="font-size:0.6rem; color:#64748b; margin-top:2px;">${g.n_fermate} tappe</div>
                        </div>
                        <div style="display:flex; align-items:center; gap:8px; flex-shrink:0;">
                            <div style="font-size:0.68rem; font-weight:800; color:${statusClr}; background:white; padding:3px 9px; border-radius:20px; border:1px solid ${borderClr};">${statusTxt}</div>
                            <span id="${chevronId}" style="font-size:0.8rem; color:#94a3b8; transition:transform 0.2s; display:inline-block; transform:${isOpen?'rotate(90deg)':'rotate(0deg)'}">▶</span>
                        </div>
                    </div>
                    <div id="${bodyId}" style="padding:${isOpen?'8px':'0'}; max-height:${isOpen?'2000px':'0'}; overflow:hidden; transition:max-height 0.3s ease, padding 0.2s;">${ferrateHtml}</div>
                </div>`;
            }).join('');

            _initSortable(giri);
        }

        function _toggleCard(gi) {
            const bodyEl = document.getElementById(`sortable-${gi}`);
            const chvEl  = document.getElementById(`chv-${gi}`);
            if (!bodyEl) return;
            const isOpen = bodyEl.style.maxHeight !== '0px' && bodyEl.style.maxHeight !== '';
            if (isOpen) {
                bodyEl.style.maxHeight = '0';
                bodyEl.style.padding   = '0';
                if (chvEl) chvEl.style.transform = 'rotate(0deg)';
            } else {
                bodyEl.style.maxHeight = '2000px';
                bodyEl.style.padding   = '8px';
                if (chvEl) chvEl.style.transform = 'rotate(90deg)';
            }
        }

        function _initSortable(giri) {
            function _activateSortables() {
                giri.forEach((g, gi) => {
                    const el = document.getElementById(`sortable-${gi}`);
                    if (!el) return;
                    const s = Sortable.create(el, {
                        animation: 150,
                        handle: '[style*="cursor:grab"]',
                        onEnd: () => {
                            _aggiornaNumeriFermate(gi);
                            // Segna che l'utente ha fatto modifiche manuali
                            _hasManualChanges = true;
                            const btnConf = document.getElementById('btn-conferma');
                            if (btnConf) {
                                btnConf.textContent = '💾 Salva ordine manuale → Rigenera mappe';
                                btnConf.style.background = '#f59e0b';
                            }
                        }
                    });
                    _sortableInstances.push(s);
                });
            }

            if (typeof Sortable !== 'undefined') {
                _activateSortables();
            } else {
                const sc = document.createElement('script');
                sc.src = 'https://cdn.jsdelivr.net/npm/sortablejs@latest/Sortable.min.js';
                sc.onload = _activateSortables;
                document.head.appendChild(sc);
            }
        }

        function _aggiornaNumeriFermate(gi) {
            const el = document.getElementById(`sortable-${gi}`);
            if (!el) return;
            el.querySelectorAll('.prv-fermata').forEach((card, idx) => {
                card.querySelectorAll('div')[0].textContent = idx + 1;
            });
        }

        function chiudiAnteprima() {
            document.getElementById('drawer-anteprima').style.display = 'none';
            document.getElementById('drawer-overlay').style.display = 'none';
            _anteprimaGiri = [];
            _sortableInstances.forEach(s => s.destroy());
            _sortableInstances = [];
        }

        async function confermaPercorsi() {
            const btn = document.getElementById('btn-conferma');
            btn.disabled = true;

            if (!_hasManualChanges) {
                // Caso A: BAT 3 è già girato con l'anteprima — solo conferma
                btn.textContent = '✅ Confermato!';
                setTimeout(() => chiudiAnteprima(), 1500);
                return;
            }

            // Caso B: l'utente ha fatto drag&drop → salva ordine manuale + rigenera
            btn.textContent = '⏳ Salvataggio ordine manuale…';

            const ordini = {};
            _anteprimaGiri.forEach((g, gi) => {
                const el = document.getElementById(`sortable-${gi}`);
                if (!el) return;
                const uids = [...el.querySelectorAll('.prv-fermata')].map(c => c.dataset.uid);
                ordini[g.id_zona] = uids;
            });

            try {
                const resp = await fetch('/conferma_percorsi', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ordini, manuale: true})
                });
                const data = await resp.json();

                if (data.status === 'ok') {
                    const body = document.getElementById('drawer-body');
                    body.innerHTML = `
                        <div style="text-align:center; padding:60px 20px;">
                            <div style="font-size:3rem; margin-bottom:16px;">✅</div>
                            <div style="font-size:1.1rem; font-weight:800; color:#10b981;">Ordine manuale salvato!</div>
                            <div style="font-size:0.8rem; color:#64748b; margin-top:10px;">Mappe rigenerate con il tuo ordine.<br>Puoi ora avviare BAT 5 per le mappe autisti.</div>
                        </div>`;
                    document.getElementById('btn-conferma').style.display = 'none';
                    // Apri riepilogo aggiornato in nuova tab
                    window.open('/serve_riepilogo', '_blank');
                } else {
                    alert('Errore: ' + data.msg);
                    btn.disabled = false;
                    btn.textContent = '💾 Salva ordine manuale → Rigenera mappe';
                }
            } catch (err) {
                alert('Errore di connessione: ' + err.message);
                btn.disabled = false;
            }
        }

        window.onload = initMap;
    </script>
</body>
</html>"""

def _carica_e_genera(data_giorno):
    global TARGET_FILE_UNIFICATO, TARGET_FILE_VIAGGI, ZONE_LIST_CACHE, DATA_GIORNO
    DATA_GIORNO = data_giorno
    output_base = CONSEGNE_DIR / f"CONSEGNE_{DATA_GIORNO}"
    TARGET_FILE_UNIFICATO = output_base / "punti_consegna_unificati.json"
    TARGET_FILE_VIAGGI = output_base / "viaggi_giornalieri.json"
    
    if not TARGET_FILE_UNIFICATO.exists(): return False
    try:
        unif_data = json.loads(TARGET_FILE_UNIFICATO.read_text(encoding="utf-8"))
        punti = unif_data.get("punti", [])
        # Raggruppa i punti per zona
        temp_dict = {}
        for p in punti:
            zid = str(p.get("zona") or "SENZA_ZONA")
            if zid not in temp_dict:
                temp_dict[zid] = []
            temp_dict[zid].append(p)

        # Ordina le chiavi numericamente/alfabeticamente (escludendo le speciali e GranChef)
        chiavi_ordinate = sorted([k for k in temp_dict.keys() if k not in ("DDT_DA_INSERIRE", "SENZA_ZONA") and not k.startswith("GranChef")])

        zone_dict = {}
        # Assegna i nomi V01, V02... in base all'ordine di zona
        for i, zid in enumerate(chiavi_ordinate, start=1):
            zone_dict[zid] = {
                "id_zona": zid,
                "lista_punti": temp_dict[zid],
                "color": _get_color(i - 1),
                "nome_giro": f"V{i:02d}"
            }

        # Aggiungi zone GranChef se presenti
        zone_gc = sorted([k for k in temp_dict.keys() if k.startswith("GranChef")])
        for idx_gc, zid in enumerate(zone_gc, start=1):
            colore_gc = _get_color(idx_gc - 1)
            zone_dict[zid] = {
                "id_zona": zid,
                "lista_punti": temp_dict[zid],
                "color": colore_gc,
                "nome_giro": zid.replace("_", " ")  # es: "GranChef V01"
            }

        # Aggiungi SENZA_ZONA alla fine se presente
        if "SENZA_ZONA" in temp_dict:
            zone_dict["SENZA_ZONA"] = {
                "id_zona": "SENZA_ZONA",
                "lista_punti": temp_dict["SENZA_ZONA"],
                "color": "#9ca3af",
                "nome_giro": "Senza Zona"
            }

        # Aggiungi DDT_DA_INSERIRE alla fine
        if "DDT_DA_INSERIRE" in temp_dict:
            zone_dict["DDT_DA_INSERIRE"] = {
                "id_zona": "DDT_DA_INSERIRE",
                "lista_punti": temp_dict["DDT_DA_INSERIRE"],
                "color": "#f59e0b",
                "nome_giro": "⚠️ DDT DA INSERIRE"
            }

        # ── Ripristino nomi personalizzati ──────────────────────────────────────
        # Se viaggi_giornalieri.json esiste già (salvato da una sessione precedente),
        # recupera i nome_giro custom e riscrivili su zone_dict prima di servire l'HTML.
        # In questo modo le rinominazioni sopravvivono a ogni riavvio del BAT 2.
        if TARGET_FILE_VIAGGI.exists():
            try:
                viaggi_salvati = json.loads(TARGET_FILE_VIAGGI.read_text(encoding="utf-8"))
                for v in viaggi_salvati:
                    zid_v = v.get("id_zona")
                    nome_v = v.get("nome_giro")
                    if zid_v and nome_v and zid_v in zone_dict:
                        zone_dict[zid_v]["nome_giro"] = nome_v
            except Exception:
                pass  # file corrotto o vuoto: si usano i nomi generati
        # ────────────────────────────────────────────────────────────────────────

        # Ricostruisci la lista cache in ordine
        zone_normali = [zone_dict[k] for k in chiavi_ordinate] + [zone_dict[k] for k in zone_gc] + ([zone_dict["SENZA_ZONA"]] if "SENZA_ZONA" in zone_dict else [])
        zone_speciali = [zone_dict["DDT_DA_INSERIRE"]] if "DDT_DA_INSERIRE" in zone_dict else []
        ZONE_LIST_CACHE = zone_normali + zone_speciali
        
        if not TARGET_FILE_VIAGGI.exists():
            TARGET_FILE_VIAGGI.write_text(json.dumps(ZONE_LIST_CACHE, indent=2, ensure_ascii=False), encoding="utf-8")
        
        _salva_html_fisico(output_base, HTML_TEMPLATE, DATA_GIORNO, json.dumps(ZONE_LIST_CACHE, ensure_ascii=False), GOOGLE_MAPS_API_KEY, unif_data.get("is_locked", True))
        _salva_kml(punti, output_base / f"zone_google_{DATA_GIORNO.replace('-', '_')}.kml", DATA_GIORNO)
        return True
    except: return False

def main():
    args = sys.argv[1:]; serve = True; data = ""
    for a in args:
        if a == "--no-serve": serve = False
        elif not a.startswith("--"): data = a
    if not data:
        f = [d for d in CONSEGNE_DIR.iterdir() if d.is_dir() and d.name.startswith("CONSEGNE_")]
        if not f: return
        f.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        data = f[0].name.replace("CONSEGNE_", "")
    
    if _carica_e_genera(data) and serve:
        _libera_porta_5000()
        print(f"\n🌐 Mappa Operativa: http://127.0.0.1:5000")
        def open_b():
            time.sleep(1.2)
            webbrowser.open("http://127.0.0.1:5000")
        threading.Thread(target=open_b, daemon=True).start()
        app.run(port=5000, debug=False)

if __name__ == "__main__": main()
